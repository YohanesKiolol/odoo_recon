"""
diagnose.py — inspect all 3 input sources and print what the script actually sees.
Run: ./run_diag.sh
"""
import sys, os, io, zipfile
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

def sep(title):
    print(f"\n{'='*60}")
    print(f"  {title}")
    print('='*60)

# ── 1. ODO FILE ───────────────────────────────────────────────────────────────
sep("ODO FILE DIAGNOSIS")
odo_path = Path(os.getenv("ODO_EXCEL_PATH", ""))
if not odo_path.exists():
    print(f"❌ File not found: {odo_path}")
else:
    import openpyxl
    wb = openpyxl.load_workbook(odo_path, data_only=True)
    ws = wb.active
    assert ws is not None

    print(f"✅ File: {odo_path.name}")
    print(f"\n--- Cell A3 (date reference) ---")
    print(f"  A3 value: {ws['A3'].value!r}")

    print(f"\n--- Row 1 headers ---")
    headers = [cell.value for cell in ws[1]]
    for i, h in enumerate(headers):
        if h: print(f"  col {i}: {h!r}")

    print(f"\n--- Column A, rows 2-60 (group detection) ---")
    amount_col = os.getenv("ODO_AMOUNT_COLUMN", "Amount Signed")
    # find amount col index
    hdrs = [str(h).strip() if h else "" for h in headers]
    try:
        amt_idx = hdrs.index(amount_col.strip())
    except ValueError:
        amt_idx = None
        print(f"  ⚠️  Amount column '{amount_col}' NOT FOUND in headers")

    group_bca     = os.getenv("ODO_GROUP_BCA", "")
    group_mandiri = os.getenv("ODO_GROUP_MANDIRI", "")
    group_bri     = os.getenv("ODO_GROUP_BRI", "")

    for row in ws.iter_rows(min_row=2, max_row=80, values_only=True):
        col_a = row[0]
        amt   = row[amt_idx] if amt_idx is not None and amt_idx < len(row) else None
        if col_a is None:
            continue
        tag = ""
        val = str(col_a).strip()
        if group_bca     and group_bca.lower()     in val.lower(): tag = "← BCA GROUP"
        if group_mandiri and group_mandiri.lower() in val.lower(): tag = "← MANDIRI GROUP"
        if group_bri     and group_bri.lower()     in val.lower(): tag = "← BRI GROUP"
        print(f"  A={val!r:45}  amt={str(amt)!r:15}  {tag}")

# ── 2. BCA FILE ───────────────────────────────────────────────────────────────
sep("BCA FILE DIAGNOSIS")
bca_dir     = Path(os.getenv("BCA_EXCEL_DIR", "input"))
bca_pattern = os.getenv("BCA_EXCEL_PATTERN", "ReportMerchantBCA_")
bca_pwd     = os.getenv("BCA_EXCEL_PASSWORD", "")
bca_date_col = os.getenv("BCA_DATE_COLUMN", "Transaction Date")

candidates = [p for p in bca_dir.iterdir()
              if p.suffix.lower() in (".xlsx", ".xls")
              and bca_pattern.lower() in p.name.lower()]

if not candidates:
    print(f"❌ No BCA file matching '{bca_pattern}' in {bca_dir}")
else:
    bca_path = candidates[0]
    print(f"✅ File: {bca_path.name}")
    try:
        import msoffcrypto
        decrypted = io.BytesIO()
        with open(bca_path, "rb") as f:
            of = msoffcrypto.OfficeFile(f)
            of.load_key(password=bca_pwd)
            of.decrypt(decrypted)
        decrypted.seek(0)
        wb2 = openpyxl.load_workbook(decrypted, data_only=True)
        ws2 = wb2.active
        assert ws2 is not None

        print(f"\n--- Row 5 headers ---")
        row5 = [cell.value for cell in ws2[5]]
        for i, h in enumerate(row5):
            if h: print(f"  col {i}: {h!r}")

        print(f"\n--- Unique dates in '{bca_date_col}' column (first 20 rows of data) ---")
        hdrs2 = [str(h).strip() if h else "" for h in row5]
        try:
            date_idx2 = hdrs2.index(bca_date_col.strip())
            dates_seen = set()
            for row in ws2.iter_rows(min_row=6, max_row=30, values_only=True):
                d = row[date_idx2] if date_idx2 < len(row) else None
                if d is not None:
                    dates_seen.add(str(d))
            print(f"  Dates found: {sorted(dates_seen)}")
            print(f"  ODO date (filter): {os.getenv('ODO_EXCEL_PATH')} → 2026-07-20")
        except ValueError:
            print(f"  ⚠️  Date column '{bca_date_col}' not found in row 5 headers")
    except Exception as e:
        print(f"❌ Error: {e}")

# ── 3. BRI PDF ────────────────────────────────────────────────────────────────
sep("BRI PDF DIAGNOSIS")
bri_dir     = Path(os.getenv("BRI_ZIP_DIR", "input"))
bri_pattern = os.getenv("BRI_ZIP_PATTERN", "_franswega7")
bri_pdf_pat = os.getenv("BRI_PDF_PATTERN", "detail_")
bri_amt_col = os.getenv("BRI_AMOUNT_COLUMN", "AMT_TRX")

zips = [p for p in bri_dir.iterdir()
        if p.suffix.lower() == ".zip" and bri_pattern.lower() in p.name.lower()]

if not zips:
    print(f"❌ No BRI ZIP matching '{bri_pattern}' in {bri_dir}")
else:
    zip_path = zips[0]
    print(f"✅ ZIP: {zip_path.name}")
    with zipfile.ZipFile(zip_path, "r") as zf:
        all_names = zf.namelist()
        detail = [n for n in all_names
                  if Path(n).suffix.lower() == ".pdf"
                  and Path(n).name.lower().startswith(bri_pdf_pat.lower())]
        print(f"   All files in ZIP: {all_names}")
        if not detail:
            print(f"❌ No PDF matching '{bri_pdf_pat}' found")
        else:
            pdf_bytes = zf.read(detail[0])
            print(f"   PDF: {detail[0]}  ({len(pdf_bytes):,} bytes)")
            try:
                import pdfplumber
                with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                    print(f"   Pages: {len(pdf.pages)}")
                    for pn, page in enumerate(pdf.pages, 1):
                        tables = page.extract_tables()
                        print(f"\n   Page {pn}: {len(tables)} table(s)")
                        for ti, table in enumerate(tables):
                            if not table: continue
                            print(f"     Table {ti+1} headers: {table[0]}")
                            print(f"     Table {ti+1} row count: {len(table)-1}")
                            if len(table) > 1:
                                print(f"     First data row: {table[1]}")
                        # Also try words if no tables
                        if not tables:
                            words = page.extract_words()
                            print(f"     Words on page: {[w['text'] for w in words[:30]]}")
            except Exception as e:
                print(f"❌ PDF parse error: {e}")

print("\n" + "="*60)
print("  DIAGNOSIS COMPLETE")
print("="*60 + "\n")
