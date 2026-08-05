from openpyxl import load_workbook
import sys

file_path = "template/Journal Entry (account.move).xlsx"
try:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(f"Row {i+1}: {row}")
        if i >= 10:
            break
except Exception as e:
    print(f"Error reading file: {e}")
