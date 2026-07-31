"""
excel_reader.py — reads transactions from an Excel file.
Returns list of dicts with normalized 'amount' (Decimal) plus raw display fields.
"""

from pathlib import Path
from decimal import Decimal
from typing import Optional
import openpyxl

from amount_utils import parse_amount, normalize_for_compare


def read_transactions(
    excel_path: Path,
    amount_col: str,
    date_col: str = "",
    desc_col: str = "",
    source_label: str = "Source",
) -> list[dict]:
    """
    Read an Excel file and return list of transaction dicts.

    Each dict contains:
        amount_raw   : original string/number from cell
        amount       : Decimal (normalized for comparison)
        date         : string or empty
        description  : string or empty
        row          : row number (1-indexed from data, 2-indexed in sheet)
        source       : source_label

    Raises:
        FileNotFoundError  if path doesn't exist
        ValueError         if amount_col not found in headers
    """
    if not excel_path.exists():
        raise FileNotFoundError(
            f"Excel not found: {excel_path}\n"
            f"Check BANK_EXCEL_PATH or ODO_EXCEL_PATH in your .env file."
        )

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    assert ws is not None, f"No active worksheet in {excel_path.name}"

    # Read header row
    headers = [cell.value for cell in ws[1]]
    headers_stripped = [str(h).strip() if h is not None else "" for h in headers]

    def _col_index(col_name: str) -> Optional[int]:
        if not col_name:
            return None
        try:
            return headers_stripped.index(col_name.strip())
        except ValueError:
            return None

    amount_idx = _col_index(amount_col)
    if amount_idx is None:
        raise ValueError(
            f"Column '{amount_col}' not found in {excel_path.name}.\n"
            f"Available columns: {[h for h in headers_stripped if h]}\n"
            f"Update the matching column key in your .env file."
        )

    date_idx = _col_index(date_col)
    desc_idx = _col_index(desc_col)

    transactions = []
    skipped = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_amount = row[amount_idx] if amount_idx < len(row) else None

        # Skip empty rows
        if raw_amount is None or str(raw_amount).strip() == "":
            skipped += 1
            continue

        try:
            parsed = parse_amount(raw_amount)
            normalized = normalize_for_compare(parsed)
        except ValueError as e:
            print(f"  [WARN] {excel_path.name} row {row_num}: {e} — skipped")
            skipped += 1
            continue

        date_val = ""
        if date_idx is not None and date_idx < len(row):
            date_val = str(row[date_idx]) if row[date_idx] is not None else ""

        desc_val = ""
        if desc_idx is not None and desc_idx < len(row):
            desc_val = str(row[desc_idx]) if row[desc_idx] is not None else ""

        transactions.append({
            "amount_raw":   raw_amount,
            "amount":       normalized,
            "date":         date_val,
            "description":  desc_val,
            "row":          row_num,
            "source":       source_label,
        })

    if skipped:
        print(f"  [INFO] {excel_path.name}: {skipped} empty/unparseable rows skipped")

    return transactions
