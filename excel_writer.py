"""
excel_writer.py — writes reconciliation results to a formatted Excel report.

Output: output/reconciliation_YYYYMMDD_HHMMSS.xlsx
  - One sheet per bank: "BCA", "Mandiri", "BRI"
  - Sheet "Selisih (Semua)"  — all discrepancies combined
  - Sheet "Legenda"          — color legend
"""

from datetime import datetime, date
from decimal import Decimal
from pathlib import Path
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reconciler import STATUS_DONE, STATUS_BANK_ONLY, STATUS_ODO_ONLY
from config import BANK_ACCOUNTS, ODOO_COMPANY_NAME

# ── Manual-match sidecar helpers ──────────────────────────────────────────────
MANUAL_MATCHES_FILE = ".manual_matches.json"  # dot-prefix = hidden on Mac/Linux


def _hide_file(path: Path) -> None:
    """Dot-prefix covers Mac/Linux. On Windows, do NOT set +H attribute because it causes PermissionError when Python tries to write/overwrite it later."""
    pass


def _write_manual_matches_json(output_dir: Path, data: list[dict]) -> None:
    """Safely write sidecar JSON with directory creation, Windows attribute un-hiding, and exception protection."""
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        path = output_dir / MANUAL_MATCHES_FILE
        import sys
        if sys.platform == "win32" and path.exists():
            try:
                import subprocess
                flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
                subprocess.run(["attrib", "-H", str(path)], check=False, capture_output=True, creationflags=flags)
            except Exception:
                pass
        path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception as e:
        print(f"  [WARN] Sidecar write skipped: {e}")


def save_manual_matches(output_dir: Path, matches: list[dict]) -> None:
    """Persist manual-match pairs to a JSON sidecar in output_dir.
    Each entry: {pair_tag, bank_date, bank_journal, bank_amount,
                 odoo_date, odoo_journal, odoo_amount}
    Merges with any pre-existing entries (deduped by pair_tag).
    """
    existing = load_manual_matches(output_dir)
    # index by pair_tag so re-saves don't duplicate
    by_tag: dict[str, dict] = {m["pair_tag"]: m for m in existing}
    for m in matches:
        by_tag[m["pair_tag"]] = m
    _write_manual_matches_json(output_dir, list(by_tag.values()))


def load_manual_matches(output_dir: Path) -> list[dict]:
    """Load persisted manual-match pairs from the JSON sidecar."""
    path = output_dir / MANUAL_MATCHES_FILE
    if not path.exists():
        return []
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []


def _normalize_date_str(val) -> str:
    """Normalize any date representation (date obj, YYYY-MM-DD, DD/MM/YYYY) to YYYY-MM-DD."""
    if not val:
        return ""
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if "/" in s:
        parts = s.split("/")
        if len(parts) == 3:
            return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
    elif "-" in s:
        parts = s.split("-")
        if len(parts) == 3:
            if len(parts[0]) == 4:
                return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
            elif len(parts[2]) == 4:
                return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
    return s


def apply_manual_matches_to_results(
    all_results: dict[str, list[dict]],
    matches: list[dict],
) -> tuple[dict[str, list[dict]], list[dict], list[dict]]:
    """Apply sidecar manual matches to all_results IN-MEMORY before report generation.

    For each sidecar entry:
    - Bank found (Only in Bank) + Odoo found (Only in Odoo):
        Merge → Bank row becomes STATUS_DONE with manual_match_tag; Odoo row removed.
    - Bank found but Odoo NOT found (Odoo payment deleted from website):
        Stale entry — leave Bank as Only in Bank; add to returned stale list.
    - Neither found (engine already matched them):
        Entry no longer needed; silently dropped from sidecar.

    Returns: (modified all_results, stale_entries, still_needed_for_sidecar)
    """
    AMT_TOL = 0.01
    still_needed: list[dict] = []
    stale_entries: list[dict] = []

    for m in matches:
        pair_tag = m["pair_tag"]
        b_date   = _normalize_date_str(m.get("bank_date", ""))
        b_j      = str(m.get("bank_journal", "")).strip().lower()
        b_amt    = float(m.get("bank_amount", 0))
        o_date   = _normalize_date_str(m.get("odoo_date", ""))
        o_amt    = float(m.get("odoo_amount", 0))
        diff     = float(m.get("diff", b_amt - o_amt))

        bank_row: dict | None = None
        odoo_row: dict | None = None
        bank_acc_key: str | None = None
        odoo_acc_key: str | None = None

        for acc_key, rows in all_results.items():
            if acc_key == "other":
                continue
            for r in rows:
                st = r.get("status", "")
                r_date = _normalize_date_str(r.get("date", ""))
                r_amt  = float(r.get("amount", 0))
                if bank_row is None and st == STATUS_BANK_ONLY and r_date == b_date and abs(r_amt - b_amt) < AMT_TOL:
                    bank_row = r
                    bank_acc_key = acc_key
                elif odoo_row is None and st == STATUS_ODO_ONLY and r_date == o_date and abs(r_amt - o_amt) < AMT_TOL:
                    odoo_row = r
                    odoo_acc_key = acc_key

        if bank_row is not None and odoo_row is not None:
            # Merge: upgrade Bank row to Match, copy Odoo metadata, remove Odoo row
            bank_row["status"] = STATUS_DONE
            bank_row["manual_match_tag"] = pair_tag
            bank_row["source"] = "Manual"
            bank_row["bank_amount"] = b_amt
            bank_row["odoo_amount"] = o_amt
            bank_row["manual_diff"] = diff
            bank_row["number_odo"] = odoo_row.get("number_odo") or m.get("odoo_number", "")
            bank_row["invoice_no"] = odoo_row.get("invoice_no") or m.get("odoo_reference", "")
            bank_row["is_reconciled"] = odoo_row.get("is_reconciled") or m.get("odoo_reconciled", "Yes")
            
            # Remove the Odoo row from results
            if odoo_acc_key and odoo_acc_key in all_results:
                all_results[odoo_acc_key] = [r for r in all_results[odoo_acc_key] if r is not odoo_row]
            still_needed.append(m)
        else:
            # Keep in sidecar for safety
            still_needed.append(m)

    return all_results, stale_entries, still_needed


def _reapply_manual_matches(wb, output_dir: Path) -> None:
    """After a fresh write_report, ensure manual match cells in bank sheets have proper green fill and values."""
    matches = load_manual_matches(output_dir)
    if not matches:
        return

    GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
    AMT_TOL = 0.01

    for m in matches:
        pair_tag = m["pair_tag"]
        b_date   = _normalize_date_str(m.get("bank_date", ""))
        b_amt    = float(m.get("bank_amount", 0))
        o_date   = _normalize_date_str(m.get("odoo_date", ""))
        o_amt    = float(m.get("odoo_amount", 0))
        o_num    = str(m.get("odoo_number", "")).strip()
        o_ref    = str(m.get("odoo_reference", "")).strip()
        o_recon  = str(m.get("odoo_reconciled", "Yes")).strip()
        diff_val = float(m.get("diff", b_amt - o_amt))

        skip_sheets = {"Daily Summary", "Differences", "Mutation Summary",
                       "Admin Fee", "Excluded Payment", "Other Payment",
                       "Other Mutation", "Legend"}
        for sname in wb.sheetnames:
            if sname in skip_sheets:
                continue
            ws_b = wb[sname]
            for rn in range(4, ws_b.max_row + 1):
                d_v = _normalize_date_str(ws_b.cell(rn, 2).value)
                b_amt_v = float(ws_b.cell(rn, 8).value or 0)
                src = str(ws_b.cell(rn, 10).value or "").strip()
                if (src in ("Bank", "Both", "Manual") or ws_b.cell(rn, 12).value == STATUS_DONE) and d_v == b_date and abs(b_amt_v - b_amt) < AMT_TOL:
                    if o_num: ws_b.cell(rn, 4).value = o_num
                    if o_ref: ws_b.cell(rn, 5).value = o_ref
                    c_o = ws_b.cell(rn, 9)
                    c_o.value = o_amt
                    c_o.number_format = '#,##0.00'
                    ws_b.cell(rn, 10).value = "Manual"
                    ws_b.cell(rn, 11).value = o_recon
                    ws_b.cell(rn, 12).value = pair_tag
                    c_diff = ws_b.cell(rn, 13)
                    c_diff.value = diff_val
                    c_diff.number_format = '#,##0.00'
                    for c in range(1, 14): ws_b.cell(rn, c).fill = GREEN_FILL


def _get_account_info(acc_key: str) -> tuple[str, str]:
    if acc_key == "other":
        return "Lainnya", "Belum Teridentifikasi"
    parts = acc_key.split("_", 1)
    if len(parts) == 2:
        bank = parts[0]
        alias = parts[1]
        acc_info = BANK_ACCOUNTS.get(bank, {}).get(alias, {})
        acc_name = acc_info.get("group") or alias
        return bank.upper(), acc_name
    return acc_key.upper(), acc_key


# ── Constants ─────────────────────────────────────────────────────────────
STATUS_DONE      = "Match"
STATUS_BANK_ONLY = "Only in Bank"
STATUS_ODO_ONLY  = "Only in Odoo"

COLOR_HEADER    = "1F4E79"
COLOR_DONE      = "E2EFDA"
COLOR_BANK_ONLY = "FCE4D6"
COLOR_ODO_ONLY  = "DDEBF7"
COLOR_WHITE     = "FFFFFF"

STATUS_COLOR = {
    STATUS_DONE:       COLOR_DONE,
    STATUS_BANK_ONLY:  COLOR_BANK_ONLY,
    STATUS_ODO_ONLY:   COLOR_ODO_ONLY,
}

THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLOR_HEADER    = "4F81BD"

def _rp(amount: Decimal) -> str:
    return f"Rp {int(amount):,}".replace(",", ".")


def _parse_date_obj(val):
    """Parse date strings (DD/MM/YYYY, YYYY-MM-DD, etc.), dates, or datetimes into (year, month, day) tuple for sorting."""
    if not val:
        return (9999, 12, 31)
    if hasattr(val, "year") and hasattr(val, "month") and hasattr(val, "day"):
        return (val.year, val.month, val.day)
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                dt = datetime.strptime(val, fmt)
                return (dt.year, dt.month, dt.day)
            except ValueError:
                pass
    return (9999, 12, 31)


def _date_sort_key(r):
    val = r.get("date") or r.get("payment_date") or r.get("txn_date") or ""
    return _parse_date_obj(val)


def _hdr(cell, text: str, bg: str = COLOR_HEADER):

    cell.value = text
    cell.font = Font(bold=True, color=COLOR_WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def _cell(cell, value, status: str, align="left"):
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=STATUS_COLOR.get(status, COLOR_WHITE))
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = BORDER
    cell.font = Font(size=10)

def _merge_title(ws, text: str, ncols: int, row: int, bg: str = COLOR_HEADER):
    """Write a full-width merged title cell."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    
    for col in range(1, ncols + 1):
        cell = ws.cell(row, col)
        if col == 1:
            cell.value = text
            cell.font = Font(size=14, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.fill = PatternFill("solid", fgColor=bg)
        
    ws.row_dimensions[row].height = 28


def _write_info_row(ws, parts: list[tuple[str, str]], ncols: int, row: int, bg: str = "DDEBF7"):
    """
    Write a merged info row containing key: value pairs separated by pipes.
    parts = [(label, value), ...]
    """
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    
    segments = "   |   ".join(f"{lbl}: {val}" for lbl, val in parts)
    
    for col in range(1, ncols + 1):
        cell = ws.cell(row, col)
        if col == 1:
            cell.value = segments
            cell.font = Font(size=10, color="1F4E79", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = Border(bottom=Side(style="medium", color="4F81BD"))
        
    ws.row_dimensions[row].height = 22

def _fmt_date(date_val) -> str:
    if not date_val:
        return ""
    if isinstance(date_val, (date, datetime)):
        return date_val.strftime("%d/%m/%Y")
    s = str(date_val).strip()
    if len(s) >= 10 and s[4] == '-' and s[7] == '-':
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except:
            pass
    return s

def _fmt_amount_str(val) -> str:
    if not val:
        return "0,00"
    try:
        s = f"{float(val):,.2f}"
        return s.replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return str(val)


def _write_bank_sheet(ws, rows: list[dict], bank_name: str, odo_date=None):
    ws.title = bank_name[:31]
    
    COL_HEADERS = [
        "No", "Date", "Payment Date", "Odoo Number", "Reference", "Bank Number", "Filename",
        "Bank Amount", "Odoo Amount", "Source", "Reconciled", "Status", "Difference"
    ]
    ncols = len(COL_HEADERS)

    done_count  = sum(1 for r in rows if r["status"] == STATUS_DONE)
    bank_count  = sum(1 for r in rows if r["status"] == STATUS_BANK_ONLY)
    odo_count   = sum(1 for r in rows if r["status"] == STATUS_ODO_ONLY)
    date_str    = odo_date.strftime("%d %B %Y") if odo_date else "-"
    total_amount = sum(float(r.get("amount", 0)) for r in rows)

    # ── Title block (rows 1-2) ──────────────────────────────────────────
    _merge_title(ws, f"RECONCILIATION {bank_name.upper()}", ncols, row=1)
    _write_info_row(ws, [
        ("Date",           date_str),
        ("Total Amount",   _fmt_amount_str(total_amount)),
        ("Total",          str(len(rows))),
        ("Match",          str(done_count)),
    ], ncols, row=2)

    # ── Column headers (row 3) ───────────────────────────────────────
    for col, h in enumerate(COL_HEADERS, 1):
        _hdr(ws.cell(row=3, column=col), h)
    ws.row_dimensions[3].height = 28

    # ── Data rows (start row 4) ───────────────────────────────────────
    rows_sorted = sorted(rows, key=_date_sort_key)
    for idx, r in enumerate(rows_sorted, 1):
        rn = idx + 3  # offset by 3 header rows
        st = r["status"]
        src = r.get("source", "")
        _cell(ws.cell(rn, 1), idx,                              st, "center")
        _cell(ws.cell(rn, 2), _fmt_date(r.get("date", "")),         st, "center")
        _cell(ws.cell(rn, 3), _fmt_date(r.get("payment_date", "")), st, "center")
        _cell(ws.cell(rn, 4), r.get("number_odo",  ""),         st)
        _cell(ws.cell(rn, 5), r.get("invoice_no",  ""),         st)
        _cell(ws.cell(rn, 6), r.get("number_bank", ""),         st)
        _cell(ws.cell(rn, 7), r.get("filename_bank", ""),       st)
        
        # Col 8 — Bank Amount
        c_b_amt = ws.cell(rn, 8)
        if st == STATUS_DONE or src == "Bank":
            b_val = float(r.get("bank_amount", r.get("amount", 0)))
            _cell(c_b_amt, b_val, st, "right")
            c_b_amt.number_format = '#,##0.00'
        else:
            _cell(c_b_amt, "", st, "right")

        # Col 9 — Odoo Amount
        c_o_amt = ws.cell(rn, 9)
        if st == STATUS_DONE or src == "Odoo":
            o_val = float(r.get("odoo_amount", r.get("amount", 0)))
            _cell(c_o_amt, o_val, st, "right")
            c_o_amt.number_format = '#,##0.00'
        else:
            _cell(c_o_amt, "", st, "right")
        
        # Determine display source
        if r.get("manual_match_tag") or src == "Manual":
            display_src = "Manual"
        elif st == STATUS_DONE:
            display_src = "Both"
        else:
            display_src = src

        _cell(ws.cell(rn, 10), display_src,                     st, "center")
        _cell(ws.cell(rn, 11), r.get("is_reconciled", ""),      st, "center")
        sc = ws.cell(rn, 12)
        # manual_match_tag present → show "Match (M01)" etc; st stays STATUS_DONE for fill color
        display_status = r.get("manual_match_tag", st)
        _cell(sc, display_status, st, "center")
        sc.font = Font(bold=True, size=10)
        
        # col 13 — Difference: manual matches carry actual diff; normal Match = 0
        diff_val = r.get("manual_diff", 0) if st == STATUS_DONE else ""
        c_diff = ws.cell(rn, 13)
        _cell(c_diff, diff_val, st, "right")
        if st == STATUS_DONE:
            c_diff.number_format = '#,##0.00'
        ws.row_dimensions[rn].height = 18

    for col, width in enumerate([6, 15, 15, 18, 20, 20, 30, 20, 20, 15, 15, 15, 18], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{len(rows) + 3}"
    ws.freeze_panes = "A4"  # freeze title + col-header rows


def _write_other_sheet(ws, rows: list[dict], odo_date=None):
    ws.title = "Other Payment"
    COL_HEADERS = ["No", "Date", "Journal", "Odoo Number", "Reference", "Amount", "Source", "Reconciled"]
    ncols = len(COL_HEADERS)
    
    date_str    = odo_date.strftime("%d %B %Y") if odo_date else "-"
    total_amount = sum(float(r.get("amount", 0)) for r in rows)

    # ── Title block (rows 1-2) ──────────────────────────────────────────
    _merge_title(ws, "OTHER PAYMENT", ncols, row=1)
    _write_info_row(ws, [
        ("Date",           date_str),
        ("Total Amount",   _fmt_amount_str(total_amount)),
        ("Total",          str(len(rows))),
    ], ncols, row=2)

    # ── Column headers (row 3) ───────────────────────────────────────
    for col, h in enumerate(COL_HEADERS, 1):
        _hdr(ws.cell(row=3, column=col), h)
    ws.row_dimensions[3].height = 28

    # ── Data rows (start row 4) ───────────────────────────────────────
    rows_sorted = sorted(rows, key=_date_sort_key)
    for idx, r in enumerate(rows_sorted, 1):
        rn = idx + 3
        st = r["status"]
        _cell(ws.cell(rn, 1), idx,                              st, "center")
        _cell(ws.cell(rn, 2), _fmt_date(r.get("date", "")),     st, "center")
        _cell(ws.cell(rn, 3), r.get("description", ""),         st, "center")
        _cell(ws.cell(rn, 4), r.get("number_odo",  ""),         st)
        _cell(ws.cell(rn, 5), r.get("invoice_no",  ""),         st)
        
        c_amt = ws.cell(rn, 6)
        _cell(c_amt, float(r["amount"]), st, "right")
        c_amt.number_format = '#,##0.00'
        
        _cell(ws.cell(rn, 7), r.get("source", ""),              st, "center")
        _cell(ws.cell(rn, 8), r.get("is_reconciled", ""),       st, "center")
        ws.row_dimensions[rn].height = 18

    for col, width in enumerate([6, 15, 20, 20, 20, 20, 15, 15], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{len(rows) + 3}"
    ws.freeze_panes = "A4"


def _write_discrepancy_sheet(ws, all_results: dict[str, list[dict]], odo_date=None):
    ws.title = "Differences"
    COL_HEADERS = [
        "No", "Date", "Bank", "Journal", "Odoo Number", "Reference", "Bank Number", "Filename",
        "Bank Amount", "Odoo Amount", "Source", "Reconciled", "Status", "Difference"
    ]
    ncols = len(COL_HEADERS)
    
    date_str   = odo_date.strftime("%d %B %Y") if odo_date else "-"
    
    rows_data = [
        r for rows in all_results.values()
        for r in rows if r["status"] != STATUS_DONE or r.get("manual_match_tag")
    ]
    total_disc = len([r for r in rows_data if not r.get("manual_match_tag")])
    total_amount = sum(float(r.get("amount", 0)) for r in rows_data if not r.get("manual_match_tag"))

    # ── Title block ────────────────────────────────────────────────────
    _merge_title(ws, "DIFFERENCES — ALL BANKS", ncols, row=1)
    _write_info_row(ws, [
        ("Date",          date_str),
        ("Total Amount",  _fmt_amount_str(total_amount)),
        ("Total Differences", str(total_disc)),
    ], ncols, row=2)

    # ── Column headers (row 3) ─────────────────────────────────────
    for col, h in enumerate(COL_HEADERS, 1):
        _hdr(ws.cell(3, col), h)
    ws.row_dimensions[3].height = 28

    idx = 0
    for acc_key, rows in all_results.items():
        if acc_key == "other":
            continue
            
        bank_label, acc_name = _get_account_info(acc_key)
        rows_sorted = sorted(
            (r for r in rows if r["status"] != STATUS_DONE or r.get("manual_match_tag")),
            key=_date_sort_key
        )
        for r in rows_sorted:
            idx += 1
            rn = idx + 3
            st = r["status"]
            src = r.get("source", "")
            is_manual = bool(r.get("manual_match_tag") or src == "Manual")

            _cell(ws.cell(rn, 1), idx,                            st, "center")
            _cell(ws.cell(rn, 2), _fmt_date(r.get("date", "")),   st, "center")
            _cell(ws.cell(rn, 3), bank_label,                     st, "center")
            _cell(ws.cell(rn, 4), acc_name,                       st, "center")
            _cell(ws.cell(rn, 5), r.get("number_odo",  ""),       st)
            _cell(ws.cell(rn, 6), r.get("invoice_no",  ""),       st)
            _cell(ws.cell(rn, 7), r.get("number_bank", ""),       st)
            _cell(ws.cell(rn, 8), r.get("filename_bank", ""),     st)
            
            # Col 9 — Bank Amount
            c_b_amt = ws.cell(rn, 9)
            if src == "Bank" or is_manual:
                b_val = float(r.get("bank_amount", r.get("amount", 0)))
                _cell(c_b_amt, b_val, st, "right")
                c_b_amt.number_format = '#,##0.00'
            else:
                _cell(c_b_amt, "", st, "right")

            # Col 10 — Odoo Amount
            c_o_amt = ws.cell(rn, 10)
            if src == "Odoo" or is_manual:
                o_val = float(r.get("odoo_amount", r.get("amount", 0)))
                _cell(c_o_amt, o_val, st, "right")
                c_o_amt.number_format = '#,##0.00'
            else:
                _cell(c_o_amt, "", st, "right")
            
            display_src = "Manual" if is_manual else src
            _cell(ws.cell(rn, 11), display_src,                   st, "center")
            _cell(ws.cell(rn, 12), r.get("is_reconciled", ""),     st, "center")
            sc = ws.cell(rn, 13)
            display_status = r.get("manual_match_tag", st)
            _cell(sc, display_status, st, "center")
            sc.font = Font(bold=True, size=10)
            
            # col 14 — Difference
            diff_val = r.get("manual_diff", "")
            c_diff = ws.cell(rn, 14)
            _cell(c_diff, diff_val, st, "right")
            if diff_val != "":
                c_diff.number_format = '#,##0.00'
            ws.row_dimensions[rn].height = 18

    for col, width in enumerate([6, 15, 10, 18, 20, 20, 20, 30, 20, 20, 15, 15, 15, 18], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{idx + 3}"
    ws.freeze_panes = "A4"


def _write_legend(ws):
    ws.title = "Legend"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    entries = [
        ("Status", "Meaning", True),
        (STATUS_DONE,      "Transaction exists in both Bank and Odoo",          False),
        (STATUS_BANK_ONLY, "Transaction exists ONLY in Bank, not in Odoo", False),
        (STATUS_ODO_ONLY,  "Transaction exists ONLY in Odoo, not in Bank", False),
    ]
    for i, (lbl, desc, is_hdr) in enumerate(entries, 1):
        a = ws.cell(i, 1, lbl)
        b = ws.cell(i, 2, desc)
        if is_hdr:
            for c in [a, b]:
                c.font = Font(bold=True, color=COLOR_WHITE)
                c.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                c.alignment = Alignment(horizontal="center")
        else:
            col = STATUS_COLOR.get(lbl, COLOR_WHITE)
            for c in [a, b]:
                c.fill = PatternFill("solid", fgColor=col)
                c.font = Font(size=10)
            a.font = Font(bold=True, size=10)
        for c in [a, b]:
            c.border = BORDER
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[i].height = 20


def _write_daily_summary_sheet(ws, all_results: dict, odo_date=None):
    from collections import defaultdict
    from decimal import Decimal

    ws.title = "Daily Summary"
    COL_HEADERS = ["No", "Date", "Payment Date", "Bank", "Journal", "Total Bank", "Total Odoo", "Difference", "Reconciled", "Status", "Journal Information", "EDC Number", "AR Number"]
    ncols = len(COL_HEADERS)
    date_str = odo_date.strftime("%d %B %Y") if odo_date else "-"

    _merge_title(ws, "DAILY SUMMARY — BANK vs ODO", ncols, row=1)
    _write_info_row(ws, [("Date", date_str)], ncols, row=2)

    for col, h in enumerate(COL_HEADERS, 1):
        _hdr(ws.cell(3, col), h)
    ws.row_dimensions[3].height = 28


    bank_sums: dict = defaultdict(lambda: defaultdict(Decimal))
    odo_sums:  dict = defaultdict(lambda: defaultdict(Decimal))
    payment_dates = {}
    reconciled_status = {}
    
    all_banks = sorted([k for k in all_results.keys() if k != "other"])

    for bank in all_banks:
        for r in sorted(all_results.get(bank, []), key=_date_sort_key):
            d   = r.get("date", "") or ""
            pd  = r.get("payment_date", "")
            if pd:
                payment_dates[(bank, d)] = pd
                
            recon = r.get("is_reconciled", "")
            if recon in ("Yes", "No"):
                prev = reconciled_status.get((bank, d))
                if prev is None:
                    reconciled_status[(bank, d)] = recon
                elif prev != recon and prev != "Mixed":
                    reconciled_status[(bank, d)] = "Mixed"
            amt = r.get("amount", Decimal(0))
            st  = r.get("status", "")
            if st == STATUS_DONE:
                bank_sums[bank][d] += amt
                odo_sums[bank][d]  += amt
            elif st == STATUS_BANK_ONLY:
                bank_sums[bank][d] += amt
            elif st == STATUS_ODO_ONLY:
                odo_sums[bank][d]  += amt

    rows = []
    for bank in all_banks:
        for d in sorted(set(list(bank_sums[bank].keys()) + list(odo_sums[bank].keys())), key=_parse_date_obj):
            b_sum = bank_sums[bank][d]
            o_sum = odo_sums[bank][d]
            pd    = payment_dates.get((bank, d), "")
            recon = reconciled_status.get((bank, d), "-")
            bank_clean, acc_name = _get_account_info(bank)
            rows.append((bank_clean, acc_name, d, pd, b_sum, o_sum, recon))

    COLOR_MATCH   = "E2EFDA"
    COLOR_DIFF    = "FCE4D6"
    COLOR_MISSING = "FFF2CC"

    for idx, (bank, acc_name, d, pd, b_sum, o_sum, recon) in enumerate(rows, 1):
        rn  = idx + 3
        sel = b_sum - o_sum

        if b_sum == 0 or o_sum == 0:
            bg = COLOR_MISSING
            label = "Incomplete Data"
        elif sel == 0:
            bg = COLOR_MATCH
            label = "Match"
        else:
            bg = COLOR_DIFF
            label = "Difference"

        def _sc(cell, val, align="left", _bg=bg):
            cell.value = val
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.fill = PatternFill("solid", fgColor=_bg)
            cell.font = Font(size=10)
            cell.border = Border(
                bottom=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
            )

        def _num(cell, val, _bg=bg):
            cell.value = float(val)
            cell.number_format = "#,##0.00"
            cell.fill = PatternFill("solid", fgColor=_bg)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = Font(size=10)
            cell.border = Border(
                bottom=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
            )

        _sc(ws.cell(rn, 1), idx, "center")
        _sc(ws.cell(rn, 2), _fmt_date(d), "center")
        _sc(ws.cell(rn, 3), _fmt_date(pd), "center")
        _sc(ws.cell(rn, 4), bank, "center")
        _sc(ws.cell(rn, 5), acc_name, "center")
        _num(ws.cell(rn, 6), b_sum)
        _num(ws.cell(rn, 7), o_sum)
        _num(ws.cell(rn, 8), sel, _bg=COLOR_DIFF if sel != 0 else bg)
        rc = ws.cell(rn, 9)
        _sc(rc, recon, "center")
        rc.font = Font(size=10, bold=True)
        lc = ws.cell(rn, 10)
        _sc(lc, label, "center")
        lc.font = Font(size=10, bold=True)
        
        # Journal status defaults to "Not Yet" until checked by journal_checker.py
        j_stat_label = "Not Yet"
            
        jc = ws.cell(rn, 11)
        _sc(jc, j_stat_label, "center")
        jc.font = Font(size=10, bold=True)
        
        c12 = ws.cell(rn, 12)
        _sc(c12, "-", "center")
        c12.font = Font(size=10, bold=True)

        c13 = ws.cell(rn, 13)
        _sc(c13, "-", "center")
        c13.font = Font(size=10, bold=True)
        ws.row_dimensions[rn].height = 18


    for col, width in enumerate([6, 15, 15, 18, 30, 20, 20, 20, 15, 20, 22, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{len(rows) + 3}"
    ws.freeze_panes = "A4"


def _auto_adjust_headers(wb):
    for ws in wb.worksheets:
        ncols = ws.max_column
        if ncols <= 1:
            continue
            
        for m in list(ws.merged_cells.ranges):
            if m.min_row in (1, 2) and m.max_row in (1, 2):
                ws.unmerge_cells(str(m))
                
        title_cell = ws.cell(row=1, column=1)
        if title_cell.value:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            bg = title_cell.fill.fgColor
            if bg and getattr(bg, "rgb", None):
                for c in range(1, ncols + 1):
                    ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=bg.rgb)
                    
        info_cell = ws.cell(row=2, column=1)
        if info_cell.value:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
            bg = info_cell.fill.fgColor
            if bg and getattr(bg, "rgb", None):
                border = Border(bottom=Side(style="medium", color="4F81BD"))
                for c in range(1, ncols + 1):
                    cell = ws.cell(row=2, column=c)
                    cell.fill = PatternFill("solid", fgColor=bg.rgb)
                    cell.border = border
                    
        if ws.auto_filter.ref:
            max_row = ws.max_row
            ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{max_row}"

def _auto_adjust_col_widths(wb):
    for ws in wb.worksheets:
        if ws.title == "Daily Summary":
            daily_widths = {
                1: 6,   # No
                2: 15,  # Date
                3: 15,  # Payment Date
                4: 15,  # Bank
                5: 18,  # Journal
                6: 20,  # Total Bank
                7: 20,  # Total Odoo
                8: 20,  # Difference
                9: 15,  # Reconciled
                10: 20, # Status
                11: 22, # Journal Information
                12: 18, # EDC Number
                13: 18, # AR Number
            }
            for col_idx, width in daily_widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            continue

        if ws.title == "Differences":
            diff_widths = {
                1: 6, 2: 15, 3: 12, 4: 22, 5: 22, 6: 22, 7: 20, 8: 30,
                9: 20, 10: 20, 11: 15, 12: 15, 13: 15, 14: 18
            }
            for col_idx, width in diff_widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            continue

        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                # Skip first two rows since they often contain merged title cells which can distort the width
                if row_idx <= 2:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            
            adjusted_width = min(max_length + 2, 60)
            if adjusted_width < 10:
                adjusted_width = 10
            # Keep whichever is larger between calculated width and current width
            curr_width = getattr(ws.column_dimensions[get_column_letter(col_idx)], "width", 0) or 0
            ws.column_dimensions[get_column_letter(col_idx)].width = max(adjusted_width, curr_width)


def write_report(
    all_results: dict[str, list[dict]],
    odo_date: date,
    output_dir: Path,
    bank_txns: dict | None = None,
    odo_bank_txns: dict | None = None,
    mutations: list[dict] = None,
    unknown_mutations: list[dict] = None,
    excluded_txns: list[dict] = None,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    company_str = ODOO_COMPANY_NAME.replace(" ", "_") if ODOO_COMPANY_NAME else "Company"

    
    banks = set()
    for acc_key in all_results.keys():
        if acc_key != "other":
            bank, _ = _get_account_info(acc_key)
            banks.add(bank)
    banks_str = "_".join(sorted(banks)) if banks else "Banks"
    
    all_dates = []
    for rows in all_results.values():
        for r in rows:
            for date_field in ["date", "payment_date"]:
                d_val = r.get(date_field)
                if not d_val: continue
                if isinstance(d_val, (date, datetime)):
                    all_dates.append(d_val)
                else:
                    s = str(d_val).strip()
                    if len(s) >= 10 and s[4] == '-' and s[7] == '-':
                        try:
                            all_dates.append(datetime.strptime(s[:10], "%Y-%m-%d").date())
                        except:
                            pass
                    elif len(s) >= 10 and s[2] == '/' and s[5] == '/':
                        try:
                            all_dates.append(datetime.strptime(s[:10], "%d/%m/%Y").date())
                        except:
                            pass
                            
    if all_dates:
        min_date = min(all_dates).strftime("%d%m%Y")
        max_date = max(all_dates).strftime("%d%m%Y")
        date_range_str = f"{min_date}_to_{max_date}"
    else:
        date_range_str = "unknown"

    odo_date_str = odo_date.strftime("%d%m%Y") if odo_date else "unknown"
    prefix = f"Reconciliation_{company_str}_{odo_date_str}__"
    out_path = output_dir / f"{prefix}{banks_str}__{date_range_str}.xlsx"

    # Close any open reconciliation workbook in Excel before deleting/overwriting
    _close_workbook_in_excel(out_path)

    if output_dir.exists():
        for existing_file in output_dir.iterdir():
            if existing_file.is_file() and existing_file.name.startswith(prefix) and existing_file.name.endswith(".xlsx"):
                try:
                    existing_file.unlink()
                except Exception:
                    pass

    wb = openpyxl.Workbook()

    # ── Apply manual matches to all_results BEFORE writing any sheet ──────────
    # This ensures Daily Summary, Differences, and bank sheets are all consistent.
    # Previously _reapply_manual_matches() ran AFTER Daily Summary was written,
    # causing Total Odoo to be stale whenever a manual match existed.
    matches = load_manual_matches(output_dir)
    if matches:
        all_results, stale_entries, still_needed = apply_manual_matches_to_results(all_results, matches)
        _write_manual_matches_json(output_dir, still_needed)
        for s in stale_entries:
            print(
                f"  [WARN] Stale manual match {s['pair_tag']}: "
                f"Odoo payment {s.get('odoo_amount', '?')} on {s.get('odoo_date', '?')} "
                f"no longer found in current Odoo data. Bank entry left as 'Only in Bank'."
            )

    # 1. Create Daily Summary as the first sheet
    ws_daily = wb.active
    ws_daily.title = "Daily Summary"
    _write_daily_summary_sheet(ws_daily, all_results, odo_date=odo_date)
    
    # 2. Create Differences as the second sheet
    ws_disc = wb.create_sheet(title="Differences")
    _write_discrepancy_sheet(ws_disc, all_results, odo_date=odo_date)
    
    # 3. Mutation Summary
    if mutations:
        ws_mut = wb.create_sheet(title="Mutation Summary")
        _write_mutation_summary(ws_mut, all_results, mutations, odo_date=odo_date)
        
    # 4. Admin Fee
    if mutations:
        ws_admin_fee = wb.create_sheet(title="Admin Fee")
        _write_admin_fee_sheet(ws_admin_fee, all_results, mutations, odo_date=odo_date)

    # 5. Excluded Payment
    if excluded_txns is not None and excluded_txns:
        ws_excluded = wb.create_sheet(title="Excluded Payment")
        _write_excluded_sheet(ws_excluded, excluded_txns, odo_date=odo_date)
    
    other_rows = all_results.pop("other", [])
    
    # 6. Bank Sheets
    for acc_key, rows in all_results.items():
        bank_label, acc_name = _get_account_info(acc_key)
        sheet_name = f"{bank_label} ({acc_name})"[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_bank_sheet(ws, rows, f"{bank_label} ({acc_name})", odo_date=odo_date)

    # 7. Other Payment
    if other_rows:
        ws_other = wb.create_sheet(title="Other Payment")
        _write_other_sheet(ws_other, other_rows, odo_date=odo_date)
        
    # 8. Other Mutation
    if unknown_mutations:
        ws_unmapped = wb.create_sheet(title="Other Mutation")
        _write_unmapped_mutations(ws_unmapped, all_results, unknown_mutations, odo_date=odo_date)
        
    ws_legend = wb.create_sheet(title="Legend")
    _write_legend(ws_legend)

    _auto_adjust_headers(wb)
    _auto_adjust_col_widths(wb)
    # Stamp green fill + Match(Mxx) tag text for manually matched rows.
    # Status changes were already applied in-memory above; this only handles
    # the Excel visual layer (fill colour + tag string in Status cell).
    if matches:
        _reapply_manual_matches(wb, output_dir)
    return _safe_save_report(wb, out_path)


def _close_workbook_in_excel(file_path: Path):
    """Close the specific open workbook in Excel on Windows/Mac before writing."""
    import sys, subprocess, os, time
    fname = file_path.name
    if sys.platform == "darwin":
        try:
            subprocess.run(
                ["osascript", "-e",
                 f'tell application "Microsoft Excel"\n'
                 f'  set wb_list to every workbook whose name contains "{fname}"\n'
                 f'  repeat with wb_item in wb_list\n'
                 f'    close wb_item saving no\n'
                 f'  end repeat\n'
                 f'end tell'],
                capture_output=True, timeout=5
            )
        except Exception:
            pass
    elif os.name == "nt" or sys.platform == "win32":
        try:
            import ctypes
            from ctypes import wintypes
            user32 = ctypes.windll.user32

            def enum_cb(hwnd, extra):
                if user32.IsWindowVisible(hwnd):
                    length = user32.GetWindowTextLengthW(hwnd)
                    if length > 0:
                        buff = ctypes.create_unicode_buffer(length + 1)
                        user32.GetWindowTextW(hwnd, buff, length + 1)
                        val = buff.value.lower()
                        if fname.lower() in val or "reconciliation" in val:
                            user32.PostMessageW(hwnd, 0x0010, 0, 0)  # WM_CLOSE
                return True

            WNDENUMPROC = ctypes.WINFUNCTYPE(ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)
            user32.EnumWindows(WNDENUMPROC(enum_cb), 0)
            time.sleep(0.5)
        except Exception:
            try:
                flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
                subprocess.run(
                    ["powershell", "-NoProfile", "-Command",
                     f'Get-Process | Where-Object {{ $_.MainWindowTitle -like "*{fname}*" -or $_.MainWindowTitle -like "*Reconciliation*" }} | ForEach-Object {{ $_.CloseMainWindow() }}'],
                    capture_output=True, creationflags=flags
                )
                time.sleep(0.5)
            except Exception:
                pass


def _safe_save_report(wb, out_path: Path) -> Path:
    """Save workbook safely, automatically closing old instance in Excel if needed."""
    try:
        wb.save(out_path)
        return out_path
    except PermissionError:
        # File is locked in Excel, close it and retry saving
        _close_workbook_in_excel(out_path)
        try:
            wb.save(out_path)
            return out_path
        except Exception:
            pass

        tmp_path = out_path.with_suffix(".tmp.xlsx")
        try:
            wb.save(tmp_path)
            os.replace(tmp_path, out_path)
            return out_path
        except Exception:
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except Exception:
                    pass

        from datetime import datetime
        ts = datetime.now().strftime("%H%M%S")
        alt_path = out_path.parent / f"{out_path.stem}_{ts}.xlsx"
        try:
            wb.save(alt_path)
            print(f"\n⚠️ Note: '{out_path.name}' was locked by Excel.")
            print(f"   Saved updated report to '{alt_path.name}' instead.\n")
            return alt_path
        except Exception:
            raise



def _build_date_lookup(all_results):
    date_lookup = {}
    if all_results:
        for acc_key, rows in all_results.items():
            if acc_key == "other": continue
            bank, acc_name = _get_account_info(acc_key)
            for r in rows:
                if r["status"] == STATUS_DONE:
                    p_date = _fmt_date(r.get("payment_date", ""))
                    o_date = _fmt_date(r.get("date", ""))
                    if p_date and o_date:
                        key = (bank, acc_name, p_date)
                        if key not in date_lookup:
                            date_lookup[key] = set()
                        date_lookup[key].add(o_date)
    return date_lookup

def _write_mutation_summary(ws, all_results, mutations, odo_date=None):
    ws.title = "Mutation Summary"
    ncols = 8
    date_str = odo_date.strftime("%d %B %Y") if odo_date else "-"
    total_amount = sum(float(m.get("amount", 0)) for m in mutations) if mutations else 0.0
    
    _merge_title(ws, "MUTATION SUMMARY", ncols, row=1)
    _write_info_row(ws, [
        ("Date", date_str),
        ("Total Amount", _fmt_amount_str(total_amount))
    ], ncols, row=2)
    
    _hdr(ws.cell(3, 1), "No")
    _hdr(ws.cell(3, 2), "Date")
    _hdr(ws.cell(3, 3), "Payment Date")
    _hdr(ws.cell(3, 4), "Bank")
    _hdr(ws.cell(3, 5), "Journal")
    _hdr(ws.cell(3, 6), "Description")
    _hdr(ws.cell(3, 7), "Transaction Category")
    _hdr(ws.cell(3, 8), "Total Amount")
    ws.row_dimensions[3].height = 28
    
    date_lookup = _build_date_lookup(all_results)
    
    from collections import defaultdict
    summary = defaultdict(float)
    for m in mutations:
        p_str = _fmt_date(m["date"])
        bank = m["bank"].upper()
        _, acc_name = _get_account_info(f"{m['bank']}_{m['alias']}")
        
        o_dates = date_lookup.get((bank, acc_name, p_str), set())
        o_str = ", ".join(sorted(o_dates)) if o_dates else "-"
        
        desc = m.get("desc", "")
        key = (o_str, p_str, bank, acc_name, desc, m.get("type", "Unknown"))
        summary[key] += m["amount"]
        
    row = 4
    idx = 1
    for key in sorted(summary.keys(), key=lambda k: (_parse_date_obj(k[0]), _parse_date_obj(k[1]), k[2])):
        o_str, p_str, bank, acc_name, desc, txn_type = key
        amount = summary[key]
        
        _cell(ws.cell(row, 1), idx, STATUS_DONE, "center")
        _cell(ws.cell(row, 2), o_str, STATUS_DONE, "center")
        _cell(ws.cell(row, 3), p_str, STATUS_DONE, "center")
        _cell(ws.cell(row, 4), bank, STATUS_DONE, "center")
        _cell(ws.cell(row, 5), acc_name, STATUS_DONE, "center")
        _cell(ws.cell(row, 6), desc, STATUS_DONE, "left")
        _cell(ws.cell(row, 7), txn_type, STATUS_DONE, "center")
        
        c_amt = ws.cell(row, 8)
        _cell(c_amt, amount, STATUS_DONE, "right")
        c_amt.number_format = '#,##0.00'
        
        row += 1
        idx += 1

    for col, width in enumerate([6, 15, 15, 12, 20, 40, 20, 20], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{row - 1}"
    ws.freeze_panes = "A4"


def _write_admin_fee_sheet(ws, all_results, mutations, odo_date=None):
    ws.title = "Admin Fee"
    ncols = 8
    date_str = odo_date.strftime("%d %B %Y") if odo_date else "-"
    
    total_admin = 0.0
    if mutations:
        total_admin += sum(float(m.get("admin_fee") or 0) for m in mutations)
    for acc_key, rows in all_results.items():
        if acc_key.startswith("bca") or acc_key.startswith("bri") or acc_key == "other":
            continue
        total_admin += sum(float(r.get("admin_fee") or 0) for r in rows)
    
    _merge_title(ws, "ADMIN FEE", ncols, row=1)
    _write_info_row(ws, [
        ("Date", date_str),
        ("Total Amount", _fmt_amount_str(total_admin))
    ], ncols, row=2)
    
    _hdr(ws.cell(3, 1), "No")
    _hdr(ws.cell(3, 2), "Date")
    _hdr(ws.cell(3, 3), "Payment Date")
    _hdr(ws.cell(3, 4), "Bank")
    _hdr(ws.cell(3, 5), "Journal")
    _hdr(ws.cell(3, 6), "Description")
    _hdr(ws.cell(3, 7), "Transaction Category")
    _hdr(ws.cell(3, 8), "Total Admin Fee")
    ws.row_dimensions[3].height = 28
    
    date_lookup = _build_date_lookup(all_results)
    
    from collections import defaultdict
    summary = defaultdict(float)
    
    if mutations:
        for m in mutations:
            adm = float(m.get("admin_fee") or 0)
            if adm > 0:
                d_str = _fmt_date(m["date"])
                bank = m["bank"].upper()
                parts = _get_account_info(f"{m['bank']}_{m['alias']}")
                acc_name = parts[1]
                
                o_dates = date_lookup.get((bank, acc_name, d_str), set())
                o_str = ", ".join(sorted(o_dates)) if o_dates else "-"
                
                desc = m.get("desc", "")
                key = (o_str, d_str, bank, acc_name, desc, m.get("type", "Unknown"))
                summary[key] += adm
                
    for acc_key, rows in all_results.items():
        if acc_key.startswith("bca") or acc_key.startswith("bri") or acc_key == "other":
            continue
            
        bank_label, acc_name = _get_account_info(acc_key)
        
        for r in rows:
            adm = float(r.get("admin_fee") or 0)
            if adm > 0:
                p_str = _fmt_date(r.get("payment_date", ""))
                o_str = _fmt_date(r.get("date", ""))
                if not p_str:
                    p_str = "Unknown"
                if not o_str:
                    o_str = "-"
                
                cat = r.get("category") or "Unknown"
                desc = r.get("description") or "-"
                key = (o_str, p_str, bank_label, acc_name, desc, cat)
                summary[key] += adm

    row = 4
    idx = 1
    for key in sorted(summary.keys(), key=lambda k: (_parse_date_obj(k[0]), _parse_date_obj(k[1]), k[2])):
        o_str, p_str, bank, acc_name, desc, txn_type = key
        amount = summary[key]
        
        _cell(ws.cell(row, 1), idx, STATUS_DONE, "center")
        _cell(ws.cell(row, 2), o_str, STATUS_DONE, "center")
        _cell(ws.cell(row, 3), p_str, STATUS_DONE, "center")
        _cell(ws.cell(row, 4), bank, STATUS_DONE, "center")
        _cell(ws.cell(row, 5), acc_name, STATUS_DONE, "center")
        _cell(ws.cell(row, 6), desc, STATUS_DONE, "left")
        _cell(ws.cell(row, 7), txn_type, STATUS_DONE, "center")
        
        c_amt = ws.cell(row, 8)
        _cell(c_amt, amount, STATUS_DONE, "right")
        c_amt.number_format = '#,##0.00'
        row += 1
        idx += 1

    for col, width in enumerate([6, 15, 15, 12, 20, 30, 20, 20], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.auto_filter.ref = f"A3:{get_column_letter(8)}{row - 1}"
    ws.freeze_panes = "A4"


def _write_unmapped_mutations(ws, all_results, unknowns, odo_date=None):
    ws.title = "Other Mutation"
    ncols = 8
    date_str = odo_date.strftime("%d %B %Y") if odo_date else "-"
    total_amount = sum(float(m.get("amount", 0)) for m in unknowns) if unknowns else 0.0
    
    _merge_title(ws, "OTHER MUTATION", ncols, row=1)
    _write_info_row(ws, [
        ("Date", date_str),
        ("Total Amount", _fmt_amount_str(total_amount))
    ], ncols, row=2)
    
    _hdr(ws.cell(3, 1), "No")
    _hdr(ws.cell(3, 2), "Date")
    _hdr(ws.cell(3, 3), "Payment Date")
    _hdr(ws.cell(3, 4), "Bank")
    _hdr(ws.cell(3, 5), "Journal")
    _hdr(ws.cell(3, 6), "Description")
    _hdr(ws.cell(3, 7), "Transaction Type")
    _hdr(ws.cell(3, 8), "Amount")
    ws.row_dimensions[3].height = 28
    
    date_lookup = _build_date_lookup(all_results)
    
    row = 4
    idx = 1
    for m in sorted(unknowns, key=_date_sort_key):
        p_str = _fmt_date(m["date"])
        bank_raw = m["bank"].upper()
        _, acc_name = _get_account_info(f"{m['bank']}_{m['alias']}")
        
        o_dates = date_lookup.get((bank_raw, acc_name, p_str), set())
        o_str = ", ".join(sorted(o_dates)) if o_dates else "-"
        
        _cell(ws.cell(row, 1), idx, STATUS_BANK_ONLY, "center")
        _cell(ws.cell(row, 2), o_str, STATUS_BANK_ONLY, "center")
        _cell(ws.cell(row, 3), p_str, STATUS_BANK_ONLY, "center")
        _cell(ws.cell(row, 4), bank_raw, STATUS_BANK_ONLY, "center")
        _cell(ws.cell(row, 5), acc_name, STATUS_BANK_ONLY, "center")
        _cell(ws.cell(row, 6), m.get("desc", ""), STATUS_BANK_ONLY, "left")
        _cell(ws.cell(row, 7), m.get("cr_db_type", ""), STATUS_BANK_ONLY, "center")
        
        c_amt = ws.cell(row, 8)
        _cell(c_amt, float(m["amount"]), STATUS_BANK_ONLY, "right")
        c_amt.number_format = '#,##0.00'
        
        row += 1
        idx += 1

    for col, width in enumerate([6, 15, 15, 12, 18, 40, 14, 20], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{row - 1}"
    ws.freeze_panes = "A4"

def _write_excluded_sheet(ws, excluded_txns: list[dict], odo_date=None):
    ws.title = "Excluded Payment"
    ncols = 8
    date_str = odo_date.strftime("%d %B %Y") if odo_date else "-"
    total_amount = sum(float(t.get("amount", 0)) for t in excluded_txns)
    
    _merge_title(ws, "EXCLUDED TRANSACTIONS", ncols, row=1)
    _write_info_row(ws, [
        ("Date", date_str),
        ("Total Excluded", str(len(excluded_txns))),
        ("Total Amount", _fmt_amount_str(total_amount))
    ], ncols, row=2)
    
    _hdr(ws.cell(3, 1), "No")
    _hdr(ws.cell(3, 2), "Date")
    _hdr(ws.cell(3, 3), "Bank")
    _hdr(ws.cell(3, 4), "Trace Number")
    _hdr(ws.cell(3, 5), "Filename")
    _hdr(ws.cell(3, 6), "Description")
    _hdr(ws.cell(3, 7), "Amount")
    _hdr(ws.cell(3, 8), "Exclusion Reason")
    ws.row_dimensions[3].height = 28
    
    row = 4
    for idx, t in enumerate(excluded_txns, 1):
        _cell(ws.cell(row, 1), idx, STATUS_BANK_ONLY, "center")
        _cell(ws.cell(row, 2), _fmt_date(t.get("date", "")), STATUS_BANK_ONLY, "center")
        bank_name = t.get("source", "").replace("Bank (", "").replace(")", "")
        _cell(ws.cell(row, 3), bank_name, STATUS_BANK_ONLY, "center")
        _cell(ws.cell(row, 4), t.get("number", ""), STATUS_BANK_ONLY, "center")
        _cell(ws.cell(row, 5), t.get("filename", ""), STATUS_BANK_ONLY, "left")
        _cell(ws.cell(row, 6), t.get("description", ""), STATUS_BANK_ONLY, "left")
        
        c_amt = ws.cell(row, 7)
        _cell(c_amt, float(t.get("amount", 0)), STATUS_BANK_ONLY, "right")
        c_amt.number_format = '#,##0.00'
        
        _cell(ws.cell(row, 8), t.get("exclusion_reason", "Voided"), STATUS_BANK_ONLY, "center")
        
        ws.row_dimensions[row].height = 18
        row += 1

    for col, width in enumerate([6, 15, 15, 20, 30, 30, 20, 25], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{row - 1}"
    ws.freeze_panes = "A4"
