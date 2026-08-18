"""
gui.py — Bank Reconciliation GUI. Cross-platform (Windows + Mac).

PyInstaller pattern:
  - GUI mode   : BankRekonsiliasi.exe            → opens this window
  - Worker mode: BankRekonsiliasi.exe --worker   → runs main logic, stdout only
"""
import sys
import os
import platform
from pathlib import Path

# ── Resolve base directory ────────────────────────────────────────────────────
if getattr(sys, "frozen", False):
    if sys.platform == "darwin" and ".app/Contents/MacOS" in str(sys.executable):
        BASE_DIR = Path(sys.executable).parents[2].parent
    else:
        BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).resolve().parent

INPUT_DIR  = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
IS_WINDOWS = platform.system() == "Windows"

# ── Worker mode ───────────────────────────────────────────────────────────────
# The GUI re-launches the same binary with --worker for a clean stdout stream.
# Must be checked BEFORE any tkinter import so the headless worker doesn't
# require a display or GUI toolkit.
if "--run-journal-creator" in sys.argv:
    sys.argv = [a for a in sys.argv if a != "--run-journal-creator"]
    os.chdir(str(BASE_DIR))
    import io as _io
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    elif hasattr(sys.stdout, "buffer"):
        sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    import runpy
    if getattr(sys, "frozen", False):
        jc_path = str(Path(getattr(sys, "_MEIPASS", BASE_DIR)) / "odoo_journal_creator.py")
    else:
        jc_path = str(BASE_DIR / "odoo_journal_creator.py")
    runpy.run_path(jc_path, run_name="__main__")
    sys.exit(0)

if "--worker" in sys.argv:
    sys.argv = [a for a in sys.argv if a != "--worker"]  # hide flag from main's argparse
    os.chdir(str(BASE_DIR))
    # Force UTF-8 — Windows CP1252 can't encode ↔, ✅, ❌, ── etc.
    import io as _io
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    elif hasattr(sys.stdout, "buffer"):
        sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    import runpy
    if getattr(sys, "frozen", False):
        main_path = str(Path(getattr(sys, "_MEIPASS", BASE_DIR)) / "main.py")
    else:
        main_path = str(BASE_DIR / "main.py")
    runpy.run_path(main_path, run_name="__main__")
    sys.exit(0)

# ── GUI-only imports (skipped in worker mode) ─────────────────────────────────
if IS_WINDOWS:
    try:
        from ctypes import windll
        # Per-Monitor V2 DPI awareness (level 2): each monitor reports its own
        # DPI and Windows does NOT bitmap-scale the window — gives sharp custom
        # fonts at any display scaling (100%, 125%, 150%, 200%).
        # Falls back to level 1 (System DPI) if shcore is unavailable.
        try:
            windll.shcore.SetProcessDpiAwareness(2)
        except Exception:
            windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

import threading
import subprocess
import shutil
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk

try:
    from tkcalendar import DateEntry
except ImportError:
    DateEntry = None

# Configure CustomTkinter — light mode
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# Force PyInstaller to bundle dynamic imports
try:
    import config
    import odoo_downloader
except Exception:
    pass

# Venv python path (dev mode only — frozen uses sys.executable)
if os.name == 'nt':
    _venv_python_path = BASE_DIR / ".venv" / "Scripts" / "python.exe"
else:
    _venv_python_path = BASE_DIR / ".venv" / "bin" / "python"

_venv_python = (
    sys.executable if getattr(sys, "frozen", False)
    else str(_venv_python_path)
)

# ── Design System & Color Palette ─────────────────────────────────────────────
FONT_FAMILY = ("Segoe UI", "Segoe UI Emoji", "Arial") if IS_WINDOWS else "Helvetica Neue"
FONT_MONO   = "Consolas" if IS_WINDOWS else "Menlo"

# Theme Palette (Modern High-Contrast Clean Light)
BG          = "#F4F5F8"  # Cool neutral app background
PANEL       = "#FFFFFF"  # Pure white card background
SIDEBAR_BG  = "#FAFAFC"  # Soft off-white sidebar background
PREVIEW_BG  = "#F8FAFC"  # Table/Sub-card background
BORDER      = "#E2E8F0"  # Crisp subtle border color
BORDER_DARK = "#CBD5E1"  # Stronger border color for inputs

ACCENT      = "#6D28D9"  # Odoo Deep Violet
ACCENT_DARK = "#5B21B6"  # Hover Violet
SUCCESS     = "#059669"  # Vibrant Emerald
SUCCESS_DARK= "#047857"  # Hover Emerald
ERROR       = "#DC2626"  # Soft Crimson
ERROR_LIGHT = "#FEE2E2"
WARN        = "#D97706"  # Warm Amber
WARN_LIGHT  = "#FEF3C7"

TEXT        = "#0F172A"  # Deep slate text
MUTED       = "#64748B"  # Muted slate text
WHITE       = "#FFFFFF"

# CustomTkinter Color Mappings
CTK_FG      = ("#0F172A", "#0F172A")
CTK_ACCENT  = ACCENT
CTK_SUCCESS = SUCCESS
CTK_ERROR   = ERROR

BANK_BADGE_COLS = {
    "BCA":     ("#1E40AF", "#DBEAFE"),
    "MANDIRI": ("#92400E", "#FEF3C7"),
    "BRI":     ("#047857", "#D1FAE5"),
    "OTHER":   ("#4B5563", "#F3F4F6"),
}

TYPE_BADGE_INFO = {
    "bank_only":         ("🏦 Bank Only", "#4338CA", "#EEF2FF"),
    "odoo_only":         ("📦 Odoo Only", "#B45309", "#FEF3C7"),
    "unreconciled_odoo": ("⚠️ Unreconciled", "#BE123C", "#FFE4E6"),
}

# ── Font System & Custom TTF Font Loading ──────────────────────────────────────
fonts_dir = (
    Path(getattr(sys, "_MEIPASS", BASE_DIR)) / "assets" / "fonts"
    if getattr(sys, "frozen", False)
    else BASE_DIR / "assets" / "fonts"
)

if fonts_dir.exists():
    for f_file in fonts_dir.glob("*.ttf"):
        try:
            ctk.FontManager.load_font(str(f_file))
        except Exception:
            pass
    # Windows: also register fonts via GDI32's AddFontResourceExW so that
    # ClearType sub-pixel rendering applies to the custom TTF stems.
    # CTkFontManager alone uses Tk's font path which GDI treats as "foreign"
    # and falls back to grayscale AA — strokes appear thin and patchy.
    if IS_WINDOWS:
        try:
            import ctypes
            _gdi32 = ctypes.windll.gdi32
            _FR_PRIVATE = 0x10  # font unloads when the process exits
            for f_file in fonts_dir.glob("*.ttf"):
                _gdi32.AddFontResourceExW(str(f_file), _FR_PRIVATE, 0)
            # Notify the system that new fonts are available
            ctypes.windll.user32.SendMessageW(0xFFFF, 0x001D, 0, 0)
        except Exception:
            pass

FONT_FAMILY = "Space Grotesk"
FONT_BODY   = "IBM Plex Sans"
FONT_MONO   = "Consolas" if IS_WINDOWS else "Menlo"
# On Windows, tk.Label widgets with emoji must use Segoe UI Emoji explicitly
# because Space Grotesk has no emoji glyphs and GDI can't apply font-linking
# to non-GDI-registered fonts. CTkLabel/CTkButton are handled by AddFontResourceExW.
_EMOJI_FONT = "Segoe UI Emoji" if IS_WINDOWS else FONT_FAMILY


def _open_path(path: str):
    """Open a file/folder in the OS default app."""
    if IS_WINDOWS:
        os.startfile(path) # type: ignore
    elif sys.platform == "darwin":
        subprocess.run(["open", path])
    else:
        subprocess.run(["xdg-open", path])


class CTkDateInput(ctk.CTkFrame):
    def __init__(self, master, variable=None, default_date=None, **kwargs):
        super().__init__(
            master, fg_color=WHITE, border_color=BORDER_DARK, border_width=1,
            corner_radius=6, height=34, **kwargs
        )
        self.pack_propagate(False)
        self._var = variable or tk.StringVar()
        
        self._cal_icon = tk.Label(
            self, text="📅", bg=WHITE, fg=MUTED,
            font=(_EMOJI_FONT, 7), cursor="hand2"
        )
        self._cal_icon.pack(side="right", padx=(0, 4), pady=2)
        
        self._entry = ctk.CTkEntry(
            self, textvariable=self._var, placeholder_text="MM/DD/YYYY",
            height=28, border_width=0, fg_color="transparent", text_color=TEXT,
            font=(FONT_FAMILY, 10, "bold")
        )
        self._entry.pack(side="left", fill="x", expand=True, padx=(4, 0), pady=2)
        
        self._cal_icon.bind("<Button-1>", lambda e: self.open_calendar())
        self._entry.bind("<Button-1>", lambda e: self.open_calendar())
        self.bind("<Button-1>", lambda e: self.open_calendar())
        
        if default_date:
            self.set_date(default_date)
            
    def get(self):
        return self._var.get().strip()
        
    def set_date(self, d):
        s = d.strftime("%m/%d/%Y") if hasattr(d, "strftime") else str(d)
        self._var.set(s)
        self._entry.delete(0, "end")
        self._entry.insert(0, s)
            
    def open_calendar(self):
        try:
            import tkcalendar
            top = ctk.CTkToplevel(self)
            top.title("Select Date")
            top.geometry("260x220")
            top.configure(fg_color=PANEL)
            top.transient(self.winfo_toplevel())
            top.grab_set()
            
            x = self.winfo_rootx()
            y = self.winfo_rooty() + self.winfo_height() + 4
            top.geometry(f"+{x}+{y}")
            
            cal = tkcalendar.Calendar(
                top, selectmode="day", date_pattern="mm/dd/yyyy",
                background=ACCENT, foreground=WHITE, headersbackground=PREVIEW_BG,
                headersforeground=TEXT, selectbackground=ACCENT, selectforeground=WHITE,
                normalbackground=WHITE, normalforeground=TEXT
            )
            cal.pack(fill="both", expand=True, padx=8, pady=8)
            
            def _select():
                self.set_date(cal.get_date())
                top.destroy()
                
            cal.bind("<<CalendarSelected>>", lambda e: _select())
        except Exception:
            pass


def _maximize_window(win):
    """Maximize a Tk/CTk window cross-platform without taskbar clipping."""
    import sys
    win.update_idletasks()
    system = sys.platform

    if system == "win32":
        try:
            import ctypes
            from ctypes import wintypes
            rect = wintypes.RECT()
            # SPI_GETWORKAREA = 48 (returns work area excluding taskbar)
            if ctypes.windll.user32.SystemParametersInfoW(48, 0, ctypes.byref(rect), 0):
                work_x = rect.left
                work_y = rect.top
                work_w = rect.right - rect.left
                work_h = rect.bottom - rect.top
                # Subtract titlebar & window border height (~45px) so client height
                # doesn't push the window bottom underneath the taskbar
                client_h = max(500, work_h - 45)
                win.geometry(f"{work_w}x{client_h}+{work_x}+{work_y}")
                return
        except Exception:
            pass
        # Fallback for Windows if SPI fails: subtract 80px from screen height
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        win.geometry(f"{sw}x{max(500, sh - 80)}+0+0")
    elif system == "darwin":
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        # Mac: menu bar ~25px, dock ~70px
        win.geometry(f"{sw}x{max(500, sh - 95)}+0+25")
    else:
        # Linux / X11
        try:
            win.attributes("-zoomed", True)
        except Exception:
            sw = win.winfo_screenwidth()
            sh = win.winfo_screenheight()
            win.geometry(f"{sw}x{max(500, sh - 80)}+0+0")


def _center_modal_on_parent(win, parent):
    """Size and center modal dialog.
    Windows: always centers inside SPI_GETWORKAREA (maximized parent coords are unreliable
             due to invisible shadow/resize borders that skew winfo_rootx/y).
    Mac/Linux: centers relative to parent window coords."""
    import sys
    win.update_idletasks()
    parent.update_idletasks()

    if sys.platform == "win32":
        # Use the desktop work area (excludes taskbar) as the reference box
        try:
            import ctypes
            from ctypes import wintypes
            rect = wintypes.RECT()
            if ctypes.windll.user32.SystemParametersInfoW(48, 0, ctypes.byref(rect), 0):
                ref_x, ref_y = rect.left, rect.top
                ref_w = rect.right - rect.left
                ref_h = rect.bottom - rect.top
            else:
                raise RuntimeError("SPI failed")
        except Exception:
            ref_x, ref_y = 0, 0
            ref_w = win.winfo_screenwidth()
            ref_h = win.winfo_screenheight()
    else:
        ref_x = parent.winfo_rootx()
        ref_y = parent.winfo_rooty()
        ref_w = parent.winfo_width()
        ref_h = parent.winfo_height()
        if ref_w < 400 or ref_h < 300:
            ref_x, ref_y = 0, 0
            ref_w = win.winfo_screenwidth()
            ref_h = win.winfo_screenheight()

    target_w = max(1000, int(ref_w * 0.88))
    target_h = max(600, int(ref_h * 0.82))

    x = ref_x + (ref_w - target_w) // 2
    y = ref_y + (ref_h - target_h) // 2

    win.geometry(f"{target_w}x{target_h}+{max(0, x)}+{max(0, y)}")




class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Bank Reconciliation Studio")
        self.configure(fg_color=BG)
        # Apply DPI scaling BEFORE building UI so every widget is created at
        # the correct physical scale from the start. Applying it after causes
        # some widgets to render at default 96-DPI scale and others at the
        # real DPI — the mismatch makes fonts look patchy/inconsistent.
        if IS_WINDOWS:
            self._apply_dpi_scaling()
        self.geometry("1280x860")
        self.minsize(1100, 720)
        self.resizable(True, True)
        # Maximise on startup — platform-correct
        self.after(50, lambda: _maximize_window(self))
        self._running = False
        self._active_proc = None
        self._last_output = None
        self._set_app_icon()
        self._build_ui()
        self._center()
        self.protocol("WM_DELETE_WINDOW", self._on_app_close)
        self._auto_scan_after_id = self.after(200, self._auto_scan_on_startup)


    def _on_app_close(self):
        """Cleanly terminate all background processes and threads before destroying window.
        Prevents PyInstaller _MEI* cleanup warnings on Windows."""
        self._running = False
        
        proc = getattr(self, "_active_proc", None)
        if proc:
            try:
                proc.terminate()
                proc.kill()
            except Exception:
                pass
            self._active_proc = None

        try:
            self.destroy()
        except Exception:
            pass
        sys.exit(0)

    def _set_app_icon(self):

        """Set high-res native window icon.
        Windows: .ico with 16/24/32/48/64/128/256 px embedded → crisp at all DPI.
        Mac:     PhotoImage from PNG → used by iconphoto()."""
        try:
            assets = (
                Path(getattr(sys, "_MEIPASS", BASE_DIR)) / "assets"
                if getattr(sys, "frozen", False)
                else BASE_DIR / "assets"
            )
            if IS_WINDOWS:
                ico = assets / "app_icon.ico"
                if ico.exists():
                    self.iconbitmap(str(ico))
            else:
                png = assets / "app_icon.png"
                if png.exists():
                    from PIL import Image as _Img, ImageTk as _ImgTk
                    _icon = _ImgTk.PhotoImage(_Img.open(png).resize((256, 256)))
                    self.iconphoto(True, _icon)
                    self._icon_ref = _icon  # keep reference — GC will blank it otherwise
        except Exception:
            pass  # Never crash over a missing icon


    def _apply_dpi_scaling(self):
        """Query the real monitor DPI and correct Tkinter's scaling factor.
        Without this, Tkinter assumes 96 DPI — at 125%/150% display scaling
        all fonts render blurry because Windows bitmap-scales the GDI output.
        With the correct scaling factor, custom TTF fonts go through ClearType
        at their true physical pixel size and stay sharp."""
        try:
            import ctypes
            # Get the window handle after the window is realized
            self.update_idletasks()
            hwnd = ctypes.windll.user32.GetParent(self.winfo_id())
            if hwnd == 0:
                hwnd = self.winfo_id()
            # GetDpiForWindow is available on Windows 10+
            try:
                dpi = ctypes.windll.user32.GetDpiForWindow(hwnd)
            except AttributeError:
                # Fall back to system DPI on older Windows
                dpi = ctypes.windll.user32.GetDpiForSystem()
            if dpi and dpi > 0:
                scale = dpi / 96.0
                # Tell Tkinter the true physical scaling so it sizes fonts correctly
                self.tk.call("tk", "scaling", scale)
        except Exception:
            pass

    def _auto_scan_on_startup(self):
        self._refresh_folder_status()
        from config import MUTATION_DIR
        all_files = []
        for d in [INPUT_DIR, MUTATION_DIR]:
            if d.exists():
                all_files.extend([f for f in d.rglob("*") if f.is_file()])
        if all_files and not self._running:
            self._on_scan()

    # ── UI Construction ───────────────────────────────────────────────────────
    def _build_ui(self):
        # DateEntry wrapper compatibility fallback
        class _DateVar:
            def __init__(self, var):       self._v = var
            def get(self):                  return self._v.get()
            def get_date(self):
                from datetime import datetime as _dt
                try:    return _dt.strptime(self._v.get(), "%m/%d/%Y")
                except: return _dt.now()
            def set_date(self, d):
                self._v.set(d.strftime("%m/%d/%Y") if hasattr(d, "strftime") else str(d))

        # Load bank logos from assets folder
        try:
            from PIL import Image as _PILImage
            assets_dir = (
                Path(getattr(sys, "_MEIPASS", BASE_DIR)) / "assets"
                if getattr(sys, "frozen", False)
                else BASE_DIR / "assets"
            )
            
            def _get_ctk_logo(filename, max_w=70, max_h=22):
                p = assets_dir / filename
                if not p.exists():
                    return None
                img = _PILImage.open(p).convert("RGBA")
                w_orig, h_orig = img.size
                scale = min(max_w / w_orig, max_h / h_orig)
                w = max(1, int(w_orig * scale))
                h = max(1, int(h_orig * scale))
                
                target_size = (w * 3, h * 3)
                img_scaled = img.resize(target_size, _PILImage.Resampling.LANCZOS)
                return ctk.CTkImage(light_image=img_scaled, dark_image=img_scaled, size=(w, h))

            self._logos = {
                "All":     _get_ctk_logo("all_logo.png"),
                "BCA":     _get_ctk_logo("bca_logo.png"),
                "Mandiri": _get_ctk_logo("mandiri_logo.png"),
                "BRI":     _get_ctk_logo("bri_logo.png"),
            }
            _has_logos = any(v is not None for v in self._logos.values())
        except Exception as e:
            print("LOGO LOAD ERROR:", e)
            self._logos = {}
            _has_logos = False

        # ── Sidebar Container ─────────────────────────────────────────────────
        sidebar_outer = ctk.CTkFrame(
            self, width=290, corner_radius=0,
            fg_color=SIDEBAR_BG, border_color=BORDER, border_width=1
        )
        sidebar_outer.pack(side="left", fill="y")
        sidebar_outer.pack_propagate(False)

        # Brand header — Pinned top
        brand = ctk.CTkFrame(sidebar_outer, fg_color="transparent", height=78)
        brand.pack(fill="x")
        brand.pack_propagate(False)

        brand_inner = ctk.CTkFrame(brand, fg_color="transparent")
        brand_inner.pack(fill="both", expand=True, padx=16, pady=14)

        ctk.CTkLabel(
            brand_inner, text="🏦 Bank Recon",
            font=(FONT_FAMILY, 18, "bold"), text_color=ACCENT,
            fg_color="transparent"
        ).pack(anchor="w")
        ctk.CTkLabel(
            brand_inner, text="Odoo Automation Tool",
            font=(FONT_FAMILY, 11, "bold"), text_color=MUTED,
            fg_color="transparent"
        ).pack(anchor="w")

        
        ctk.CTkFrame(sidebar_outer, height=1, fg_color=BORDER).pack(fill="x")

        # Scrollable Sidebar Content
        scroll = ctk.CTkScrollableFrame(
            sidebar_outer, fg_color="transparent",
            scrollbar_button_color=BORDER,
            scrollbar_button_hover_color=MUTED
        )
        scroll.pack(fill="both", expand=True, padx=4, pady=4)

        # ── Credentials Section ───────────────────────────────────────────────
        sec_cred = ctk.CTkFrame(scroll, fg_color=PANEL, corner_radius=8, border_color=BORDER, border_width=1)
        sec_cred.pack(fill="x", padx=6, pady=(6, 8))
        
        ctk.CTkLabel(
            sec_cred, text="C R E D E N T I A L S", font=(FONT_FAMILY, 9, "bold"),
            text_color=MUTED, fg_color="transparent"
        ).pack(anchor="w", padx=10, pady=(8, 4))

        from config import PREDEFINED_ACCOUNTS
        self._predefined_accounts = PREDEFINED_ACCOUNTS

        self._email_var = tk.StringVar()
        self._password_var = tk.StringVar()

        if self._predefined_accounts:
            ctk.CTkLabel(
                sec_cred, text="Saved Account", font=(FONT_FAMILY, 10, "bold"),
                text_color=TEXT, fg_color="transparent"
            ).pack(anchor="w", padx=10, pady=(2, 0))

            preset_names = list(self._predefined_accounts.keys())
            first_key = preset_names[0]

            # Custom styled dropdown button matching CTk rounded inputs
            btn_preset = ctk.CTkButton(
                sec_cred,
                text=f"👤  {first_key}",
                anchor="w",
                height=34,
                corner_radius=6,
                border_width=1,
                border_color=BORDER_DARK,
                fg_color=WHITE,
                hover_color=PREVIEW_BG,
                text_color=TEXT,
                font=(FONT_FAMILY, 10, "bold"),
            )
            btn_preset.pack(fill="x", padx=10, pady=(2, 8))

            lbl_chevron = ctk.CTkLabel(
                btn_preset, text="▾", text_color=MUTED,
                font=(FONT_FAMILY, 11, "bold"), fg_color="transparent"
            )
            lbl_chevron.place(relx=1.0, rely=0.5, anchor="e", x=-10)

            menu_preset = tk.Menu(
                btn_preset, tearoff=0, bg=WHITE, fg=TEXT,
                activebackground=PREVIEW_BG, activeforeground=ACCENT,
                font=(FONT_FAMILY, 10, "bold"), bd=1, relief="solid"
            )

            def _select_account(name):
                if name in self._predefined_accounts:
                    acc = self._predefined_accounts[name]
                    self._email_var.set(acc["username"])
                    self._password_var.set(acc["password"])
                    btn_preset.configure(text=f"👤  {name}")
                elif name == "Custom":
                    self._email_var.set("")
                    self._password_var.set("")
                    btn_preset.configure(text="✏️  Custom (Manual)")

            for name in preset_names:
                menu_preset.add_command(
                    label=f"  👤  {name}  ",
                    command=lambda n=name: _select_account(n)
                )
            menu_preset.add_separator()
            menu_preset.add_command(
                label="  ✏️  Custom (Manual)  ",
                command=lambda: _select_account("Custom")
            )

            def _open_menu(event=None):
                try:
                    x = btn_preset.winfo_rootx()
                    y = btn_preset.winfo_rooty() + btn_preset.winfo_height() + 2
                    menu_preset.tk_popup(x, y)
                finally:
                    menu_preset.grab_release()

            btn_preset.configure(command=_open_menu)
            lbl_chevron.bind("<Button-1>", _open_menu)

            # Pre-fill first account
            first_acc = self._predefined_accounts[first_key]
            self._email_var.set(first_acc["username"])
            self._password_var.set(first_acc["password"])

        ctk.CTkLabel(
            sec_cred, text="Username", font=(FONT_FAMILY, 10, "bold"),
            text_color=TEXT, fg_color="transparent"
        ).pack(anchor="w", padx=10)
        
        ctk.CTkEntry(
            sec_cred, textvariable=self._email_var,
            placeholder_text="odoo@example.com",
            height=32, corner_radius=6, border_color=BORDER_DARK, fg_color=WHITE, text_color=TEXT,
            font=(FONT_FAMILY, 10, "bold")
        ).pack(fill="x", padx=10, pady=(2, 6))

        ctk.CTkLabel(
            sec_cred, text="Password", font=(FONT_FAMILY, 10, "bold"),
            text_color=TEXT, fg_color="transparent"
        ).pack(anchor="w", padx=10)
        
        pass_frame = ctk.CTkFrame(sec_cred, fg_color=WHITE, border_color=BORDER_DARK, border_width=1, corner_radius=6, height=34)
        pass_frame.pack(fill="x", padx=10, pady=(2, 10))
        pass_frame.pack_propagate(False)

        def _show_pass(event=None):
            self._password_entry.configure(show="")

        def _hide_pass(event=None):
            self._password_entry.configure(show="•")
            
        _btn_eye = tk.Label(
            pass_frame, text="👁", bg=WHITE, fg=MUTED,
            font=(FONT_FAMILY, 9, "bold"), cursor="hand2"
        )
        _btn_eye.pack(side="right", padx=(4, 10), pady=4)
        _btn_eye.bind("<ButtonPress-1>", _show_pass)
        _btn_eye.bind("<ButtonRelease-1>", _hide_pass)
        
        self._password_entry = ctk.CTkEntry(
            pass_frame, textvariable=self._password_var,
            placeholder_text="Password", show="•",
            height=32, border_width=0, fg_color="transparent", text_color=TEXT,
            font=(FONT_FAMILY, 10, "bold")
        )
        self._password_entry.pack(side="left", fill="x", expand=True, padx=(8, 4))

        # ── Bank Target Section ───────────────────────────────────────────────
        sec_bank = ctk.CTkFrame(scroll, fg_color=PANEL, corner_radius=8, border_color=BORDER, border_width=1)
        sec_bank.pack(fill="x", padx=6, pady=4)

        ctk.CTkLabel(
            sec_bank, text="B A N K   T A R G E T", font=(FONT_FAMILY, 9, "bold"),
            text_color=MUTED, fg_color="transparent"
        ).pack(anchor="w", padx=10, pady=(8, 6))

        self._bank_vars = {
            "BCA":     tk.BooleanVar(value=False),
            "Mandiri": tk.BooleanVar(value=False),
            "BRI":     tk.BooleanVar(value=False),
            "All":     tk.BooleanVar(value=True),
        }

        # Build dynamic subtitles from BANK_ACCOUNTS config
        try:
            from config import BANK_ACCOUNTS as _BA

            def _bank_subtitle(bank_key):
                accs = _BA.get(bank_key.lower(), {})
                if not accs:
                    return "Not configured"

                is_single_main = list(accs.keys()) == ["main"]

                if is_single_main:
                    # Detect supported payment types from AR keys
                    info = accs["main"]
                    types = []
                    if any("creditcard" in k for k in info):
                        types.append("Credit")
                    if any("debitcard" in k for k in info):
                        types.append("Debit")
                    if any("_qr" in k for k in info):
                        types.append("QR")
                    return "  ·  ".join(types) if types else "EDC"
                else:
                    # Multi-account → very short alias (≤3 chars, e.g. 'lbf') → UPPER, longer → Title
                    parts = [a.upper() if len(a) <= 3 else a.capitalize() for a in accs.keys()]
                    return "  ·  ".join(parts)

            _bank_sub = {
                "All":     "All configured banks",
                "BCA":     _bank_subtitle("BCA"),
                "Mandiri": _bank_subtitle("Mandiri"),
                "BRI":     _bank_subtitle("BRI"),
            }
        except Exception:
            _bank_sub = {"All": "All configured banks", "BCA": "Credit Card · Debit · QR",
                         "Mandiri": "Debit Card · QR", "BRI": "LBF · Frans · Nara"}


        # Brand accent strip colors per bank
        _strip_col  = {"All": ACCENT,      "BCA": "#0066AE", "Mandiri": "#F0A500", "BRI": "#004B87"}
        _card_refs: dict = {}   # bname → (outer_frame, strip_frame, name_lbl, sub_lbl, check_lbl)

        def _on_bank_toggle(name):
            if name == "All" and self._bank_vars["All"].get():
                for b in ["BCA", "Mandiri", "BRI"]:
                    self._bank_vars[b].set(False)
            elif name in ["BCA", "Mandiri", "BRI"] and self._bank_vars[name].get():
                self._bank_vars["All"].set(False)
            _refresh_bank_cards()

        def _refresh_bank_cards():
            for bname, refs in _card_refs.items():
                outer, strip, name_lbl, sub_lbl, check_lbl = refs
                sel     = self._bank_vars[bname].get()
                strip_c = _strip_col[bname]
                outer.configure(
                    fg_color=WHITE if sel else PREVIEW_BG,
                    border_color=strip_c if sel else BORDER,
                    border_width=2 if sel else 1,
                )
                strip.configure(fg_color=strip_c if sel else BORDER_DARK)
                name_lbl.configure(text_color=strip_c if sel else TEXT)
                sub_lbl.configure(text_color=MUTED)
                check_lbl.configure(text="✓" if sel else "", text_color=strip_c)


        bank_grid = ctk.CTkFrame(sec_bank, fg_color="transparent")
        bank_grid.pack(fill="x", padx=8, pady=(0, 10))
        bank_grid.columnconfigure(0, weight=1)
        bank_grid.columnconfigure(1, weight=1)

        for bname, grow, gcol in [("All", 0, 0), ("BCA", 0, 1), ("Mandiri", 1, 0), ("BRI", 1, 1)]:
            sel     = self._bank_vars[bname].get()
            strip_c = _strip_col[bname]

            def _make_cmd(n=bname):
                def _cmd():
                    self._bank_vars[n].set(not self._bank_vars[n].get())
                    _on_bank_toggle(n)
                return _cmd

            # Outer card — 2-line with accent strip
            outer = ctk.CTkFrame(
                bank_grid, height=50, corner_radius=7,
                fg_color=WHITE if sel else PREVIEW_BG,
                border_color=strip_c if sel else BORDER,
                border_width=2 if sel else 1,
                cursor="hand2",
            )
            outer.grid(row=grow, column=gcol, padx=3, pady=3, sticky="ew")
            outer.pack_propagate(False)
            outer.bind("<Button-1>", lambda e, n=bname: _make_cmd(n)())

            # Left accent strip (4 px)
            strip = ctk.CTkFrame(outer, width=4, corner_radius=0,
                                 fg_color=strip_c if sel else BORDER_DARK)
            strip.pack(side="left", fill="y")
            strip.pack_propagate(False)
            strip.bind("<Button-1>", lambda e, n=bname: _make_cmd(n)())

            # Text block
            txt = ctk.CTkFrame(outer, fg_color="transparent")
            txt.pack(side="left", fill="both", expand=True, padx=(7, 2), pady=5)
            txt.bind("<Button-1>", lambda e, n=bname: _make_cmd(n)())

            name_lbl = ctk.CTkLabel(
                txt, text=bname if bname != "All" else "All Banks",
                font=(FONT_FAMILY, 10, "bold"),
                text_color=strip_c if sel else TEXT,
                fg_color="transparent", anchor="w",
            )
            name_lbl.pack(anchor="w")
            name_lbl.bind("<Button-1>", lambda e, n=bname: _make_cmd(n)())

            sub_lbl = ctk.CTkLabel(
                txt, text=_bank_sub[bname],
                font=(FONT_FAMILY, 8), text_color=MUTED,
                fg_color="transparent", anchor="w",
            )
            sub_lbl.pack(anchor="w")
            sub_lbl.bind("<Button-1>", lambda e, n=bname: _make_cmd(n)())

            # Checkmark badge
            check_lbl = ctk.CTkLabel(
                outer, text="✓" if sel else "",
                font=(FONT_FAMILY, 11, "bold"),
                text_color=strip_c, fg_color="transparent", width=20,
            )
            check_lbl.pack(side="right", padx=(0, 8))
            check_lbl.bind("<Button-1>", lambda e, n=bname: _make_cmd(n)())

            _card_refs[bname] = (outer, strip, name_lbl, sub_lbl, check_lbl)

        _refresh_bank_cards()



        # ── Date Range Section ────────────────────────────────────────────────
        sec_date = ctk.CTkFrame(scroll, fg_color=PANEL, corner_radius=8, border_color=BORDER, border_width=1)
        sec_date.pack(fill="x", padx=6, pady=4)

        ctk.CTkLabel(
            sec_date, text="D A T E   R A N G E", font=(FONT_FAMILY, 9, "bold"),
            text_color=MUTED, fg_color="transparent"
        ).pack(anchor="w", padx=10, pady=(8, 4))

        yesterday = datetime.now() - timedelta(days=1)
        import tkcalendar
        
        self._date_from_var = tk.StringVar()
        self._date_to_var = tk.StringVar()

        date_grid = ctk.CTkFrame(sec_date, fg_color="transparent")
        date_grid.pack(fill="x", padx=8, pady=(0, 8))
        date_grid.columnconfigure(0, weight=1)
        date_grid.columnconfigure(1, weight=1)
        
        ctk.CTkLabel(date_grid, text="From", font=(FONT_FAMILY, 9, "bold"), text_color=MUTED,
                     fg_color="transparent").grid(row=0, column=0, sticky="w", padx=2)
        ctk.CTkLabel(date_grid, text="To", font=(FONT_FAMILY, 9, "bold"), text_color=MUTED,
                     fg_color="transparent").grid(row=0, column=1, sticky="w", padx=2)
                     
        self._date_from_widget = CTkDateInput(
            date_grid, variable=self._date_from_var, default_date=yesterday
        )
        self._date_from_widget.grid(row=1, column=0, sticky="ew", padx=(0, 3))
        
        self._date_to_widget = CTkDateInput(
            date_grid, variable=self._date_to_var, default_date=yesterday
        )
        self._date_to_widget.grid(row=1, column=1, sticky="ew", padx=(3, 0))

        # ── Quick Folder Links Section ────────────────────────────────────────
        sec_links = ctk.CTkFrame(scroll, fg_color=PANEL, corner_radius=8, border_color=BORDER, border_width=1)
        sec_links.pack(fill="x", padx=6, pady=4)

        ctk.CTkLabel(
            sec_links, text="Q U I C K   A C C E S S", font=(FONT_FAMILY, 9, "bold"),
            text_color=MUTED, fg_color="transparent"
        ).pack(anchor="w", padx=10, pady=(6, 2))

        def _link_btn(parent, icon, label_text, cmd):
            row = ctk.CTkFrame(parent, fg_color="transparent", height=28)
            row.pack(fill="x", padx=4, pady=1)
            row.pack_propagate(False)

            ico_font = ("Segoe UI Emoji", 10) if IS_WINDOWS else (FONT_FAMILY, 10)
            lbl_ico = tk.Label(
                row, text=icon, font=ico_font,
                bg=PANEL, fg=MUTED, width=3, anchor="center"
            )
            lbl_ico.pack(side="left", padx=(4, 2))
            lbl_ico.bind("<Button-1>", lambda e: cmd())

            btn = ctk.CTkButton(
                row, text=label_text, height=28,
                fg_color="transparent", hover_color=PREVIEW_BG,
                text_color=ACCENT, font=(FONT_FAMILY, 10, "bold"),
                anchor="w", command=cmd
            )
            btn.pack(side="left", fill="both", expand=True)
            return btn

        self._qa_result_btn   = _link_btn(sec_links, "📊", "Open Result", self._open_output)
        self._qa_pdf_btn      = _link_btn(sec_links, "📄", "Summary PDF", self._on_export_summary_pdf)
        self._qa_merchant_btn = _link_btn(sec_links, "📂", "Open Merchant", self._open_input)
        self._qa_mutation_btn = _link_btn(sec_links, "📁", "Open Mutation", self._open_mutation)
        self._qa_payment_btn  = _link_btn(sec_links, "💳", "Open Payment", self._open_odoo_file)
        self._qa_journal_btn  = _link_btn(sec_links, "📑", "Open Journal Entries", self._open_journal_file)
        self._qa_recap_btn    = _link_btn(sec_links, "🗄", "Open Recap", self._open_recap)
        self._open_btn_sidebar = self._qa_result_btn

        # ── Primary CTA Action Stack — Pinned Bottom ─────────────────────────
        ctk.CTkFrame(sidebar_outer, height=1, fg_color=BORDER).pack(fill="x")
        _cta = ctk.CTkFrame(sidebar_outer, fg_color="transparent")
        _cta.pack(fill="x", padx=12, pady=12)

        self._offline_var = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            _cta, text="Offline Mode (Skip Downloader)",
            variable=self._offline_var,
            font=(FONT_FAMILY, 9, "bold"),
            fg_color=ACCENT, hover_color=ACCENT_DARK,
            text_color=TEXT, checkbox_width=16, checkbox_height=16
        ).pack(anchor="w", padx=2, pady=(0, 8))

        self._run_btn = ctk.CTkButton(
            _cta, text="⚡  Run Reconciliation",
            height=40, fg_color=WHITE, hover_color="#F8FAFC",
            border_color=BORDER_DARK, border_width=1,
            text_color=TEXT, font=(FONT_FAMILY, 11, "bold"),
            corner_radius=8, command=self._on_run
        )
        self._run_btn.pack(fill="x", pady=(0, 6))

        self._match_btn = ctk.CTkButton(
            _cta, text="🧩  Manual Match",
            height=40, fg_color=WHITE, hover_color="#F8FAFC",
            border_color=BORDER_DARK, border_width=1,
            text_color=ACCENT, font=(FONT_FAMILY, 11, "bold"),
            corner_radius=8, command=self._on_manual_match
        )
        self._match_btn.pack(fill="x", pady=(0, 6))

        self._sync_cloud_btn = ctk.CTkButton(
            _cta, text="☁️  Sync Sales Portal",
            height=40, fg_color=WHITE, hover_color="#F8FAFC",
            border_color=BORDER_DARK, border_width=1,
            text_color=TEXT, font=(FONT_FAMILY, 11, "bold"),
            corner_radius=8, command=self._on_sync_cloud
        )
        self._sync_cloud_btn.pack(fill="x", pady=(0, 6))

        self._journal_btn = ctk.CTkButton(
            _cta, text="📋  Generate Journal",
            height=42, fg_color=ACCENT, hover_color=ACCENT_DARK,
            text_color=WHITE, font=(FONT_FAMILY, 11, "bold"),
            corner_radius=8, command=self._on_journal
        )
        self._journal_btn.pack(fill="x")

        self._stop_btn = ctk.CTkButton(
            _cta, text="⏹  Stop Process",
            height=38, fg_color=ERROR, hover_color="#B91C1C",
            text_color=WHITE, font=(FONT_FAMILY, 11, "bold"),
            corner_radius=8, command=self._on_stop
        )

        # ── Main Content Area ─────────────────────────────────────────────────
        main_area = ctk.CTkFrame(self, fg_color=BG, corner_radius=0)
        main_area.pack(side="left", fill="both", expand=True)

        # ── Top Header Toolbar (Input Folder Status) ──────────────────────────
        topbar = ctk.CTkFrame(main_area, fg_color=SIDEBAR_BG, corner_radius=0, height=78)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)
        
        tb_wrap = ctk.CTkFrame(topbar, fg_color="transparent")
        tb_wrap.pack(fill="both", expand=True, padx=20, pady=22)

        self._folder_label = ctk.CTkLabel(
            tb_wrap, text="",
            font=(FONT_FAMILY, 11, "bold"), text_color=SUCCESS,
            fg_color="transparent", anchor="w"
        )
        self._folder_label.pack(side="left", anchor="w")
        self._update_dashboard_summary(skip_drill=True)  # fast init; drill fires after startup

        # ── Action Buttons & Console Log Area ─────────────────────────────────
        card_wrap = ctk.CTkFrame(main_area, fg_color="transparent")
        card_wrap.pack(fill="both", expand=True, padx=16, pady=14)

        # ── Action Buttons Toolbar ───────────────────────────────────────────
        action_bar = ctk.CTkFrame(card_wrap, fg_color="transparent")
        action_bar.pack(fill="x", pady=(0, 10))

        def _sec_btn(parent, text, cmd, color=TEXT, hover=PREVIEW_BG):
            return ctk.CTkButton(
                parent, text=text, height=34,
                fg_color=PANEL, hover_color=hover,
                border_color=BORDER_DARK, border_width=1,
                text_color=color, font=(FONT_FAMILY, 10, "bold"),
                corner_radius=6, command=cmd
            )

        self._upload_btn   = _sec_btn(action_bar, "⬆ Upload", self._on_upload)
        self._upload_btn.pack(side="left", padx=(0, 6))
        self._download_btn = _sec_btn(action_bar, "⬇ Download", self._on_download)
        self._download_btn.pack(side="left", padx=(0, 6))
        self._scan_btn     = _sec_btn(action_bar, "🔍 Scan", self._on_scan)
        self._scan_btn.pack(side="left", padx=(0, 6))
        self._pdf_btn      = _sec_btn(action_bar, "📑 Summary PDF", self._on_export_summary_pdf)
        self._pdf_btn.pack(side="left", padx=(0, 6))
        self._cleanse_btn  = _sec_btn(action_bar, "🗑 Clean", self._on_cleanse, color=ERROR, hover=ERROR_LIGHT)
        self._cleanse_btn.pack(side="left")

        # ── Live Input Summary Dashboard ──────────────────────────────────────
        dash_card = ctk.CTkFrame(card_wrap, fg_color=PANEL, corner_radius=10, border_color=BORDER, border_width=1)
        dash_card.pack(fill="x", pady=(0, 10))

        dash_hdr = ctk.CTkFrame(dash_card, fg_color="transparent")
        dash_hdr.pack(fill="x", padx=16, pady=(10, 6))

        ctk.CTkLabel(
            dash_hdr, text="📊 Live Input Data Summary",
            font=(FONT_FAMILY, 13, "bold"), text_color=TEXT,
            fg_color="transparent"
        ).pack(side="left")

        self._dash_last_update = ctk.CTkLabel(
            dash_hdr, text="Updated: —",
            font=(FONT_FAMILY, 10, "bold"), text_color=MUTED,
            fg_color="transparent"
        )
        self._dash_last_update.pack(side="right")

        ctk.CTkFrame(dash_card, height=1, fg_color=BORDER).pack(fill="x", padx=16)

        kpi_grid = ctk.CTkFrame(dash_card, fg_color="transparent")
        kpi_grid.pack(fill="x", padx=12, pady=(8, 8))
        kpi_grid.rowconfigure(0, weight=1)
        for c in range(4):
            kpi_grid.columnconfigure(c, weight=1, uniform="kpi_col")

        def _make_kpi_card(parent, col, icon, title, val_attr, sub_attr, click_cmd=None):
            card = ctk.CTkFrame(parent, fg_color=PREVIEW_BG, corner_radius=8, border_color=BORDER, border_width=1, height=96)
            card.grid(row=0, column=col, padx=4, pady=2, sticky="nsew")
            card.pack_propagate(False)
            
            top_f = ctk.CTkFrame(card, fg_color="transparent")
            top_f.pack(fill="x", padx=10, pady=(8, 2))
            
            ico_font = ("Segoe UI Emoji", 13) if IS_WINDOWS else (FONT_FAMILY, 13)
            ico_lbl = tk.Label(top_f, text=icon, font=ico_font, bg=PREVIEW_BG, fg=ACCENT)
            ico_lbl.pack(side="left")

            t_lbl = ctk.CTkLabel(top_f, text=title, font=(FONT_FAMILY, 10, "bold"), text_color=MUTED)
            t_lbl.pack(side="left", padx=(4, 0))

            val_lbl = tk.Label(card, text="0 Files", font=(FONT_FAMILY, 13, "bold"), fg=TEXT, bg=PREVIEW_BG, anchor="w")
            val_lbl.pack(fill="x", padx=10, pady=(2, 1))
            setattr(self, val_attr, val_lbl)

            sub_lbl = tk.Label(card, text="—", font=(FONT_FAMILY, 10, "bold"), fg=MUTED, bg=PREVIEW_BG, anchor="w")
            sub_lbl.pack(fill="x", padx=10, pady=(1, 6))
            setattr(self, sub_attr, sub_lbl)

            if click_cmd:
                for widget in (card, top_f, ico_lbl, t_lbl, val_lbl, sub_lbl):
                    try:
                        widget.bind("<Button-1>", lambda e, cmd=click_cmd: cmd())
                        widget.configure(cursor="hand2")
                    except Exception:
                        pass

        _make_kpi_card(kpi_grid, 0, "🏦", "MERCHANT REPORT", "_kpi_bank_val", "_kpi_bank_sub", click_cmd=self._open_input)
        _make_kpi_card(kpi_grid, 1, "💳", "ODOO PAYMENTS",   "_kpi_odoo_val", "_kpi_odoo_sub", click_cmd=self._open_odoo_file)
        _make_kpi_card(kpi_grid, 2, "📊", "MUTATIONS & FEES", "_kpi_mut_val", "_kpi_mut_sub", click_cmd=self._open_mutation)
        _make_kpi_card(kpi_grid, 3, "⚡", "ENGINE STATUS",   "_kpi_eng_val",  "_kpi_eng_sub",  click_cmd=self._open_output)

        # ── Executive Reconciliation & Settlement Insights ───────────────────
        ctk.CTkFrame(dash_card, height=1, fg_color=BORDER).pack(fill="x", padx=16, pady=(4, 6))
        
        ins_hdr = ctk.CTkFrame(dash_card, fg_color="transparent")
        ins_hdr.pack(fill="x", padx=16, pady=(0, 4))
        ctk.CTkLabel(
            ins_hdr, text="🎯 Reconciliation & Accounting Insights",
            font=(FONT_FAMILY, 12, "bold"), text_color=TEXT,
            fg_color="transparent"
        ).pack(side="left")

        ins_grid = ctk.CTkFrame(dash_card, fg_color="transparent")
        ins_grid.pack(fill="x", padx=12, pady=(0, 10))
        ins_grid.rowconfigure(0, weight=1)
        for c in range(4):
            ins_grid.columnconfigure(c, weight=1, uniform="kpi_col")

        def _make_insight_card(parent, col, icon, title, val_attr, sub_attr, click_cmd=None):
            card = ctk.CTkFrame(parent, fg_color="#F8FAFC", corner_radius=8, border_color=BORDER, border_width=1, height=96)
            card.grid(row=0, column=col, padx=4, pady=2, sticky="nsew")
            card.pack_propagate(False)

            top_f = ctk.CTkFrame(card, fg_color="transparent")
            top_f.pack(fill="x", padx=10, pady=(8, 2))

            ico_font = ("Segoe UI Emoji", 13) if IS_WINDOWS else (FONT_FAMILY, 13)
            ico_lbl = tk.Label(top_f, text=icon, font=ico_font, bg="#F8FAFC", fg=ACCENT)
            ico_lbl.pack(side="left")

            t_lbl = ctk.CTkLabel(top_f, text=title, font=(FONT_FAMILY, 10, "bold"), text_color=MUTED)
            t_lbl.pack(side="left", padx=(4, 0))

            val_lbl = tk.Label(card, text="Checking...", font=(FONT_FAMILY, 13, "bold"), fg=TEXT, bg="#F8FAFC", anchor="w")
            val_lbl.pack(fill="x", padx=10, pady=(2, 1))
            setattr(self, val_attr, val_lbl)

            sub_lbl = tk.Label(card, text="—", font=(FONT_FAMILY, 10, "bold"), fg=MUTED, bg="#F8FAFC", anchor="w")
            sub_lbl.pack(fill="x", padx=10, pady=(1, 6))
            setattr(self, sub_attr, sub_lbl)

            if click_cmd:
                for widget in (card, top_f, ico_lbl, t_lbl, val_lbl, sub_lbl):
                    try:
                        widget.bind("<Button-1>", lambda e, cmd=click_cmd: cmd())
                        widget.configure(cursor="hand2")
                    except Exception:
                        pass

        _make_insight_card(ins_grid, 0, "🎯", "RECON MATCH HEALTH", "_ins_health_val", "_ins_health_sub", click_cmd=self._open_output)
        _make_insight_card(ins_grid, 1, "📑", "SETTLEMENT JOURNALS", "_ins_jrn_val", "_ins_jrn_sub", click_cmd=self._on_journal)
        _make_insight_card(ins_grid, 2, "🏪", "SALES PORTAL TICKETS", "_ins_sales_val", "_ins_sales_sub", click_cmd=self._on_sync_cloud)
        _make_insight_card(ins_grid, 3, "📋", "REPORT COVERAGE",    "_ins_cov_val",    "_ins_cov_sub",    click_cmd=self._open_output)

        # ── Per-Account Drill-Down Panel ─────────────────────────────────────
        drill_card = ctk.CTkFrame(card_wrap, fg_color=PANEL, corner_radius=10, border_color=BORDER, border_width=1)
        drill_card.pack(fill="x", pady=(0, 10))

        drill_hdr = ctk.CTkFrame(drill_card, fg_color="transparent")
        drill_hdr.pack(fill="x", padx=16, pady=(10, 6))

        ctk.CTkLabel(
            drill_hdr, text="🔍 Per-Account Date Coverage",
            font=(FONT_FAMILY, 13, "bold"), text_color=TEXT,
            fg_color="transparent"
        ).pack(side="left")

        self._drill_status_lbl = ctk.CTkLabel(
            drill_hdr, text="",
            font=(FONT_FAMILY, 10, "bold"), text_color=MUTED,
            fg_color="transparent"
        )
        self._drill_status_lbl.pack(side="right", padx=(0, 4))

        ctk.CTkFrame(drill_card, height=1, fg_color=BORDER).pack(fill="x", padx=16)

        # Grid header row with compact close column spacing
        drill_grid = ctk.CTkFrame(drill_card, fg_color="transparent")
        drill_grid.pack(fill="x", padx=16, pady=(6, 8))
        drill_grid.columnconfigure(0, weight=0, minsize=140)
        drill_grid.columnconfigure(1, weight=0, minsize=240)
        drill_grid.columnconfigure(2, weight=0, minsize=240)
        drill_grid.columnconfigure(3, weight=1)

        for ci, h in enumerate(["ACCOUNT / ALIAS", "MERCHANT REPORT", "MUTATION FILE"]):
            ctk.CTkLabel(
                drill_grid, text=h,
                font=(FONT_FAMILY, 10, "bold"), text_color=MUTED,
                fg_color="transparent", anchor="w"
            ).grid(row=0, column=ci, padx=(0, 20), pady=(2, 4), sticky="w")

        ctk.CTkFrame(drill_grid, height=1, fg_color=BORDER).grid(
            row=1, column=0, columnspan=4, sticky="ew", padx=(0, 0), pady=(2, 6)
        )

        # Placeholder rows (will be populated by _update_drill_rows)
        self._drill_grid  = drill_grid
        self._drill_rows  = []  # list of (account_lbl, stmt_lbl, mut_lbl)

        # Console Log Card
        log_card = ctk.CTkFrame(card_wrap, fg_color=PANEL, corner_radius=10, border_color=BORDER, border_width=1)
        log_card.pack(fill="both", expand=True)

        log_hdr = ctk.CTkFrame(log_card, fg_color="transparent")
        log_hdr.pack(fill="x", padx=16, pady=(12, 6))
        
        ctk.CTkLabel(
            log_hdr, text="Logs",
            font=(FONT_FAMILY, 13, "bold"), text_color=TEXT,
            fg_color="transparent"
        ).pack(side="left")

        self._clear_log_btn = ctk.CTkButton(
            log_hdr, text="↻ Clear Log", height=26, width=70,
            fg_color="transparent", hover_color=PREVIEW_BG,
            border_color=BORDER_DARK, border_width=1,
            text_color=MUTED, font=(FONT_FAMILY, 10, "bold"),
            corner_radius=5, command=self._clear_log
        )
        self._clear_log_btn.pack(side="right")
        
        ctk.CTkFrame(log_card, height=1, fg_color=BORDER).pack(fill="x", padx=16)

        log_frame = ctk.CTkFrame(log_card, fg_color="transparent")
        log_frame.pack(fill="both", expand=True, padx=10, pady=(6, 10))

        self._log = tk.Text(
            log_frame, bg=PANEL, fg=TEXT,
            font=(FONT_MONO, 10, "bold"), relief="flat", state="disabled",
            wrap="word", borderwidth=0, highlightthickness=0,
            padx=10, pady=10, insertbackground=TEXT
        )
        _log_scroll = ctk.CTkScrollbar(
            log_frame, command=self._log.yview,
            button_color=BORDER,
            button_hover_color=MUTED,
            corner_radius=4, width=12,
        )
        self._log.configure(yscrollcommand=_log_scroll.set)
        _log_scroll.pack(side="right", fill="y", padx=(0, 2), pady=4)
        self._log.pack(fill="both", expand=True)

        self._log.tag_config("ok",   foreground=SUCCESS)
        self._log.tag_config("err",  foreground=ERROR)
        self._log.tag_config("warn", foreground=WARN)
        self._log.tag_config("head", foreground=ACCENT, font=(FONT_MONO, 9, "bold"))
        self._log.tag_config("dim",  foreground=MUTED)

        # ── Integrated Status Bar ─────────────────────────────────────────────
        statusbar = ctk.CTkFrame(log_card, fg_color=PREVIEW_BG, corner_radius=0, border_color=BORDER, border_width=1)
        statusbar.pack(fill="x", side="bottom")
        
        st = ctk.CTkFrame(statusbar, fg_color="transparent")
        st.pack(fill="x", padx=16, pady=6)
        
        self._dot = ctk.CTkLabel(
            st, text="●", width=14,
            font=(FONT_FAMILY, 12, "bold"), text_color=SUCCESS,
            fg_color="transparent"
        )
        self._dot.pack(side="left")
        
        self._status_var = tk.StringVar(value="Ready")
        ctk.CTkLabel(
            st, textvariable=self._status_var,
            font=(FONT_FAMILY, 10, "bold"), text_color=TEXT,
            fg_color="transparent"
        ).pack(side="left", padx=(4, 0))

        # Alias for completion callback compatibility
        self._open_btn = self._open_btn_sidebar

    def _center(self):
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w,  h  = self.winfo_width(),       self.winfo_height()
        self.geometry(f"+{(sw-w)//2}+{(sh-h)//2}")

    def _refresh_folder_status(self):
        files = [
            f for f in (INPUT_DIR.rglob("*") if INPUT_DIR.exists() else [])
            if f.is_file() and not f.name.startswith(".") and not f.name.startswith("~$")
        ]
        n = len(files)
        color = SUCCESS if n > 0 else WARN
        self._folder_label.configure(
            text=f"📂 Input folder: {INPUT_DIR}   ({n} file{'s' if n != 1 else ''} found)",
            text_color=color,
        )
        self._update_dashboard_summary()

    def _update_dashboard_summary(self, skip_drill: bool = False):
        def _set_lbl(lbl, text, color):
            if hasattr(lbl, "config"):
                try:
                    lbl.config(text=text, fg=color)
                    return
                except Exception:
                    pass
            if hasattr(lbl, "configure"):
                try:
                    lbl.configure(text=text, text_color=color)
                except Exception:
                    pass

        try:
            def _bank_files(bank_dir):
                if not bank_dir.exists():
                    return []
                return [
                    f for f in bank_dir.rglob("*")
                    if f.is_file()
                    and not f.name.startswith(".")
                    and not f.name.startswith("~$")
                ]
            # 1. Bank Files
            c_bca = _bank_files(INPUT_DIR / "bca")
            c_man = _bank_files(INPUT_DIR / "mandiri")
            c_bri = _bank_files(INPUT_DIR / "bri")
            
            tot_bank = len(c_bca) + len(c_man) + len(c_bri)
            if hasattr(self, "_kpi_bank_val"):
                _set_lbl(self._kpi_bank_val, f"{tot_bank} File{'s' if tot_bank!=1 else ''}", SUCCESS if tot_bank > 0 else TEXT)
                parts = []
                if c_bca: parts.append(f"BCA: {len(c_bca)}")
                if c_man: parts.append(f"Mandiri: {len(c_man)}")
                if c_bri: parts.append(f"BRI: {len(c_bri)}")
                _set_lbl(self._kpi_bank_sub, " | ".join(parts) if parts else "No merchant files", MUTED)

            # 2. Odoo Payments File
            from config import ODO_EXCEL_PATH
            if hasattr(self, "_kpi_odoo_val"):
                if ODO_EXCEL_PATH.exists():
                    sz = ODO_EXCEL_PATH.stat().st_size / 1024
                    _set_lbl(self._kpi_odoo_val, "Ready", SUCCESS)
                    _set_lbl(self._kpi_odoo_sub, f"{ODO_EXCEL_PATH.name} ({sz:.1f} KB)", MUTED)
                else:
                    _set_lbl(self._kpi_odoo_val, "Not Found", WARN)
                    _set_lbl(self._kpi_odoo_sub, "Will download via Odoo", MUTED)

            # 3. Mutations & Fees
            from config import MUTATION_DIR
            mut_by_bank = {"bca": 0, "mandiri": 0, "bri": 0}
            for bank in ["bca", "mandiri", "bri"]:
                bd = MUTATION_DIR / bank
                if bd.exists():
                    mut_by_bank[bank] = sum(1 for f in bd.rglob("*.csv") if f.is_file())
            tot_mut = sum(mut_by_bank.values())
            if hasattr(self, "_kpi_mut_val"):
                if tot_mut > 0:
                    _set_lbl(self._kpi_mut_val, f"{tot_mut} CSV File{'s' if tot_mut!=1 else ''}", SUCCESS)
                    mut_parts = []
                    if mut_by_bank["bca"]:     mut_parts.append(f"BCA: {mut_by_bank['bca']}")
                    if mut_by_bank["mandiri"]: mut_parts.append(f"Mandiri: {mut_by_bank['mandiri']}")
                    if mut_by_bank["bri"]:     mut_parts.append(f"BRI: {mut_by_bank['bri']}")
                    _set_lbl(self._kpi_mut_sub, " | ".join(mut_parts) if mut_parts else "Mutation files", MUTED)
                else:
                    _set_lbl(self._kpi_mut_val, "None Loaded", MUTED)
                    _set_lbl(self._kpi_mut_sub, "No mutation CSV files found", MUTED)

            # Drill-down: per-account date coverage (run in background thread)
            if hasattr(self, "_drill_grid") and not skip_drill:
                self._start_drill_update(MUTATION_DIR)

            # 4. Recon Report Details
            import glob, openpyxl
            # 4. Engine Status
            if hasattr(self, "_kpi_eng_val"):
                if getattr(self, "_running", False):
                    _set_lbl(self._kpi_eng_val, "Running...", ACCENT)
                    _set_lbl(self._kpi_eng_sub, "Reconciling data...", MUTED)
                else:
                    _set_lbl(self._kpi_eng_val, "Ready", SUCCESS)
                    output_files = glob.glob(str(OUTPUT_DIR / "[Rr]econciliation_*.xlsx"))
                    if output_files:
                        latest = max(output_files, key=os.path.getmtime)
                        mtime = datetime.fromtimestamp(os.path.getmtime(latest)).strftime("%d/%m %H:%M")
                        _set_lbl(self._kpi_eng_sub, f"Last: {mtime}", MUTED)
                    else:
                        _set_lbl(self._kpi_eng_sub, "Engine idle", MUTED)

            if hasattr(self, "_dash_last_update"):
                _set_lbl(self._dash_last_update, f"Updated: {datetime.now().strftime('%H:%M:%S')}", MUTED)

            # ── Async Insights Extractor ──────────────────────────────────────
            def _fetch_insights():
                health_text = "No report yet"
                health_sub = "Run reconciliation to view match health"
                health_color = MUTED

                jrn_text = "No journal data"
                jrn_sub = "Download settlement journals from Odoo"
                jrn_color = MUTED

                cov_text = "No report"
                cov_sub = "Reconciliation output missing"
                cov_color = MUTED

                import glob, openpyxl
                output_files = glob.glob(str(OUTPUT_DIR / "[Rr]econciliation_*.xlsx"))
                if output_files:
                    latest = max(output_files, key=os.path.getmtime)
                    try:
                        wb = openpyxl.load_workbook(latest, read_only=True, data_only=True)
                        
                        complete_keys = set()
                        incomplete_count = 0
                        settlement_total = 0
                        settlement_resolved = 0
                        b_set = set()
                        d_list = []

                        if "Daily Summary" in wb.sheetnames:
                            ws_ds = wb["Daily Summary"]
                            for row in ws_ds.iter_rows(min_row=4, values_only=True):
                                if not row or not row[0]: continue
                                d_str = str(row[1] or "").strip()
                                p_d = str(row[2] or "").strip()
                                b_n = str(row[3] or "").strip()
                                j_n = str(row[4] or b_n).strip()
                                status = str(row[9] or "").strip()

                                if b_n: b_set.add(b_n)
                                if d_str: d_list.append(d_str)

                                if "incomplete" in status.lower():
                                    incomplete_count += 1
                                else:
                                    complete_keys.add((d_str, j_n))

                                if len(row) > 10:
                                    j_info = str(row[10] or "").strip()
                                    if j_info and j_info != "-":
                                        settlement_total += 1
                                        if "EDC" in j_info or "AR" in j_info or "SETTLEMENT" in j_info.upper():
                                            settlement_resolved += 1

                        actionable_diff_count = 0
                        manual_match_count = 0

                        if "Differences" in wb.sheetnames:
                            ws_diff = wb["Differences"]
                            for row in ws_diff.iter_rows(min_row=4, values_only=True):
                                if not row or not row[0]: continue
                                d_str = str(row[1] or "").strip()
                                b_name = str(row[2] or "").strip()
                                j_name = str(row[3] or "").strip()
                                raw_status = str(row[12] or "").strip()
                                key = (d_str, j_name)

                                is_manual_match = "Match" in raw_status or "M0" in raw_status or "M1" in raw_status

                                if is_manual_match:
                                    manual_match_count += 1
                                elif key in complete_keys:
                                    actionable_diff_count += 1

                        wb.close()

                        b_str = ", ".join(sorted(list(b_set))) if b_set else "Summary"
                        from datetime import datetime as _dt
                        parsed_dates = []
                        for ds in d_list:
                            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y", "%m/%d/%Y"):
                                try:
                                    parsed_dates.append(_dt.strptime(str(ds).strip()[:10], fmt).date())
                                    break
                                except Exception:
                                    continue
                        if parsed_dates:
                            mn_d, mx_d = min(parsed_dates), max(parsed_dates)
                            cov_text = f"{mn_d.strftime('%d/%m')} – {mx_d.strftime('%d/%m')}"
                            cov_sub = f"{b_str} ({len(b_set)} Banks)" if b_set else "Recon report range"
                            cov_color = SUCCESS
                        else:
                            cov_text = "Report Ready"
                            cov_sub = Path(latest).name
                            cov_color = SUCCESS

                        if actionable_diff_count == 0:
                            health_text = "100% Matched"
                            health_sub = f"0 Differences • {len(complete_keys)} Batches Audited"
                            health_color = SUCCESS
                        else:
                            health_text = f"{actionable_diff_count} Discrepancies"
                            health_sub = f"Audited dates ({len(complete_keys)} Batches) • Click"
                            health_color = WARN

                        if settlement_total > 0:
                            pct = int((settlement_resolved / settlement_total) * 100)
                            jrn_text = f"{pct}% Resolved ({settlement_resolved}/{settlement_total})"
                            jrn_sub = "EDC & AR settlement journals linked" if pct == 100 else f"{settlement_total - settlement_resolved} missing journal numbers"
                            jrn_color = SUCCESS if pct == 100 else WARN
                        else:
                            jrn_text = "Ready to Check"
                            jrn_sub = "Run journal checker or generate entries"
                            jrn_color = MUTED
                    except Exception:
                        pass

                sales_text = "Cloud not configured"
                sales_sub = "Set Supabase credentials to sync"
                sales_color = MUTED

                try:
                    from cloud_sync import fetch_discrepancies, is_cloud_configured
                    if is_cloud_configured():
                        items = fetch_discrepancies(limit=200)
                        if items:
                            resolved = sum(1 for i in items if i.get("status") == "resolved")
                            pending = len(items) - resolved
                            sales_text = f"{len(items)} Tickets ({resolved} Done)"
                            sales_sub = f"{pending} pending store review • Click" if pending > 0 else "All tickets resolved • Click"
                            sales_color = SUCCESS if pending == 0 else WARN
                        else:
                            sales_text = "0 Active Tickets"
                            sales_sub = "All store discrepancies clear"
                            sales_color = SUCCESS
                except Exception:
                    pass

                self.after(0, lambda: _set_lbl(self._ins_health_val, health_text, health_color) if hasattr(self, "_ins_health_val") else None)
                self.after(0, lambda: _set_lbl(self._ins_health_sub, health_sub, MUTED) if hasattr(self, "_ins_health_sub") else None)
                self.after(0, lambda: _set_lbl(self._ins_jrn_val, jrn_text, jrn_color) if hasattr(self, "_ins_jrn_val") else None)
                self.after(0, lambda: _set_lbl(self._ins_jrn_sub, jrn_sub, MUTED) if hasattr(self, "_ins_jrn_sub") else None)
                self.after(0, lambda: _set_lbl(self._ins_sales_val, sales_text, sales_color) if hasattr(self, "_ins_sales_val") else None)
                self.after(0, lambda: _set_lbl(self._ins_sales_sub, sales_sub, MUTED) if hasattr(self, "_ins_sales_sub") else None)
                self.after(0, lambda: _set_lbl(self._ins_cov_val, cov_text, cov_color) if hasattr(self, "_ins_cov_val") else None)
                self.after(0, lambda: _set_lbl(self._ins_cov_sub, cov_sub, MUTED) if hasattr(self, "_ins_cov_sub") else None)

                # Update states of recon-dependent CTA buttons
                has_recon = bool(output_files)
                is_running = getattr(self, "_running", False)
                recon_state = "normal" if (has_recon and not is_running) else "disabled"

                def _apply_btn_state(btn_attr, normal_fg, normal_txt, dis_fg="#F1F5F9", dis_txt=MUTED):
                    if hasattr(self, btn_attr):
                        b = getattr(self, btn_attr)
                        try:
                            b.configure(
                                state=recon_state,
                                fg_color=normal_fg if recon_state == "normal" else dis_fg,
                                text_color=normal_txt if recon_state == "normal" else dis_txt
                            )
                        except Exception:
                            pass

                self.after(0, lambda: _apply_btn_state("_match_btn", WHITE, ACCENT))
                self.after(0, lambda: _apply_btn_state("_sync_cloud_btn", WHITE, TEXT))
                self.after(0, lambda: _apply_btn_state("_journal_btn", ACCENT, WHITE, dis_fg="#E2E8F0"))
                self.after(0, lambda: _apply_btn_state("_pdf_btn", PANEL, TEXT))

                # Update states of Quick Access links (all transparent background)
                from config import ODO_EXCEL_PATH, ODO_JOURNAL_EXCEL_PATH
                has_merchant = tot_bank > 0
                has_mutation = tot_mut > 0
                has_payment = ODO_EXCEL_PATH.exists()
                has_journal_entry = ODO_JOURNAL_EXCEL_PATH.exists()
                recap_dir = BASE_DIR / "recap"
                has_recap = recap_dir.exists() and bool(list(recap_dir.glob("*.xlsx")))

                def _set_qa_link(attr, exists):
                    if hasattr(self, attr):
                        b = getattr(self, attr)
                        try:
                            b.configure(
                                state="normal" if exists else "disabled",
                                fg_color="transparent",
                                text_color=ACCENT if exists else MUTED
                            )
                        except Exception:
                            pass

                self.after(0, lambda: _set_qa_link("_qa_result_btn", has_recon))
                self.after(0, lambda: _set_qa_link("_qa_pdf_btn", has_recon))
                self.after(0, lambda: _set_qa_link("_qa_merchant_btn", has_merchant))
                self.after(0, lambda: _set_qa_link("_qa_mutation_btn", has_mutation))
                self.after(0, lambda: _set_qa_link("_qa_payment_btn", has_payment))
                self.after(0, lambda: _set_qa_link("_qa_journal_btn", has_journal_entry))
                self.after(0, lambda: _set_qa_link("_qa_recap_btn", has_recap))

            threading.Thread(target=_fetch_insights, daemon=True).start()

        except Exception:
            pass

    # ── Drill-down threading helpers ─────────────────────────────────────────
    def _start_drill_update(self, mutation_dir):
        """Debounced drill scan: coalesces rapid successive calls into one.
        Actual I/O starts after a 400 ms idle window — if another call arrives
        before the timer fires, the timer resets and only ONE thread ever runs."""
        # Cancel any pending debounce timer
        pending = getattr(self, "_drill_pending_id", None)
        if pending:
            try:
                self.after_cancel(pending)
            except Exception:
                pass
        # Bump gen immediately so any already-running thread sees itself as stale
        self._drill_gen = getattr(self, "_drill_gen", 0) + 1
        # Show "Scanning" hint right away (cheap, main-thread)
        if hasattr(self, "_drill_status_lbl"):
            self._drill_status_lbl.configure(text="⏳ Scanning...", text_color=WARN)
        # Debounce: launch the real work after 50 ms of silence
        self._drill_pending_id = self.after(50, lambda: self._do_start_drill(mutation_dir))

    def _do_start_drill(self, mutation_dir):
        """Actually launches the background I/O thread after the debounce settles."""
        import threading
        self._drill_pending_id = None
        my_gen = self._drill_gen  # snapshot: thread uses this to detect stale runs

        # ── Main-thread: rebuild loading state ──────────────────────────────
        for trio in getattr(self, "_drill_rows", []):
            for w in trio:
                try: w.destroy()
                except Exception: pass
        self._drill_rows = []
        try:
            self._drill_loading_lbl.destroy()
        except Exception:
            pass
        loading_lbl = ctk.CTkLabel(
            self._drill_grid, text="Reading files, please wait...",
            font=(FONT_FAMILY, 11, "bold"), text_color=MUTED,
            fg_color="transparent", anchor="w"
        )
        loading_lbl.grid(row=2, column=0, columnspan=3, padx=8, pady=4, sticky="w")
        self._drill_loading_lbl = loading_lbl
        self.update_idletasks()

        def _compute():
            """Background: abort early if our gen is no longer current."""
            rows_data = self._compute_drill_rows(mutation_dir, my_gen)
            self.after(0, lambda: self._render_drill_rows(rows_data, my_gen))

        threading.Thread(target=_compute, daemon=True).start()

    def _render_drill_rows(self, rows_data, gen):
        """Main-thread: render the computed rows_data into the grid."""
        # Stale scan — a newer one is already running
        if gen != getattr(self, "_drill_gen", 0):
            return

        # Destroy loading placeholder
        try:
            self._drill_loading_lbl.destroy()
        except Exception:
            pass

        # Destroy any leftover rows from previous render
        for trio in getattr(self, "_drill_rows", []):
            for w in trio:
                try: w.destroy()
                except Exception: pass
        self._drill_rows = []

        def _bank_color(label):
            if label.startswith("BCA"):     return "#1565C0"
            if label.startswith("MANDIRI"): return "#E65100"
            if label.startswith("BRI"):     return "#2E7D32"
            return ACCENT

        for i, (label, stmt_r, mut_r) in enumerate(rows_data):
            row_num = i + 2
            bcolor = _bank_color(label)

            acc_lbl = ctk.CTkLabel(
                self._drill_grid, text=label,
                font=(FONT_FAMILY, 11, "bold"), text_color=bcolor,
                fg_color="transparent", anchor="w"
            )
            acc_lbl.grid(row=row_num, column=0, padx=(0, 20), pady=3, sticky="w")

            stmt_lbl = ctk.CTkLabel(
                self._drill_grid, text=stmt_r or "No files",
                font=(FONT_FAMILY, 11, "bold"), text_color=SUCCESS if stmt_r else MUTED,
                fg_color="transparent", anchor="w"
            )
            stmt_lbl.grid(row=row_num, column=1, padx=(0, 20), pady=3, sticky="w")

            mut_lbl = ctk.CTkLabel(
                self._drill_grid, text=mut_r or "No mutation CSV",
                font=(FONT_FAMILY, 11, "bold"), text_color=SUCCESS if mut_r else WARN,
                fg_color="transparent", anchor="w"
            )
            mut_lbl.grid(row=row_num, column=2, padx=(0, 20), pady=3, sticky="w")

            self._drill_rows.append((acc_lbl, stmt_lbl, mut_lbl))

        if hasattr(self, "_drill_status_lbl"):
            self._drill_status_lbl.configure(text="✓ Up to date", text_color=SUCCESS)
            self.after(3000, lambda: self._drill_status_lbl.configure(text="", text_color=MUTED) if hasattr(self, "_drill_status_lbl") else None)

    # ── Dashboard drill cache ─────────────────────────────────────────────────
    def _drill_cache_key(self):
        """Fast fingerprint of all input/mutation dirs via mtime.
        If nothing on disk changed, the key is identical → cache hit."""
        try:
            from config import BCA_EXCEL_DIR, MANDIRI_ZIP_DIR, BRI_ZIP_DIR, MUTATION_DIR
            parts = []
            for d in (BCA_EXCEL_DIR, MANDIRI_ZIP_DIR, BRI_ZIP_DIR, MUTATION_DIR):
                if d.exists():
                    for f in sorted(d.rglob("*")):
                        if f.is_file() and not f.name.startswith("."):
                            parts.append(f"{f.stat().st_mtime:.2f}")
            return "|".join(parts)
        except Exception:
            return None


    def _compute_drill_rows(self, mutation_dir, my_gen: int = 0):
        """Pure data computation - no Tkinter calls, safe for background threads.
        mtime cache: if files unchanged, returns previous result instantly.
        parallel reads: BCA/Mandiri/BRI run concurrently - wall-clock = max not sum.
        stale checks: each parallel worker returns early if superseded.
        """
        def _stale():
            return my_gen != getattr(self, "_drill_gen", my_gen)

        import io as _io, contextlib, re, zipfile
        from concurrent.futures import ThreadPoolExecutor
        rows_data = []                       # always defined — safe fallback if exception
        _sink = _io.StringIO()

        with contextlib.redirect_stdout(_sink), contextlib.redirect_stderr(_sink):
            try:
                # ── Cache check ───────────────────────────────────────────────
                cache_key = self._drill_cache_key()
                cached = getattr(self, "_drill_cache", None)
                if cache_key and cached and cached.get("key") == cache_key:
                    return cached["data"]   # instant, zero I/O

                from config import (
                    BCA_EXCEL_DIR, BCA_EXCEL_PATTERN, BCA_EXCEL_PASSWORD,
                    BCA_AMOUNT_COLUMN, BCA_DATE_COLUMN, BCA_NUMBER_COLUMN,
                    MANDIRI_ZIP_DIR, MANDIRI_ZIP_PASSWORD,
                    MANDIRI_AMOUNT_COLUMN, MANDIRI_NUMBER_COLUMN,
                    BRI_ZIP_DIR, BRI_PDF_PATTERN,
                    BRI_AMOUNT_COLUMN, BRI_NUMBER_COLUMN,
                    BANK_ACCOUNTS,
                )

                # ── Date helpers (shared across all workers) ──────────────────
                def _parse_dates(dates):
                    from datetime import date as _d, datetime as _dt
                    clean = set()
                    for d in dates:
                        if d is None: continue
                        if hasattr(d, "date") and callable(d.date) and not isinstance(d, _d):
                            d = d.date()
                        if isinstance(d, str):
                            try: d = _dt.strptime(d[:10], "%Y-%m-%d").date()
                            except Exception:
                                try: d = _dt.strptime(d[:10], "%d/%m/%Y").date()
                                except Exception: continue
                        if hasattr(d, "strftime"): clean.add(d)
                    return sorted(clean)

                def _date_range(dates):
                    clean = _parse_dates(dates)
                    if not clean: return None
                    mn, mx = clean[0], clean[-1]
                    if mn == mx: return mn.strftime("%d/%m/%y")
                    return f"{mn.strftime('%d/%m/%y')} \u2013 {mx.strftime('%d/%m/%y')}"

                def _stmt_dates(dates):
                    clean = _parse_dates(dates)
                    if not clean: return None
                    if len(clean) == 1: return clean[0].strftime("%d/%m/%y")
                    is_contiguous = all((clean[i+1]-clean[i]).days == 1 for i in range(len(clean)-1))
                    if is_contiguous:
                        return f"{clean[0].strftime('%d/%m/%y')} \u2013 {clean[-1].strftime('%d/%m/%y')}"
                    if len(clean) <= 5:
                        return " | ".join(d.strftime("%d/%m/%y") for d in clean)
                    return f"{clean[0].strftime('%d/%m/%y')} | \u2026 | {clean[-1].strftime('%d/%m/%y')} ({len(clean)} dates)"

                def _mut_range_alias(alias_dir):
                    if not alias_dir.exists(): return None
                    try:
                        from readers.mutation_reader import read_mutation_bca, read_mutation_mandiri, read_mutation_bri
                        bank_name = alias_dir.parent.name.lower()
                        alias = alias_dir.name
                        dates = []
                        csv_files = set(list(alias_dir.glob("*.csv")) + list(alias_dir.glob("*.CSV")))
                        if alias_dir.is_dir() and not csv_files:
                            csv_files = set(list(alias_dir.parent.glob("*.csv")) + list(alias_dir.parent.glob("*.CSV")))
                        for csv_f in csv_files:
                            try:
                                if bank_name == "bca":       rows2, unks2 = read_mutation_bca(csv_f, alias)
                                elif bank_name == "mandiri": rows2, unks2 = read_mutation_mandiri(csv_f, alias)
                                else:                        rows2, unks2 = read_mutation_bri(csv_f, alias)
                                for r in (rows2 + unks2):
                                    d = r.get("date")
                                    if d: dates.append(d)
                            except Exception: pass
                        return _date_range(dates)
                    except Exception:
                        return None

                # ── Per-bank workers (run in parallel) ───────────────────────
                def _work_bca():
                    if _stale(): return []
                    bca_stmt_dates = []
                    try:
                        from readers.bca_reader import _find_bca_excels, _read_one_bca
                        search_dirs = []
                        if BCA_EXCEL_DIR.exists():
                            for item in BCA_EXCEL_DIR.iterdir():
                                if item.is_dir(): search_dirs.append(item)
                            if not search_dirs: search_dirs = [BCA_EXCEL_DIR]
                        for sdir in search_dirs:
                            if _stale(): break
                            try:
                                bca_files = _find_bca_excels(sdir, BCA_EXCEL_PATTERN)
                                with ThreadPoolExecutor(max_workers=min(4, len(bca_files) or 1)) as bca_pool:
                                    bca_futs = [bca_pool.submit(_read_one_bca, f, BCA_EXCEL_PASSWORD, BCA_AMOUNT_COLUMN, BCA_DATE_COLUMN, BCA_NUMBER_COLUMN) for f in bca_files]
                                    for fut in bca_futs:
                                        if _stale(): break
                                        for r in fut.result():
                                            d = r.get("date") or r.get("txn_date")
                                            if d: bca_stmt_dates.append(d)
                            except Exception: pass
                    except Exception: pass
                    if _stale(): return []
                    bca_mut_dir = mutation_dir / "bca"
                    bca_mut_r = _mut_range_alias(bca_mut_dir / "main") if (bca_mut_dir / "main").exists() else None
                    if not bca_mut_r and bca_mut_dir.exists():
                        for a in bca_mut_dir.iterdir():
                            if a.is_dir():
                                bca_mut_r = _mut_range_alias(a)
                                break
                    return [("BCA", _stmt_dates(bca_stmt_dates), bca_mut_r)]

                def _work_mandiri():
                    if _stale(): return []
                    mandiri_stmt_by_alias = {}
                    try:
                        from readers.mandiri_reader import extract_mandiri_dates_from_zip
                        for alias_d in sorted(MANDIRI_ZIP_DIR.iterdir()):
                            if _stale(): break
                            if alias_d.is_dir():
                                alias_dates = []
                                for z_path in alias_d.glob("*.zip"):
                                    alias_dates.extend(extract_mandiri_dates_from_zip(z_path, MANDIRI_ZIP_PASSWORD))
                                if alias_dates:
                                    mandiri_stmt_by_alias[alias_d.name] = alias_dates
                        if not mandiri_stmt_by_alias and MANDIRI_ZIP_DIR.exists():
                            alias_dates = []
                            for z_path in MANDIRI_ZIP_DIR.glob("*.zip"):
                                alias_dates.extend(extract_mandiri_dates_from_zip(z_path, MANDIRI_ZIP_PASSWORD))
                            if alias_dates:
                                mandiri_stmt_by_alias["main"] = alias_dates
                    except Exception: pass
                    if _stale(): return []
                    result = []
                    mandiri_mut_dir = mutation_dir / "mandiri"
                    seen_man = set()
                    for alias, dates in sorted(mandiri_stmt_by_alias.items()):
                        seen_man.add(alias)
                        label = "MANDIRI" if alias == "main" else f"MANDIRI / {alias}"
                        result.append((label, _stmt_dates(dates), _mut_range_alias(mandiri_mut_dir / alias)))
                    if mandiri_mut_dir.exists():
                        for a in sorted(mandiri_mut_dir.iterdir()):
                            if a.is_dir() and a.name not in seen_man:
                                label = "MANDIRI" if a.name == "main" else f"MANDIRI / {a.name}"
                                result.append((label, None, _mut_range_alias(a)))
                    return result

                def _work_bri():
                    if _stale(): return []
                    bri_stmt_by_alias = {}
                    try:
                        for alias_d in sorted(BRI_ZIP_DIR.iterdir()):
                            if _stale(): break
                            if alias_d.is_dir():
                                alias_dates = []
                                for z_path in alias_d.glob("*.zip"):
                                    try:
                                        with zipfile.ZipFile(z_path, "r") as zf:
                                            for name in zf.namelist():
                                                m = re.search(r'(\d{4}-\d{2}-\d{2})', name)
                                                if m:
                                                    alias_dates.append(m.group(1))
                                                else:
                                                    m2 = re.search(r'((?:20\d{2})(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01]))', name)
                                                    if m2:
                                                        ds = m2.group(1)
                                                        alias_dates.append(f"{ds[:4]}-{ds[4:6]}-{ds[6:8]}")
                                    except Exception: pass
                                if alias_dates:
                                    bri_stmt_by_alias[alias_d.name] = alias_dates
                        if not bri_stmt_by_alias and BRI_ZIP_DIR.exists():
                            alias_dates = []
                            for z_path in BRI_ZIP_DIR.glob("*.zip"):
                                try:
                                    with zipfile.ZipFile(z_path, "r") as zf:
                                        for name in zf.namelist():
                                            m = re.search(r'(\d{4}-\d{2}-\d{2})', name)
                                            if m:
                                                alias_dates.append(m.group(1))
                                            else:
                                                m2 = re.search(r'((?:20\d{2})(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01]))', name)
                                                if m2:
                                                    ds = m2.group(1)
                                                    alias_dates.append(f"{ds[:4]}-{ds[4:6]}-{ds[6:8]}")
                                except Exception: pass
                            if alias_dates:
                                bri_stmt_by_alias["main"] = alias_dates
                    except Exception: pass
                    if _stale(): return []
                    result = []
                    bri_mut_dir = mutation_dir / "bri"
                    seen_bri = set()
                    for alias, dates in sorted(bri_stmt_by_alias.items()):
                        seen_bri.add(alias)
                        label = "BRI" if alias == "main" else f"BRI / {alias}"
                        result.append((label, _stmt_dates(dates), _mut_range_alias(bri_mut_dir / alias)))
                    if bri_mut_dir.exists():
                        for a in sorted(bri_mut_dir.iterdir()):
                            if a.is_dir() and a.name not in seen_bri:
                                label = "BRI" if a.name == "main" else f"BRI / {a.name}"
                                result.append((label, None, _mut_range_alias(a)))
                    return result

                # ── Run all three banks concurrently ─────────────────────────
                with ThreadPoolExecutor(max_workers=3, thread_name_prefix="drill") as pool:
                    f_bca     = pool.submit(_work_bca)
                    f_mandiri = pool.submit(_work_mandiri)
                    f_bri     = pool.submit(_work_bri)
                    bca_rows     = f_bca.result()
                    mandiri_rows = f_mandiri.result()
                    bri_rows     = f_bri.result()

                if not _stale():
                    rows_data = bca_rows + mandiri_rows + bri_rows
                    # Save to cache so next scan (if files unchanged) is instant
                    if cache_key:
                        self._drill_cache = {"key": cache_key, "data": rows_data}

            except Exception:
                pass

        return rows_data


    def _log_write(self, text: str, tag: str = ""):
        self._log.config(state="normal")
        self._log.insert("end", text, tag)
        self._log.see("end")
        self._log.config(state="disabled")

    def _set_status(self, text: str, color: str):
        self._status_var.set(text)
        self._dot.configure(text_color=color)

    def _clear_log(self):
        self._log.config(state="normal")
        self._log.delete("1.0", tk.END)
        self._log.config(state="disabled")

    # ── Feature Operations ────────────────────────────────────────────────────
    def _on_upload(self):
        try:
            from readers.file_detector import detect_file, copy_file

            files = filedialog.askopenfilenames(title="Select Bank or Odoo File", filetypes=[("All Files", "*.*")])
            if not files:
                return

            self._log_write("\n── Uploading Files ──\n", "head")

            # ── Pre-scan for conflicts (#1) ────────────────────────────────
            detections: list[tuple[Path, object]] = []
            for f in files:
                path = Path(f)
                if path.suffix.lower() == ".pdf":
                    # Give helpful message for raw PDFs that aren't BRI EDC
                    result = detect_file(path)
                    if result is None:
                        self._log_write(
                            f"⚠️ Ignored: {path.name} — BRI PDFs must be inside their ZIP, or must contain an AMT_TRX table to be detected as BRI EDC.\n", "warn"
                        )
                        continue
                    detections.append((path, result))
                else:
                    result = detect_file(path)
                    if result is None:
                        self._log_write(f"⚠️ Ignored: {path.name} (No bank pattern matched)\n", "warn")
                    else:
                        detections.append((path, result))

            if not detections:
                self._refresh_folder_status()
                return

            # Check for conflicts (destination already exists)
            conflicts = []
            for path, result in detections:
                dest_name = (path.name + ".zip") if result.wrap_as_zip else path.name
                dest = result.target_dir / dest_name
                if dest.exists():
                    conflicts.append((path, result, dest))

            # Single upfront dialog for all conflicts (#1)
            overwrite_action = "replace"  # default if no conflicts
            if conflicts:
                import tkinter.simpledialog as _sd
                names = "\n".join(f"  • {p.name}" for p, _, _ in conflicts[:8])
                if len(conflicts) > 8:
                    names += f"\n  ... and {len(conflicts) - 8} more"
                dlg = ctk.CTkToplevel(self)
                dlg.title("File Conflict")
                dlg.geometry("480x280")
                dlg.resizable(False, False)
                dlg.transient(self)
                dlg.grab_set()
                dlg.configure(fg_color=BG)
                sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
                dlg.geometry(f"480x280+{max(0,sw//2-240)}+{max(0,sh//2-140)}")

                choice = tk.StringVar(value="")
                panel = ctk.CTkFrame(dlg, fg_color=PANEL, corner_radius=10, border_color=BORDER, border_width=1)
                panel.pack(fill="both", expand=True, padx=14, pady=14)
                ctk.CTkLabel(panel, text=f"⚠️  {len(conflicts)} file(s) already exist in destination:", font=(FONT_FAMILY, 12, "bold"), text_color=WARN).pack(anchor="w", padx=14, pady=(14, 4))
                ctk.CTkLabel(panel, text=names, font=(FONT_MONO, 10), text_color=MUTED, justify="left").pack(anchor="w", padx=14)
                btn_frame = ctk.CTkFrame(panel, fg_color="transparent")
                btn_frame.pack(fill="x", side="bottom", padx=14, pady=14)

                def _pick(v):
                    choice.set(v)
                    dlg.destroy()

                ctk.CTkButton(btn_frame, text="Replace All", width=120, height=34, fg_color=ERROR, hover_color="#B91C1C", text_color=WHITE, font=(FONT_FAMILY, 11, "bold"), command=lambda: _pick("replace")).pack(side="left", padx=(0, 8))
                ctk.CTkButton(btn_frame, text="Keep Both", width=120, height=34, fg_color=ACCENT, hover_color=ACCENT_HOVER, text_color=WHITE, font=(FONT_FAMILY, 11, "bold"), command=lambda: _pick("keep")).pack(side="left", padx=(0, 8))
                ctk.CTkButton(btn_frame, text="Skip Conflicts", width=120, height=34, fg_color=WHITE, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1, text_color=TEXT, font=(FONT_FAMILY, 11, "bold"), command=lambda: _pick("skip")).pack(side="left")
                dlg.wait_window()
                overwrite_action = choice.get() or "skip"

            # ── Copy files ────────────────────────────────────────────────
            from datetime import datetime as _dt
            ts = _dt.now().strftime("%Y%m%d_%H%M%S")
            for path, result in detections:
                dest_name = (path.name + ".zip") if result.wrap_as_zip else path.name
                dest = result.target_dir / dest_name
                is_conflict = dest.exists()

                if is_conflict:
                    if overwrite_action == "skip":
                        self._log_write(f"⏭️  Skipped: {path.name} (already exists)\n", "dim")
                        continue
                    elif overwrite_action == "keep":
                        # Rename incoming: insert timestamp before extension
                        stem = Path(dest_name).stem
                        suf = Path(dest_name).suffix
                        dest_name = f"{stem}_{ts}{suf}"
                        dest = result.target_dir / dest_name

                try:
                    final_dest = copy_file(result, path)
                    if overwrite_action == "keep" and is_conflict:
                        # copy_file wrote to original name, rename it
                        final_dest.rename(dest)
                        final_dest = dest
                    verb = "Wrapped & Copied" if result.wrap_as_zip else "Copied"
                    folder_display = f"{result.target_dir.parent.name}/{result.target_dir.name}/" if result.alias else f"{result.target_dir.name}/"
                    self._log_write(f"✅ {verb}: {path.name} → {folder_display}{final_dest.name}\n", "ok")
                except Exception as fe:
                    self._log_write(f"❌ Failed: {path.name} — {fe}\n", "err")

            self._refresh_folder_status()
        except Exception as e:
            self._log_write(f"\n❌ Error during Upload: {e}\nEnsure your .env file is configured correctly!\n", "err")


    def _on_download(self):
        if DateEntry:
            date_from = self._date_from_widget.get()
            date_to = self._date_to_widget.get()
        else:
            date_from = self._date_from_var.get().strip()
            date_to = self._date_to_var.get().strip()
        
        self._log_write(f"\n── Downloading Odoo Payment (From {date_from} to {date_to}) ──\n", "head")
        self._set_status("Downloading Odoo...", WARN)
        self._running = True
        if hasattr(self, "_stop_btn"):
            self._stop_btn.pack(fill="x", pady=(6, 0))
        
        def run():
            try:
                if getattr(sys, "frozen", False):
                    cmd = [sys.executable, "--run-downloader", "--date-from", date_from, "--date-to", date_to]
                else:
                    cmd = [_venv_python, "odoo_downloader.py", "--date-from", date_from, "--date-to", date_to]
                    
                email = self._email_var.get().strip()
                password = self._password_var.get()
                if email and password:
                    cmd.extend(["--email", email, "--password", password])
                    
                selected_banks = [b for b, var in self._bank_vars.items() if var.get()]
                if selected_banks:
                    cmd.extend(["--banks", ",".join(selected_banks)])

                env = os.environ.copy()
                env.pop("TCL_LIBRARY", None)
                env.pop("TK_LIBRARY", None)
                env["PYTHONIOENCODING"] = "utf-8"
                flags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if IS_WINDOWS else 0
                process = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=1,
                    encoding="utf-8",
                    errors="replace",
                    creationflags=flags,
                    env=env
                )
                self._active_proc = process

                for line in process.stdout:
                    self.after(0, self._log_write, line)

                process.wait()
                if process.returncode == 0:
                    self.after(0, self._log_write, "\n✅ Odoo Download Finished!\n", "ok")
                else:
                    self.after(0, self._log_write, f"\n❌ Odoo Download failed with code {process.returncode}\n", "err")

            except Exception as e:
                self.after(0, self._log_write, f"\n❌ Error: {e}\n", "err")
            finally:
                self._active_proc = None
                self._running = False
                self.after(0, self._refresh_folder_status)
                if hasattr(self, "_stop_btn"):
                    self.after(0, self._stop_btn.pack_forget)
                self.after(0, self._set_status, "Ready", SUCCESS)

        threading.Thread(target=run, daemon=True).start()

    def _on_cleanse(self):
        try:
            from config import INPUT_DIR, MUTATION_DIR, OUTPUT_DIR, ODO_EXCEL_PATH
            
            recap_base = BASE_DIR / "recap"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            target_recap = recap_base / timestamp
            
            files_to_move = []
            for base_dir in [INPUT_DIR, MUTATION_DIR, OUTPUT_DIR]:
                if base_dir.exists() and base_dir.is_dir():
                    for f in base_dir.rglob("*"):
                        if not f.is_file():
                            continue
                        n = f.name
                        # Exclude: journal log, hidden sidecars, DS_Store, Excel temp files
                        if n == "journal_creation_log.xlsx":
                            continue
                        if n.startswith("."):  # .manual_matches.json, .journal_config.json, .DS_Store
                            continue
                        if n.startswith("~$"):  # Excel temp/lock files
                            continue
                        files_to_move.append(f)
            if ODO_EXCEL_PATH.exists() and ODO_EXCEL_PATH.is_file() and ODO_EXCEL_PATH not in files_to_move:
                files_to_move.append(ODO_EXCEL_PATH)
                
            if not files_to_move:
                self._log_write("\nℹ️ No data files to clean.\n", "dim")
                self._set_status("No data files to clean", MUTED)
                return

            dlg = ctk.CTkToplevel(self)
            dlg.title("Confirm Data Cleanup")
            dlg.geometry("540x370")
            dlg.resizable(False, False)
            dlg.transient(self)
            dlg.grab_set()
            
            sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
            cx, cy = max(0, int(sw/2 - 540/2)), max(0, int(sh/2 - 370/2))
            dlg.geometry(f"540x370+{cx}+{cy}")
            dlg.configure(fg_color=BG)
            
            content = ctk.CTkFrame(dlg, fg_color=PANEL, corner_radius=10, border_color=BORDER, border_width=1)
            content.pack(fill="both", expand=True, padx=16, pady=16)
            
            hdr = ctk.CTkFrame(content, fg_color="transparent")
            hdr.pack(fill="x", padx=16, pady=(16, 8))
            ctk.CTkLabel(hdr, text=f"This will relocate {len(files_to_move)} file(s) out of active input, mutation & output folders.", font=(FONT_FAMILY, 10, "bold"), text_color=MUTED).pack(anchor="w", pady=(2, 0))

            info_box = ctk.CTkFrame(content, fg_color=PREVIEW_BG, corner_radius=8, border_color=BORDER, border_width=1)
            info_box.pack(fill="x", padx=16, pady=12)

            ctk.CTkLabel(info_box, text="🗄️ Archive Destination Folder:", font=(FONT_FAMILY, 10, "bold"), text_color=TEXT).pack(anchor="w", padx=12, pady=(10, 2))
            ctk.CTkLabel(info_box, text=str(target_recap), font=(FONT_MONO, 10, "bold"), text_color=ACCENT, wraplength=460, justify="left").pack(anchor="w", padx=12, pady=(0, 10))

            ctk.CTkLabel(content, text="Cleaned files will be stored safely in Recap storage and can be accessed via Quick Access.", font=(FONT_FAMILY, 10, "bold"), text_color=MUTED, wraplength=460, justify="left").pack(anchor="w", padx=16)

            btn_frame = ctk.CTkFrame(content, fg_color="transparent")
            btn_frame.pack(fill="x", side="bottom", padx=16, pady=(12, 16))

            def _do_clean():
                dlg.destroy()
                try:
                    self._log_write("\n── Cleaning Workspace Data ──\n", "head")
                    moved_count = 0
                    for f in files_to_move:
                        try:
                            rel_path = f.relative_to(BASE_DIR)
                        except ValueError:
                            rel_path = Path(f.name)
                        target_file = target_recap / rel_path
                        target_file.parent.mkdir(parents=True, exist_ok=True)
                        shutil.move(str(f), str(target_file))
                        moved_count += 1

                    self._log_write(f"✅ {moved_count} file(s) archived to: recap/{timestamp}/\n", "ok")
                    self._refresh_folder_status()
                    self._set_status(f"Cleaned {moved_count} files", SUCCESS)
                except Exception as ex:
                    self._log_write(f"\n❌ Error during Data Cleanup: {ex}\n", "err")

            ctk.CTkButton(
                btn_frame, text="Cancel", height=36, width=100,
                fg_color=WHITE, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1,
                text_color=TEXT, font=(FONT_FAMILY, 11, "bold"), command=dlg.destroy
            ).pack(side="right", padx=(8, 0))

            ctk.CTkButton(
                btn_frame, text="Confirm Clean", height=36, width=130,
                fg_color=ERROR, hover_color="#B91C1C", text_color=WHITE,
                font=(FONT_FAMILY, 11, "bold"), command=_do_clean
            ).pack(side="right")

        except Exception as e:
            self._log_write(f"\n❌ Error preparing Data Cleanup: {e}\n", "err")

    def _get_selected_banks(self) -> list[str]:
        """Return clean list of active bank keys ('bca', 'mandiri', 'bri')."""
        if hasattr(self, "_bank_vars"):
            if self._bank_vars.get("All") and self._bank_vars["All"].get():
                return ["bca", "mandiri", "bri"]
            sel = [b.lower() for b in ["BCA", "Mandiri", "BRI"] if self._bank_vars.get(b) and self._bank_vars[b].get()]
            if sel:
                return sel
        return ["bca", "mandiri", "bri"]

    def _on_scan(self):
        if self._running:
            return
        self._running = True
        selected_banks = self._get_selected_banks()
        if not selected_banks:
            self._set_status("Select at least 1 bank!", ERROR)
            self._running = False
            return

        self._scan_btn.configure(state="disabled")
        self._run_btn.configure(state="disabled")
        self._open_btn.configure(state="disabled")
        self._upload_btn.configure(state="disabled")
        self._download_btn.configure(state="disabled")
        self._log.config(state="normal")
        self._log.delete("1.0", "end")
        self._log.config(state="disabled")
        self._set_status("Scanning data...", WARN)
        self.update_idletasks()

        def _deferred_scan():
            self._refresh_folder_status()
            threading.Thread(target=self._run_script, args=(selected_banks, True), daemon=True).start()
        self.after(50, _deferred_scan)

    def _on_run(self):
        self._running = True
        if hasattr(self, "_stop_btn"):
            self._stop_btn.pack(fill="x", pady=(6, 0))

        selected_banks = self._get_selected_banks()
        if not selected_banks:
            self._set_status("Select at least 1 bank!", ERROR)
            self._running = False
            if hasattr(self, "_stop_btn"):
                self._stop_btn.pack_forget()
            return

        # Check if merchant files exist for selected banks
        has_merchant_files = False
        for b in selected_banks:
            b_dir = INPUT_DIR / b
            if b_dir.exists():
                files = [f for f in b_dir.rglob("*") if f.is_file() and not f.name.startswith(".") and not f.name.startswith("~$")]
                if files:
                    has_merchant_files = True
                    break

        if not has_merchant_files:
            self._set_status("No merchant report files found in input folder!", ERROR)
            self._log_write("\n❌ Cannot run reconciliation: No merchant report files found in 'input/' for selected banks.\nPlease upload bank statement / merchant report files first.\n", "err")
            self._running = False
            if hasattr(self, "_stop_btn"):
                self._stop_btn.pack_forget()
            return

        self._scan_btn.configure(state="disabled")
        self._run_btn.configure(state="disabled", text="⏳ Processing...")
        if hasattr(self, "_match_btn"): self._match_btn.configure(state="disabled", fg_color="#F1F5F9", text_color=MUTED)
        if hasattr(self, "_sync_cloud_btn"): self._sync_cloud_btn.configure(state="disabled", fg_color="#F1F5F9", text_color=MUTED)
        if hasattr(self, "_journal_btn"): self._journal_btn.configure(state="disabled", fg_color="#E2E8F0", text_color=MUTED)
        if hasattr(self, "_pdf_btn"): self._pdf_btn.configure(state="disabled", fg_color="#F1F5F9", text_color=MUTED)
        self._open_btn.configure(state="disabled")
        self._upload_btn.configure(state="disabled")
        self._download_btn.configure(state="disabled")
        self._log.config(state="normal")
        self._log.delete("1.0", "end")
        self._log.config(state="disabled")
        self._set_status("Processing...", WARN)
        self.update_idletasks()

        def run_all():
            try:
                self.after(0, self._log_write, f"\n── Starting Single Browser Auto-Recon ──\n", "head")

                if getattr(sys, "frozen", False):
                    cmd = [sys.executable, "--run-downloader", "--mode", "auto_recon"]
                else:
                    cmd = [_venv_python, "odoo_downloader.py", "--mode", "auto_recon"]

                email = self._email_var.get().strip()
                password = self._password_var.get()
                if email and password:
                    cmd.extend(["--email", email, "--password", password])

                banks_for_dl = [b for b, var in self._bank_vars.items() if var.get()]
                if banks_for_dl:
                    cmd.extend(["--banks", ",".join(banks_for_dl)])

                env = os.environ.copy()
                env.pop("TCL_LIBRARY", None)
                env.pop("TK_LIBRARY", None)
                env["PYTHONIOENCODING"] = "utf-8"
                flags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if IS_WINDOWS else 0
                process = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=1,
                    encoding="utf-8",
                    errors="replace",
                    creationflags=flags,
                    env=env
                )
                self._active_proc = process



                for line in process.stdout:
                    self.after(0, self._log_write, line)

                process.wait()
                if process.returncode != 0:
                    self.after(0, self._log_write, f"\n\u274c Auto-Recon Failed.\n", "err")
                    self.after(0, self._on_done, process.returncode, None)
                    return

                self.after(0, self._log_write, "\n\u2705 Auto-Recon Completed Successfully!\n", "ok")
                self.after(0, self._on_done, 0, None)

            except Exception as e:
                self.after(0, self._log_write, f"\n\u274c Error during Auto-Recon: {e}\n", "err")
                self.after(0, self._on_done, 1, None)

        def _deferred_run():
            self._refresh_folder_status()
            if self._offline_var.get():
                self._log_write("\n\u2500\u2500 Running Offline Reconciliation (Skipping Downloader) \u2500\u2500\n", "head")
                threading.Thread(target=self._run_script, args=(selected_banks, False), daemon=True).start()
                return
            threading.Thread(target=run_all, daemon=True).start()

        self.after(50, _deferred_run)

    def _run_script(self, selected_banks, is_scan=False):
        if is_scan:
            self._running = True
            def _log(msg, tag=None):
                self.after(0, self._log_write, msg, tag or self._tag(msg))

            try:
                from main import scan_bank_date_range
                self.after(0, self._log_write, "\n────────────────────────────────────────────────────────────\n  SCAN SUMMARY\n────────────────────────────────────────────────────────────\n\n", "head")
                dates_range = scan_bank_date_range(selected_banks, log_fn=_log)
                if dates_range:
                    min_d, max_d = dates_range
                    self.after(0, self._set_dates, min_d, max_d)
                    self.after(0, self._log_write, f"  [+] Detected Date Range: {min_d} to {max_d}\n\n", "ok")
                else:
                    self.after(0, self._log_write, "  [-] No valid bank dates detected in input folders.\n\n", "warn")
                self.after(0, self._on_done, 0, None)
                return
            except Exception as e:
                self.after(0, self._log_write, f"\n❌ Scan Error: {e}\n", "err")
                self.after(0, self._on_done, 1, None)
                return

        if getattr(sys, "frozen", False):
            cmd = [sys.executable, "--worker"]
        elif _venv_python_path.exists():
            cmd = [_venv_python, str(BASE_DIR / "main.py")]
        else:
            cmd = [sys.executable, str(BASE_DIR / "main.py")]

        if "all" in selected_banks:
            cmd.append("--all")
            selected_banks.remove("all")
            
        if selected_banks:
            cmd.extend(["--bank"] + selected_banks)

        self._running = True
        if hasattr(self, "_stop_btn"):
            self._stop_btn.pack(fill="x", pady=(6, 0))

        env = os.environ.copy()
        env.pop("TCL_LIBRARY", None)
        env.pop("TK_LIBRARY", None)
        env["PYTHONIOENCODING"] = "utf-8"
        flags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if IS_WINDOWS else 0
        proc = subprocess.Popen(
            cmd,
            cwd=str(BASE_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            creationflags=flags,
            env=env
        )
        self._active_proc = proc

        last_output_file = None
        if proc.stdout:
            for line in proc.stdout:
                ls = line.rstrip()
                if ls.startswith("[DATE_RANGE]|"):
                    try:
                        _, min_d, max_d = ls.split("|")
                        self.after(0, self._set_dates, min_d, max_d)
                    except Exception:
                        pass
                    continue
                    
                ls_lower = ls.lower()
                if "reconciliation_" in ls_lower and ".xlsx" in ls_lower:
                    for part in ls.split():
                        if "reconciliation_" in part.lower() and ".xlsx" in part.lower():
                            last_output_file = part.strip()
                self.after(0, self._log_write, ls + "\n", self._tag(ls))

        proc.wait()
        
        if proc.returncode == 0 and last_output_file and not is_scan:
            from config import ODO_JOURNAL_EXCEL_PATH
            if not ODO_JOURNAL_EXCEL_PATH.exists():
                self.after(0, self._log_write, f"\n⚠️ File '{ODO_JOURNAL_EXCEL_PATH.name}' tidak ditemukan. Skipping Journal Entries check.\n", "warn")
            else:
                self.after(0, self._log_write, f"\n── Checking Journal Entries ──\n", "head")
                
                j_cmd = [sys.executable, "journal_checker.py", last_output_file] if getattr(sys, "frozen", False) else [_venv_python, "journal_checker.py", last_output_file]
                j_proc = subprocess.Popen(
                    j_cmd,
                    cwd=str(BASE_DIR),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    creationflags=flags,
                    env=env
                )
                self._active_proc = j_proc
                
                if j_proc.stdout:
                    for line in j_proc.stdout:
                        self.after(0, self._log_write, line)
                j_proc.wait()
                
                if j_proc.returncode != 0:
                    self.after(0, self._log_write, f"\n⚠️ Journal Entries Check ended with warnings.\n", "warn")
                
        self.after(0, self._on_done, proc.returncode, last_output_file)

    @staticmethod
    def _tag(line: str) -> str:
        if "✅" in line:                          return "ok"
        if "❌" in line or "[ERROR]" in line:     return "err"
        if "⚠️" in line or "[WARN]" in line:      return "warn"
        if line.startswith("─") or line.startswith("="): return "head"
        if "[ODO]" in line or "skipped" in line.lower(): return "dim"
        return ""

    def _on_done(self, code: int, output_path: str | None):
        self._active_proc = None
        self._running = False
        if hasattr(self, "_stop_btn"):
            self._stop_btn.pack_forget()
        self._scan_btn.configure(state="normal")
        self._run_btn.configure(state="normal", text="▶  Reconciliation")
        self._journal_btn.configure(state="normal")
        self._open_btn.configure(state="normal")
        self._upload_btn.configure(state="normal")
        self._download_btn.configure(state="normal")

        if code == 0:
            if output_path:
                self._set_status("Finished ✓", SUCCESS)
                self._last_output = output_path
                if Path(output_path).exists():
                    _open_path(output_path)
            else:
                self._set_status("Scan Complete ✓", SUCCESS)
        else:
            self._set_status("Failed — check logs below", ERROR)

        # Refresh KPI cards and engine status card
        self._update_dashboard_summary(skip_drill=True)

    def _on_stop(self):
        proc = getattr(self, "_active_proc", None)
        if proc:
            try:
                proc.terminate()
                self.after(500, lambda: proc.kill() if proc and proc.poll() is None else None)
            except Exception:
                pass
            self._active_proc = None
        self._running = False
        if hasattr(self, "_stop_btn"):
            self._stop_btn.pack_forget()
        self._log_write("\n🛑 Process stopped by user.\n", "err")
        self._set_status("Process Stopped", ERROR)
        self._scan_btn.configure(state="normal")
        self._run_btn.configure(state="normal", text="▶  Reconciliation")
        self._journal_btn.configure(state="normal")
        self._open_btn.configure(state="normal")
        self._upload_btn.configure(state="normal")
        self._download_btn.configure(state="normal")
        self._update_dashboard_summary(skip_drill=True)

    def _set_dates(self, min_d: str, max_d: str):
        try:
            from datetime import datetime
            d_from = datetime.strptime(min_d, "%Y-%m-%d").strftime("%d/%m/%Y")
            d_to = datetime.strptime(max_d, "%Y-%m-%d").strftime("%d/%m/%Y")
            if DateEntry:
                self._date_from_widget.set_date(datetime.strptime(d_from, "%d/%m/%Y"))
                self._date_to_widget.set_date(datetime.strptime(d_to, "%d/%m/%Y"))
            else:
                self._date_from_var.set(d_from)
                self._date_to_var.set(d_to)
            self._log_write(f"\n📅 [Auto-Detect] Dates updated: {d_from} - {d_to}\n", "ok")
        except Exception as e:
            self._log_write(f"\n⚠️ Failed to update dates: {e}\n", "warn")

    # ── Cloud Sync Sales Portal ──────────────────────────────────────────
    def _on_sync_cloud(self):
        import glob
        import os
        from openpyxl import load_workbook
        from cloud_sync import is_cloud_configured, push_bank_discrepancies, fetch_discrepancies

        if not is_cloud_configured():
            self._log_write("\n⚠️ Supabase is not configured in .env. Please set SUPABASE_URL and SUPABASE_KEY first.\n", "warn")
            self._set_status("Cloud sync not configured", WARN)
            return

        self._sync_cloud_btn.configure(text="⏳ Loading...", state="disabled")
        self.update_idletasks()

        try:
            output_files = glob.glob(str(OUTPUT_DIR / "[Rr]econciliation_*.xlsx"))
            if not output_files:
                self._set_status("No reconciliation file found. Run recon first!", ERROR)
                return
            latest_file = max(output_files, key=os.path.getctime)

            try:
                wb = load_workbook(latest_file, read_only=True, data_only=True)
                if "Differences" not in wb.sheetnames:
                    self._set_status("No 'Differences' sheet found", ERROR)
                    wb.close()
                    return

                # Build active recon days where BOTH sources exist (>0 bank AND >0 odoo)
                valid_recon_days = set()
                if "Daily Summary" in wb.sheetnames:
                    ws_sum = wb["Daily Summary"]
                    for r_sum in ws_sum.iter_rows(values_only=True):
                        if not r_sum or len(r_sum) < 7:
                            continue
                        d_s = str(r_sum[1] or "").strip()
                        b_s = str(r_sum[3] or "").strip().upper()
                        j_s = str(r_sum[4] or "").strip()
                        try:
                            tot_b = float(r_sum[5] or 0.0)
                            tot_o = float(r_sum[6] or 0.0)
                        except (ValueError, TypeError):
                            continue
                        if tot_b > 0 and tot_o > 0:
                            valid_recon_days.add((d_s, b_s, j_s))
                            valid_recon_days.add((d_s, j_s))

                ws = wb["Differences"]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
            except Exception as e:
                self._set_status(f"Error loading recon file: {e}", ERROR)
                return

            hdr_row = [str(cell or "").strip().lower() for cell in (rows[2] if len(rows) > 2 else [])]
            def _find_c(name, fallback):
                for i, h in enumerate(hdr_row):
                    if name in h: return i
                return fallback

            c_date     = _find_c("date", 1)
            c_bank     = _find_c("bank", 2)
            c_journal  = _find_c("journal", 3)
            c_odo_num  = _find_c("odoo number", 4)
            c_inv_ref  = _find_c("reference", 5)
            c_bank_num = _find_c("bank number", 6)
            c_file     = _find_c("filename", 7)
            c_bank_amt = _find_c("bank amount", 8)
            c_odo_amt  = _find_c("odoo amount", 9)
            c_src      = _find_c("source", 10)
            c_recon    = _find_c("reconciled", 11)
            c_status   = _find_c("status", 12)

            discrepancy_items = []
            for r_idx, r in enumerate(rows[3:], 4):
                if not r or len(r) <= c_status:
                    continue

                r_date = str(r[c_date] or "").strip()
                r_bank = str(r[c_bank] or "").strip().upper()
                r_journal = str(r[c_journal] or "").strip()

                # ONLY include if BOTH bank file and Odoo payments existed for this date/journal!
                if valid_recon_days and (r_date, r_bank, r_journal) not in valid_recon_days and (r_date, r_journal) not in valid_recon_days:
                    continue

                st = str(r[c_status] or "").strip()
                src = str(r[c_src] or "").strip()
                recon = str(r[c_recon] or "").strip().lower() if len(r) > c_recon else ""
                b_amt = float(r[c_bank_amt] or 0.0) if len(r) > c_bank_amt and r[c_bank_amt] else 0.0
                o_amt = float(r[c_odo_amt] or 0.0) if len(r) > c_odo_amt and r[c_odo_amt] else 0.0

                # 1. Bank Only: Exists in bank, missing in Odoo
                if st == "Only in Bank" or src == "Bank":
                    disc_type = "bank_only"
                    amt = b_amt or o_amt
                # 3. Unreconciled in Odoo: Exists in Odoo where Reconciled is 'No'
                elif recon in ("no", "false") and st != "Only in Bank" and src != "Bank":
                    disc_type = "unreconciled_odoo"
                    amt = o_amt or b_amt
                # 2. Odoo Only: Exists in Odoo, missing in Bank
                elif st == "Only in Odoo" or src == "Odoo":
                    disc_type = "odoo_only"
                    amt = o_amt or b_amt
                else:
                    continue

                if amt <= 0.0:
                    continue

                odo_num = str(r[c_odo_num] or "").strip() if len(r) > c_odo_num else ""
                inv_ref = str(r[c_inv_ref] or "").strip() if len(r) > c_inv_ref else ""
                bank_num = str(r[c_bank_num] or "").strip() if len(r) > c_bank_num else ""

                discrepancy_items.append({
                    "row_idx": r_idx,
                    "date": str(r[c_date] or "").strip(),
                    "bank": str(r[c_bank] or "").strip(),
                    "journal": str(r[c_journal] or "").strip(),
                    "number_bank": bank_num,
                    "number_odo": odo_num,
                    "invoice_no": inv_ref,
                    "is_reconciled": str(r[c_recon] or "Yes").strip() if len(r) > c_recon else "Yes",
                    "discrepancy_type": disc_type,
                    "filename": str(r[c_file] or "").strip() if len(r) > c_file else "",
                    "amount": amt,
                })

            def _disc_sort_key(it):
                b = str(it.get("bank", "")).strip().upper()
                d_raw = str(it.get("date", "")).strip()
                if "/" in d_raw:
                    pts = d_raw.split("/")
                    if len(pts) == 3:
                        d_norm = f"{pts[2]}{pts[1].zfill(2)}{pts[0].zfill(2)}"
                    else:
                        d_norm = d_raw
                else:
                    d_norm = d_raw.replace("-", "")
                return (b, d_norm)

            discrepancy_items.sort(key=_disc_sort_key)

            # Initialize each item
            for item in discrepancy_items:
                item["is_synced"] = False

            # ── Open Modal Window ──
            top = ctk.CTkToplevel(self)
            top.title("Share Discrepancies to Sales Portal")
            top.minsize(1050, 640)
            top.resizable(True, True)
            top.configure(fg_color=BG)
            top.transient(self)

            def _safe_grab():
                try:
                    if top.winfo_exists():
                        top.grab_set()
                except Exception:
                    pass
            top.after(100, _safe_grab)
            top.after(200, lambda: _center_modal_on_parent(top, self))

            # Header
            hdr_frame = ctk.CTkFrame(top, fg_color=PANEL, corner_radius=0, height=76, border_color=BORDER, border_width=1)
            hdr_frame.pack(fill="x")
            hdr_frame.pack_propagate(False)

            hi = ctk.CTkFrame(hdr_frame, fg_color="transparent")
            hi.pack(fill="both", expand=True, padx=24, pady=12)

            ctk.CTkLabel(hi, text="Share Discrepancies to Sales Portal", font=(FONT_FAMILY, 16, "bold"), text_color=ACCENT).pack(anchor="w")
            ctk.CTkLabel(hi, text="Review Bank Only, Odoo Only, and Unreconciled discrepancies to upload to the Sales Portal.", font=(FONT_FAMILY, 11, "bold"), text_color=MUTED).pack(anchor="w")

            # Filter & Search Bar
            filt_frame = ctk.CTkFrame(top, fg_color=SIDEBAR_BG, corner_radius=0, height=62, border_color=BORDER, border_width=1)
            filt_frame.pack(fill="x")
            filt_frame.pack_propagate(False)

            fi = ctk.CTkFrame(filt_frame, fg_color="transparent")
            fi.pack(fill="both", expand=True, padx=16, pady=10)

            # Left controls: Type tabs + Bank pills
            f_left = ctk.CTkFrame(fi, fg_color="transparent")
            f_left.pack(side="left", fill="y")

            cur_type = ["ALL"]       # 'ALL', 'bank_only', 'odoo_only', 'unreconciled_odoo'
            cur_bank = ["ALL"]       # 'ALL', 'BCA', 'MANDIRI', 'BRI'
            cur_status = ["ALL"]     # 'ALL', 'READY', 'SYNCED' - default to ALL so all data is always visible immediately
            search_q = [""]

            type_btn_map = {}
            for t_key, t_lbl in [
                ("ALL", "All Types"),
                ("bank_only", "🏦 Bank Only"),
                ("odoo_only", "📦 Odoo Only"),
                ("unreconciled_odoo", "⚠️ Unreconciled"),
            ]:
                tb = ctk.CTkButton(
                    f_left, text=t_lbl, height=36, width=135 if t_key == "unreconciled_odoo" else (125 if t_key == "bank_only" else (120 if t_key == "odoo_only" else 95)),
                    fg_color=ACCENT if t_key == "ALL" else WHITE,
                    text_color=WHITE if t_key == "ALL" else TEXT,
                    border_color=ACCENT if t_key == "ALL" else BORDER_DARK,
                    border_width=1,
                    hover_color=ACCENT_DARK if t_key == "ALL" else PREVIEW_BG,
                    font=(FONT_FAMILY, 11, "bold"), corner_radius=6
                )
                tb.pack(side="left", padx=(0, 6))
                type_btn_map[t_key] = tb

            # Separator
            sep_m = tk.Frame(f_left, bg=BORDER_DARK, width=1, height=24)
            sep_m.pack(side="left", padx=8, pady=4)

            # Bank Selector Pills
            bank_btn_map = {}
            for bname in ["ALL", "BCA", "MANDIRI", "BRI"]:
                bb = ctk.CTkButton(
                    f_left, text=bname, height=36, width=55 if bname != "MANDIRI" else 80,
                    fg_color=ACCENT if bname == "ALL" else WHITE,
                    text_color=WHITE if bname == "ALL" else TEXT,
                    border_color=ACCENT if bname == "ALL" else BORDER_DARK,
                    border_width=1, hover_color=ACCENT_DARK if bname == "ALL" else PREVIEW_BG,
                    font=(FONT_FAMILY, 11, "bold"), corner_radius=6
                )
                bb.pack(side="left", padx=(0, 4))
                bank_btn_map[bname] = bb

            # Right controls: Status tabs + Search box
            f_right = ctk.CTkFrame(fi, fg_color="transparent")
            f_right.pack(side="right", fill="y")

            # Search Box
            search_v = tk.StringVar()
            search_ent = ctk.CTkEntry(
                f_right, textvariable=search_v, placeholder_text="🔍 Search...",
                height=36, width=180, corner_radius=6, border_color=BORDER_DARK, fg_color=WHITE, text_color=TEXT,
                font=(FONT_FAMILY, 11, "bold")
            )
            search_ent.pack(side="right", padx=(8, 0))

            # Status Tabs on right
            btn_st_all = ctk.CTkButton(
                f_right, text="All", height=36, width=65,
                fg_color=ACCENT, text_color=WHITE, border_color=ACCENT, border_width=1,
                hover_color=ACCENT_DARK, font=(FONT_FAMILY, 11, "bold"), corner_radius=6
            )
            btn_st_all.pack(side="left", padx=(0, 4))
            btn_st_ready = ctk.CTkButton(
                f_right, text="Ready to Send", height=36, width=120,
                fg_color=WHITE, text_color=TEXT, border_color=BORDER_DARK, border_width=1,
                hover_color=PREVIEW_BG, font=(FONT_FAMILY, 11, "bold"), corner_radius=6
            )
            btn_st_ready.pack(side="left", padx=(0, 4))
            btn_st_synced = ctk.CTkButton(
                f_right, text="Synced", height=36, width=85,
                fg_color=WHITE, text_color=TEXT, border_color=BORDER_DARK, border_width=1,
                hover_color=PREVIEW_BG, font=(FONT_FAMILY, 11, "bold"), corner_radius=6
            )
            btn_st_synced.pack(side="left", padx=(0, 4))

            # Scrollable Cards Area
            scroll_cards = ctk.CTkScrollableFrame(top, fg_color="transparent", scrollbar_button_color=BORDER, scrollbar_button_hover_color=MUTED)
            scroll_cards.pack(fill="both", expand=True, padx=20, pady=8)

            # Footer Action Bar
            footer = ctk.CTkFrame(top, fg_color=PANEL, corner_radius=0, height=64, border_color=BORDER, border_width=1)
            footer.pack(fill="x", side="bottom")
            footer.pack_propagate(False)

            foot_inner = ctk.CTkFrame(footer, fg_color="transparent")
            foot_inner.pack(fill="both", expand=True, padx=20, pady=10)

            btn_select_all = ctk.CTkButton(foot_inner, text="Select All Available", height=38, width=175, fg_color=WHITE, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1, text_color=TEXT, font=(FONT_FAMILY, 11, "bold"), corner_radius=6)
            btn_select_all.pack(side="left")

            lbl_summary = ctk.CTkLabel(foot_inner, text="Selected 0 items", font=(FONT_FAMILY, 12, "bold"), text_color=TEXT)
            lbl_summary.pack(side="left", padx=(16, 0))

            def _close_modal():
                try:
                    top.grab_release()
                except Exception:
                    pass
                try:
                    top.destroy()
                except Exception:
                    pass

            top.protocol("WM_DELETE_WINDOW", _close_modal)
            btn_close = ctk.CTkButton(foot_inner, text="Close", height=38, width=90, fg_color=WHITE, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1, text_color=TEXT, font=(FONT_FAMILY, 11, "bold"), corner_radius=6, command=_close_modal)
            btn_close.pack(side="right")

            btn_submit = ctk.CTkButton(foot_inner, text="Send Selected to Sales Portal (0)", height=38, fg_color=BORDER_DARK, hover_color=ACCENT_DARK, text_color=WHITE, font=(FONT_FAMILY, 12, "bold"), corner_radius=6, state="disabled")
            btn_submit.pack(side="right", padx=(0, 12))

            # Checkbox variables per item (default all deselected on initial load)
            item_checks = {}
            for idx, item in enumerate(discrepancy_items):
                item_checks[idx] = tk.BooleanVar(value=False)

            BANK_BADGE_COLS = {
                "BCA":     ("#1E40AF", "#DBEAFE"),
                "MANDIRI": ("#92400E", "#FEF3C7"),
                "BRI":     ("#047857", "#D1FAE5"),
                "OTHER":   ("#4B5563", "#F3F4F6"),
            }

            TYPE_BADGE_INFO = {
                "bank_only":         ("🏦 Bank Only", "#4338CA", "#EEF2FF"),
                "odoo_only":         ("📦 Odoo Only", "#B45309", "#FEF3C7"),
                "unreconciled_odoo": ("⚠️ Unreconciled", "#BE123C", "#FFE4E6"),
            }

            def _update_counts():
                selected_count = 0
                selected_sum = 0.0
                available_count = 0
                for idx, item in enumerate(discrepancy_items):
                    if not item["is_synced"]:
                        available_count += 1
                        if item_checks[idx].get():
                            selected_count += 1
                            selected_sum += item["amount"]
                lbl_summary.configure(text=f"Selected {selected_count} item(s) • Total: Rp {selected_sum:,.2f}")
                btn_submit.configure(text=f"Send Selected to Sales Portal ({selected_count})")
                if selected_count == 0:
                    btn_submit.configure(state="disabled", fg_color=BORDER_DARK)
                else:
                    btn_submit.configure(state="normal", fg_color=ACCENT)

                if available_count > 0 and selected_count == available_count:
                    btn_select_all.configure(text="Deselect All")
                else:
                    btn_select_all.configure(text="Select All Available")

            def _toggle_select_all():
                available = [i for i, item in enumerate(discrepancy_items) if not item["is_synced"]]
                all_on = all(item_checks[i].get() for i in available) if available else False
                new_val = not all_on
                for i in available:
                    item_checks[i].set(new_val)
                _update_counts()

            btn_select_all.configure(command=_toggle_select_all)

            display_limit = [100]

            def _render_modal_cards():
                for w in scroll_cards.winfo_children():
                    w.destroy()

                # Filter items
                visible = []
                for idx, item in enumerate(discrepancy_items):
                    # Discrepancy Type filter
                    if cur_type[0] != "ALL" and item["discrepancy_type"] != cur_type[0]:
                        continue

                    # Status filter
                    if cur_status[0] == "READY" and item["is_synced"]:
                        continue
                    if cur_status[0] == "SYNCED" and not item["is_synced"]:
                        continue

                    # Bank filter
                    bname = item["bank"].upper()
                    if cur_bank[0] != "ALL" and bname != cur_bank[0]:
                        continue

                    # Search filter
                    if search_q[0]:
                        s_str = f"{item['bank']} {item['journal']} {item['number_bank']} {item['number_odo']} {item['invoice_no']} {item['discrepancy_type']} {item['amount']}".lower()
                        if search_q[0] not in s_str:
                            continue

                    visible.append((idx, item))

                # Update count badges
                ready_cnt = sum(1 for it in discrepancy_items if not it["is_synced"])
                synced_cnt = sum(1 for it in discrepancy_items if it["is_synced"])
                btn_st_ready.configure(text=f"Ready ({ready_cnt})")
                btn_st_synced.configure(text=f"Synced ({synced_cnt})")
                btn_st_all.configure(text=f"All ({len(discrepancy_items)})")

                if not visible:
                    emp = tk.Frame(scroll_cards, bg=PANEL, highlightbackground=BORDER, highlightthickness=1)
                    emp.pack(fill="x", padx=16, pady=20)
                    tk.Label(emp, text="No items match current filter criteria.", bg=PANEL, fg=MUTED, font=(FONT_FAMILY, 10, "bold")).pack(pady=16)
                    return

                slice_items = visible[:display_limit[0]]

                for idx, item in slice_items:
                    card = tk.Frame(scroll_cards, bg=PANEL, highlightbackground=BORDER, highlightthickness=1, bd=0)
                    card.pack(fill="x", padx=4, pady=3)
                    
                    # Rigid column widths for perfect alignment
                    card.columnconfigure(0, minsize=34)   # Col 0: Checkbox
                    card.columnconfigure(1, minsize=118)  # Col 1: Live Odoo Button
                    card.columnconfigure(2, minsize=120)  # Col 2: Discrepancy Type Badge
                    card.columnconfigure(3, minsize=80)   # Col 3: Bank Badge
                    card.columnconfigure(4, minsize=95)   # Col 4: Date
                    card.columnconfigure(5, minsize=140)  # Col 5: Journal
                    card.columnconfigure(6, weight=1)     # Col 6: Ref / Doc (Expands & absorbs available width)
                    card.columnconfigure(7, minsize=155)  # Col 7: Amount
                    card.columnconfigure(8, minsize=120)  # Col 8: Status

                    # Col 0: Checkbox
                    if not item["is_synced"]:
                        cb = tk.Checkbutton(
                            card, variable=item_checks[idx], bg=PANEL, activebackground=PANEL,
                            highlightthickness=0, bd=0, command=_update_counts
                        )
                        cb.grid(row=0, column=0, padx=(10, 4), pady=7, sticky="w")
                    else:
                        lbl_lock = tk.Label(card, text="Synced", bg=PANEL, fg=MUTED, font=(FONT_FAMILY, 9, "bold"))
                        lbl_lock.grid(row=0, column=0, padx=(8, 4), pady=7, sticky="w")

                    # Col 1: Live Odoo Inspection Button (Second Column)
                    btn_chk = ctk.CTkButton(
                        card, text="🔍 Check Odoo", width=104, height=28,
                        fg_color=WHITE, hover_color=PREVIEW_BG,
                        border_color=BORDER_DARK, border_width=1,
                        text_color=TEXT, font=(FONT_FAMILY, 9, "bold"), corner_radius=6,
                        command=lambda it=item: self._open_discrepancy_inspection_modal(it, parent_win=top)
                    )
                    btn_chk.grid(row=0, column=1, padx=(0, 8), pady=4, sticky="w")

                    # Col 2: Type badge
                    t_lbl, t_fg, t_bg = TYPE_BADGE_INFO.get(item["discrepancy_type"], ("Discrepancy", TEXT, BG))
                    lbl_tbadge = tk.Label(card, text=f" {t_lbl} ", bg=t_bg, fg=t_fg, font=(FONT_FAMILY, 9, "bold"), padx=5, pady=3, width=14, anchor="center")
                    lbl_tbadge.grid(row=0, column=2, padx=(0, 6), pady=7, sticky="w")

                    # Col 3: Bank badge
                    bname = item["bank"].upper()
                    fg_b, bg_b = BANK_BADGE_COLS.get(bname, BANK_BADGE_COLS["OTHER"])
                    lbl_badge = tk.Label(card, text=f" {bname} ", bg=bg_b, fg=fg_b, font=(FONT_FAMILY, 9, "bold"), padx=5, pady=3, width=7, anchor="center")
                    lbl_badge.grid(row=0, column=3, padx=(0, 6), pady=7, sticky="w")

                    # Col 4: Date
                    lbl_d = tk.Label(card, text=item["date"], bg=PANEL, fg=TEXT, font=(FONT_FAMILY, 10, "bold"), width=10, anchor="w")
                    lbl_d.grid(row=0, column=4, padx=(0, 6), pady=7, sticky="w")

                    # Col 5: Journal
                    lbl_j = tk.Label(card, text=item["journal"], bg=PANEL, fg=MUTED, font=(FONT_FAMILY, 9), width=16, anchor="w")
                    lbl_j.grid(row=0, column=5, padx=(0, 6), pady=7, sticky="w")

                    # Col 6: Bank Ref or Odoo Doc (Unclipped, expands into available space)
                    if item["discrepancy_type"] == "bank_only":
                        b_ref = item["number_bank"] or "-"
                        ref_text = f"Ref: {b_ref}"
                    elif item["discrepancy_type"] == "unreconciled_odoo":
                        o_doc = item["number_odo"] or item["invoice_no"] or "-"
                        ref_text = f"Unrecon: {o_doc}"
                    else:
                        o_doc = item["number_odo"] or item["invoice_no"] or "-"
                        ref_text = f"Doc: {o_doc}"

                    lbl_ref = tk.Label(card, text=ref_text, bg=PANEL, fg=TEXT, font=(FONT_MONO, 10, "bold"), anchor="w")
                    lbl_ref.grid(row=0, column=6, padx=(0, 10), pady=7, sticky="ew")

                    # Col 7: Amount
                    lbl_amt = tk.Label(card, text=f"Rp {item['amount']:,.2f}", bg=PANEL, fg=TEXT, font=(FONT_FAMILY, 11, "bold"), width=16, anchor="e")
                    lbl_amt.grid(row=0, column=7, padx=(0, 10), pady=7, sticky="e")

                    # Col 8: Status
                    if item["is_synced"]:
                        lbl_st = tk.Label(card, text="● Synced", bg=PANEL, fg="#059669", font=(FONT_FAMILY, 9, "bold"), width=12, anchor="w")
                    else:
                        lbl_st = tk.Label(card, text="● Ready to Sync", bg=PANEL, fg="#D97706", font=(FONT_FAMILY, 9, "bold"), width=12, anchor="w")
                    lbl_st.grid(row=0, column=8, padx=(0, 10), pady=7, sticky="w")

                    # Hover styling
                    labels_hover = [lbl_d, lbl_j, lbl_ref, lbl_amt, lbl_st]
                    if item["is_synced"]:
                        labels_hover.append(lbl_lock)

                    def _make_hover(c_widget=card, labels=labels_hover, it=item):
                        def _enter(e):
                            c_widget.config(bg=PREVIEW_BG, highlightbackground=ACCENT)
                            for l in labels: l.config(bg=PREVIEW_BG)
                        def _leave(e):
                            c_widget.config(bg=PANEL, highlightbackground=BORDER)
                            for l in labels: l.config(bg=PANEL)
                        def _on_double_click(e):
                            self._open_discrepancy_inspection_modal(it, parent_win=top)
                        c_widget.bind("<Enter>", _enter)
                        c_widget.bind("<Leave>", _leave)
                        c_widget.bind("<Double-Button-1>", _on_double_click)
                        for l in labels:
                            l.bind("<Double-Button-1>", _on_double_click)
                    _make_hover()

                # Load More button if more items exist
                if len(visible) > len(slice_items):
                    remaining = len(visible) - len(slice_items)
                    more_frame = tk.Frame(scroll_cards, bg=BG)
                    more_frame.pack(fill="x", pady=10)
                    def _show_more():
                        display_limit[0] += 100
                        _render_modal_cards()
                    btn_more = ctk.CTkButton(
                        more_frame, text=f"Load Next 100 Items (Showing {len(slice_items)} of {len(visible)} items — {remaining} remaining)",
                        height=32, fg_color=WHITE, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1,
                        text_color=TEXT, font=(FONT_FAMILY, 9, "bold"), command=_show_more
                    )
                    btn_more.pack(pady=4)

                _update_counts()

                # Reset scroll position to top
                def _reset_scroll():
                    try:
                        scroll_cards._parent_canvas.yview_moveto(0.0)
                    except Exception:
                        pass
                _reset_scroll()
                top.after(10, _reset_scroll)
                top.after(50, _reset_scroll)

            def _switch_type(tk_val):
                display_limit[0] = 100
                cur_type[0] = tk_val
                for k, btn in type_btn_map.items():
                    if k == tk_val:
                        btn.configure(fg_color=ACCENT, text_color=WHITE, border_color=ACCENT, hover_color=ACCENT_DARK)
                    else:
                        btn.configure(fg_color=WHITE, text_color=TEXT, border_color=BORDER_DARK, hover_color=PREVIEW_BG)
                _render_modal_cards()

            for tk_val, btn in type_btn_map.items():
                btn.configure(command=lambda t=tk_val: _switch_type(t))

            def _switch_status(st):
                display_limit[0] = 100
                cur_status[0] = st
                btn_st_ready.configure(
                    fg_color=ACCENT if st == "READY" else WHITE,
                    text_color=WHITE if st == "READY" else TEXT,
                    border_color=ACCENT if st == "READY" else BORDER_DARK,
                    hover_color=ACCENT_DARK if st == "READY" else PREVIEW_BG
                )
                btn_st_synced.configure(
                    fg_color=ACCENT if st == "SYNCED" else WHITE,
                    text_color=WHITE if st == "SYNCED" else TEXT,
                    border_color=ACCENT if st == "SYNCED" else BORDER_DARK,
                    hover_color=ACCENT_DARK if st == "SYNCED" else PREVIEW_BG
                )
                btn_st_all.configure(
                    fg_color=ACCENT if st == "ALL" else WHITE,
                    text_color=WHITE if st == "ALL" else TEXT,
                    border_color=ACCENT if st == "ALL" else BORDER_DARK,
                    hover_color=ACCENT_DARK if st == "ALL" else PREVIEW_BG
                )
                _render_modal_cards()

            btn_st_ready.configure(command=lambda: _switch_status("READY"))
            btn_st_synced.configure(command=lambda: _switch_status("SYNCED"))
            btn_st_all.configure(command=lambda: _switch_status("ALL"))

            def _switch_bank(bn):
                display_limit[0] = 100
                cur_bank[0] = bn
                for k, btn in bank_btn_map.items():
                    if k == bn:
                        btn.configure(fg_color=ACCENT, text_color=WHITE, border_color=ACCENT)
                    else:
                        btn.configure(fg_color=WHITE, text_color=TEXT, border_color=BORDER_DARK)
                _render_modal_cards()

            for bn, btn in bank_btn_map.items():
                btn.configure(command=lambda b=bn: _switch_bank(b))

            def _on_search_change(*args):
                display_limit[0] = 100
                search_q[0] = search_v.get().strip().lower()
                _render_modal_cards()

            search_v.trace_add("write", _on_search_change)

            # Send Action
            def _do_push_to_sales():
                to_send = []
                for idx, item in enumerate(discrepancy_items):
                    if not item["is_synced"] and item_checks[idx].get():
                        to_send.append(item)

                if not to_send:
                    return

                btn_submit.configure(state="disabled", text="⏳ Sending to Cloud...")
                top.update_idletasks()

                recon_date = to_send[0]["date"] if to_send else ""
                res = push_bank_discrepancies(to_send, recon_date=recon_date)

                if res.get("success"):
                    self._log_write(f"\n✅ Uploaded {res.get('count', len(to_send))} discrepancies to Sales Portal!\n", "ok")
                    self._set_status(f"Sent {res.get('count')} items to Sales Portal", SUCCESS)
                    # Mark as synced locally
                    for item in to_send:
                        item["is_synced"] = True
                    _switch_status("SYNCED")
                    _render_modal_cards()
                else:
                    self._log_write(f"\n❌ Failed to send to cloud: {res.get('error')}\n", "err")
                    self._set_status(f"Upload failed: {res.get('error')}", ERROR)
                    btn_submit.configure(state="normal", text="📤 Send Selected to Sales Portal")

            btn_submit.configure(command=_do_push_to_sales)
            _render_modal_cards()
            _update_counts()

            # Background cloud check to flag already-synced items without freezing the UI
            import threading
            def _bg_check_synced():
                try:
                    existing = fetch_discrepancies(limit=1000)
                    if not existing or not top.winfo_exists():
                        return
                    cloud_keys = set()
                    for ci in existing:
                        d = str(ci.get("transaction_date", "")).strip()
                        b = str(ci.get("bank_name", "")).strip().upper()
                        num = str(ci.get("bank_number", "")).strip()
                        onum = str(ci.get("odoo_number", "")).strip()
                        dtype = str(ci.get("discrepancy_type", "bank_only")).strip()
                        a = round(float(ci.get("amount", 0.0)), 2)
                        if num or onum:
                            cloud_keys.add((d, b, num, onum, dtype, a))
                            if num: cloud_keys.add((d, b, num, "", dtype, a))
                            if onum: cloud_keys.add((d, b, "", onum, dtype, a))
                        else:
                            cloud_keys.add((d, b, "", "", dtype, a))

                    changed = False
                    for item in discrepancy_items:
                        d = item["date"]
                        b = item["bank"].upper()
                        num = item["number_bank"]
                        onum = item["number_odo"]
                        dtype = item["discrepancy_type"]
                        a = round(item["amount"], 2)
                        is_match = False
                        if (d, b, num, onum, dtype, a) in cloud_keys:
                            is_match = True
                        elif num and (d, b, num, "", dtype, a) in cloud_keys:
                            is_match = True
                        elif onum and (d, b, "", onum, dtype, a) in cloud_keys:
                            is_match = True

                        if is_match:
                            item["is_synced"] = True
                            changed = True

                    if changed and top.winfo_exists():
                        top.after(0, _render_modal_cards)
                        top.after(0, _update_counts)
                except Exception:
                    pass

            threading.Thread(target=_bg_check_synced, daemon=True).start()

        finally:
            self._sync_cloud_btn.configure(text="☁️  Sync Sales Portal", state="normal")

    # ── Live Odoo Diagnostics Modal for Finance ──────────────────────────────
    def _open_discrepancy_inspection_modal(self, item: dict, parent_win=None):
        import odoo_inspector
        import threading

        parent = parent_win if (parent_win and parent_win.winfo_exists()) else self
        dlg = ctk.CTkToplevel(parent)
        dlg.title("Odoo Discrepancy Diagnostics (Finance)")
        dlg.geometry("820x460")
        dlg.minsize(760, 400)
        dlg.resizable(True, True)
        dlg.transient(parent)

        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        cx, cy = max(0, int(sw / 2 - 820 / 2)), max(0, int(sh / 2 - 460 / 2))
        dlg.geometry(f"820x460+{cx}+{cy}")
        dlg.configure(fg_color=BG)

        def _safe_grab():
            try:
                if dlg.winfo_exists():
                    dlg.grab_set()
            except Exception:
                pass
        dlg.after(100, _safe_grab)

        scroll_root = ctk.CTkScrollableFrame(dlg, fg_color="transparent")
        scroll_root.pack(fill="both", expand=True, padx=16, pady=16)

        content = ctk.CTkFrame(scroll_root, fg_color=PANEL, corner_radius=10, border_color=BORDER, border_width=1)
        content.pack(fill="both", expand=True)

        # ── 1. Header Info ──
        hdr = ctk.CTkFrame(content, fg_color=PREVIEW_BG, corner_radius=8, border_color=BORDER, border_width=1)
        hdr.pack(fill="x", padx=16, pady=(16, 12))

        amt = float(item.get("amount", 0.0))
        t_date = str(item.get("date", item.get("transaction_date", ""))).strip()
        b_name = str(item.get("bank", item.get("bank_name", ""))).upper()
        journal = str(item.get("journal", "")).strip()
        dtype = str(item.get("discrepancy_type") or "bank_only").strip()
        badge_map = {
            "bank_only":         ("🏦 Bank Only", "#4338CA", "#EEF2FF"),
            "odoo_only":         ("📦 Odoo Only", "#B45309", "#FEF3C7"),
            "unreconciled_odoo": ("⚠️ Unreconciled", "#BE123C", "#FFE4E6"),
        }
        t_lbl, t_fg, t_bg = badge_map.get(dtype, ("Discrepancy", TEXT, BG))

        ctk.CTkLabel(hdr, text=f"Rp {amt:,.2f}  •  {t_lbl}", font=(FONT_FAMILY, 16, "bold"), text_color=TEXT).pack(anchor="w", padx=14, pady=(10, 3))
        ctk.CTkLabel(hdr, text=f"{b_name} • {journal} • Date: {t_date}", font=(FONT_FAMILY, 11, "bold"), text_color=ACCENT).pack(anchor="w", padx=14)

        b_num = str(item.get("number_bank") or item.get("bank_number") or "-").strip()
        o_num = str(item.get("number_odo") or item.get("odoo_number") or item.get("invoice_no") or "").strip()
        ref_txt = f"Bank Ref: {b_num}"
        if o_num:
            ref_txt += f"  •  Odoo Doc: {o_num}"
        ctk.CTkLabel(hdr, text=ref_txt, font=(FONT_MONO, 11), text_color=MUTED).pack(anchor="w", padx=14, pady=(3, 10))

        # ── 2. Odoo Live Diagnostics Assistant ──
        diag_frame = ctk.CTkFrame(content, fg_color="#F8FAFC", corner_radius=8, border_color=BORDER_DARK, border_width=1)
        diag_frame.pack(fill="x", padx=16, pady=(0, 16))

        diag_hdr = ctk.CTkFrame(diag_frame, fg_color="transparent")
        diag_hdr.pack(fill="x", padx=14, pady=(10, 6))

        ctk.CTkLabel(diag_hdr, text="🔍 Odoo Live Diagnostics (Read-Only)", font=(FONT_FAMILY, 12, "bold"), text_color=ACCENT_DARK).pack(side="left")
        lbl_status = ctk.CTkLabel(diag_hdr, text="⏳ Checking Odoo...", font=(FONT_FAMILY, 11, "bold"), text_color=MUTED)
        lbl_status.pack(side="right")

        diag_body = ctk.CTkFrame(diag_frame, fg_color="transparent")
        diag_body.pack(fill="x", padx=14, pady=(0, 14))

        btn_bar = ctk.CTkFrame(content, fg_color="transparent")
        btn_bar.pack(fill="x", side="bottom", padx=16, pady=(0, 16))

        btn_close = ctk.CTkButton(
            btn_bar, text="Close", height=32, width=90,
            fg_color=WHITE, hover_color=PREVIEW_BG,
            border_color=BORDER_DARK, border_width=1,
            text_color=TEXT, font=(FONT_FAMILY, 10, "bold"),
            command=dlg.destroy
        )
        btn_close.pack(side="right")

        inspect_item = {
            "discrepancy_type": dtype,
            "amount": amt,
            "transaction_date": t_date,
            "odoo_number": o_num,
            "bank_number": b_num,
            "bank_name": b_name,
            "journal": journal
        }

        def _apply_diag_ui(res: dict):
            if not dlg.winfo_exists():
                return
            for w in diag_body.winfo_children():
                w.destroy()

            if not res.get("success"):
                lbl_status.configure(text="⚠️ Odoo Check Offline", text_color=WARN)
                err_msg = res.get("error") or "Could not connect to Odoo server."
                ctk.CTkLabel(diag_body, text=err_msg, font=(FONT_FAMILY, 11), text_color=MUTED, wraplength=640, justify="left").pack(anchor="w", pady=4)
                return

            lbl_status.configure(text="✅ Checked Live", text_color=SUCCESS)

            if dtype == "bank_only":
                invoices = res.get("invoices", [])
                summary = res.get("summary", "")
                ctk.CTkLabel(diag_body, text=summary, font=(FONT_FAMILY, 12, "bold"), text_color=TEXT, wraplength=640, justify="left").pack(anchor="w", pady=(0, 8))

                if invoices:
                    for inv in invoices:
                        inv_card = ctk.CTkFrame(diag_body, fg_color=WHITE, corner_radius=8, border_color=BORDER, border_width=1)
                        inv_card.pack(fill="x", pady=4)

                        top_row = ctk.CTkFrame(inv_card, fg_color="transparent")
                        top_row.pack(fill="x", padx=12, pady=(10, 4))

                        badge_color = SUCCESS if inv["status_code"] == "AVAILABLE_OPEN" else (WARN if inv["status_code"] == "DRAFT_INVOICE" else ERROR)
                        ctk.CTkLabel(top_row, text=f" {inv['badge']} ", font=(FONT_FAMILY, 10, "bold"), text_color=WHITE, fg_color=badge_color, corner_radius=4).pack(side="left", padx=(0, 8))
                        ctk.CTkLabel(top_row, text=inv['name'], font=(FONT_MONO, 13, "bold"), text_color=TEXT).pack(side="left")
                        ctk.CTkLabel(top_row, text=f"Rp {inv['amount']:,.2f}", font=(FONT_FAMILY, 13, "bold"), text_color=SUCCESS_DARK).pack(side="right")

                        desc = f"Customer: {inv['customer']}  •  Date: {inv['date']}"
                        ctk.CTkLabel(inv_card, text=desc, font=(FONT_FAMILY, 11, "bold"), text_color=MUTED).pack(anchor="w", padx=12, pady=(0, 4))
                        if inv.get("detail"):
                            ctk.CTkLabel(inv_card, text=f"↳ {inv['detail']}", font=(FONT_FAMILY, 11), text_color=TEXT, wraplength=620, justify="left").pack(anchor="w", padx=12, pady=(0, 10))

            elif dtype == "odoo_only":
                summary = res.get("summary", "")
                linked_invoices = res.get("linked_invoices", [])
                ctk.CTkLabel(diag_body, text=summary, font=(FONT_FAMILY, 12, "bold"), text_color=TEXT, wraplength=640, justify="left").pack(anchor="w", pady=(0, 8))

                if linked_invoices:
                    for li in linked_invoices:
                        li_box = ctk.CTkFrame(diag_body, fg_color=WHITE, corner_radius=8, border_color=BORDER, border_width=1)
                        li_box.pack(fill="x", pady=4)

                        r1 = ctk.CTkFrame(li_box, fg_color="transparent")
                        r1.pack(fill="x", padx=12, pady=(10, 4))
                        ctk.CTkLabel(r1, text="🔗 Linked Invoice:", font=(FONT_FAMILY, 11, "bold"), text_color=ACCENT).pack(side="left", padx=(0, 6))
                        ctk.CTkLabel(r1, text=li['name'], font=(FONT_MONO, 13, "bold"), text_color=TEXT).pack(side="left")
                        ctk.CTkLabel(r1, text=f"Rp {li['amount']:,.2f}", font=(FONT_FAMILY, 13, "bold"), text_color=TEXT).pack(side="right")

                        desc = f"Customer: {li['customer']}  •  Invoice Date: {li['date']}  •  State: {li['state']}"
                        ctk.CTkLabel(li_box, text=desc, font=(FONT_FAMILY, 11, "bold"), text_color=MUTED).pack(anchor="w", padx=12, pady=(0, 10))
                else:
                    no_inv_box = ctk.CTkFrame(diag_body, fg_color=WHITE, corner_radius=6, border_color=BORDER, border_width=1)
                    no_inv_box.pack(fill="x", pady=3)
                    ctk.CTkLabel(no_inv_box, text="⚪ No linked invoice found on this payment document in Odoo.", font=(FONT_FAMILY, 11, "bold"), text_color=MUTED).pack(anchor="w", padx=12, pady=10)

            elif dtype == "unreconciled_odoo":
                summary = res.get("summary", "")
                invoices = res.get("invoices", [])
                ctk.CTkLabel(diag_body, text=summary, font=(FONT_FAMILY, 12, "bold"), text_color=TEXT, wraplength=640, justify="left").pack(anchor="w", pady=(0, 8))

                if invoices:
                    for inv in invoices:
                        inv_card = ctk.CTkFrame(diag_body, fg_color=WHITE, corner_radius=8, border_color=BORDER, border_width=1)
                        inv_card.pack(fill="x", pady=4)

                        r1 = ctk.CTkFrame(inv_card, fg_color="transparent")
                        r1.pack(fill="x", padx=12, pady=(10, 4))
                        ctk.CTkLabel(r1, text=f" {inv['badge']} ", font=(FONT_FAMILY, 10, "bold"), text_color=WHITE, fg_color=ERROR if inv.get("other_payments") else SUCCESS, corner_radius=4).pack(side="left", padx=(0, 8))
                        ctk.CTkLabel(r1, text=f"Invoice: {inv['name']}", font=(FONT_MONO, 13, "bold"), text_color=TEXT).pack(side="left")
                        ctk.CTkLabel(r1, text=f"Rp {inv['amount']:,.2f}", font=(FONT_FAMILY, 13, "bold"), text_color=TEXT).pack(side="right")

                        desc = f"Customer: {inv['customer']}  •  Date: {inv['date']}"
                        ctk.CTkLabel(inv_card, text=desc, font=(FONT_FAMILY, 11, "bold"), text_color=MUTED).pack(anchor="w", padx=12, pady=(0, 4))

                        if inv.get("other_payments"):
                            other_box = ctk.CTkFrame(inv_card, fg_color=ERROR_LIGHT, corner_radius=6, border_color=ERROR, border_width=1)
                            other_box.pack(fill="x", padx=12, pady=(4, 10))
                            ctk.CTkLabel(other_box, text="⚠️ Already linked to other payment(s):", font=(FONT_FAMILY, 11, "bold"), text_color=ERROR).pack(anchor="w", padx=8, pady=(6, 2))
                            for op in inv["other_payments"]:
                                op_txt = f"• {op['ref']}  |  Date: {op['date']}  |  Journal: {op['journal']}  |  Rp {op['amount']:,.2f}"
                                ctk.CTkLabel(other_box, text=op_txt, font=(FONT_FAMILY, 11, "bold"), text_color=TEXT, justify="left").pack(anchor="w", padx=12, pady=1)
                        else:
                            ctk.CTkLabel(inv_card, text="🟢 Invoice is not linked to any other customer payment.", font=(FONT_FAMILY, 11, "bold"), text_color=SUCCESS).pack(anchor="w", padx=12, pady=(2, 10))

            else:
                summary = res.get("summary", "")
                ctk.CTkLabel(diag_body, text=summary, font=(FONT_FAMILY, 9), text_color=MUTED).pack(anchor="w")

        def _bg_inspect():
            try:
                res = odoo_inspector.inspect_discrepancy(inspect_item)
            except Exception as e:
                res = {"success": False, "error": str(e)}
            if dlg.winfo_exists():
                dlg.after(0, lambda: _apply_diag_ui(res))

        threading.Thread(target=_bg_inspect, daemon=True).start()

    # ── Manual Match & Discrepancy Reconciler Modal ─────────────────────────
    def _on_manual_match(self):
        import glob
        import os
        from openpyxl import load_workbook

        output_files = glob.glob(str(OUTPUT_DIR / "[Rr]econciliation_*.xlsx"))
        if not output_files:
            self._set_status("No reconciliation file found. Run scan/recon first!", ERROR)
            return
        latest_file = max(output_files, key=os.path.getctime)

        try:
            wb = load_workbook(latest_file, data_only=True)
            try:
                if "Differences" not in wb.sheetnames:
                    self._set_status("No 'Differences' sheet found in recon file", ERROR)
                    return

                ws = wb["Differences"]
                rows = list(ws.iter_rows(values_only=True))
            finally:
                wb.close()
        except Exception as e:
            self._set_status(f"Error loading recon file: {e}", ERROR)
        # Dynamic header mapping for backwards and forwards compatibility
        hdr_row = [str(cell.value or "").strip().lower() for cell in ws[3]]
        def _find_c(name, fallback):
            for i, h in enumerate(hdr_row):
                if name in h: return i
            return fallback

        c_date     = _find_c("date", 1)
        c_bank     = _find_c("bank", 2)
        c_journal  = _find_c("journal", 3)
        c_odo_num  = _find_c("odoo number", 4)
        c_ref      = _find_c("reference", 5)
        c_bank_num = _find_c("bank number", 6)
        c_file     = _find_c("filename", 7)
        c_bank_amt = _find_c("bank amount", 8)
        c_odo_amt  = _find_c("odoo amount", 9 if len(hdr_row) > 13 else 8)
        c_src      = _find_c("source", 10 if len(hdr_row) > 13 else 9)
        c_recon    = _find_c("reconciled", 11 if len(hdr_row) > 13 else 10)
        c_status   = _find_c("status", 12 if len(hdr_row) > 13 else 11)

        bank_items = []
        odo_items = []
        from collections import defaultdict
        odo_by_key = defaultdict(list)

        for r_idx, r in enumerate(rows[3:], 4):
            if not r or len(r) <= c_status:
                continue
            st = str(r[c_status] or "").strip()
            if not st or st.startswith("DIFFERENCES") or st.startswith("Date:"):
                continue

            src = str(r[c_src] or "").strip()
            # Extract amount based on source column
            if src == "Bank":
                amt = float(r[c_bank_amt] or r[c_odo_amt] or 0.0)
            else:
                amt = float(r[c_odo_amt] or r[c_bank_amt] or 0.0)

            row_dict = {
                "row_idx": r_idx,
                "no": r[0],
                "date": str(r[c_date] or "").strip(),
                "bank": str(r[c_bank] or "").strip(),
                "journal": str(r[c_journal] or "").strip(),
                "number_odo": str(r[c_odo_num] or "").strip(),
                "reference": str(r[c_ref] or "").strip(),
                "number_bank": str(r[c_bank_num] or "").strip(),
                "filename": str(r[c_file] or "").strip(),
                "amount": amt,
                "source": src,
                "reconciled": str(r[c_recon] or "").strip(),
                "status": st,
            }
            if st == "Only in Bank":
                bank_items.append(row_dict)
            elif st == "Only in Odoo":
                odo_items.append(row_dict)
                odo_by_key[(row_dict["date"], row_dict["journal"].lower())].append(row_dict)

        if not bank_items and not odo_items:
            self._set_status("No discrepancy items found in Differences sheet!", SUCCESS)
            return

        candidate_pairs = []
        for b in bank_items:
            key = (b["date"], b["journal"].lower())
            for o in odo_by_key.get(key, []):
                diff = b["amount"] - o["amount"]
                abs_diff = abs(diff)
                max_val = max(b["amount"], o["amount"], 1.0)
                pct = (abs_diff / max_val) * 100.0
                
                s1 = str(int(round(b["amount"])))
                s2 = str(int(round(o["amount"])))
                is_transposition = (len(s1) == len(s2) and sorted(s1) == sorted(s2))

                # Strict Pattern Detection: low variance (pct <= 3.0%), fee cap (abs_diff <= 50,000), or digit transposition
                if pct <= 3.0 or (pct <= 8.0 and abs_diff <= 50000) or is_transposition:
                    candidate_pairs.append({
                        "bank": b,
                        "odoo": o,
                        "diff": diff,
                        "abs_diff": abs_diff,
                        "pct": pct,
                        "date": b["date"],
                        "journal": b["journal"],
                    })

        def _parse_date_key(d_str: str):
            try:
                if "/" in d_str:
                    parts = d_str.split("/")
                    if len(parts) == 3:
                        return (int(parts[2]), int(parts[1]), int(parts[0]))
                elif "-" in d_str:
                    parts = d_str.split("-")
                    if len(parts) == 3:
                        return (int(parts[0]), int(parts[1]), int(parts[2]))
            except Exception:
                pass
            return (9999, 12, 31)

        bank_items.sort(key=lambda r: (_parse_date_key(r["date"]), r["row_idx"]))
        odo_items.sort(key=lambda r: (_parse_date_key(r["date"]), r["row_idx"]))
        # Rank by Date ASC -> Variance % ASC (smallest typo % first) -> abs_diff ASC
        candidate_pairs.sort(key=lambda x: (_parse_date_key(x["date"]), x["pct"], x["abs_diff"]))

        top = ctk.CTkToplevel(self)
        top.withdraw()
        top.title("Manual Match — Reconcile Differences")

        top.minsize(1050, 640)
        top.configure(fg_color=BG)
        top.transient(self)
        top.grab_set()
        top.after(200, lambda: _center_modal_on_parent(top, self))

        active_matched = []
        auto_page = [0]
        CANDIDATES_PER_PAGE = 10

        hdr = ctk.CTkFrame(top, fg_color=PANEL, height=76, corner_radius=0, border_color=BORDER, border_width=1)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)

        hdr_in = ctk.CTkFrame(hdr, fg_color="transparent")
        hdr_in.pack(fill="both", expand=True, padx=24, pady=14)

        ctk.CTkLabel(
            hdr_in, text="🧩 Manual Reconciliation & Pair Matching",
            font=(FONT_FAMILY, 16, "bold"), text_color=ACCENT
        ).pack(anchor="w")
        ctk.CTkLabel(
            hdr_in, text=f"Review candidate pairs or manually match Bank vs Odoo discrepancies. Loaded {len(bank_items)} Bank & {len(odo_items)} Odoo items.",
            font=(FONT_FAMILY, 10, "bold"), text_color=MUTED
        ).pack(anchor="w", pady=(2, 0))

        lbl_modal_msg = ctk.CTkLabel(
            hdr_in, text="",
            font=(FONT_FAMILY, 10, "bold"), text_color=ERROR
        )
        lbl_modal_msg.pack(anchor="w", pady=(2, 0))

        def show_modal_msg(msg: str, is_err=True):
            lbl_modal_msg.configure(text=msg, text_color=ERROR if is_err else SUCCESS)
            top.after(5000, lambda: lbl_modal_msg.configure(text=""))

        # Modal Footer Toolbar (packed at bottom before body expansion so it never gets clipped)
        footer = ctk.CTkFrame(top, fg_color=SIDEBAR_BG, height=60, corner_radius=0, border_color=BORDER, border_width=1)
        footer.pack(fill="x", side="bottom")

        body_frame = ctk.CTkFrame(top, fg_color=BG, corner_radius=0)
        body_frame.pack(fill="both", expand=True, padx=20, pady=12)

        tab_bar = ctk.CTkFrame(body_frame, fg_color="transparent")
        tab_bar.pack(fill="x", pady=(0, 8))

        active_tab = {"mode": "auto"}

        btn_tab_auto = ctk.CTkButton(
            tab_bar, text=f"⚡ Smart Candidates ({len(candidate_pairs)})", height=32,
            fg_color=WHITE, hover_color=PREVIEW_BG, border_color=ACCENT, border_width=2,
            text_color=ACCENT, font=(FONT_FAMILY, 10, "bold")
        )
        btn_tab_auto.pack(side="left", padx=(0, 6))

        btn_tab_manual = ctk.CTkButton(
            tab_bar, text=f"🔍 Free-Form Selector (Bank: {len(bank_items)} | Odoo: {len(odo_items)})", height=32,
            fg_color=PANEL, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1,
            text_color=TEXT, font=(FONT_FAMILY, 10, "bold")
        )
        btn_tab_manual.pack(side="left")

        content_box = ctk.CTkFrame(body_frame, fg_color=PANEL, corner_radius=10, border_color=BORDER, border_width=1)
        content_box.pack(fill="both", expand=True)

        queue_card = ctk.CTkFrame(body_frame, fg_color=PANEL, corner_radius=10, border_color=BORDER, border_width=1, height=230)
        queue_card.pack(fill="x", pady=(10, 0))
        queue_card.pack_propagate(False)

        q_hdr = ctk.CTkFrame(queue_card, fg_color="transparent")
        q_hdr.pack(fill="x", padx=16, pady=(8, 2))

        lbl_q_title = ctk.CTkLabel(q_hdr, text="📋 Confirmed Matched Pairs Queue (0 pairs)", font=(FONT_FAMILY, 10, "bold"), text_color=TEXT)
        lbl_q_title.pack(side="left")

        q_scroll = ctk.CTkScrollableFrame(queue_card, fg_color=PREVIEW_BG, corner_radius=6, border_color=BORDER, border_width=1)
        q_scroll.pack(fill="both", expand=True, padx=12, pady=(0, 8))

        def render_queue():
            for w in q_scroll.winfo_children():
                w.destroy()
            lbl_q_title.configure(text=f"📋 Confirmed Matched Pairs Queue ({len(active_matched)} pair{'s' if len(active_matched)!=1 else ''})")
            btn_submit_matched.configure(text=f"💾 Update Recon & Open Journal ({len(active_matched)})")

            col_widths = [110, 220, 160, 160, 160, 90]
            for col, w in enumerate(col_widths):
                q_scroll.grid_columnconfigure(col, weight=1, minsize=w)

            if not active_matched:
                tk.Label(q_scroll, text="No matched pairs queued yet. Click '🔗 Match Pair' above to add items here.", bg=PREVIEW_BG, fg=MUTED, font=(FONT_FAMILY, 9, "bold")).pack(pady=10)
                return

            q_headers = ["Date", "Journal", "Bank Amount", "Odoo Amount", "Net Difference", "Action"]
            for col, h in enumerate(q_headers):
                tk.Label(q_scroll, text=h, bg=PREVIEW_BG, fg=MUTED, font=(FONT_FAMILY, 9, "bold")).grid(row=0, column=col, padx=12, pady=(4, 6), sticky="w")
            
            tk.Frame(q_scroll, bg=BORDER_DARK, height=1).grid(row=1, column=0, columnspan=len(q_headers), sticky="ew")

            for idx, pair in enumerate(active_matched, 2):
                b, o, diff = pair["bank"], pair["odoo"], pair["diff"]
                impact_color = ERROR if diff < 0 else WARN if diff > 0 else SUCCESS

                tk.Label(q_scroll, text=f"📅 {b['date']}", bg=PREVIEW_BG, fg=TEXT, font=(_EMOJI_FONT, 9, "bold")).grid(row=idx, column=0, padx=12, pady=2, sticky="w")
                tk.Label(q_scroll, text=f"🏦 {b['journal']}", bg=PREVIEW_BG, fg=ACCENT, font=(_EMOJI_FONT, 9, "bold")).grid(row=idx, column=1, padx=12, pady=2, sticky="w")
                tk.Label(q_scroll, text=f"Rp {b['amount']:,.0f}", bg=PREVIEW_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=idx, column=2, padx=12, pady=2, sticky="w")
                tk.Label(q_scroll, text=f"Rp {o['amount']:,.0f}", bg=PREVIEW_BG, fg=ACCENT, font=(FONT_FAMILY, 10, "bold")).grid(row=idx, column=3, padx=12, pady=2, sticky="w")
                tk.Label(q_scroll, text=f"Rp {diff:,.0f}", bg=PREVIEW_BG, fg=impact_color, font=(FONT_FAMILY, 10, "bold")).grid(row=idx, column=4, padx=12, pady=2, sticky="w")

                def _unpair(p=pair):
                    active_matched.remove(p)
                    render_queue()
                    if active_tab["mode"] == "auto":
                        render_auto_tab()
                    else:
                        render_manual_tab()

                btn_unp = tk.Label(q_scroll, text="❌ Unpair", bg=PREVIEW_BG, fg=ERROR, cursor="hand2", font=(_EMOJI_FONT, 9, "bold"))
                btn_unp.bind("<Button-1>", lambda e, func=_unpair: func())
                btn_unp.grid(row=idx, column=5, padx=12, pady=2, sticky="w")

        tab_auto_frame = ctk.CTkFrame(content_box, fg_color="transparent")
        tab_manual_frame = ctk.CTkFrame(content_box, fg_color="transparent")

        # Initial pack
        tab_auto_frame.pack(fill="both", expand=True)

        def switch_to_auto():
            active_tab["mode"] = "auto"
            btn_tab_auto.configure(fg_color=WHITE, border_color=ACCENT, border_width=2, text_color=ACCENT)
            btn_tab_manual.configure(fg_color=PANEL, border_color=BORDER_DARK, border_width=1, text_color=TEXT)
            tab_manual_frame.pack_forget()
            tab_auto_frame.pack(fill="both", expand=True)

        def switch_to_auto():
            active_tab["mode"] = "auto"
            btn_tab_auto.configure(fg_color=WHITE, border_color=ACCENT, border_width=2, text_color=ACCENT)
            btn_tab_manual.configure(fg_color=PANEL, border_color=BORDER_DARK, border_width=1, text_color=TEXT)
            tab_manual_frame.pack_forget()
            tab_auto_frame.pack(fill="both", expand=True)

        def switch_to_manual():
            active_tab["mode"] = "manual"
            btn_tab_auto.configure(fg_color=PANEL, border_color=BORDER_DARK, border_width=1, text_color=TEXT)
            btn_tab_manual.configure(fg_color=WHITE, border_color=ACCENT, border_width=2, text_color=ACCENT)
            tab_auto_frame.pack_forget()
            tab_manual_frame.pack(fill="both", expand=True)
            render_manual_tab()

        btn_tab_auto.configure(command=switch_to_auto)
        btn_tab_manual.configure(command=switch_to_manual)

        def render_auto_tab():
            matched_bank_rows = {p["bank"]["row_idx"] for p in active_matched}
            matched_odo_rows = {p["odoo"]["row_idx"] for p in active_matched}

            avail_candidates = [
                c for c in candidate_pairs
                if c["bank"]["row_idx"] not in matched_bank_rows and c["odoo"]["row_idx"] not in matched_odo_rows
            ]

            # Clear tab_auto_frame cleanly
            for w in tab_auto_frame.winfo_children():
                w.destroy()

            if not avail_candidates:
                ctk.CTkLabel(tab_auto_frame, text="No candidate pairs remaining to match.", font=(FONT_FAMILY, 11, "bold"), text_color=MUTED).pack(pady=40)
                return

            total_pages = max(1, (len(avail_candidates) + CANDIDATES_PER_PAGE - 1) // CANDIDATES_PER_PAGE)
            if auto_page[0] >= total_pages:
                auto_page[0] = total_pages - 1

            # Pagination Bar Header
            p_bar = ctk.CTkFrame(tab_auto_frame, fg_color="transparent")
            p_bar.pack(fill="x", padx=14, pady=(8, 4))

            ctk.CTkLabel(p_bar, text=f"Showing candidates {auto_page[0]*CANDIDATES_PER_PAGE + 1}–{min((auto_page[0]+1)*CANDIDATES_PER_PAGE, len(avail_candidates))} of {len(avail_candidates)}", font=(FONT_FAMILY, 9, "bold"), text_color=MUTED).pack(side="left")

            p_controls = ctk.CTkFrame(p_bar, fg_color="transparent")
            p_controls.pack(side="right")

            def _prev_p():
                if auto_page[0] > 0:
                    auto_page[0] -= 1
                    render_auto_tab()

            def _next_p():
                if auto_page[0] < total_pages - 1:
                    auto_page[0] += 1
                    render_auto_tab()

            ctk.CTkButton(p_controls, text="◀ Prev", width=65, height=26, fg_color=PANEL, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1, text_color=TEXT, font=(FONT_FAMILY, 9, "bold"), state="normal" if auto_page[0]>0 else "disabled", command=_prev_p).pack(side="left", padx=2)
            ctk.CTkLabel(p_controls, text=f"Page {auto_page[0]+1} of {total_pages}", font=(FONT_FAMILY, 9, "bold"), text_color=TEXT).pack(side="left", padx=8)
            ctk.CTkButton(p_controls, text="Next ▶", width=65, height=26, fg_color=PANEL, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1, text_color=TEXT, font=(FONT_FAMILY, 9, "bold"), state="normal" if auto_page[0]<total_pages-1 else "disabled", command=_next_p).pack(side="left", padx=2)

            c_scroll = ctk.CTkScrollableFrame(tab_auto_frame, fg_color="transparent")
            c_scroll.pack(fill="both", expand=True, padx=12, pady=(0, 8))

            def _reset_scroll(scroll_frame):
                def _apply():
                    try:
                        scroll_frame._parent_canvas.yview_moveto(0.0)
                    except Exception:
                        pass
                _apply()
                top.after(10, _apply)
                top.after(50, _apply)

            _reset_scroll(c_scroll)

            page_items = avail_candidates[auto_page[0]*CANDIDATES_PER_PAGE : (auto_page[0]+1)*CANDIDATES_PER_PAGE]

            for c in page_items:
                b, o, diff = c["bank"], c["odoo"], c["diff"]
                
                card = tk.Frame(c_scroll, bg=WHITE, highlightbackground=BORDER_DARK, highlightcolor=BORDER_DARK, highlightthickness=1)
                card.pack(fill="x", pady=3, padx=4)

                c_inner = tk.Frame(card, bg=WHITE)
                c_inner.pack(fill="x", padx=12, pady=6)

                info_left = tk.Frame(c_inner, bg=WHITE)
                info_left.pack(side="left", fill="both", expand=True)

                txt_hdr = f"📅 {b['date']}  |  🏦 {b['journal']}"
                txt_dtl = f"Bank: Rp {b['amount']:,.0f} (Ref: {b['number_bank'] or 'N/A'})   •   Odoo: Rp {o['amount']:,.0f} (Doc: {o['number_odo'] or 'N/A'})"
                
                lbl_h = tk.Label(info_left, text=txt_hdr, font=(_EMOJI_FONT, 10, "bold"), fg=ACCENT, bg=WHITE, anchor="w")
                lbl_h.pack(fill="x", pady=(0, 2))

                lbl_d = tk.Label(info_left, text=txt_dtl, font=(FONT_FAMILY, 11, "bold"), fg=TEXT, bg=WHITE, anchor="w")
                lbl_d.pack(fill="x", pady=(0, 0))

                info_right = tk.Frame(c_inner, bg=WHITE)
                info_right.pack(side="right", anchor="e")

                var_color = SUCCESS if diff == 0 else WARN
                var_text = f"Variance: Rp {diff:,.0f}" if diff != 0 else "Exact Match"
                lbl_v = tk.Label(info_right, text=var_text, font=(FONT_FAMILY, 10, "bold"), fg=var_color, bg=WHITE, anchor="e")
                lbl_v.pack(side="left", padx=(0, 14))

                def _pair_auto(cand=c):
                    o_r = cand["odoo"]
                    o_recon = str(o_r.get("reconciled", "")).strip().lower()
                    if o_recon in ("no", "false"):
                        show_modal_msg(
                            f"⚠️ Cannot match: Odoo payment ({o_r.get('number_odo') or o_r.get('amount')}) is not reconciled with invoice in Odoo (Reconciled: No)."
                        )
                        return
                    active_matched.append(cand)
                    show_modal_msg("✅ Matched pair added to queue.", is_err=False)
                    render_queue()
                    render_auto_tab()

                ctk.CTkButton(
                    info_right, text="🔗 Match Pair", height=28, width=105,
                    fg_color=ACCENT, hover_color=ACCENT_DARK, text_color=WHITE,
                    font=(FONT_FAMILY, 10, "bold"), corner_radius=5,
                    command=_pair_auto
                ).pack(side="left")

                for w in (card, c_inner, info_left, lbl_h, lbl_d, info_right, lbl_v):
                    w.bind("<Enter>", lambda e, c=card, ci=c_inner, il=info_left, lh=lbl_h, ld=lbl_d, ir=info_right, lv=lbl_v: [
                        c.config(bg=PREVIEW_BG), ci.config(bg=PREVIEW_BG), il.config(bg=PREVIEW_BG), lh.config(bg=PREVIEW_BG), ld.config(bg=PREVIEW_BG), ir.config(bg=PREVIEW_BG), lv.config(bg=PREVIEW_BG)
                    ])
                    w.bind("<Leave>", lambda e, c=card, ci=c_inner, il=info_left, lh=lbl_h, ld=lbl_d, ir=info_right, lv=lbl_v: [
                        c.config(bg=WHITE), ci.config(bg=WHITE), il.config(bg=WHITE), lh.config(bg=WHITE), ld.config(bg=WHITE), ir.config(bg=WHITE), lv.config(bg=WHITE)
                    ])

        def render_manual_tab():
            for w in tab_manual_frame.winfo_children():
                w.destroy()

            bar_sel = ctk.CTkFrame(tab_manual_frame, fg_color=PREVIEW_BG, height=52, corner_radius=0, border_color=BORDER, border_width=1)
            bar_sel.pack(fill="x", side="bottom")
            bar_sel.pack_propagate(False)

            split = ctk.CTkFrame(tab_manual_frame, fg_color="transparent")
            split.pack(fill="both", expand=True, padx=10, pady=10)
            split.rowconfigure(0, weight=1)
            split.columnconfigure(0, weight=1, uniform="equal_cols")
            split.columnconfigure(1, weight=1, uniform="equal_cols")

            matched_b_ids = {p["bank"]["row_idx"] for p in active_matched}
            matched_o_ids = {p["odoo"]["row_idx"] for p in active_matched}

            avail_b = [b for b in bank_items if b["row_idx"] not in matched_b_ids]
            avail_o = [o for o in odo_items if o["row_idx"] not in matched_o_ids]

            sel_b = [None]
            sel_o = [None]
            ITEMS_PAGE_SIZE = 25
            page_b = [0]
            page_o = [0]

            box_b = ctk.CTkFrame(split, fg_color=PREVIEW_BG, corner_radius=8, border_color=BORDER, border_width=1)
            box_b.grid(row=0, column=0, sticky="nsew", padx=(0, 4))

            hdr_box_b = ctk.CTkFrame(box_b, fg_color="transparent")
            hdr_box_b.pack(fill="x", padx=10, pady=(6, 2))
            lbl_title_b = ctk.CTkLabel(hdr_box_b, text=f"Bank Items ({len(avail_b)})", font=(FONT_FAMILY, 10, "bold"), text_color=TEXT)
            lbl_title_b.pack(side="left")

            p_bar_b = ctk.CTkFrame(hdr_box_b, fg_color="transparent")
            p_bar_b.pack(side="right")
            lbl_page_b = ctk.CTkLabel(p_bar_b, text="Page 1/1", font=(FONT_FAMILY, 9), text_color=MUTED)
            lbl_page_b.pack(side="left", padx=4)
            btn_prev_b = ctk.CTkButton(p_bar_b, text="◀", width=26, height=22, font=(FONT_FAMILY, 9, "bold"), fg_color=PANEL, hover_color=PREVIEW_BG, text_color=TEXT, border_width=1, border_color=BORDER_DARK)
            btn_prev_b.pack(side="left", padx=1)
            btn_next_b = ctk.CTkButton(p_bar_b, text="▶", width=26, height=22, font=(FONT_FAMILY, 9, "bold"), fg_color=PANEL, hover_color=PREVIEW_BG, text_color=TEXT, border_width=1, border_color=BORDER_DARK)
            btn_next_b.pack(side="left", padx=1)

            entry_search_b = ctk.CTkEntry(box_b, placeholder_text="🔍 Filter Bank by date/amount/ref...", height=28, font=(FONT_FAMILY, 9.5))
            entry_search_b.pack(fill="x", padx=8, pady=(0, 4))

            s_b = ctk.CTkScrollableFrame(box_b, fg_color="transparent")
            s_b.pack(fill="both", expand=True, padx=4, pady=4)

            box_o = ctk.CTkFrame(split, fg_color=PREVIEW_BG, corner_radius=8, border_color=BORDER, border_width=1)
            box_o.grid(row=0, column=1, sticky="nsew", padx=(4, 0))

            hdr_box_o = ctk.CTkFrame(box_o, fg_color="transparent")
            hdr_box_o.pack(fill="x", padx=10, pady=(6, 2))
            lbl_title_o = ctk.CTkLabel(hdr_box_o, text=f"Odoo Items ({len(avail_o)})", font=(FONT_FAMILY, 10, "bold"), text_color=TEXT)
            lbl_title_o.pack(side="left")

            p_bar_o = ctk.CTkFrame(hdr_box_o, fg_color="transparent")
            p_bar_o.pack(side="right")
            lbl_page_o = ctk.CTkLabel(p_bar_o, text="Page 1/1", font=(FONT_FAMILY, 9), text_color=MUTED)
            lbl_page_o.pack(side="left", padx=4)
            btn_prev_o = ctk.CTkButton(p_bar_o, text="◀", width=26, height=22, font=(FONT_FAMILY, 9, "bold"), fg_color=PANEL, hover_color=PREVIEW_BG, text_color=TEXT, border_width=1, border_color=BORDER_DARK)
            btn_prev_o.pack(side="left", padx=1)
            btn_next_o = ctk.CTkButton(p_bar_o, text="▶", width=26, height=22, font=(FONT_FAMILY, 9, "bold"), fg_color=PANEL, hover_color=PREVIEW_BG, text_color=TEXT, border_width=1, border_color=BORDER_DARK)
            btn_next_o.pack(side="left", padx=1)

            entry_search_o = ctk.CTkEntry(box_o, placeholder_text="🔍 Filter Odoo by date/amount/doc...", height=28, font=(FONT_FAMILY, 9.5))
            entry_search_o.pack(fill="x", padx=8, pady=(0, 4))

            s_o = ctk.CTkScrollableFrame(box_o, fg_color="transparent")
            s_o.pack(fill="both", expand=True, padx=4, pady=4)

            bar_in = ctk.CTkFrame(bar_sel, fg_color="transparent")
            bar_in.pack(fill="both", expand=True, padx=14, pady=6)

            lbl_sel_info = ctk.CTkLabel(bar_in, text="Select 1 Bank item and 1 Odoo item to calculate variance and pair.", font=(FONT_FAMILY, 10.5, "bold"), text_color=MUTED)
            lbl_sel_info.pack(side="left")

            def _update_sel_bar():
                if sel_b[0] and sel_o[0]:
                    b_amt = sel_b[0]["amount"]
                    o_amt = sel_o[0]["amount"]
                    diff = b_amt - o_amt
                    lbl_sel_info.configure(
                        text=f"Selected: Bank Rp {b_amt:,.0f}  |  Odoo Rp {o_amt:,.0f}   →   Net Difference: Rp {diff:,.0f}",
                        font=(FONT_FAMILY, 11.5, "bold"),
                        text_color=ACCENT
                    )
                    btn_pair_sel.configure(state="normal", fg_color=ACCENT)
                elif sel_b[0]:
                    lbl_sel_info.configure(
                        text=f"Selected Bank: Rp {sel_b[0]['amount']:,.0f} ({sel_b[0]['date']} | {sel_b[0]['journal']}) — Select matching Odoo item",
                        font=(FONT_FAMILY, 10.5, "bold"),
                        text_color=ACCENT
                    )
                    btn_pair_sel.configure(state="disabled", fg_color=BORDER_DARK)
                elif sel_o[0]:
                    lbl_sel_info.configure(
                        text=f"Selected Odoo: Rp {sel_o[0]['amount']:,.0f} ({sel_o[0]['date']} | {sel_o[0]['journal']}) — Select matching Bank item",
                        font=(FONT_FAMILY, 10.5, "bold"),
                        text_color=ACCENT
                    )
                    btn_pair_sel.configure(state="disabled", fg_color=BORDER_DARK)
                else:
                    lbl_sel_info.configure(
                        text="Select 1 Bank item and 1 Odoo item to calculate variance and pair.",
                        font=(FONT_FAMILY, 10.5, "bold"),
                        text_color=MUTED
                    )
                    btn_pair_sel.configure(state="disabled", fg_color=BORDER_DARK)

            def _confirm_manual_pair():
                if sel_b[0] and sel_o[0]:
                    b, o = sel_b[0], sel_o[0]
                    o_recon = str(o.get("reconciled", "")).strip().lower()
                    if o_recon in ("no", "false"):
                        show_modal_msg(
                            f"⚠️ Cannot match: Selected Odoo payment ({o.get('number_odo') or o.get('amount')}) is not reconciled with invoice in Odoo (Reconciled: No)."
                        )
                        return
                    diff = b["amount"] - o["amount"]
                    active_matched.append({
                        "bank": b,
                        "odoo": o,
                        "diff": diff,
                        "abs_diff": abs(diff),
                        "pct": (abs(diff)/max(b["amount"], 1.0))*100.0,
                        "date": b["date"],
                        "journal": b["journal"],
                    })
                    show_modal_msg("✅ Matched pair added to queue.", is_err=False)
                    sel_b[0] = None
                    sel_o[0] = None
                    render_queue()
                    render_auto_tab()
                    render_manual_tab()

            btn_pair_sel = ctk.CTkButton(
                bar_in, text="🔗 Pair Selected Items", height=36,
                fg_color=BORDER_DARK, state="disabled", text_color=WHITE,
                font=(FONT_FAMILY, 10, "bold"), command=_confirm_manual_pair
            )
            btn_pair_sel.pack(side="right")

            def _toggle_b(bi):
                if sel_b[0] == bi:
                    sel_b[0] = None
                else:
                    sel_b[0] = bi
                _rebuild_lists()
                _update_sel_bar()

            def _toggle_o(oi):
                if sel_o[0] == oi:
                    sel_o[0] = None
                else:
                    sel_o[0] = oi
                _rebuild_lists()
                _update_sel_bar()

            def _render_native_card(parent, item, is_sel, is_bank, on_click):
                bg_col = ACCENT if is_sel else WHITE
                fg_hdr = WHITE if is_sel else ACCENT
                fg_txt = WHITE if is_sel else TEXT
                border_col = ACCENT_DARK if is_sel else BORDER_DARK

                card = tk.Frame(parent, bg=bg_col, highlightbackground=border_col, highlightcolor=border_col, highlightthickness=1, cursor="hand2")
                card.pack(fill="x", pady=2, padx=2)

                top_txt = f"📅 {item['date']}  |  🏦 {item['journal']}"
                if is_bank:
                    bot_txt = f"Rp {item['amount']:,.0f}  (Ref: {item.get('number_bank') or 'N/A'})"
                else:
                    recon_st = str(item.get("reconciled") or "Yes")
                    recon_tag = f"  •  Reconciled: {recon_st}"
                    bot_txt = f"Rp {item['amount']:,.0f}  (Doc: {item.get('number_odo') or 'N/A'}{recon_tag})"

                l_top = tk.Label(card, text=top_txt, font=(_EMOJI_FONT, 10, "bold"), fg=fg_hdr, bg=bg_col, anchor="w", cursor="hand2")
                l_top.pack(fill="x", padx=8, pady=(4, 1))

                l_bot = tk.Label(card, text=bot_txt, font=(FONT_FAMILY, 11, "bold"), fg=fg_txt, bg=bg_col, anchor="w", cursor="hand2")
                l_bot.pack(fill="x", padx=8, pady=(0, 4))

                for w in (card, l_top, l_bot):
                    w.bind("<Button-1>", lambda e, func=on_click: func())
                    if not is_sel:
                        w.bind("<Enter>", lambda e, c=card, t=l_top, b=l_bot: (c.config(bg=PREVIEW_BG), t.config(bg=PREVIEW_BG), b.config(bg=PREVIEW_BG)))
                        w.bind("<Leave>", lambda e, c=card, t=l_top, b=l_bot: (c.config(bg=WHITE), t.config(bg=WHITE), b.config(bg=WHITE)))

            def _rebuild_lists():
                q_b = entry_search_b.get().strip().lower()
                q_o = entry_search_o.get().strip().lower()

                # Cross-filter Odoo list based on selected Bank item
                if sel_b[0]:
                    target_d = sel_b[0]["date"]
                    target_j = sel_b[0]["journal"].strip().lower()
                    base_o = [o for o in avail_o if o["date"] == target_d and o["journal"].strip().lower() == target_j]
                else:
                    base_o = avail_o

                # Cross-filter Bank list based on selected Odoo item
                if sel_o[0]:
                    target_d = sel_o[0]["date"]
                    target_j = sel_o[0]["journal"].strip().lower()
                    base_b = [b for b in avail_b if b["date"] == target_d and b["journal"].strip().lower() == target_j]
                else:
                    base_b = avail_b

                # Apply text search filter
                if q_b:
                    filtered_b = [b for b in base_b if q_b in f"{b['date']} {b['journal']} {b['amount']} {b.get('number_bank','')}".lower()]
                else:
                    filtered_b = base_b

                if q_o:
                    filtered_o = [o for o in base_o if q_o in f"{o['date']} {o['journal']} {o['amount']} {o.get('number_odo','')}".lower()]
                else:
                    filtered_o = base_o

                # Calculate pages
                tot_p_b = max(1, (len(filtered_b) + ITEMS_PAGE_SIZE - 1) // ITEMS_PAGE_SIZE)
                if page_b[0] >= tot_p_b: page_b[0] = tot_p_b - 1
                if page_b[0] < 0: page_b[0] = 0

                tot_p_o = max(1, (len(filtered_o) + ITEMS_PAGE_SIZE - 1) // ITEMS_PAGE_SIZE)
                if page_o[0] >= tot_p_o: page_o[0] = tot_p_o - 1
                if page_o[0] < 0: page_o[0] = 0

                # Update container titles & pagination
                lbl_title_b.configure(text=f"Bank Items ({len(filtered_b)} shown of {len(avail_b)})")
                lbl_page_b.configure(text=f"P {page_b[0]+1}/{tot_p_b}")
                btn_prev_b.configure(state="normal" if page_b[0] > 0 else "disabled", command=lambda: [_change_page_b(-1)])
                btn_next_b.configure(state="normal" if page_b[0] < tot_p_b - 1 else "disabled", command=lambda: [_change_page_b(1)])

                lbl_title_o.configure(text=f"Odoo Items ({len(filtered_o)} shown of {len(avail_o)})")
                lbl_page_o.configure(text=f"P {page_o[0]+1}/{tot_p_o}")
                btn_prev_o.configure(state="normal" if page_o[0] > 0 else "disabled", command=lambda: [_change_page_o(-1)])
                btn_next_o.configure(state="normal" if page_o[0] < tot_p_o - 1 else "disabled", command=lambda: [_change_page_o(1)])

                # Fast Native Render Bank Items (Page slice)
                for w in s_b.winfo_children(): w.destroy()
                slice_b = filtered_b[page_b[0]*ITEMS_PAGE_SIZE : (page_b[0]+1)*ITEMS_PAGE_SIZE]
                if not slice_b:
                    tk.Label(s_b, text="No matching Bank items", bg=PANEL, fg=MUTED, font=(FONT_FAMILY, 9)).pack(pady=12)
                else:
                    for b_item in slice_b:
                        is_sel = (sel_b[0] == b_item)
                        _render_native_card(s_b, b_item, is_sel, True, lambda bi=b_item: _toggle_b(bi))

                # Fast Native Render Odoo Items (Page slice)
                for w in s_o.winfo_children(): w.destroy()
                slice_o = filtered_o[page_o[0]*ITEMS_PAGE_SIZE : (page_o[0]+1)*ITEMS_PAGE_SIZE]
                if not slice_o:
                    tk.Label(s_o, text="No matching Odoo items", bg=PANEL, fg=MUTED, font=(FONT_FAMILY, 9)).pack(pady=12)
                else:
                    for o_item in slice_o:
                        is_sel = (sel_o[0] == o_item)
                        _render_native_card(s_o, o_item, is_sel, False, lambda oi=o_item: _toggle_o(oi))

            def _reset_scroll_frame(scroll_frame):
                def _apply():
                    try:
                        scroll_frame._parent_canvas.yview_moveto(0.0)
                    except Exception:
                        pass
                _apply()
                top.after(10, _apply)
                top.after(50, _apply)

            def _change_page_b(delta):
                page_b[0] += delta
                _rebuild_lists()
                _reset_scroll_frame(s_b)

            def _change_page_o(delta):
                page_o[0] += delta
                _rebuild_lists()
                _reset_scroll_frame(s_o)

            entry_search_b.bind("<KeyRelease>", lambda e: [setattr(page_b, '__setitem__', (0, 0)) if hasattr(page_b, '__setitem__') else None, _rebuild_lists(), _reset_scroll_frame(s_b)])
            entry_search_o.bind("<KeyRelease>", lambda e: [setattr(page_o, '__setitem__', (0, 0)) if hasattr(page_o, '__setitem__') else None, _rebuild_lists(), _reset_scroll_frame(s_o)])

            _rebuild_lists()

        f_in = ctk.CTkFrame(footer, fg_color="transparent")
        f_in.pack(fill="both", expand=True, padx=24, pady=12)

        ctk.CTkButton(
            f_in, text="Close", height=36, width=90,
            fg_color=WHITE, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1,
            text_color=TEXT, font=(FONT_FAMILY, 10, "bold"), command=top.destroy
        ).pack(side="left")

        right_btn_frame = ctk.CTkFrame(f_in, fg_color="transparent")
        right_btn_frame.pack(side="right")

        def _apply_matched_journals():
            if not active_matched:
                return
            top.destroy()
            self._log_write(f"\n── Processing {len(active_matched)} Manually Matched Pairs ──\n", "head")
            for pair in active_matched:
                b, o, diff = pair["bank"], pair["odoo"], pair["diff"]
                self._log_write(f"✅ Matched [{b['date']} {b['journal']}] Bank: Rp {b['amount']:,.0f} <-> Odoo: Rp {o['amount']:,.0f} (Diff: Rp {diff:,.0f})\n", "ok")
            self._set_status(f"Matched {len(active_matched)} pairs", SUCCESS)
            self._on_journal()

        def _update_recon_file(open_journal_modal=False):
            if not active_matched:
                show_modal_msg("⚠️ No matched pairs in queue! Click '🔗 Match Pair' or '🔗 Pair Selected Items' first.")
                return
            try:
                import openpyxl
                from openpyxl.styles import PatternFill, Font as XLFont
                from excel_writer import _normalize_date_str
                from reconciler import STATUS_DONE, STATUS_BANK_ONLY, STATUS_ODO_ONLY

                wb = openpyxl.load_workbook(latest_file)
                GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")

                # Find highest existing match sequence number so reopening the modal doesn't restart at M01
                next_pair_idx = 1
                if "Differences" in wb.sheetnames:
                    import re as _re
                    ws_scan = wb["Differences"]
                    for r_scan in range(4, ws_scan.max_row + 1):
                        v = ws_scan.cell(r_scan, 13).value or ws_scan.cell(r_scan, 12).value
                        if v:
                            m = _re.search(r"M(\d+)", str(v))
                            if m:
                                next_pair_idx = max(next_pair_idx, int(m.group(1)) + 1)

                SKIP_SHEETS = {"Daily Summary", "Differences", "Mutation Summary",
                               "Admin Fee", "Excluded Payment", "Other Payment",
                               "Other Mutation", "Legend"}

                def _ensure_diff_header(ws, col_no, label="Difference"):
                    """Add 'Difference' header to col_no row 3 if not already present."""
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter as _gcl
                    hdr_cell = ws.cell(3, col_no)
                    if not hdr_cell.value:
                        hdr_cell.value = label
                        hdr_cell.font = Font(bold=True, color="FFFFFF", size=10)
                        hdr_cell.fill = PatternFill("solid", fgColor="1F4E79")
                        hdr_cell.alignment = Alignment(horizontal="center", vertical="center")
                        _thin = Side(style="thin", color="CCCCCC")
                        hdr_cell.border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
                        ws.column_dimensions[_gcl(col_no)].width = 18

                # rows to delete in Differences sheet (collect, then delete bottom-up to preserve indices)
                diff_rows_to_delete = []

                for pair_idx, pair in enumerate(active_matched, next_pair_idx):
                    pair_tag = f"Match (M{pair_idx:02d})"
                    b_item = pair["bank"]
                    o_item = pair["odoo"]
                    diff   = pair["diff"]  # Bank - Odoo
                    b_date_norm = _normalize_date_str(b_item["date"])
                    o_date_norm = _normalize_date_str(o_item["date"])

                    # ── 1. Differences Sheet ──────────────────────────────────────
                    if "Differences" in wb.sheetnames:
                        ws_diff = wb["Differences"]
                        _ensure_diff_header(ws_diff, 14)
                        b_row_d = None
                        o_row_d = None
                        for r_idx in range(4, ws_diff.max_row + 1):
                            src_v = str(ws_diff.cell(r_idx, 11).value or ws_diff.cell(r_idx, 10).value or "").strip()
                            d_v   = _normalize_date_str(ws_diff.cell(r_idx, 2).value)
                            # Check col 9 (Bank Amt / Amt) and col 10 (Odoo Amt)
                            a_v9  = float(ws_diff.cell(r_idx, 9).value or 0)
                            a_v10 = float(ws_diff.cell(r_idx, 10).value or 0)
                            if src_v == "Bank" and d_v == b_date_norm and (abs(a_v9 - b_item["amount"]) < 0.01 or abs(a_v10 - b_item["amount"]) < 0.01):
                                b_row_d = r_idx
                            elif src_v == "Odoo" and d_v == o_date_norm and (abs(a_v9 - o_item["amount"]) < 0.01 or abs(a_v10 - o_item["amount"]) < 0.01):
                                o_row_d = r_idx

                        # Write merged row on the Bank row; mark Odoo row for deletion
                        if b_row_d:
                            ws_diff.cell(b_row_d, 5).value = o_item.get("number_odo", "")
                            ws_diff.cell(b_row_d, 6).value = o_item.get("reference", "")
                            c_o = ws_diff.cell(b_row_d, 10)
                            c_o.value = o_item["amount"]
                            c_o.number_format = '#,##0.00'
                            ws_diff.cell(b_row_d, 11).value = "Manual"
                            ws_diff.cell(b_row_d, 12).value = o_item.get("reconciled", "Yes")
                            ws_diff.cell(b_row_d, 13).value = pair_tag
                            ws_diff.cell(b_row_d, 13).font = XLFont(bold=True, size=10)
                            c_diff = ws_diff.cell(b_row_d, 14)
                            c_diff.value = diff
                            c_diff.number_format = '#,##0.00'
                            for c in range(1, 15): ws_diff.cell(b_row_d, c).fill = GREEN_FILL
                        if o_row_d:
                            diff_rows_to_delete.append(o_row_d)

                    # ── 2. Individual Bank Sheet ──────────────────────────────────
                    target_j = b_item["journal"].strip().lower()
                    for sname in wb.sheetnames:
                        if sname in SKIP_SHEETS:
                            continue
                        if target_j not in sname.lower() and sname.lower() not in target_j:
                            continue
                        ws_b = wb[sname]
                        _ensure_diff_header(ws_b, 13)
                        b_row_b = None
                        o_row_b = None
                        bank_rows_to_delete = []
                        for r in range(4, ws_b.max_row + 1):
                            d_v = _normalize_date_str(ws_b.cell(r, 2).value)
                            src_v = str(ws_b.cell(r, 10).value or ws_b.cell(r, 9).value or "").strip()
                            a_v8 = float(ws_b.cell(r, 8).value or 0.0)
                            a_v9 = float(ws_b.cell(r, 9).value or 0.0)
                            if (src_v in ("Bank", "Both", "Manual") or ws_b.cell(r, 12).value == STATUS_DONE) and d_v == b_date_norm and (abs(a_v8 - b_item["amount"]) < 0.01 or abs(a_v9 - b_item["amount"]) < 0.01):
                                b_row_b = r
                            elif src_v == "Odoo" and d_v == o_date_norm and (abs(a_v8 - o_item["amount"]) < 0.01 or abs(a_v9 - o_item["amount"]) < 0.01):
                                o_row_b = r

                        if b_row_b:
                            ws_b.cell(b_row_b, 4).value = o_item.get("number_odo", "")
                            ws_b.cell(b_row_b, 5).value = o_item.get("reference", "")
                            c_o = ws_b.cell(b_row_b, 9)
                            c_o.value = o_item["amount"]
                            c_o.number_format = '#,##0.00'
                            ws_b.cell(b_row_b, 10).value = "Manual"
                            ws_b.cell(b_row_b, 11).value = o_item.get("reconciled", "Yes")
                            ws_b.cell(b_row_b, 12).value = pair_tag
                            ws_b.cell(b_row_b, 12).font = XLFont(bold=True, size=10)
                            c_diff = ws_b.cell(b_row_b, 13)
                            c_diff.value = diff
                            c_diff.number_format = '#,##0.00'
                            for c in range(1, 14): ws_b.cell(b_row_b, c).fill = GREEN_FILL
                        if o_row_b:
                            bank_rows_to_delete.append(o_row_b)

                        # Delete Odoo rows bottom-up so indices stay valid
                        for r_del in sorted(bank_rows_to_delete, reverse=True):
                            ws_b.delete_rows(r_del)

                # Delete Odoo rows in Differences sheet bottom-up
                for r_del in sorted(diff_rows_to_delete, reverse=True):
                    ws_diff.delete_rows(r_del)

                from pathlib import Path as _Path
                from odoo_journal_creator import safe_save_workbook
                saved_ok = safe_save_workbook(wb, _Path(latest_file))
                if saved_ok:
                    self._log_write(f"\n\u2705 Updated reconciliation report ({os.path.basename(latest_file)}) with {len(active_matched)} matched pairs!\n", "ok")
                    self._set_status(f"Updated recon report with {len(active_matched)} matched pairs", SUCCESS)
                    # Persist matches to sidecar so next rerun re-applies them
                    try:
                        from excel_writer import save_manual_matches
                        output_dir = _Path(latest_file).parent
                        sidecar_entries = []
                        for pi, pair in enumerate(active_matched, next_pair_idx):
                            sidecar_entries.append({
                                "pair_tag":        f"Match (M{pi:02d})",
                                "bank_date":       pair["bank"]["date"],
                                "bank_journal":    pair["bank"]["journal"],
                                "bank_amount":     pair["bank"]["amount"],
                                "bank_number":     pair["bank"].get("number_bank", ""),
                                "odoo_date":       pair["odoo"]["date"],
                                "odoo_journal":    pair["odoo"]["journal"],
                                "odoo_amount":     pair["odoo"]["amount"],
                                "odoo_number":     pair["odoo"].get("number_odo", ""),
                                "odoo_reference":  pair["odoo"].get("reference", ""),
                                "odoo_reconciled": pair["odoo"].get("reconciled", "Yes"),
                                "diff":            pair["diff"],
                            })

                        save_manual_matches(output_dir, sidecar_entries)
                        self._log_write(f"\u2705 Saved {len(sidecar_entries)} match(es) to .manual_matches.json\n", "ok")
                    except Exception as _se:
                        self._log_write(f"\u26a0\ufe0f Could not save .manual_matches.json: {_se}\n", "warn")
                    top.destroy()
                    if open_journal_modal:
                        self._on_journal()
                else:
                    self._log_write(f"\n\u26a0\ufe0f Could not save reconciliation report ({os.path.basename(latest_file)}). Please close Excel and try again.\n", "warn")
                    show_modal_msg("⚠️ Could not save Excel file. Please close Excel and try again.")

            except Exception as e:
                self._set_status(f"Failed to update recon report: {e}", ERROR)
                self._log_write(f"\n❌ Error updating recon report: {e}\n", "err")
                show_modal_msg(f"❌ Error updating recon report: {e}")
                self._log_write(f"\n\u274c Error updating recon file: {e}\n", "err")
                self._set_status(f"Error updating recon file: {e}", ERROR)

        btn_update_recon = ctk.CTkButton(
            right_btn_frame, text=f"💾 Update Recon Only", height=38,
            fg_color=WHITE, hover_color=PREVIEW_BG, border_color=ACCENT, border_width=2,
            text_color=ACCENT, font=(FONT_FAMILY, 11, "bold"), command=lambda: _update_recon_file(open_journal_modal=False)
        )
        btn_update_recon.pack(side="left", padx=(0, 8))

        btn_submit_matched = ctk.CTkButton(
            right_btn_frame, text=f"💾 Update Recon & Open Journal (0)", height=38,
            fg_color=ACCENT, hover_color=ACCENT_DARK, text_color=WHITE,
            font=(FONT_FAMILY, 11, "bold"), command=lambda: _update_recon_file(open_journal_modal=True)
        )
        btn_submit_matched.pack(side="left")

        render_queue()
        render_auto_tab()
        switch_to_auto()
        top.deiconify()

    # ── Journal Confirmation Modal Overhaul ───────────────────────────────────
    def _on_journal(self):
        if getattr(self, "_journal_window", None) and self._journal_window.winfo_exists():
            try:
                self._journal_window.deiconify()
                self._journal_window.lift()
                self._journal_window.focus_force()
                return
            except Exception:
                pass

        if self._running:
            return
        
        import glob
        import os
        from openpyxl import load_workbook
        
        output_files = glob.glob(str(OUTPUT_DIR / "[Rr]econciliation_*.xlsx"))
        if not output_files:
            self._set_status("No reconciliation file found", ERROR)
            return
        latest_file = max(output_files, key=os.path.getctime)
        
        top = ctk.CTkToplevel(self)
        self._journal_window = top

        def _on_close_modal():
            self._journal_window = None
            top.destroy()

        top.protocol("WM_DELETE_WINDOW", _on_close_modal)
        top.title("Confirm Journal Creation")
        top.minsize(1050, 640)
        top.resizable(True, True)
        top.configure(fg_color=BG)
        top.transient(self)
        top.grab_set()
        top.after(200, lambda: _center_modal_on_parent(top, self))


        items = []
        journal_state = []

        
        def _load_data():
            nonlocal items, journal_state
            import glob
            import os
            output_files = glob.glob(str(OUTPUT_DIR / "[Rr]econciliation_*.xlsx"))
            if not output_files: return
            latest_file = max(output_files, key=os.path.getctime)
            try:
                wb = load_workbook(latest_file, data_only=True)
                if "Daily Summary" not in wb.sheetnames:
                    self._set_status("'Daily Summary' sheet not found", ERROR)
                    return
                def get_col_map(sheet, row_idx=3):
                    return {str(sheet.cell(row=row_idx, column=c).value).strip().lower(): c for c in range(1, sheet.max_column + 1) if sheet.cell(row=row_idx, column=c).value}

                if "Mutation Summary" in wb.sheetnames:
                    ws_mut = wb["Mutation Summary"]
                    col_map = get_col_map(ws_mut)
                    c_date = col_map.get("payment date", 3)
                    c_bank = col_map.get("bank", 4)
                    c_group = col_map.get("journal", 5)
                    c_cat = col_map.get("transaction category", 7)
                    c_amt = col_map.get("total amount", 8)
                
                    from collections import defaultdict
                    self._mutation_totals = defaultdict(float)
                    self._mutation_raw = []
                    for row in range(4, ws_mut.max_row + 1):
                        tanggal = ws_mut.cell(row=row, column=c_date).value
                        bank = ws_mut.cell(row=row, column=c_bank).value
                        group = ws_mut.cell(row=row, column=c_group).value
                        cat = ws_mut.cell(row=row, column=c_cat).value
                        amount = ws_mut.cell(row=row, column=c_amt).value
                        if tanggal and group and amount:
                            self._mutation_raw.append({"payment_date": tanggal, "bank": bank, "group": group, "category": cat, "amount": float(amount)})
                            try:
                                self._mutation_totals[(str(tanggal).strip(), str(group).strip())] += float(amount)
                            except ValueError:
                                pass
                else:
                    self._mutation_totals = {}
                
                if "Admin Fee" in wb.sheetnames:
                    ws_adm = wb["Admin Fee"]
                    col_map = get_col_map(ws_adm)
                    c_date = col_map.get("payment date", 3)
                    c_bank = col_map.get("bank", 4)
                    c_group = col_map.get("journal", 5)
                    c_cat = col_map.get("transaction category", 7)
                    c_amt = col_map.get("total amount", 8)
                
                    from collections import defaultdict
                    self._admin_totals = defaultdict(float)
                    self._admin_raw = []
                    for row in range(4, ws_adm.max_row + 1):
                        tanggal = ws_adm.cell(row=row, column=c_date).value
                        bank = ws_adm.cell(row=row, column=c_bank).value
                        group = ws_adm.cell(row=row, column=c_group).value
                        cat = ws_adm.cell(row=row, column=c_cat).value
                        amount = ws_adm.cell(row=row, column=c_amt).value
                        if tanggal and group and amount:
                            self._admin_raw.append({"payment_date": tanggal, "bank": bank, "group": group, "category": cat, "amount": float(amount)})
                            try:
                                self._admin_totals[(str(tanggal).strip(), str(group).strip())] += float(amount)
                            except ValueError:
                                pass
                else:
                    self._admin_totals = {}

                ws = wb["Daily Summary"]
                col_map = get_col_map(ws)
                c_date = col_map.get("date", 2)
                c_pdate = col_map.get("payment date", 3)
                c_bank = col_map.get("bank", 4)
                c_group = col_map.get("journal", 5)
                c_tbank = col_map.get("total bank", 6)
                c_todoo = col_map.get("total odoo", 7)
                c_diff = col_map.get("difference", 8)
                c_recon = col_map.get("reconciled", 9)
                c_status = col_map.get("status", 10)
                c_jstatus = col_map.get("journal information", 11)
            
                items.clear()
                for row in range(4, ws.max_row + 1):
                    bank = ws.cell(row=row, column=c_bank).value
                    group = ws.cell(row=row, column=c_group).value
                    tanggal = ws.cell(row=row, column=c_date).value
                    payment_date = ws.cell(row=row, column=c_pdate).value
                    total_bank = ws.cell(row=row, column=c_tbank).value
                    total_odoo = ws.cell(row=row, column=c_todoo).value
                    selisih = ws.cell(row=row, column=c_diff).value
                    reconciled = ws.cell(row=row, column=c_recon).value
                    status = ws.cell(row=row, column=c_status).value
                    journal_status = ws.cell(row=row, column=c_jstatus).value
                    if not bank or not status: continue
                    if "incomplete" in str(status).lower(): continue
                
                    try:
                        diff = abs(float(selisih)) if selisih is not None else 0
                    except:
                        diff = 0
                    
                    from config import JOURNAL_TOLERANCE
                
                    status_str = str(status).strip()
                    status_valid = ("Match" in status_str) or (diff <= JOURNAL_TOLERANCE)
                
                    mutation_found = False
                    mutation_matched = False
                    mut_total = 0.0
                    try:
                        d_mut_str = str(payment_date).strip()
                    
                        mut_amount = self._mutation_totals.get((d_mut_str, str(group).strip()), 0.0)
                        adm_amount = self._admin_totals.get((d_mut_str, str(group).strip()), 0.0)
                        mut_total = mut_amount + adm_amount
                    
                        if mut_amount > 0:
                            mutation_found = True
                            mut_diff = abs(mut_total - float(total_bank)) if total_bank else 0
                            if mut_diff <= JOURNAL_TOLERANCE:
                                mutation_matched = True
                    except Exception as e:
                        pass

                    items.append({
                        "row": row,
                        "bank": bank,
                        "group": group,
                        "tanggal": tanggal,
                        "payment_date": payment_date,
                        "merchant_amount": total_bank,
                        "odoo_amount": total_odoo,
                        "amount": total_odoo,
                        "selisih": selisih,
                        "mutation_found": mutation_found,
                        "mutation_matched": mutation_matched,
                        "mutation_amount": mut_total,
                        "reconciled": reconciled,
                        "journal_status": journal_status,
                        "status_valid": status_valid
                    })
                if not items:
                    self._set_status(f"No data with difference <= {JOURNAL_TOLERANCE}", ERROR)
                    return
            except Exception as e:
                self._set_status(f"Error reading excel: {e}", ERROR)
                return
            finally:
                try:
                    wb.close()
                except Exception:
                    pass



            journal_state.clear()
            for item in items:
                is_reconciled = str(item.get("reconciled", "")).strip().lower() == "yes"
                status_valid = item.get("status_valid", True)
            
                disabled_edc = False
                disabled_ar = False
            
                if not is_reconciled or not status_valid:
                    disabled_edc = True
                    disabled_ar = True
            
                if not item.get("mutation_matched", False):
                    disabled_ar = True
                
                j_status = item.get("journal_status")
                if j_status:
                    j_status_str = str(j_status).strip()
                    if j_status_str not in ["", "None", "Not Yet"]:
                        parts = [p.strip() for p in j_status_str.split("|")]
                        for p in parts:
                            if "(Both" in p:
                                if "Posted" in p:
                                    disabled_edc = True
                                    disabled_ar = True
                            elif "(EDC" in p:
                                if "Posted" in p:
                                    disabled_edc = True
                            elif "(AR" in p:
                                if "Posted" in p:
                                    disabled_ar = True

                var_item = tk.BooleanVar(value=not (disabled_edc and disabled_ar))
                var_edc = tk.BooleanVar(value=not disabled_edc)
                var_ar = tk.BooleanVar(value=not disabled_ar)

                journal_state.append({
                    "item": item,
                    "var_item": var_item,
                    "var_edc": var_edc,
                    "var_ar": var_ar,
                    "disabled_edc": disabled_edc,
                    "disabled_ar": disabled_ar
                })

        _load_data()

        ITEMS_PER_PAGE = 15
        current_page = [0]
        
        # Modal Header Bar
        header_frame = ctk.CTkFrame(top, fg_color=PANEL, corner_radius=0, height=80, border_color=BORDER, border_width=1)
        header_frame.pack(fill="x", side="top")
        header_frame.pack_propagate(False)
        
        left_header = ctk.CTkFrame(header_frame, fg_color="transparent")
        left_header.pack(side="left", padx=24, pady=16)
        
        ctk.CTkLabel(
            left_header, text="Confirm Journal Creation",
            font=(FONT_FAMILY, 15, "bold"), text_color=TEXT
        ).pack(anchor="w")
        ctk.CTkLabel(
            left_header, text="Review and select transactions to post. Expand any row to preview journal entries.",
            font=(FONT_FAMILY, 10, "bold"), text_color=MUTED
        ).pack(anchor="w", pady=(2, 0))
        
        def _refresh_modal():
            _load_data()
            render_page(current_page[0])
            
        ctk.CTkButton(
            header_frame, text="🧩 Manual Match", height=32,
            fg_color=PANEL, hover_color=PREVIEW_BG,
            border_color=BORDER_DARK, border_width=1,
            text_color=ACCENT, font=(FONT_FAMILY, 10, "bold"),
            corner_radius=6, command=lambda: [top.destroy(), self._on_manual_match()]
        ).pack(side="right", padx=(0, 12))

        ctk.CTkButton(
            header_frame, text="↻ Refresh Data", height=32,
            fg_color=PANEL, hover_color=PREVIEW_BG,
            border_color=BORDER_DARK, border_width=1,
            text_color=TEXT, font=(FONT_FAMILY, 10, "bold"),
            corner_radius=6, command=_refresh_modal
        ).pack(side="right", padx=(0, 24))
        
        # Modal Footer Toolbar (packed at bottom before body expansion so action buttons never clip)
        footer_frame = ctk.CTkFrame(top, fg_color=PANEL, corner_radius=0, height=70, border_color=BORDER, border_width=1)
        footer_frame.pack(fill="x", side="bottom")
        footer_frame.pack_propagate(False)

        # Main Scrollable Body
        body_frame = ctk.CTkFrame(top, fg_color=BG, corner_radius=0)
        body_frame.pack(fill="both", expand=True, padx=24, pady=16)
        
        list_frame = ctk.CTkFrame(body_frame, fg_color=PANEL, corner_radius=8, border_color=BORDER, border_width=1)
        list_frame.pack(fill="both", expand=True)
        
        canvas = tk.Canvas(list_frame, bg=PANEL, highlightthickness=0)
        scrollbar = ctk.CTkScrollbar(
            list_frame, orientation="vertical", command=canvas.yview,
            button_color="#94A3B8", button_hover_color=ACCENT
        )
        scrollable_frame = tk.Frame(canvas, bg=PANEL)
        
        def _update_scrollregion(e=None):
            canvas.update_idletasks()
            bbox = canvas.bbox("all")
            if bbox:
                canvas.configure(scrollregion=(0, 0, bbox[2], bbox[3]))

        scrollable_frame.bind("<Configure>", _update_scrollregion)
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_window, width=e.width))
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y", padx=(0, 4), pady=4)
        
        footer_inner = ctk.CTkFrame(footer_frame, fg_color="transparent")
        footer_inner.pack(fill="both", expand=True, padx=24, pady=14)

        pagination_frame = ctk.CTkFrame(footer_inner, fg_color="transparent")
        pagination_frame.pack(side="left")
        
        btn_prev = ctk.CTkButton(
            pagination_frame, text="◄ Prev", width=80, height=34,
            fg_color=PANEL, hover_color=PREVIEW_BG,
            border_color=BORDER_DARK, border_width=1,
            text_color=TEXT, font=(FONT_FAMILY, 11, "bold"), corner_radius=6
        )
        btn_prev.pack(side="left")
        
        lbl_page = ctk.CTkLabel(pagination_frame, text="", font=(FONT_FAMILY, 11, "bold"), text_color=TEXT)
        lbl_page.pack(side="left", padx=12)
        
        btn_next = ctk.CTkButton(
            pagination_frame, text="Next ►", width=80, height=34,
            fg_color=PANEL, hover_color=PREVIEW_BG,
            border_color=BORDER_DARK, border_width=1,
            text_color=TEXT, font=(FONT_FAMILY, 11, "bold"), corner_radius=6
        )
        btn_next.pack(side="left")
        
        def _on_mousewheel(event):
            try:
                if not canvas.winfo_exists():
                    return
                y_top, y_bot = canvas.yview()
                delta = event.delta
                if abs(delta) >= 120:
                    delta = int(delta / 120)
                
                # Boundary clamping: stop scrolling if at bounds
                if delta > 0 and y_top <= 0.001:
                    return
                if delta < 0 and y_bot >= 0.999:
                    return
                    
                if delta != 0:
                    canvas.yview_scroll(int(-1 * delta), "units")
            except Exception:
                pass
                
        top.bind("<MouseWheel>", _on_mousewheel)
        canvas.bind("<MouseWheel>", _on_mousewheel)
        scrollable_frame.bind("<MouseWheel>", _on_mousewheel)

        
        def render_page(page_idx):
            active_det = {"btn": None, "frm": None}
            for widget in scrollable_frame.winfo_children():
                widget.destroy()
                
            # Column Minwidth Specifications for Clean Breathing Room
            col_widths = {
                0: 40,   # Expand ►
                1: 80,   # Select
                2: 110,  # Date
                3: 140,  # Journal  ← stretchy
                4: 130,  # Merchant Amt
                5: 130,  # Odoo Amt
                6: 145,  # Mutation + Admin
                7: 100,  # Difference
                8: 80,   # EDC
                9: 115,  # EDC Status  ← stretchy
                10: 80,  # AR
                11: 115  # AR Status   ← stretchy
            }
            stretchy = {3, 9, 11}
            for col, w in col_widths.items():
                scrollable_frame.grid_columnconfigure(
                    col, minsize=w, weight=2 if col in stretchy else 0
                )
            # Trailing dummy column soaks up remaining slack
            scrollable_frame.grid_columnconfigure(len(col_widths), weight=1)

            headers = ["", "Select", "Date", "Journal", "Merchant Amt", "Odoo Amt", "Mutation + Admin", "Difference", "EDC", "EDC Status", "AR", "AR Status"]
            for col, h in enumerate(headers):
                lbl_anchor = "w" if col in [2, 3] else "e" if col in [4, 5, 6, 7] else "center"
                tk.Label(
                    scrollable_frame, text=h, bg=PREVIEW_BG, fg=MUTED,
                    font=(FONT_FAMILY, 10, "bold"), anchor=lbl_anchor, padx=14
                ).grid(row=0, column=col, sticky="nsew", pady=(0, 4), ipady=7)

            # Header Bottom Border Line
            tk.Frame(scrollable_frame, bg=BORDER_DARK, height=1).grid(row=1, column=0, columnspan=len(headers)+1, sticky="ew", pady=(0, 4))
            # Trailing header cell to fill full width
            tk.Label(scrollable_frame, text="", bg=PREVIEW_BG).grid(row=0, column=len(headers), sticky="nsew", ipady=7)

                
            start_idx = page_idx * ITEMS_PER_PAGE
            end_idx = min(start_idx + ITEMS_PER_PAGE, len(journal_state))
            
            for i, state in enumerate(journal_state[start_idx:end_idx]):
                r_main = i * 3 + 2
                r_det  = i * 3 + 3
                r_sep  = i * 3 + 4
                bg_row = PANEL if i % 2 == 0 else "#FAFAFC"
                
                item = state["item"]
                var_item = state["var_item"]
                var_edc = state["var_edc"]
                var_ar = state["var_ar"]
                
                amt = float(item['amount']) if item['amount'] else 0
                sel = float(item['selisih']) if item['selisih'] else 0
                
                def _on_jurnal_toggle(vi=var_item, ve=var_edc, va=var_ar, de=state["disabled_edc"], da=state["disabled_ar"]):
                    if not de:
                        ve.set(vi.get())
                    if not da:
                        va.set(vi.get())
                
                import config
                from collections import defaultdict
                
                b_name = str(item['bank']).lower()
                b_group = str(item['group']).lower()
                
                alias = ""
                props = {}
                for a, p in config.BANK_ACCOUNTS.get(b_name, {}).items():
                    if p.get("group", "").lower() == b_group:
                        alias = a
                        props = p
                        break
                        
                det_frame = tk.Frame(scrollable_frame, bg=PREVIEW_BG)
                
                # EDC Section Preview
                edc_debit = props.get("edc_debit") or f"{str(item['bank']).upper()} EDC Debit"
                edc_credit = props.get("edc_credit") or f"{str(item['group'])} Credit"
                
                edc_frame = tk.Frame(det_frame, bg=PREVIEW_BG)
                edc_frame.pack(side="left", anchor="n", padx=20, pady=10)
                
                tk.Label(edc_frame, text="EDC Journal:", bg=PREVIEW_BG, fg=ACCENT, font=(FONT_FAMILY, 9, "bold")).grid(row=0, column=0, columnspan=3, sticky="w")
                tk.Label(edc_frame, text="Debit:", bg=PREVIEW_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=1, column=0, sticky="w")
                tk.Label(edc_frame, text=edc_debit, bg=PREVIEW_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=1, column=1, sticky="w", padx=(10, 30))
                tk.Label(edc_frame, text=f"Rp {amt:,.0f}", bg=PREVIEW_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=1, column=2, sticky="e")
                tk.Label(edc_frame, text="Credit:", bg=PREVIEW_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=2, column=0, sticky="w")
                tk.Label(edc_frame, text=edc_credit, bg=PREVIEW_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=2, column=1, sticky="w", padx=(10, 30))
                tk.Label(edc_frame, text=f"Rp {amt:,.0f}", bg=PREVIEW_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=2, column=2, sticky="e")

                # Separator & Totals Footer for EDC
                tk.Frame(edc_frame, bg=BORDER_DARK, height=1).grid(row=3, column=0, columnspan=3, sticky="ew", pady=(4, 4))
                tk.Label(edc_frame, text=f"Total Debit: Rp {amt:,.0f}", bg=PREVIEW_BG, fg=SUCCESS, font=(FONT_FAMILY, 10, "bold")).grid(row=4, column=0, columnspan=2, sticky="w")
                tk.Label(edc_frame, text=f"Total Credit: Rp {amt:,.0f}", bg=PREVIEW_BG, fg=SUCCESS, font=(FONT_FAMILY, 10, "bold")).grid(row=4, column=2, sticky="e")

                
                # AR Section Preview
                from config import ODOO_ACCOUNT_BANK_DIFF_INCOME, ODOO_ACCOUNT_BANK_DIFF_LOSS
                if item.get("mutation_matched", False) or (item.get("mutation_found", False) and abs(sel) > 0.01):
                    m_date = item['payment_date']
                    m_group = item['group']
                    m_raw = [m for m in getattr(self, "_mutation_raw", []) if m["payment_date"] == m_date and m["group"] == m_group]
                    a_raw = [a for a in getattr(self, "_admin_raw", []) if a["payment_date"] == m_date and a["group"] == m_group]
                    
                    ar_debits = []
                    for m in m_raw:
                        cat = (m["category"] or "").replace(" ", "").lower()
                        acc = props.get(f"ar_debit_{cat}") or props.get("ar_debit") or f"AR DEBIT ({cat})"
                        ar_debits.append({"account": acc, "amount": m["amount"]})
                        
                    admin_sums = defaultdict(float)
                    for a in a_raw:
                        cat = (a["category"] or "").replace(" ", "").lower()
                        acc = props.get(f"admin_debit_{cat}") or props.get("admin_debit") or f"ADMIN DEBIT ({cat})"
                        admin_sums[acc] += a["amount"]
                        
                    for acc, a_amt in admin_sums.items():
                        if a_amt > 0:
                            ar_debits.append({"account": acc, "amount": a_amt})
                    
                    t_debit = sum(d["amount"] for d in ar_debits)
                    t_credit = float(amt)
                    t_diff = t_debit - t_credit
                    
                    ar_rows = []
                    for d in ar_debits:
                        ar_rows.append(("Debit:", d['account'], d['amount']))
                        
                    if t_credit > 0:
                        ar_rows.append(("Credit:", edc_debit, t_credit))
                    
                    if round(t_diff, 2) > 0:
                        ar_rows.append(("Credit:", ODOO_ACCOUNT_BANK_DIFF_INCOME, abs(t_diff)))
                    elif round(t_diff, 2) < 0:
                        ar_rows.append(("Debit:", ODOO_ACCOUNT_BANK_DIFF_LOSS, abs(t_diff)))
                        
                    ar_frame = tk.Frame(det_frame, bg=PREVIEW_BG)
                    ar_frame.pack(side="left", anchor="n", padx=30, pady=10)
                    
                    tk.Label(ar_frame, text="AR Journal:", bg=PREVIEW_BG, fg=ACCENT, font=(FONT_FAMILY, 9, "bold")).grid(row=0, column=0, columnspan=3, sticky="w")
                    
                    tot_deb = sum(amt_v for typ, acc, amt_v in ar_rows if typ == "Debit:")
                    tot_crd = sum(amt_v for typ, acc, amt_v in ar_rows if typ == "Credit:")

                    for idx, (typ, acc, amt_val) in enumerate(ar_rows):
                        r = idx + 1
                        tk.Label(ar_frame, text=typ, bg=PREVIEW_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=r, column=0, sticky="w")
                        tk.Label(ar_frame, text=acc, bg=PREVIEW_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=r, column=1, sticky="w", padx=(10, 30))
                        tk.Label(ar_frame, text=f"Rp {amt_val:,.0f}", bg=PREVIEW_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=r, column=2, sticky="e")
                    
                    # Separator & Totals Footer for AR
                    last_r = len(ar_rows) + 1
                    tot_color = SUCCESS if abs(tot_deb - tot_crd) < 0.01 else WARN
                    tk.Frame(ar_frame, bg=BORDER_DARK, height=1).grid(row=last_r, column=0, columnspan=3, sticky="ew", pady=(4, 4))
                    tk.Label(ar_frame, text=f"Total Debit: Rp {tot_deb:,.0f}", bg=PREVIEW_BG, fg=tot_color, font=(FONT_FAMILY, 10, "bold")).grid(row=last_r+1, column=0, columnspan=2, sticky="w")
                    tk.Label(ar_frame, text=f"Total Credit: Rp {tot_crd:,.0f}", bg=PREVIEW_BG, fg=tot_color, font=(FONT_FAMILY, 10, "bold")).grid(row=last_r+1, column=2, sticky="e")

                
                def _toggle_det(btn, frm=det_frame, row_idx=r_det):
                    if frm.winfo_ismapped():
                        frm.grid_remove()
                        btn.config(text="►")
                        if active_det["frm"] == frm:
                            active_det["btn"] = None
                            active_det["frm"] = None
                    else:
                        prev_frm = active_det["frm"]
                        prev_btn = active_det["btn"]
                        if prev_frm and prev_frm.winfo_exists() and prev_frm.winfo_ismapped():
                            prev_frm.grid_remove()
                            if prev_btn and prev_btn.winfo_exists():
                                prev_btn.config(text="►")
                        
                        frm.grid(row=row_idx, column=1, columnspan=10, sticky="w", pady=(8, 10))
                        btn.config(text="▼")
                        active_det["btn"] = btn
                        active_det["frm"] = frm
                        
                btn_expand = tk.Label(scrollable_frame, text="►", bg=bg_row, fg=ACCENT, cursor="hand2", font=(FONT_FAMILY, 10, "bold"))
                btn_expand.bind("<Button-1>", lambda e, b=btn_expand, f=det_frame: _toggle_det(b, f))
                btn_expand.grid(row=r_main, column=0, padx=8, ipady=6, sticky="nsew")
                
                def _make_cb(parent, variable, bg_color, command=None):
                    """18x18 canvas checkbox with proper rounded corners."""
                    s = 18
                    r = 3   # corner radius
                    m = 1   # margin from canvas edge
                    x, y = m, m
                    w, h = s - 2*m, s - 2*m  # 16x16 box
                    d = 2 * r                 # arc diameter

                    cv = tk.Canvas(parent, width=s, height=s, bg=bg_color,
                                   highlightthickness=0, cursor="hand2")

                    def _rrect(fill, outline, bw):
                        cv.delete("all")
                        if fill and fill != bg_color:
                            # Flood-fill using overlapping rects + ovals
                            cv.create_rectangle(x+r, y, x+w-r, y+h,
                                                fill=fill, outline="")
                            cv.create_rectangle(x, y+r, x+w, y+h-r,
                                                fill=fill, outline="")
                            cv.create_oval(x, y, x+d, y+d,
                                           fill=fill, outline="")
                            cv.create_oval(x+w-d, y, x+w, y+d,
                                           fill=fill, outline="")
                            cv.create_oval(x, y+h-d, x+d, y+h,
                                           fill=fill, outline="")
                            cv.create_oval(x+w-d, y+h-d, x+w, y+h,
                                           fill=fill, outline="")
                        # Border arcs (4 corners)
                        cv.create_arc(x, y, x+d, y+d,
                                      start=90,  extent=90, style="arc",
                                      outline=outline, width=bw)
                        cv.create_arc(x+w-d, y, x+w, y+d,
                                      start=0,   extent=90, style="arc",
                                      outline=outline, width=bw)
                        cv.create_arc(x, y+h-d, x+d, y+h,
                                      start=180, extent=90, style="arc",
                                      outline=outline, width=bw)
                        cv.create_arc(x+w-d, y+h-d, x+w, y+h,
                                      start=270, extent=90, style="arc",
                                      outline=outline, width=bw)
                        # Straight edges
                        cv.create_line(x+r, y,   x+w-r, y,   fill=outline, width=bw)
                        cv.create_line(x+r, y+h, x+w-r, y+h, fill=outline, width=bw)
                        cv.create_line(x,   y+r, x,   y+h-r, fill=outline, width=bw)
                        cv.create_line(x+w, y+r, x+w, y+h-r, fill=outline, width=bw)

                    def _draw(*_):
                        cv.delete("all")
                        if variable.get():
                            _rrect(fill=ACCENT, outline=ACCENT_DARK, bw=1)
                            cv.create_line(4, 9, 7, 13, fill=WHITE, width=2,
                                           capstyle="round", joinstyle="round")
                            cv.create_line(7, 13, 14, 5, fill=WHITE, width=2,
                                           capstyle="round", joinstyle="round")
                        else:
                            _rrect(fill=bg_color, outline="#94A3B8", bw=2)

                    def _toggle(e):
                        variable.set(1 - variable.get())
                        _draw()
                        if command:
                            command()

                    cv.bind("<Button-1>", _toggle)
                    variable.trace_add("write", _draw)
                    _draw()
                    return cv

                # Select checkbox
                if state["disabled_edc"] and state["disabled_ar"]:
                    _cell1 = tk.Frame(scrollable_frame, bg=bg_row)
                    _cell1.grid(row=r_main, column=1, sticky="nsew", pady=4)
                    _cell1.grid_rowconfigure(0, weight=1)
                    _cell1.grid_columnconfigure(0, weight=1)
                    tk.Label(_cell1, text="—", bg=bg_row, fg=MUTED,
                             font=(FONT_FAMILY, 11, "bold")).grid(row=0, column=0)
                else:
                    _make_cb(scrollable_frame, var_item, bg_row,
                             command=_on_jurnal_toggle
                             ).grid(row=r_main, column=1, pady=4)

                tk.Label(scrollable_frame, text=str(item['tanggal']), bg=bg_row, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=r_main, column=2, sticky="w", padx=14, ipady=6)
                tk.Label(scrollable_frame, text=str(item['group']), bg=bg_row, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=r_main, column=3, sticky="w", padx=14, ipady=6)
                amt_merch = float(item.get('merchant_amount') or 0)
                amt_odoo = float(item.get('odoo_amount') or 0)
                tk.Label(scrollable_frame, text=f"Rp {amt_merch:,.0f}", bg=bg_row, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=r_main, column=4, sticky="e", padx=14, ipady=6)
                tk.Label(scrollable_frame, text=f"Rp {amt_odoo:,.0f}", bg=bg_row, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=r_main, column=5, sticky="e", padx=14, ipady=6)
                mut_amt = float(item.get("mutation_amount", 0))
                tk.Label(scrollable_frame, text=f"Rp {mut_amt:,.0f}", bg=bg_row, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=r_main, column=6, sticky="e", padx=14, ipady=6)
                sel_color = WARN if sel != 0 else TEXT
                tk.Label(scrollable_frame, text=f"Rp {sel:,.0f}", bg=bg_row, fg=sel_color, font=(FONT_FAMILY, 10, "bold")).grid(row=r_main, column=7, sticky="e", padx=14, ipady=6)

                # EDC checkbox
                if state["disabled_edc"]:
                    _cell8 = tk.Frame(scrollable_frame, bg=bg_row)
                    _cell8.grid(row=r_main, column=8, sticky="nsew", pady=4)
                    _cell8.grid_rowconfigure(0, weight=1)
                    _cell8.grid_columnconfigure(0, weight=1)
                    tk.Label(_cell8, text="—", bg=bg_row, fg=MUTED,
                             font=(FONT_FAMILY, 11, "bold")).grid(row=0, column=0)
                else:
                    _make_cb(scrollable_frame, var_edc, bg_row
                             ).grid(row=r_main, column=8, pady=4)

                # AR checkbox
                if state["disabled_ar"]:
                    _cell10 = tk.Frame(scrollable_frame, bg=bg_row)
                    _cell10.grid(row=r_main, column=10, sticky="nsew", pady=4)
                    _cell10.grid_rowconfigure(0, weight=1)
                    _cell10.grid_columnconfigure(0, weight=1)
                    tk.Label(_cell10, text="—", bg=bg_row, fg=MUTED,
                             font=(FONT_FAMILY, 11, "bold")).grid(row=0, column=0)
                else:
                    _make_cb(scrollable_frame, var_ar, bg_row
                             ).grid(row=r_main, column=10, pady=4)
                
                edc_info_texts = []
                ar_info_texts = []
                is_reconciled = str(item.get("reconciled", "")).strip().lower() == "yes"
                status_valid = item.get("status_valid", True)

                # Build status lists: (label_text, is_good)
                edc_badges = []
                ar_badges = []

                if not is_reconciled:
                    edc_badges.append(("Unreconciled", False))
                elif not status_valid:
                    edc_badges.append(("Difference", False))

                if not is_reconciled:
                    ar_badges.append(("Unreconciled", False))
                elif not status_valid:
                    ar_badges.append(("Difference", False))
                elif not item.get("mutation_matched", False):
                    if not item.get("mutation_found", False):
                        ar_badges.append(("No Mutation", False))
                    else:
                        ar_badges.append(("Mut Difference", False))

                j_status = item.get("journal_status")
                if j_status:
                    j_status_str = str(j_status).strip()
                    if j_status_str not in ["", "None", "Not Yet"]:
                        parts = [p.strip() for p in j_status_str.split("|")]
                        for p in parts:
                            if "Posted" in p:
                                label = "Posted (Diff)" if "Difference" in p else "Posted"
                                is_good = "Difference" not in p
                            elif "Draft" in p:
                                label = "Draft"
                                is_good = True
                            else:
                                label = p
                                is_good = False
                            if "(Both" in p:
                                if not edc_badges: edc_badges.append((label, is_good))
                                if not ar_badges: ar_badges.append((label, is_good))
                            elif "(EDC" in p:
                                if not edc_badges: edc_badges.append((label, is_good))
                            elif "(AR" in p:
                                if not ar_badges: ar_badges.append((label, is_good))

                def _render_status(parent, col, badges, bg_row):
                    if not badges:
                        return
                    txt = "\n".join(t for t, _ in badges)
                    good = all(g for _, g in badges)
                    tk.Label(
                        parent, text=txt, bg=bg_row,
                        fg=SUCCESS if good else WARN,
                        font=(FONT_FAMILY, 10, "bold"),
                        justify="center", anchor="center"
                    ).grid(row=r_main, column=col, sticky="nsew", padx=6, ipady=6)

                _render_status(scrollable_frame, 9, edc_badges, bg_row)
                _render_status(scrollable_frame, 11, ar_badges, bg_row)

                # Row Divider Line
                tk.Frame(scrollable_frame, bg=BORDER, height=1).grid(row=r_sep, column=0, columnspan=len(headers)+1, sticky="ew")


                
            total_pages = max(1, (len(journal_state) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE)
            lbl_page.configure(text=f"Page {page_idx + 1} of {total_pages}")
            
            btn_prev.configure(state="normal" if page_idx > 0 else "disabled")
            btn_next.configure(state="normal" if page_idx < total_pages - 1 else "disabled")
            
            canvas.yview_moveto(0)
            
        def _prev_page():
            if current_page[0] > 0:
                current_page[0] -= 1
                render_page(current_page[0])
                
        def _next_page():
            total_pages = (len(journal_state) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE
            if current_page[0] < total_pages - 1:
                current_page[0] += 1
                render_page(current_page[0])
                
        btn_prev.configure(command=_prev_page)
        btn_next.configure(command=_next_page)
        
        render_page(0)
        
        def _get_selected_config():
            selected = []
            for v in journal_state:
                if v["var_item"].get():
                    selected.append({
                        "row": v["item"]["row"],
                        "edc": v["var_edc"].get(),
                        "ar": v["var_ar"].get()
                    })
            return selected
            
        def _export(mode="edc"):
            selected = _get_selected_config()
            if not selected:
                messagebox.showwarning("Warning", "Select at least 1 transaction")
                return

            # Guard: check that the requested mode has at least 1 row to export
            if mode == "ar":
                ar_items = [item for item in selected if item.get("ar", False)]
                if not ar_items:
                    messagebox.showinfo(
                        "No AR Entries",
                        "None of the selected transactions have AR checked.\n\n"
                        "To export an AR journal, tick the AR checkbox on at least one transaction first."
                    )
                    return
            elif mode == "edc":
                edc_items = [item for item in selected if item.get("edc", False)]
                if not edc_items:
                    messagebox.showinfo(
                        "No EDC Entries",
                        "None of the selected transactions have EDC checked.\n\n"
                        "To export an EDC journal, tick the EDC checkbox on at least one transaction first."
                    )
                    return

            import json
            from journal_generator import generate_journal_import
            config_path = BASE_DIR / ".journal_config.json"
            if sys.platform == "win32" and config_path.exists():
                try:
                    import subprocess
                    flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
                    subprocess.run(["attrib", "-H", str(config_path)], check=False, capture_output=True, creationflags=flags)
                except Exception:
                    pass
            config_path.write_text(json.dumps(selected))
            try:
                out_path = generate_journal_import(latest_file, config_path, mode=mode, is_preview=True)
                if out_path:
                    messagebox.showinfo("Success", f"{mode.upper()} Journal exported:\n{out_path.name}")
                    _open_path(str(out_path))
                else:
                    messagebox.showwarning(
                        "Nothing Exported",
                        f"No {mode.upper()} journal rows were generated.\n\n"
                        "This can happen if all matching transactions are within the rounding tolerance "
                        "or if the reconciliation file has no data in the Daily Summary sheet."
                    )
            except Exception as e:
                messagebox.showerror("Error", f"Error exporting {mode.upper()}: {e}")


        def _process():
            selected = _get_selected_config()
            if not selected:
                messagebox.showwarning("Warning", "Select at least 1 transaction")
                return
            import json
            config_path = BASE_DIR / ".journal_config.json"
            if sys.platform == "win32" and config_path.exists():
                try:
                    import subprocess
                    flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
                    subprocess.run(["attrib", "-H", str(config_path)], check=False, capture_output=True, creationflags=flags)
                except Exception:
                    pass
            config_path.write_text(json.dumps(selected))

            
            from journal_generator import generate_journal_import
            recon_file = Path(latest_file)
            if not recon_file.exists():
                messagebox.showerror("Error", f"Reconciliation file not found at {recon_file.name}. Did you rename it while this window was open?")
                return
                
            try:
                out_path = generate_journal_import(recon_file, config_path, mode="both", is_preview=True)
                if not out_path:
                    messagebox.showerror("Error", "Failed to generate combined preview file.")
                    return
            except Exception as e:
                messagebox.showerror("Error", f"Error generating combined preview: {e}")
                return
                
            ar_count = sum(1 for item in selected if item.get("ar", False))
            edc_count = sum(1 for item in selected if item.get("edc", False))
            
            def _show_custom_confirm():
                dlg = ctk.CTkToplevel(top)
                dlg.title("Confirm Upload")
                dlg.geometry("520x290")
                dlg.configure(fg_color=PANEL)
                dlg.transient(top)
                dlg.grab_set()
                
                dlg.update_idletasks()
                x = top.winfo_x() + (top.winfo_width() - 520) // 2
                y = top.winfo_y() + (top.winfo_height() - 290) // 2
                dlg.geometry(f"+{x}+{y}")
                
                ctk.CTkLabel(
                    dlg, text="Confirm Upload to Odoo",
                    font=(FONT_FAMILY, 15, "bold"), text_color=TEXT
                ).pack(pady=(24, 8))
                
                ctk.CTkLabel(
                    dlg, 
                    text=f"Ready to import to Odoo.\n\nAR Journals: {ar_count}   |   EDC Journals: {edc_count}\n\nTo make manual edits before importing, click 'Edit Excel'.", 
                    justify="center", text_color=MUTED, font=(FONT_FAMILY, 10, "bold")
                ).pack(pady=(0, 20), padx=20)
                
                result = {"confirm": False}
                
                def on_upload():
                    result["confirm"] = True
                    dlg.destroy()
                    
                def on_cancel():
                    dlg.destroy()
                
                btn_frame = ctk.CTkFrame(dlg, fg_color="transparent")
                btn_frame.pack(fill="x", padx=20, pady=10)
                
                ctk.CTkButton(
                    btn_frame, text="Cancel", height=36,
                    fg_color=PANEL, hover_color=PREVIEW_BG,
                    border_color=BORDER_DARK, border_width=1,
                    text_color=TEXT, font=(FONT_FAMILY, 10, "bold"),
                    corner_radius=6, command=on_cancel
                ).pack(side="left", padx=4, expand=True, fill="x")
                
                ctk.CTkButton(
                    btn_frame, text="Edit Excel", height=36,
                    fg_color=PANEL, hover_color=PREVIEW_BG,
                    border_color=BORDER_DARK, border_width=1,
                    text_color=TEXT, font=(FONT_FAMILY, 10, "bold"),
                    corner_radius=6, command=lambda: _open_path(str(out_path))
                ).pack(side="left", padx=4, expand=True, fill="x")
                
                ctk.CTkButton(
                    btn_frame, text="Upload to Odoo", height=36,
                    fg_color=SUCCESS, hover_color=SUCCESS_DARK,
                    text_color=WHITE, font=(FONT_FAMILY, 10, "bold"),
                    corner_radius=6, command=on_upload
                ).pack(side="left", padx=4, expand=True, fill="x")
                
                top.wait_window(dlg)
                return result["confirm"]
                
            confirm = _show_custom_confirm()
            
            if not confirm:
                return
                
            self._journal_window = None
            top.destroy()

            
            self._running = True
            self._set_status("Uploading Edited Journal to Odoo...", WARN)
            self._journal_btn.configure(state="disabled")
            
            def run_script():
                try:
                    env = os.environ.copy()
                    env.pop("TCL_LIBRARY", None)
                    env.pop("TK_LIBRARY", None)
                    env["PYTHONIOENCODING"] = "utf-8"
                    env["PYTHONUTF8"] = "1"
                    
                    flags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if IS_WINDOWS else 0
                    
                    if getattr(sys, "frozen", False):
                        cmd = [
                            sys.executable, "--run-journal-creator",
                            "--file", str(recon_file),
                            "--import-file", str(out_path),
                            "--config", str(config_path)
                        ]
                    else:
                        cmd = [
                            _venv_python, "odoo_journal_creator.py",
                            "--file", str(recon_file),
                            "--import-file", str(out_path),
                            "--config", str(config_path)
                        ]

                    if hasattr(self, "_stop_btn"):
                        self.after(0, lambda: self._stop_btn.pack(fill="x", pady=(6, 0)))

                    proc = subprocess.Popen(
                        cmd, cwd=str(BASE_DIR),
                        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                        text=True, bufsize=1, encoding="utf-8",
                        env=env, creationflags=flags
                    )
                    self._active_proc = proc

                    if proc.stdout:
                        for line in iter(proc.stdout.readline, ''):
                            self.after(0, self._log_write, line, "")
                    proc.wait()

                    if proc.returncode == 0:
                        self.after(0, self._on_done, 0, None)
                    else:
                        self.after(0, self._on_done, proc.returncode, None)

                except Exception as e:
                    self.after(0, self._log_write, f"ERROR: {str(e)}\n", "err")
                    self.after(0, self._on_done, 1, None)
            
            threading.Thread(target=run_script, daemon=True).start()
            
        footer_right = ctk.CTkFrame(footer_inner, fg_color="transparent")
        footer_right.pack(side="right")
        
        def _select_all():
            for v in journal_state:
                if not (v["disabled_edc"] and v["disabled_ar"]):
                    v["var_item"].set(True)
                if not v["disabled_edc"]:
                    v["var_edc"].set(True)
                if not v["disabled_ar"]:
                    v["var_ar"].set(True)
            render_page(current_page[0])
                
        def _deselect_all():
            for v in journal_state:
                v["var_item"].set(False)
                if not v["disabled_edc"]:
                    v["var_edc"].set(False)
                if not v["disabled_ar"]:
                    v["var_ar"].set(False)
            render_page(current_page[0])
                
        tools_frame = ctk.CTkFrame(pagination_frame, fg_color="transparent")
        tools_frame.pack(side="left", padx=(20, 0))
        
        ctk.CTkButton(
            tools_frame, text="Select All", height=32, width=80,
            fg_color=PANEL, hover_color=PREVIEW_BG,
            border_color=BORDER_DARK, border_width=1,
            text_color=TEXT, font=(FONT_FAMILY, 10, "bold"),
            corner_radius=6, command=_select_all
        ).pack(side="left", padx=(0, 6))

        ctk.CTkButton(
            tools_frame, text="Deselect All", height=32, width=85,
            fg_color=PANEL, hover_color=PREVIEW_BG,
            border_color=BORDER_DARK, border_width=1,
            text_color=TEXT, font=(FONT_FAMILY, 10, "bold"),
            corner_radius=6, command=_deselect_all
        ).pack(side="left")
        
        ctk.CTkButton(
            footer_right, text="Cancel", height=36, width=80,
            fg_color=PANEL, hover_color=PREVIEW_BG,
            border_color=BORDER_DARK, border_width=1,
            text_color=TEXT, font=(FONT_FAMILY, 10, "bold"),
            corner_radius=6, command=top.destroy
        ).pack(side="left", padx=(0, 8))
        
        ctk.CTkButton(
            footer_right, text="Export EDC", height=36, width=95,
            fg_color=PANEL, hover_color=PREVIEW_BG,
            border_color=BORDER_DARK, border_width=1,
            text_color=TEXT, font=(FONT_FAMILY, 10, "bold"),
            corner_radius=6, command=lambda: _export("edc")
        ).pack(side="left", padx=(0, 6))
        
        ctk.CTkButton(
            footer_right, text="Export AR", height=36, width=90,
            fg_color=PANEL, hover_color=PREVIEW_BG,
            border_color=BORDER_DARK, border_width=1,
            text_color=TEXT, font=(FONT_FAMILY, 10, "bold"),
            corner_radius=6, command=lambda: _export("ar")
        ).pack(side="left", padx=(0, 12))
        
        ctk.CTkButton(
            footer_right, text="Submit", height=36, width=100,
            fg_color=SUCCESS, hover_color=SUCCESS_DARK,
            text_color=WHITE, font=(FONT_FAMILY, 10, "bold"),
            corner_radius=6, command=_process
        ).pack(side="left")

    def _open_output(self):
        path = getattr(self, "_last_output", None)
        if path and Path(path).exists() and Path(path).is_file():
            _open_path(str(path))
            return

        if OUTPUT_DIR.exists():
            files = [f for f in OUTPUT_DIR.glob("Reconciliation_*.xlsx") if f.is_file() and not f.name.startswith("~$")]
            if files:
                latest = max(files, key=lambda p: p.stat().st_mtime)
                _open_path(str(latest))
                return
            _open_path(str(OUTPUT_DIR))
        else:
            self._set_status("No reconciliation file found. Run recon first!", WARN)

    def _open_odoo_file(self):
        from config import ODO_EXCEL_PATH
        if ODO_EXCEL_PATH.exists():
            _open_path(str(ODO_EXCEL_PATH))
        else:
            self._set_status("Odoo payment file not found. Click Download first!", WARN)

    def _open_journal_file(self):
        from config import ODO_JOURNAL_EXCEL_PATH
        if ODO_JOURNAL_EXCEL_PATH.exists():
            _open_path(str(ODO_JOURNAL_EXCEL_PATH))
        else:
            self._set_status("Odoo journal entries file not found.", WARN)

    def _open_input(self):
        input_dir = BASE_DIR / "input"
        input_dir.mkdir(exist_ok=True)
        _open_path(str(input_dir))

    def _open_mutation(self):
        from config import MUTATION_DIR
        MUTATION_DIR.mkdir(exist_ok=True)
        _open_path(str(MUTATION_DIR))

    def _open_recap(self):
        recap_dir = BASE_DIR / "recap"
        recap_dir.mkdir(parents=True, exist_ok=True)
        _open_path(str(recap_dir))

    def _on_export_summary_pdf(self):
        def _run():
            self._set_status("Generating PDF Summary...", WARN)
            self._log_write("\n── Generating Executive PDF Summary ──\n", "head")
            try:
                from pdf_summary_generator import generate_executive_summary_pdf
                pdf_path = generate_executive_summary_pdf()
                self.after(0, self._log_write, f"✅ Executive Summary PDF Generated: {pdf_path.name}\n", "ok")
                self.after(0, self._set_status, "PDF Ready", SUCCESS)
                _open_path(str(pdf_path))
            except Exception as e:
                self.after(0, self._log_write, f"❌ PDF generation failed: {e}\n", "err")
                self.after(0, self._set_status, "PDF Error", ERROR)

        threading.Thread(target=_run, daemon=True).start()


if __name__ == "__main__":
    import sys
    import io

    if len(sys.argv) > 1:
        try:
            if sys.stdout is not None and hasattr(sys.stdout, "buffer"):
                sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
            if sys.stderr is not None and hasattr(sys.stderr, "buffer"):
                sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
        except (AttributeError, TypeError):
            pass
            
        if sys.argv[1] == "--run-downloader":
            import odoo_downloader
            sys.argv = [sys.argv[0]] + sys.argv[2:]
            odoo_downloader.run_downloader()
            sys.exit(0)
        elif sys.argv[1] == "--run-main":
            import runpy
            runpy.run_module('main', run_name='__main__')
            sys.exit(0)
        elif sys.argv[1] == "--install-playwright":
            import runpy
            sys.argv = [sys.argv[0], "install", "chromium"]
            runpy.run_module('playwright', run_name='__main__')
            sys.exit(0)
        elif sys.argv[1] == "-m":
            import runpy
            module_name = sys.argv[2]
            sys.argv = [sys.argv[0]] + sys.argv[3:]
            runpy.run_module(module_name, run_name='__main__')
            sys.exit(0)
        elif sys.argv[1].endswith(".py"):
            import runpy
            script_name = sys.argv[1][:-3]
            sys.argv = [sys.argv[0]] + sys.argv[2:]
            runpy.run_module(script_name, run_name='__main__')
            sys.exit(0)

    app = App()
    app.mainloop()
