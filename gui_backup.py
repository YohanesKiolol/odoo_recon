"""
gui.py — Bank Reconciliation GUI. Cross-platform (Windows + Mac).

PyInstaller pattern:
  - GUI mode   : BankRekonsiliasi.exe            → opens this window
  - Worker mode: BankRekonsiliasi.exe --worker   → runs main logic, stdout only
"""
import sys
import os
import platform
from pathlib import Path

# ── Resolve base directory ────────────────────────────────────────────────────
BASE_DIR = (
    Path(sys.executable).parent   # frozen: next to the .exe
    if getattr(sys, "frozen", False)
    else Path(__file__).parent    # dev: next to gui.py
)
INPUT_DIR  = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
IS_WINDOWS = platform.system() == "Windows"

# ── Worker mode ───────────────────────────────────────────────────────────────
# The GUI re-launches the same binary with --worker for a clean stdout stream.
# Must be checked BEFORE any tkinter import so the headless worker doesn't
# require a display or GUI toolkit.
if "--worker" in sys.argv:
    sys.argv = [a for a in sys.argv if a != "--worker"]  # hide flag from main's argparse
    os.chdir(str(BASE_DIR))
    import runpy
    if getattr(sys, "frozen", False):
        main_path = str(Path(getattr(sys, "_MEIPASS", BASE_DIR)) / "main.py")
    else:
        main_path = str(BASE_DIR / "main.py")
    runpy.run_path(main_path, run_name="__main__")
    sys.exit(0)

# ── GUI-only imports (skipped in worker mode) ─────────────────────────────────
if IS_WINDOWS:
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

import threading
import subprocess
import shutil
import fnmatch
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import scrolledtext, filedialog, messagebox
try:
    from tkcalendar import DateEntry
except ImportError:
    DateEntry = None

# Force PyInstaller to bundle dynamic imports
try:
    import config
    import odoo_downloader
except Exception:
    pass

# Venv python path (dev mode only — frozen uses sys.executable)
if os.name == 'nt':
    _venv_python_path = BASE_DIR / ".venv" / "Scripts" / "python.exe"
else:
    _venv_python_path = BASE_DIR / ".venv" / "bin" / "python"

_venv_python = (
    sys.executable if getattr(sys, "frozen", False) 
    else str(_venv_python_path)
)

# ── Color palette ─────────────────────────────────────────────────────────────
BG          = "#F9FAFB" # Very light modern gray background
PANEL       = "#FFFFFF" # Pure white for inputs/cards
PREVIEW_BG  = "#F3F4F6" # Subtle light gray for nested panels
BORDER      = "#E5E7EB" # Very subtle gray border
ACCENT      = "#71639E" # Odoo Purple
ACCENT_DARK = "#5B4F80" # Darker Odoo Purple for hover
SUCCESS     = "#017E84" # Odoo Teal
SUCCESS_DARK= "#016064" # Darker Odoo Teal for hover
ERROR       = "#EF4444" # Modern flat red
WARN        = "#F59E0B" # Modern flat amber
TEXT        = "#111827" # Crisp near-black
MUTED       = "#6B7280" # Standard gray for less important text
WHITE       = "#FFFFFF"


def _open_path(path: str):
    """Open a file/folder in the OS default app."""
    if IS_WINDOWS:
        os.startfile(path) # type: ignore
    elif sys.platform == "darwin":
        subprocess.run(["open", path])
    else:
        subprocess.run(["xdg-open", path])


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Bank Reconciliation Tool")
        self.configure(bg=BG)
        self.geometry("900x700")
        self.minsize(700, 500)
        self.resizable(True, True)
        self._running = False
        self._last_output: str | None = None
        self._build_ui()
        self._center()

    # ── UI ────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        # Main container with padding
        main_container = tk.Frame(self, bg=BG)
        main_container.pack(fill="both", expand=True, padx=40, pady=30)
        
        # Header
        hdr = tk.Frame(main_container, bg=BG)
        hdr.pack(fill="x", pady=(0, 20))
        tk.Label(hdr, text="Bank Reconciliation", bg=BG, fg=TEXT, font=("Segoe UI", 16, "bold")).pack(anchor="w")
        tk.Label(hdr, text="Compare Bank transactions with Odoo automatically", bg=BG, fg=MUTED, font=("Segoe UI", 9)).pack(anchor="w", pady=(4, 0))

        # --- Card 1: Configuration ---
        config_card = tk.Frame(main_container, bg=PANEL, highlightbackground=BORDER, highlightthickness=1)
        config_card.pack(fill="x", pady=(0, 15))
        
        config_inner = tk.Frame(config_card, bg=PANEL, padx=25, pady=20)
        config_inner.pack(fill="both", expand=True)
        
        # tk.Label(config_inner, text="Configuration", bg=PANEL, fg=TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 15))
        
        form_frame = tk.Frame(config_inner, bg=PANEL)
        form_frame.pack(fill="x")
        
        entry_style = {"bg": WHITE, "fg": TEXT, "insertbackground": TEXT, "borderwidth": 1, "highlightthickness": 1, "highlightbackground": BORDER, "highlightcolor": ACCENT, "relief": "flat"}
        
        # Row 1: Credentials
        tk.Label(form_frame, text="Email", bg=PANEL, fg=MUTED, font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", pady=(0, 5))
        self._email_var = tk.StringVar()
        tk.Entry(form_frame, textvariable=self._email_var, width=35, **entry_style).grid(row=1, column=0, sticky="w", padx=(0, 25), pady=(0, 15), ipady=4)
        
        tk.Label(form_frame, text="Password", bg=PANEL, fg=MUTED, font=("Segoe UI", 9)).grid(row=0, column=1, sticky="w", pady=(0, 5))
        self._password_var = tk.StringVar()
        pass_entry_frame = tk.Frame(form_frame, bg=PANEL)
        pass_entry_frame.grid(row=1, column=1, sticky="w", pady=(0, 15))
        self._password_entry = tk.Entry(pass_entry_frame, textvariable=self._password_var, width=35, show="*", **entry_style)
        self._password_entry.pack(side="left", ipady=4)
        
        eye_lbl = tk.Label(pass_entry_frame, text="👁", bg=PANEL, fg=MUTED, cursor="hand2")
        eye_lbl.pack(side="left", padx=(10, 0))
        eye_lbl.bind("<ButtonPress-1>", lambda e: self._password_entry.config(show=""))
        eye_lbl.bind("<ButtonRelease-1>", lambda e: self._password_entry.config(show="*"))
        
        # Row 2: Bank Target and Date Range
        tk.Label(form_frame, text="Bank Target", bg=PANEL, fg=MUTED, font=("Segoe UI", 9)).grid(row=2, column=0, sticky="w", pady=(0, 5))
        bank_frame = tk.Frame(form_frame, bg=PANEL)
        bank_frame.grid(row=3, column=0, sticky="w", pady=(0, 5))
        
        self._bank_vars = {
            "BCA": tk.BooleanVar(value=False),
            "Mandiri": tk.BooleanVar(value=False),
            "BRI": tk.BooleanVar(value=False),
            "All": tk.BooleanVar(value=True)
        }
        
        def _on_bank_toggle(bank_name):
            if bank_name == "All" and self._bank_vars["All"].get():
                for b in ["BCA", "Mandiri", "BRI"]:
                    self._bank_vars[b].set(False)
            elif bank_name in ["BCA", "Mandiri", "BRI"] and self._bank_vars[bank_name].get():
                self._bank_vars["All"].set(False)
        
        for b in ["All", "BCA", "Mandiri", "BRI"]:
            cb = tk.Checkbutton(
                bank_frame, text=b, variable=self._bank_vars[b],
                bg=PANEL, fg=TEXT, selectcolor=PANEL, activebackground=PANEL, activeforeground=TEXT,
                font=("Segoe UI", 10), cursor="hand2", command=lambda name=b: _on_bank_toggle(name)
            )
            cb.pack(side="left", padx=(0, 15))
            
        tk.Label(form_frame, text="Date Range", bg=PANEL, fg=MUTED, font=("Segoe UI", 9)).grid(row=2, column=1, sticky="w", pady=(0, 5))
        date_frame = tk.Frame(form_frame, bg=PANEL)
        date_frame.grid(row=3, column=1, sticky="w", pady=(0, 5))
        
        yesterday = datetime.now() - timedelta(days=1)
        if DateEntry:
            self._date_from_widget = DateEntry(date_frame, width=12, background=ACCENT, foreground=WHITE, fieldbackground=WHITE, borderwidth=1, date_pattern='mm/dd/yyyy')
            self._date_from_widget.pack(side="left", padx=(0, 10))
            self._date_from_widget.set_date(yesterday)
            
            tk.Label(date_frame, text="to", bg=PANEL, fg=MUTED, font=("Segoe UI", 9)).pack(side="left", padx=(0, 10))
            
            self._date_to_widget = DateEntry(date_frame, width=12, background=ACCENT, foreground=WHITE, fieldbackground=WHITE, borderwidth=1, date_pattern='mm/dd/yyyy')
            self._date_to_widget.pack(side="left")
            self._date_to_widget.set_date(yesterday)
            
            def _validate_dates(event=None):
                d_from = self._date_from_widget.get_date()
                d_to = self._date_to_widget.get_date()
                if d_to < d_from:
                    if event and event.widget == self._date_from_widget:
                        self._date_to_widget.set_date(d_from)
                    else:
                        self._date_from_widget.set_date(d_to)
            
            self._date_from_widget.bind("<<DateEntrySelected>>", _validate_dates)
            self._date_to_widget.bind("<<DateEntrySelected>>", _validate_dates)
        else:
            self._date_from_var = tk.StringVar(value=yesterday.strftime("%m/%d/%Y"))
            tk.Entry(date_frame, textvariable=self._date_from_var, width=12, **entry_style).pack(side="left", padx=(0, 10))
            tk.Label(date_frame, text="to", bg=PANEL, fg=MUTED, font=("Segoe UI", 9)).pack(side="left", padx=(0, 10))
            self._date_to_var = tk.StringVar(value=yesterday.strftime("%m/%d/%Y"))
            tk.Entry(date_frame, textvariable=self._date_to_var, width=12, **entry_style).pack(side="left")

        # --- Card 2: Operations ---
        ops_card = tk.Frame(main_container, bg=PANEL, highlightbackground=BORDER, highlightthickness=1)
        ops_card.pack(fill="x", pady=(0, 15))
        ops_inner = tk.Frame(ops_card, bg=PANEL, padx=25, pady=20)
        ops_inner.pack(fill="both", expand=True)
        
        self._folder_label = tk.Label(ops_inner, bg=PANEL, fg=MUTED, font=("Segoe UI", 9), anchor="w")
        self._folder_label.pack(fill="x", pady=(0, 15))
        self._refresh_folder_status()
        
        action_frame = tk.Frame(ops_inner, bg=PANEL)
        action_frame.pack(fill="x")
        
        sec_btn_style = {"font": ("Segoe UI", 9, "bold"), "relief": "solid", "cursor": "hand2", "padx": 16, "pady": 6, "borderwidth": 1, "highlightbackground": BORDER}
        prim_btn_style = {"font": ("Segoe UI", 9, "bold"), "relief": "flat", "cursor": "hand2", "padx": 16, "pady": 7, "borderwidth": 0}
        
        left_actions = tk.Frame(action_frame, bg=PANEL)
        left_actions.pack(side="left")
        
        right_actions = tk.Frame(action_frame, bg=PANEL)
        right_actions.pack(side="right")
        
        self._upload_btn = tk.Button(left_actions, text="Upload", bg=WHITE, fg=TEXT, activebackground=PREVIEW_BG, command=self._on_upload, **sec_btn_style)
        self._upload_btn.pack(side="left", padx=(0, 8))
        self._download_btn = tk.Button(left_actions, text="Download", bg=WHITE, fg=TEXT, activebackground=PREVIEW_BG, command=self._on_download, **sec_btn_style)
        self._download_btn.pack(side="left", padx=(0, 8))
        self._cleanse_btn = tk.Button(left_actions, text="Clean", bg=WHITE, fg=ERROR, activebackground=PREVIEW_BG, command=self._on_cleanse, **sec_btn_style)
        self._cleanse_btn.pack(side="left", padx=(0, 8))
        self._scan_btn = tk.Button(left_actions, text="Scan Data", bg=WHITE, fg=TEXT, activebackground=PREVIEW_BG, command=self._on_scan, **sec_btn_style)
        self._scan_btn.pack(side="left", padx=(0, 8))
        
        self._run_btn = tk.Button(right_actions, text="Reconciliation", bg=SUCCESS, fg=WHITE, activebackground=SUCCESS_DARK, activeforeground=WHITE, command=self._on_run, **prim_btn_style)
        self._run_btn.pack(side="left", padx=(0, 8))
        self._journal_btn = tk.Button(right_actions, text="Generate Journal", bg=ACCENT, fg=WHITE, activebackground=ACCENT_DARK, activeforeground=WHITE, command=self._on_journal, **prim_btn_style)
        self._journal_btn.pack(side="left")

        # --- Card 3: Output Logs ---
        log_card = tk.Frame(main_container, bg=PANEL, highlightbackground=BORDER, highlightthickness=1)
        log_card.pack(fill="both", expand=True)
        log_inner = tk.Frame(log_card, bg=PANEL, padx=0, pady=0)
        log_inner.pack(fill="both", expand=True)
        
        status_row = tk.Frame(log_inner, bg=PANEL, padx=20, pady=12)
        status_row.pack(fill="x", side="bottom")
        
        tk.Frame(log_inner, bg=BORDER, height=1).pack(fill="x", side="bottom")
        
        self._log = scrolledtext.ScrolledText(
            log_inner, bg=PANEL, fg=TEXT, insertbackground=TEXT,
            font=("Consolas", 10), relief="flat", state="disabled",
            wrap="word", borderwidth=0, highlightthickness=0, padx=20, pady=20
        )
        self._log.pack(fill="both", expand=True)
        self._log.tag_config("ok",   foreground=SUCCESS)
        self._log.tag_config("err",  foreground=ERROR)
        self._log.tag_config("warn", foreground=WARN)
        self._log.tag_config("head", foreground=ACCENT, font=("Consolas", 10, "bold"))
        self._log.tag_config("dim",  foreground=MUTED)
        
        self._status_var = tk.StringVar(value="Ready")
        self._dot = tk.Label(status_row, text="●", bg=PANEL, fg=MUTED, font=("Segoe UI", 12))
        self._dot.pack(side="left")
        tk.Label(status_row, textvariable=self._status_var, bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold")).pack(side="left", padx=(6, 0))
        
        util_links = tk.Frame(status_row, bg=PANEL)
        util_links.pack(side="right")
        
        def _make_link(parent, text, cmd):
            lbl = tk.Label(parent, text=text, bg=PANEL, fg=ACCENT, font=("Segoe UI", 9, "underline"), cursor="hand2")
            lbl.pack(side="left", padx=(15, 0))
            def _on_click(e):
                if lbl.cget("state") != "disabled":
                    cmd()
            lbl.bind("<Button-1>", _on_click)
            return lbl
            
        self._clear_log_btn = _make_link(util_links, "↻ Clear Logs", self._clear_log)
        self._open_input_btn = _make_link(util_links, "Open Merchant", self._open_input)
        self._open_mutation_btn = _make_link(util_links, "Open Mutation", self._open_mutation)
        self._open_btn = _make_link(util_links, "Open Result", self._open_output)

    def _center(self):
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w,  h  = self.winfo_width(),       self.winfo_height()
        self.geometry(f"+{(sw-w)//2}+{(sh-h)//2}")

    def _refresh_folder_status(self):
        files = list(INPUT_DIR.rglob("*")) if INPUT_DIR.exists() else []
        n = sum(1 for f in files if f.is_file())
        self._folder_label.config(
            text=f"📂  Input folder: {INPUT_DIR}   ({n} file found)",
            fg=SUCCESS if n > 0 else WARN,
        )

    def _log_write(self, text: str, tag: str = ""):
        self._log.config(state="normal")
        self._log.insert("end", text, tag)
        self._log.see("end")
        self._log.config(state="disabled")

    def _set_status(self, text: str, color: str):
        self._status_var.set(text)
        self._dot.config(fg=color)

    def _clear_log(self):
        self._log.config(state="normal")
        self._log.delete("1.0", tk.END)
        self._log.config(state="disabled")

    # ── New Feature Stubs ─────────────────────────────────────────────────────
    def _on_upload(self):
        try:
            from config import (
                BCA_EXCEL_DIR, BCA_EXCEL_PATTERN, 
                MANDIRI_ZIP_DIR, MANDIRI_ZIP_PATTERN, 
                BRI_ZIP_DIR, BRI_PDF_PATTERN, ODO_EXCEL_PATH,
                MUTATION_DIR, BCA_MUTATION_PATTERN,
                MANDIRI_MUTATION_PATTERN, BRI_MUTATION_PATTERN
            )
            
            files = filedialog.askopenfilenames(title="Select Bank or Odoo File", filetypes=[("All Files", "*.*")])
            if not files:
                return
                
            self._log_write("\n── Uploading Files ──\n", "head")
            for f in files:
                path = Path(f)
                name = path.name.lower()
                
                target_dir = None
                matched_bank = None
                is_mutation = False
                
                # 1. Check Mutations First
                if BCA_MUTATION_PATTERN and BCA_MUTATION_PATTERN.lower() in name:
                    target_dir = MUTATION_DIR / "bca"
                    matched_bank = "bca"
                    is_mutation = True
                elif MANDIRI_MUTATION_PATTERN and MANDIRI_MUTATION_PATTERN.lower() in name:
                    target_dir = MUTATION_DIR / "mandiri"
                    matched_bank = "mandiri"
                    is_mutation = True
                elif BRI_MUTATION_PATTERN and BRI_MUTATION_PATTERN.lower() in name:
                    target_dir = MUTATION_DIR / "bri"
                    matched_bank = "bri"
                    is_mutation = True
                
                # 2. If not mutation, check Input (EDC)
                elif BCA_EXCEL_PATTERN and BCA_EXCEL_PATTERN.lower() in name:
                    target_dir = BCA_EXCEL_DIR
                    matched_bank = "bca"
                elif fnmatch.fnmatch(name, MANDIRI_ZIP_PATTERN.lower()):
                    target_dir = MANDIRI_ZIP_DIR
                    matched_bank = "mandiri"
                elif "payments" in name and name.endswith(".xlsx"):
                    target_dir = ODO_EXCEL_PATH.parent
                else:
                    # Detect BRI based on MID and .zip extension, fallback to BRI_ZIP_PATTERN
                    from config import BANK_ACCOUNTS, BRI_ZIP_PATTERN
                    if name.endswith(".zip") or name.endswith(".pdf") or name.endswith(".csv"):
                        for alias, acc_info in BANK_ACCOUNTS.get("bri", {}).items():
                            mid = acc_info.get("mid", "")
                            if mid and mid.lower() in name:
                                target_dir = BRI_ZIP_DIR
                                matched_bank = "bri"
                                break
                        
                        # Fallback for legacy single account
                        if not matched_bank and BRI_ZIP_PATTERN and BRI_ZIP_PATTERN.lower() in name:
                            target_dir = BRI_ZIP_DIR
                            matched_bank = "bri"
                
                if matched_bank:
                    from config import BANK_ACCOUNTS
                    matched_alias = None
                    accounts = BANK_ACCOUNTS.get(matched_bank, {})
                    for alias, acc_info in accounts.items():
                        identifier = acc_info.get("acc", "") if is_mutation else acc_info.get("mid", "")
                        if identifier and identifier.lower() in name:
                            matched_alias = alias
                            break
                    
                    if not matched_alias:
                        if "main" in accounts:
                            matched_alias = "main"
                        elif accounts:
                            matched_alias = list(accounts.keys())[0]
                        else:
                            matched_alias = "main"
                    target_dir = target_dir / matched_alias

                if target_dir:
                    target_dir.mkdir(parents=True, exist_ok=True)
                    dest = target_dir / path.name
                    shutil.copy2(path, dest)
                    self._log_write(f"✅ Copied: {path.name} -> {target_dir.parent.name}/{target_dir.name}/\n", "ok")
                else:
                    self._log_write(f"⚠️ Ignored: {path.name} (No bank pattern matched)\n", "warn")
                    
            self._refresh_folder_status()
        except Exception as e:
            self._log_write(f"\n❌ Error during Upload: {e}\nEnsure your .env file is configured correctly!\n", "error")

    def _on_download(self):
        if DateEntry:
            date_from = self._date_from_widget.get()
            date_to = self._date_to_widget.get()
        else:
            date_from = self._date_from_var.get().strip()
            date_to = self._date_to_var.get().strip()
        
        self._log_write(f"\n── Downloading Odoo Payment (From {date_from} to {date_to}) ──\n", "head")
        self._set_status("Downloading Odoo...", "orange")
        self._running = True
        
        def run():
            try:
                if getattr(sys, "frozen", False):
                    cmd = [sys.executable, "--run-downloader", "--date-from", date_from, "--date-to", date_to]
                else:
                    cmd = [_venv_python, "odoo_downloader.py", "--date-from", date_from, "--date-to", date_to]
                    
                email = self._email_var.get().strip()
                password = self._password_var.get()
                if email and password:
                    cmd.extend(["--email", email, "--password", password])
                    
                selected_banks = [b for b, var in self._bank_vars.items() if var.get()]
                if selected_banks:
                    cmd.extend(["--banks", ",".join(selected_banks)])

                env = os.environ.copy()
                env["PYTHONIOENCODING"] = "utf-8"
                process = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=1,
                    encoding="utf-8",
                    errors="replace",
                    env=env
                )
                
                for line in process.stdout:
                    self._log_write(line)
                    
                process.wait()
                if process.returncode == 0:
                    self._log_write("\n✅ Odoo Download Finished!\n", "ok")
                else:
                    self._log_write(f"\n❌ Odoo Download failed with code {process.returncode}\n", "error")
                    
            except Exception as e:
                self._log_write(f"\n❌ Error: {e}\n", "error")
            finally:
                self._running = False
                self._refresh_folder_status()
                self._set_status("Ready", SUCCESS)
                
        threading.Thread(target=run, daemon=True).start()

    def _on_cleanse(self):
        try:
            from config import BCA_EXCEL_DIR, MANDIRI_ZIP_DIR, BRI_ZIP_DIR, ODO_EXCEL_PATH, OUTPUT_DIR
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            recap_dir = BASE_DIR / "recap" / timestamp
            
            self._log_write("\n── Cleaning Data ──\n", "head")
            
            moved_count = 0
            dirs_to_clean = [BCA_EXCEL_DIR, MANDIRI_ZIP_DIR, BRI_ZIP_DIR, OUTPUT_DIR]
            
            for d in dirs_to_clean:
                if d.exists() and d.is_dir():
                    for f in d.glob("*"):
                        if f.name == "journal_creation_log.xlsx":
                            continue
                        target_subdir = recap_dir / d.name
                        target_subdir.mkdir(parents=True, exist_ok=True)
                        shutil.move(str(f), str(target_subdir / f.name))
                        moved_count += 1
                            
            if ODO_EXCEL_PATH.exists():
                recap_dir.mkdir(parents=True, exist_ok=True)
                shutil.move(str(ODO_EXCEL_PATH), str(recap_dir / ODO_EXCEL_PATH.name))
                moved_count += 1
                
            if moved_count > 0:
                self._log_write(f"✅ {moved_count} files moved to: recap/{timestamp}/\n", "ok")
            else:
                self._log_write("ℹ️ No data to clean.\n", "dim")
                
            self._refresh_folder_status()
        except Exception as e:
            self._log_write(f"\n❌ Error during Data Cleanup: {e}\nEnsure your .env file is configured correctly!\n", "error")

    # ── Run / Scan ────────────────────────────────────────────────────────────
    def _on_scan(self):
        if self._running:
            return
        self._running = True
        selected_banks = [b.lower() for b, var in self._bank_vars.items() if var.get()]
        if not selected_banks:
            self._set_status("Select at least 1 bank!", ERROR)
            self._running = False
            return

        self._scan_btn.config(state="disabled")
        self._run_btn.config(state="disabled")
        self._open_btn.config(state="disabled")
        self._log.config(state="normal")
        self._log.delete("1.0", "end")
        self._log.config(state="disabled")
        self._set_status("Scanning data...", WARN)
        self._refresh_folder_status()
        threading.Thread(target=self._run_script, args=(selected_banks, True), daemon=True).start()

    def _on_run(self):
        if self._running:
            return
        self._running = True
        selected_banks = [b.lower() for b, var in self._bank_vars.items() if var.get()]
        if not selected_banks:
            self._set_status("Select at least 1 bank!", ERROR)
            self._running = False
            return

        self._scan_btn.config(state="disabled")
        self._run_btn.config(state="disabled", text="⏳  Processing...")
        self._open_btn.config(state="disabled")
        self._log.config(state="normal")
        self._log.delete("1.0", "end")
        self._log.config(state="disabled")
        self._set_status("Processing...", WARN)
        self._refresh_folder_status()
        
        def run_all():
            try:
                self.after(0, self._log_write, f"\n── Starting Single Browser Auto-Recon ──\n", "head")
                
                if getattr(sys, "frozen", False):
                    cmd = [sys.executable, "--run-downloader", "--mode", "auto_recon"]
                else:
                    cmd = [_venv_python, "odoo_downloader.py", "--mode", "auto_recon"]
                    
                email = self._email_var.get().strip()
                password = self._password_var.get()
                if email and password:
                    cmd.extend(["--email", email, "--password", password])
                    
                banks_for_dl = [b for b, var in self._bank_vars.items() if var.get()]
                if banks_for_dl:
                    cmd.extend(["--banks", ",".join(banks_for_dl)])
                    
                env = os.environ.copy()
                env["PYTHONIOENCODING"] = "utf-8"
                process = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=1,
                    encoding="utf-8",
                    errors="replace",
                    env=env
                )
                
                for line in process.stdout:
                    self.after(0, self._log_write, line)
                    
                process.wait()
                if process.returncode != 0:
                    self.after(0, self._log_write, f"\n❌ Auto-Recon Failed.\n", "error")
                    self.after(0, self._on_done, process.returncode, None)
                    return
                
                self.after(0, self._log_write, "\n✅ Auto-Recon Completed Successfully!\n", "ok")
                self.after(0, self._on_done, 0, None)
                
            except Exception as e:
                self.after(0, self._log_write, f"\n❌ Error during Auto-Recon: {e}\n", "error")
                self.after(0, self._on_done, 1, None)
                return
            
        threading.Thread(target=run_all, daemon=True).start()

    def _run_script(self, selected_banks, is_scan=False):
        # When frozen: re-launch the same .exe with --worker flag for clean stdout
        # When in dev:  launch main.py via the venv python
        if getattr(sys, "frozen", False):
            cmd = [sys.executable, "--worker"]
        elif _venv_python_path.exists():
            cmd = [_venv_python, str(BASE_DIR / "main.py")]
        else:
            cmd = [sys.executable, str(BASE_DIR / "main.py")]

        if "all" in selected_banks:
            cmd.append("--all")
            selected_banks.remove("all")
            
        if selected_banks:
            cmd.extend(["--bank"] + selected_banks)
        
        if is_scan:
            cmd.append("--scan")

        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        flags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if IS_WINDOWS else 0
        proc = subprocess.Popen(
            cmd,
            cwd=str(BASE_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            creationflags=flags,
            env=env
        )

        last_output_file = None
        if proc.stdout:
            for line in proc.stdout:
                ls = line.rstrip()
                if ls.startswith("[DATE_RANGE]|"):
                    try:
                        _, min_d, max_d = ls.split("|")
                        self.after(0, self._set_dates, min_d, max_d)
                    except Exception:
                        pass
                    continue
                    
                ls_lower = ls.lower()
                if "reconciliation_" in ls_lower and ".xlsx" in ls_lower:
                    for part in ls.split():
                        if "reconciliation_" in part.lower() and ".xlsx" in part.lower():
                            last_output_file = part.strip()
                self.after(0, self._log_write, ls + "\n", self._tag(ls))

        proc.wait()
        
        if proc.returncode == 0 and last_output_file and not is_scan:
            self.after(0, self._log_write, f"\n── Checking Journal Entries ──\n", "head")
            
            j_cmd = [sys.executable, "journal_checker.py", last_output_file] if getattr(sys, "frozen", False) else [_venv_python, "journal_checker.py", last_output_file]
            j_proc = subprocess.Popen(
                j_cmd,
                cwd=str(BASE_DIR),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                creationflags=flags,
                env=env
            )
            
            if j_proc.stdout:
                for line in j_proc.stdout:
                    self.after(0, self._log_write, line)
            j_proc.wait()
            
            if j_proc.returncode != 0:
                self.after(0, self._log_write, f"\n❌ Journal Entries Check failed.\n", "error")
                
        self.after(0, self._on_done, proc.returncode, last_output_file)

    @staticmethod
    def _tag(line: str) -> str:
        if "✅" in line:                          return "ok"
        if "❌" in line or "[ERROR]" in line:     return "err"
        if "⚠️" in line or "[WARN]" in line:      return "warn"
        if line.startswith("─") or line.startswith("="): return "head"
        if "[ODO]" in line or "skipped" in line.lower(): return "dim"
        return ""

    def _on_done(self, code: int, output_path: str | None):
        self._running = False
        self._scan_btn.config(state="normal")
        self._run_btn.config(state="normal", text="▶ Reconciliation")
        self._journal_btn.config(state="normal")
        self._open_btn.config(state="normal")
        
        if code == 0:
            self._set_status("Finished ✓", SUCCESS)
            self._last_output = output_path
            if output_path and Path(output_path).exists():
                _open_path(output_path)
        else:
            self._set_status("Failed — check logs below", ERROR)
            
    def _set_dates(self, min_d: str, max_d: str):
        try:
            from datetime import datetime
            d_from = datetime.strptime(min_d, "%Y-%m-%d").strftime("%d/%m/%Y")
            d_to = datetime.strptime(max_d, "%Y-%m-%d").strftime("%d/%m/%Y")
            if DateEntry:
                self._date_from_widget.set_date(datetime.strptime(d_from, "%d/%m/%Y"))
                self._date_to_widget.set_date(datetime.strptime(d_to, "%d/%m/%Y"))
            else:
                self._date_from_var.set(d_from)
                self._date_to_var.set(d_to)
            self._log_write(f"\n📅 [Auto-Detect] Dates updated: {d_from} - {d_to}\n", "ok")
        except Exception as e:
            self._log_write(f"\n⚠️ Failed to update dates: {e}\n", "warn")

    def _on_journal(self):
        if self._running:
            return
        
        import glob
        import os
        from openpyxl import load_workbook
        
        output_files = glob.glob(str(OUTPUT_DIR / "[Rr]econciliation_*.xlsx"))
        if not output_files:
            self._set_status("No reconciliation file found", ERROR)
            return
        latest_file = max(output_files, key=os.path.getctime)
        
        try:
            wb = load_workbook(latest_file, data_only=True)
            if "Daily Summary" not in wb.sheetnames:
                self._set_status("'Daily Summary' sheet not found", ERROR)
                return
            def get_col_map(sheet, row_idx=3):
                return {str(sheet.cell(row=row_idx, column=c).value).strip().lower(): c for c in range(1, sheet.max_column + 1) if sheet.cell(row=row_idx, column=c).value}

            if "Mutation Summary" in wb.sheetnames:
                ws_mut = wb["Mutation Summary"]
                col_map = get_col_map(ws_mut)
                c_date = col_map.get("payment date", 3)
                c_bank = col_map.get("bank", 4)
                c_group = col_map.get("journal", 5)
                c_cat = col_map.get("transaction category", 7)
                c_amt = col_map.get("total amount", 8)
                
                # Dictionary to hold mutation sums: (tanggal_string, group_string) -> total_amount
                from collections import defaultdict
                self._mutation_totals = defaultdict(float)
                self._mutation_raw = []
                for row in range(4, ws_mut.max_row + 1):
                    tanggal = ws_mut.cell(row=row, column=c_date).value
                    bank = ws_mut.cell(row=row, column=c_bank).value
                    group = ws_mut.cell(row=row, column=c_group).value
                    cat = ws_mut.cell(row=row, column=c_cat).value
                    amount = ws_mut.cell(row=row, column=c_amt).value
                    if tanggal and group and amount:
                        self._mutation_raw.append({"payment_date": tanggal, "bank": bank, "group": group, "category": cat, "amount": float(amount)})
                        try:
                            self._mutation_totals[(str(tanggal).strip(), str(group).strip())] += float(amount)
                        except ValueError:
                            pass
            else:
                self._mutation_totals = {}
                
            if "Admin Fee" in wb.sheetnames:
                ws_adm = wb["Admin Fee"]
                col_map = get_col_map(ws_adm)
                c_date = col_map.get("payment date", 3)
                c_bank = col_map.get("bank", 4)
                c_group = col_map.get("journal", 5)
                c_cat = col_map.get("transaction category", 7)
                c_amt = col_map.get("total amount", 8)
                
                from collections import defaultdict
                self._admin_totals = defaultdict(float)
                self._admin_raw = []
                for row in range(4, ws_adm.max_row + 1):
                    tanggal = ws_adm.cell(row=row, column=c_date).value
                    bank = ws_adm.cell(row=row, column=c_bank).value
                    group = ws_adm.cell(row=row, column=c_group).value
                    cat = ws_adm.cell(row=row, column=c_cat).value
                    amount = ws_adm.cell(row=row, column=c_amt).value
                    if tanggal and group and amount:
                        self._admin_raw.append({"payment_date": tanggal, "bank": bank, "group": group, "category": cat, "amount": float(amount)})
                        try:
                            self._admin_totals[(str(tanggal).strip(), str(group).strip())] += float(amount)
                        except ValueError:
                            pass
            else:
                self._admin_totals = {}

            ws = wb["Daily Summary"]
            col_map = get_col_map(ws)
            c_date = col_map.get("date", 2)
            c_pdate = col_map.get("payment date", 3)
            c_bank = col_map.get("bank", 4)
            c_group = col_map.get("journal", 5)
            c_tbank = col_map.get("total bank", 6)
            c_todoo = col_map.get("total odoo", 7)
            c_diff = col_map.get("difference", 8)
            c_recon = col_map.get("reconciled", 9)
            c_status = col_map.get("status", 10)
            c_jstatus = col_map.get("journal status", 11)
            
            items = []
            for row in range(4, ws.max_row + 1):
                bank = ws.cell(row=row, column=c_bank).value
                group = ws.cell(row=row, column=c_group).value
                tanggal = ws.cell(row=row, column=c_date).value
                payment_date = ws.cell(row=row, column=c_pdate).value
                total_bank = ws.cell(row=row, column=c_tbank).value
                total_odoo = ws.cell(row=row, column=c_todoo).value
                selisih = ws.cell(row=row, column=c_diff).value
                reconciled = ws.cell(row=row, column=c_recon).value
                status = ws.cell(row=row, column=c_status).value
                journal_status = ws.cell(row=row, column=c_jstatus).value
                if not bank or not status: continue
                
                try:
                    diff = abs(float(selisih)) if selisih is not None else 0
                except:
                    diff = 0
                    
                from config import JOURNAL_TOLERANCE
                
                status_str = str(status).strip()
                status_valid = ("Match" in status_str) or (diff <= JOURNAL_TOLERANCE)
                
                # Use actual Payment Date for mutation check
                from datetime import datetime, timedelta
                mutation_found = False
                mutation_matched = False
                mut_total = 0.0
                try:
                    d_mut_str = str(payment_date).strip()
                    
                    mut_amount = self._mutation_totals.get((d_mut_str, str(group).strip()), 0.0)
                    adm_amount = self._admin_totals.get((d_mut_str, str(group).strip()), 0.0)
                    mut_total = mut_amount + adm_amount
                    
                    if mut_amount > 0:
                        mutation_found = True
                        mut_diff = abs(mut_total - float(total_bank)) if total_bank else 0
                        if mut_diff <= JOURNAL_TOLERANCE:
                            mutation_matched = True
                except Exception as e:
                    pass

                items.append({
                    "row": row,
                    "bank": bank,
                    "group": group,
                    "tanggal": tanggal,
                    "payment_date": payment_date,
                    "merchant_amount": total_bank,
                    "odoo_amount": total_odoo,
                    "amount": total_bank,
                    "selisih": selisih,
                    "mutation_found": mutation_found,
                    "mutation_matched": mutation_matched,
                    "mutation_amount": mut_total,
                    "reconciled": reconciled,
                    "journal_status": journal_status,
                    "status_valid": status_valid
                })
            if not items:
                self._set_status(f"No data with difference <= {JOURNAL_TOLERANCE}", ERROR)
                return
        except Exception as e:
            self._set_status(f"Error reading excel: {e}", ERROR)
            return
            
        top = tk.Toplevel(self)
        top.title("Confirm Journal Creation")
        
        window_width = 1200
        window_height = 850
        screen_width = top.winfo_screenwidth()
        screen_height = top.winfo_screenheight()
        center_x = int(screen_width/2 - window_width / 2)
        center_y = int(screen_height/2 - window_height / 2)
        top.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
        top.configure(bg=BG)
        top.transient(self)
        top.grab_set()
        
        # State Initialization
        journal_state = []
        for item in items:
            is_reconciled = str(item.get("reconciled", "")).strip().lower() == "yes"
            status_valid = item.get("status_valid", True)
            
            disabled_edc = False
            disabled_ar = False
            
            if not is_reconciled or not status_valid:
                disabled_edc = True
                disabled_ar = True
            
            if not item.get("mutation_matched", False):
                disabled_ar = True
                
            j_status = item.get("journal_status")
            if j_status:
                j_status_str = str(j_status).strip()
                if j_status_str not in ["", "None", "Not Yet", "⏳ Not Yet"]:
                    parts = [p.strip() for p in j_status_str.split("|")]
                    for p in parts:
                        if "(Both" in p or p == "✅ Both":
                            if "✅ Posted" in p or "✅ Both" in p:
                                disabled_edc = True
                                disabled_ar = True
                        elif "(EDC" in p or p == "✅ EDC":
                            if "✅ Posted" in p or "✅ EDC" in p:
                                disabled_edc = True
                        elif "(AR" in p or p == "✅ AR":
                            if "✅ Posted" in p or "✅ AR" in p:
                                disabled_ar = True

            var_item = tk.BooleanVar(value=not (disabled_edc and disabled_ar))
            var_edc = tk.BooleanVar(value=not disabled_edc)
            var_ar = tk.BooleanVar(value=not disabled_ar)

            journal_state.append({
                "item": item,
                "var_item": var_item,
                "var_edc": var_edc,
                "var_ar": var_ar,
                "disabled_edc": disabled_edc,
                "disabled_ar": disabled_ar
            })
            
        ITEMS_PER_PAGE = 15
        current_page = [0]
        
        header_frame = tk.Frame(top, bg=PANEL, padx=40, pady=25, highlightbackground=BORDER, highlightthickness=1)
        header_frame.pack(fill="x")
        
        left_header = tk.Frame(header_frame, bg=PANEL)
        left_header.pack(side="left")
        
        tk.Label(left_header, text="Journal Creation", bg=PANEL, fg=TEXT, font=("Segoe UI", 16, "bold")).pack(anchor="w")
        tk.Label(left_header, text="Review and select transactions to post. Expand a row to preview journal entries.", bg=PANEL, fg=MUTED, font=("Segoe UI", 9)).pack(anchor="w", pady=(4, 0))
        
        def _refresh_modal():
            top.destroy()
            self.after(50, self._on_journal)
            
        tk.Button(header_frame, text="↻ Refresh", bg=WHITE, fg=TEXT, relief="solid", borderwidth=1, highlightbackground=BORDER, font=("Segoe UI", 9, "bold"), cursor="hand2", command=_refresh_modal, padx=12, pady=4).pack(side="right")
        
        # Main body container
        body_frame = tk.Frame(top, bg=BG)
        body_frame.pack(fill="both", expand=True, padx=40, pady=20)
        
        # Scrollable Data Area
        list_frame = tk.Frame(body_frame, bg=PANEL, highlightbackground=BORDER, highlightthickness=1)
        list_frame.pack(fill="both", expand=True)
        
        canvas = tk.Canvas(list_frame, bg=PANEL, highlightthickness=0)
        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=PANEL)
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_window, width=e.width))
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Bottom Action Bar
        footer_frame = tk.Frame(top, bg=PANEL, padx=40, pady=20, highlightbackground=BORDER, highlightthickness=1)
        footer_frame.pack(fill="x", side="bottom")
        
        # Pagination (Left in footer)
        pagination_frame = tk.Frame(footer_frame, bg=PANEL)
        pagination_frame.pack(side="left")
        
        sec_btn_style = {"font": ("Segoe UI", 9, "bold"), "relief": "solid", "cursor": "hand2", "padx": 12, "pady": 4, "borderwidth": 1, "highlightbackground": BORDER}
        
        btn_prev = tk.Button(pagination_frame, text="< Prev", bg=WHITE, fg=TEXT, **sec_btn_style)
        btn_prev.pack(side="left")
        
        lbl_page = tk.Label(pagination_frame, text="", bg=PANEL, fg=TEXT, font=("Segoe UI", 9, "bold"))
        lbl_page.pack(side="left", padx=15)
        
        btn_next = tk.Button(pagination_frame, text="Next >", bg=WHITE, fg=TEXT, **sec_btn_style)
        btn_next.pack(side="left")
        
        # We will add process buttons to footer later
        
        def _on_mousewheel(event):
            delta = event.delta
            if abs(delta) >= 120:
                delta = int(delta / 120)
            if delta != 0:
                canvas.yview_scroll(int(-1 * delta), "units")
                
        top.bind("<MouseWheel>", _on_mousewheel)
        canvas.bind("<MouseWheel>", _on_mousewheel)
        scrollable_frame.bind("<MouseWheel>", _on_mousewheel)
        
        def render_page(page_idx):
            active_det = {"btn": None, "frm": None}
            for widget in scrollable_frame.winfo_children():
                widget.destroy()
                
            # Render Table Header
            headers = ["", "Select", "Date", "Journal", "Merchant Amt", "Odoo Amt", "Mutation + Admin", "Difference", "EDC", "EDC Info", "AR", "AR Info"]
            for col, h in enumerate(headers):
                lbl_anchor = "w" if col in [2, 3] else "e" if col in [4, 5, 6, 7] else "center"
                lbl = tk.Label(scrollable_frame, text=h, bg=PREVIEW_BG, fg=MUTED, font=("Segoe UI", 9, "bold"), anchor=lbl_anchor, padx=10)
                lbl.grid(row=0, column=col, sticky="nsew", pady=(0, 8), ipady=8)
                
            # Fill the remaining space on the right with the header color
            dummy = tk.Label(scrollable_frame, text="", bg=PREVIEW_BG)
            dummy.grid(row=0, column=len(headers), sticky="nsew", pady=(0, 8), ipady=8)
            scrollable_frame.grid_columnconfigure(len(headers), weight=1)
                
            start_idx = page_idx * ITEMS_PER_PAGE
            end_idx = min(start_idx + ITEMS_PER_PAGE, len(journal_state))
            
            for i, state in enumerate(journal_state[start_idx:end_idx]):
                r_main = i * 2 + 1
                r_det = i * 2 + 2
                
                item = state["item"]
                var_item = state["var_item"]
                var_edc = state["var_edc"]
                var_ar = state["var_ar"]
                
                amt = float(item['amount']) if item['amount'] else 0
                sel = float(item['selisih']) if item['selisih'] else 0
                
                def _on_jurnal_toggle(vi=var_item, ve=var_edc, va=var_ar, de=state["disabled_edc"], da=state["disabled_ar"]):
                    if not de:
                        ve.set(vi.get())
                    if not da:
                        va.set(vi.get())
                
                import config
                from collections import defaultdict
                
                b_name = str(item['bank']).lower()
                b_group = str(item['group']).lower()
                
                # Find alias and properties
                alias = ""
                props = {}
                for a, p in config.BANK_ACCOUNTS.get(b_name, {}).items():
                    if p.get("group", "").lower() == b_group:
                        alias = a
                        props = p
                        break
                        
                det_frame = tk.Frame(scrollable_frame, bg=PREVIEW_BG, highlightbackground=BORDER, highlightthickness=1)
                
                # EDC Section
                edc_debit = props.get("edc_debit") or f"{str(item['bank']).upper()} EDC Debit"
                edc_credit = props.get("edc_credit") or f"{str(item['group'])} Credit"
                
                edc_frame = tk.Frame(det_frame, bg=PREVIEW_BG)
                edc_frame.pack(side="left", anchor="n", padx=20, pady=5)
                
                tk.Label(edc_frame, text="EDC Journal:", bg=PREVIEW_BG, fg=TEXT, font=("Segoe UI", 9)).grid(row=0, column=0, columnspan=3, sticky="w")
                tk.Label(edc_frame, text="Debit:", bg=PREVIEW_BG, fg=TEXT, font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w")
                tk.Label(edc_frame, text=edc_debit, bg=PREVIEW_BG, fg=TEXT, font=("Segoe UI", 9)).grid(row=1, column=1, sticky="w", padx=(10, 40))
                tk.Label(edc_frame, text=f"Rp {amt:,.0f}", bg=PREVIEW_BG, fg=TEXT, font=("Segoe UI", 9)).grid(row=1, column=2, sticky="e")
                tk.Label(edc_frame, text="Credit:", bg=PREVIEW_BG, fg=TEXT, font=("Segoe UI", 9)).grid(row=2, column=0, sticky="w")
                tk.Label(edc_frame, text=edc_credit, bg=PREVIEW_BG, fg=TEXT, font=("Segoe UI", 9)).grid(row=2, column=1, sticky="w", padx=(10, 40))
                tk.Label(edc_frame, text=f"Rp {amt:,.0f}", bg=PREVIEW_BG, fg=TEXT, font=("Segoe UI", 9)).grid(row=2, column=2, sticky="e")
                
                # AR Section Preview Logic (Only if AR is valid)
                if item.get("mutation_matched", False):
                    ar_text_lines = ["AR Journal:"]
                    
                    m_date = item['payment_date']
                    m_group = item['group']
                    m_raw = [m for m in getattr(self, "_mutation_raw", []) if m["payment_date"] == m_date and m["group"] == m_group]
                    a_raw = [a for a in getattr(self, "_admin_raw", []) if a["payment_date"] == m_date and a["group"] == m_group]
                    
                    ar_debits = []
                    for m in m_raw:
                        cat = (m["category"] or "").replace(" ", "").lower()
                        acc = props.get(f"ar_debit_{cat}") or props.get("ar_debit") or f"AR DEBIT ({cat})"
                        ar_debits.append({"account": acc, "amount": m["amount"]})
                        
                    admin_sums = defaultdict(float)
                    for a in a_raw:
                        cat = (a["category"] or "").replace(" ", "").lower()
                        acc = props.get(f"admin_debit_{cat}") or props.get("admin_debit") or f"ADMIN DEBIT ({cat})"
                        admin_sums[acc] += a["amount"]
                        
                    for acc, a_amt in admin_sums.items():
                        if a_amt > 0:
                            ar_debits.append({"account": acc, "amount": a_amt})
                    
                    t_debit = sum(d["amount"] for d in ar_debits)
                    t_credit = float(amt)
                    t_diff = t_debit - t_credit
                    
                    ar_rows = []
                    for d in ar_debits:
                        ar_rows.append(("Debit:", d['account'], d['amount']))
                        
                    ar_rows.append(("Credit:", edc_debit, t_credit))
                    
                    if round(t_diff, 2) > 0:
                        ar_rows.append(("Credit:", "82005 Bank Difference Income", abs(t_diff)))
                    elif round(t_diff, 2) < 0:
                        ar_rows.append(("Debit:", "8107 Bank Difference Loss", abs(t_diff)))
                        
                    ar_frame = tk.Frame(det_frame, bg=PREVIEW_BG)
                    ar_frame.pack(side="left", anchor="n", padx=40, pady=5)
                    
                    tk.Label(ar_frame, text="AR Journal:", bg=PREVIEW_BG, fg=TEXT, font=("Segoe UI", 9)).grid(row=0, column=0, columnspan=3, sticky="w")
                    
                    for i, (typ, acc, amt_val) in enumerate(ar_rows):
                        r = i + 1
                        tk.Label(ar_frame, text=typ, bg=PREVIEW_BG, fg=TEXT, font=("Segoe UI", 9)).grid(row=r, column=0, sticky="w")
                        tk.Label(ar_frame, text=acc, bg=PREVIEW_BG, fg=TEXT, font=("Segoe UI", 9)).grid(row=r, column=1, sticky="w", padx=(10, 40))
                        tk.Label(ar_frame, text=f"Rp {amt_val:,.0f}", bg=PREVIEW_BG, fg=TEXT, font=("Segoe UI", 9)).grid(row=r, column=2, sticky="e")
                
                def _toggle_det(btn, frm=det_frame, row_idx=r_det):
                    if frm.winfo_ismapped():
                        frm.grid_remove()
                        btn.config(text="►")
                        if active_det["frm"] == frm:
                            active_det["btn"] = None
                            active_det["frm"] = None
                    else:
                        # Close previous if open
                        prev_frm = active_det["frm"]
                        prev_btn = active_det["btn"]
                        if prev_frm and prev_frm.winfo_exists() and prev_frm.winfo_ismapped():
                            prev_frm.grid_remove()
                            if prev_btn and prev_btn.winfo_exists():
                                prev_btn.config(text="►")
                        
                        frm.grid(row=row_idx, column=1, columnspan=10, sticky="w", pady=(15, 10))
                        btn.config(text="▼")
                        active_det["btn"] = btn
                        active_det["frm"] = frm
                        
                btn_expand = tk.Label(scrollable_frame, text="►", bg=PANEL, fg=TEXT, cursor="hand2", font=("Segoe UI", 10))
                btn_expand.bind("<Button-1>", lambda e, b=btn_expand, f=det_frame: _toggle_det(b, f))
                btn_expand.grid(row=r_main, column=0, padx=5)
                
                if state["disabled_edc"] and state["disabled_ar"]:
                    cb_item = tk.Label(scrollable_frame, text="—", bg=PANEL, fg=MUTED, font=("Segoe UI", 10))
                else:
                    cb_item = tk.Checkbutton(scrollable_frame, variable=var_item, bg=PANEL, selectcolor=PANEL, command=_on_jurnal_toggle)
                cb_item.grid(row=r_main, column=1, pady=2)
                
                tk.Label(scrollable_frame, text=str(item['tanggal']), bg=PANEL, fg=TEXT, font=("Segoe UI", 9)).grid(row=r_main, column=2, sticky="w", padx=10)
                tk.Label(scrollable_frame, text=str(item['group']), bg=PANEL, fg=TEXT, font=("Segoe UI", 9)).grid(row=r_main, column=3, sticky="w", padx=10)
                amt_merch = float(item.get('merchant_amount') or 0)
                amt_odoo = float(item.get('odoo_amount') or 0)
                tk.Label(scrollable_frame, text=f"Rp {amt_merch:,.0f}", bg=PANEL, fg=TEXT, font=("Segoe UI", 9)).grid(row=r_main, column=4, sticky="e", padx=10)
                tk.Label(scrollable_frame, text=f"Rp {amt_odoo:,.0f}", bg=PANEL, fg=TEXT, font=("Segoe UI", 9)).grid(row=r_main, column=5, sticky="e", padx=10)
                
                mut_amt = float(item.get("mutation_amount", 0))
                tk.Label(scrollable_frame, text=f"Rp {mut_amt:,.0f}", bg=PANEL, fg=TEXT, font=("Segoe UI", 9)).grid(row=r_main, column=6, sticky="e", padx=10)
                
                sel_color = WARN if sel != 0 else TEXT
                tk.Label(scrollable_frame, text=f"Rp {sel:,.0f}", bg=PANEL, fg=sel_color, font=("Segoe UI", 9)).grid(row=r_main, column=7, sticky="e", padx=10)
                
                if state["disabled_edc"]:
                    cb_edc = tk.Label(scrollable_frame, text="—", bg=PANEL, fg=MUTED, font=("Segoe UI", 10))
                else:
                    cb_edc = tk.Checkbutton(scrollable_frame, variable=var_edc, bg=PANEL, selectcolor=PANEL)
                cb_edc.grid(row=r_main, column=8, padx=15)
                
                if state["disabled_ar"]:
                    cb_ar = tk.Label(scrollable_frame, text="—", bg=PANEL, fg=MUTED, font=("Segoe UI", 10))
                else:
                    cb_ar = tk.Checkbutton(scrollable_frame, variable=var_ar, bg=PANEL, selectcolor=PANEL)
                cb_ar.grid(row=r_main, column=10, padx=15)
                
                edc_info_texts = []
                ar_info_texts = []
                is_reconciled = str(item.get("reconciled", "")).strip().lower() == "yes"
                status_valid = item.get("status_valid", True)
                
                # Base EDC Validation
                if not is_reconciled:
                    edc_info_texts.append("⚠️ Unreconciled")
                elif not status_valid:
                    edc_info_texts.append("⚠️ Difference")
                
                # Base AR Validation
                if not is_reconciled:
                    ar_info_texts.append("⚠️ Unreconciled")
                elif not status_valid:
                    ar_info_texts.append("⚠️ Difference")
                elif not item.get("mutation_matched", False):
                    if not item.get("mutation_found", False):
                        ar_info_texts.append("⚠️ No Mutation")
                    else:
                        ar_info_texts.append("⚠️ Mut Difference")
                
                j_status = item.get("journal_status")
                if j_status:
                    j_status_str = str(j_status).strip()
                    if j_status_str not in ["", "None", "Not Yet", "⏳ Not Yet"]:
                        parts = [p.strip() for p in j_status_str.split("|")]
                        for p in parts:
                            if "(Both" in p or p == "✅ Both":
                                stripped = p.replace("(Both Difference)", "(Diff)").replace("(Both)", "").replace("✅ Both", "✅ Imported").strip()
                                if not edc_info_texts:
                                    edc_info_texts.append(stripped)
                                if not ar_info_texts:
                                    ar_info_texts.append(stripped)
                            elif "(EDC" in p or p == "✅ EDC":
                                stripped = p.replace("(EDC Difference)", "(Diff)").replace("(EDC)", "").replace("✅ EDC", "✅ Imported").strip()
                                if not edc_info_texts:
                                    edc_info_texts.append(stripped)
                            elif "(AR" in p or p == "✅ AR":
                                stripped = p.replace("(AR Difference)", "(Diff)").replace("(AR)", "").replace("✅ AR", "✅ Imported").strip()
                                if not ar_info_texts:
                                    ar_info_texts.append(stripped)
                    
                if edc_info_texts:
                    lbl_color = WARN if any("⚠️" in t for t in edc_info_texts) else SUCCESS
                    tk.Label(scrollable_frame, text="\n".join(edc_info_texts), bg=PANEL, fg=lbl_color, font=("Segoe UI", 8)).grid(row=r_main, column=9, sticky="w", padx=10)
                    
                cb_ar.grid(row=r_main, column=10, padx=15)
                
                if ar_info_texts:
                    lbl_color = WARN if any("⚠️" in t for t in ar_info_texts) else SUCCESS
                    tk.Label(scrollable_frame, text="\n".join(ar_info_texts), bg=PANEL, fg=lbl_color, font=("Segoe UI", 8)).grid(row=r_main, column=11, sticky="w", padx=10)
                
            total_pages = max(1, (len(journal_state) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE)
            lbl_page.config(text=f"Page {page_idx + 1} of {total_pages}")
            
            btn_prev.config(state="normal" if page_idx > 0 else "disabled")
            btn_next.config(state="normal" if page_idx < total_pages - 1 else "disabled")
            
            canvas.yview_moveto(0)
            
        def _prev_page():
            if current_page[0] > 0:
                current_page[0] -= 1
                render_page(current_page[0])
                
        def _next_page():
            total_pages = (len(journal_state) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE
            if current_page[0] < total_pages - 1:
                current_page[0] += 1
                render_page(current_page[0])
                
        btn_prev.config(command=_prev_page)
        btn_next.config(command=_next_page)
        
        render_page(0)
        
        def _get_selected_config():
            selected = []
            for v in journal_state:
                if v["var_item"].get():
                    selected.append({
                        "row": v["item"]["row"],
                        "edc": v["var_edc"].get(),
                        "ar": v["var_ar"].get()
                    })
            return selected
            
        def _export(mode="edc"):
            selected = _get_selected_config()
            if not selected:
                messagebox.showwarning("Warning", "Select at least 1 transaction")
                return
            import json
            from journal_generator import generate_journal_import
            config_path = BASE_DIR / "journal_config.json"
            config_path.write_text(json.dumps(selected))
            try:
                out_path = generate_journal_import(latest_file, config_path, mode=mode, is_preview=True)
                if out_path:
                    messagebox.showinfo("Success", f"{mode.upper()} Journal exported:\n{out_path.name}")
                    _open_path(str(out_path))
                else:
                    messagebox.showerror("Error", f"Failed to generate {mode.upper()} journal.")
            except Exception as e:
                messagebox.showerror("Error", f"Error exporting {mode.upper()}: {e}")

        def _process():
            selected = _get_selected_config()
            if not selected:
                messagebox.showwarning("Warning", "Select at least 1 transaction")
                return
            import json
            config_path = BASE_DIR / "journal_config.json"
            config_path.write_text(json.dumps(selected))
            
            from journal_generator import generate_journal_import
            recon_file = Path(latest_file)
            if not recon_file.exists():
                messagebox.showerror("Error", f"Reconciliation file not found at {recon_file.name}. Did you rename it while this window was open?")
                return
                
            try:
                out_path = generate_journal_import(recon_file, config_path, mode="both", is_preview=True)
                if not out_path:
                    messagebox.showerror("Error", "Failed to generate combined preview file.")
                    return
            except Exception as e:
                messagebox.showerror("Error", f"Error generating combined preview: {e}")
                return
                
            ar_count = sum(1 for item in selected if item.get("ar", False))
            edc_count = sum(1 for item in selected if item.get("edc", False))
            
            def _show_custom_confirm():
                dlg = tk.Toplevel(top)
                dlg.title("Confirm Upload")
                dlg.geometry("500x280")
                dlg.configure(bg=PANEL)
                dlg.transient(top)
                dlg.grab_set()
                
                dlg.update_idletasks()
                x = top.winfo_x() + (top.winfo_width() - 500) // 2
                y = top.winfo_y() + (top.winfo_height() - 280) // 2
                dlg.geometry(f"+{x}+{y}")
                
                tk.Label(dlg, text="Confirm Upload", bg=PANEL, fg=TEXT, font=("Segoe UI", 16, "bold")).pack(pady=(25, 10))
                
                lbl = tk.Label(
                    dlg, 
                    text=f"Ready to import to Odoo.\n\nAR Journals to be created: {ar_count}\nEDC Journals to be created: {edc_count}\n\nIf you need to make manual edits before importing, click 'Edit Excel'.", 
                    justify="center", bg=PANEL, fg=MUTED, font=("Segoe UI", 10)
                )
                lbl.pack(pady=(0, 20), padx=20)
                
                result = {"confirm": False}
                
                def on_upload():
                    result["confirm"] = True
                    dlg.destroy()
                    
                def on_cancel():
                    dlg.destroy()
                
                btn_frame = tk.Frame(dlg, bg=PANEL)
                btn_frame.pack(fill="x", pady=10)
                
                cancel_btn = tk.Button(btn_frame, text="Cancel", bg=WHITE, fg=TEXT, relief="solid", borderwidth=1, highlightbackground=BORDER, font=("Segoe UI", 9, "bold"), cursor="hand2", command=on_cancel, padx=20, pady=8)
                cancel_btn.pack(side="left", padx=(30, 10), expand=True)
                
                edit_btn = tk.Button(btn_frame, text="Edit Excel", bg=WHITE, fg=TEXT, relief="solid", borderwidth=1, highlightbackground=BORDER, font=("Segoe UI", 9, "bold"), cursor="hand2", command=lambda: _open_path(str(out_path)), padx=20, pady=8)
                edit_btn.pack(side="left", padx=10, expand=True)
                
                upload_btn = tk.Button(btn_frame, text="Upload to Odoo", bg=SUCCESS, fg=WHITE, relief="flat", borderwidth=0, font=("Segoe UI", 9, "bold"), cursor="hand2", command=on_upload, padx=20, pady=8)
                upload_btn.pack(side="right", padx=(10, 30), expand=True)
                
                top.wait_window(dlg)
                return result["confirm"]
                
            confirm = _show_custom_confirm()
            
            if not confirm:
                return
                
            top.destroy()
            
            self._running = True
            self._set_status("Uploading Edited Journal to Odoo...", WARN)
            self._journal_btn.config(state="disabled")
            
            def run_script():
                try:
                    cmd = [_venv_python, "odoo_journal_creator.py", "--file", str(recon_file), "--import-file", str(out_path), "--config", str(config_path)]
                    proc = subprocess.Popen(
                        cmd, cwd=str(BASE_DIR),
                        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                        text=True, bufsize=1, encoding="utf-8"
                    )
                    if proc.stdout:
                        for line in iter(proc.stdout.readline, ''):
                            self.after(0, self._log_write, line, "")
                    proc.wait()
                    if proc.returncode == 0:
                        self.after(0, self._on_done, 0, None)
                    else:
                        self.after(0, self._on_done, proc.returncode, None)
                except Exception as e:
                    self.after(0, self._log_write, f"ERROR: {str(e)}\n", "err")
                    self.after(0, self._on_done, 1, None)
            
            threading.Thread(target=run_script, daemon=True).start()
            
        footer_right = tk.Frame(footer_frame, bg=PANEL)
        footer_right.pack(side="right")
        
        def _select_all():
            for v in journal_state:
                if not (v["disabled_edc"] and v["disabled_ar"]):
                    v["var_item"].set(True)
                if not v["disabled_edc"]:
                    v["var_edc"].set(True)
                if not v["disabled_ar"]:
                    v["var_ar"].set(True)
            render_page(current_page[0])
                
        def _deselect_all():
            for v in journal_state:
                v["var_item"].set(False)
                if not v["disabled_edc"]:
                    v["var_edc"].set(False)
                if not v["disabled_ar"]:
                    v["var_ar"].set(False)
            render_page(current_page[0])
                
        mod_btn_style = {"bg": WHITE, "fg": TEXT, "font": ("Segoe UI", 9, "bold"), "relief": "solid", "borderwidth": 1, "cursor": "hand2", "padx": 12, "pady": 4, "highlightbackground": BORDER}
        primary_mod_style = {"bg": SUCCESS, "fg": WHITE, "font": ("Segoe UI", 9, "bold"), "relief": "flat", "borderwidth": 0, "cursor": "hand2", "padx": 16, "pady": 5}
        
        tools_frame = tk.Frame(pagination_frame, bg=PANEL)
        tools_frame.pack(side="left", padx=(30, 0))
        tk.Button(tools_frame, text="Select All", command=_select_all, **mod_btn_style).pack(side="left", padx=(0, 8))
        tk.Button(tools_frame, text="Deselect All", command=_deselect_all, **mod_btn_style).pack(side="left")
        
        tk.Button(footer_right, text="Cancel", command=top.destroy, **mod_btn_style).pack(side="left", padx=(0, 15))
        tk.Button(footer_right, text="Export EDC", command=lambda: _export("edc"), **mod_btn_style).pack(side="left", padx=(0, 8))
        tk.Button(footer_right, text="Export AR", command=lambda: _export("ar"), **mod_btn_style).pack(side="left", padx=(0, 15))
        
        tk.Button(footer_right, text="Submit", command=_process, **primary_mod_style).pack(side="left")

    def _open_output(self):
        path = self._last_output
        if path and Path(path).exists():
            _open_path(path)
        elif OUTPUT_DIR.exists():
            _open_path(str(OUTPUT_DIR))

    def _open_input(self):
        input_dir = BASE_DIR / "input"
        input_dir.mkdir(exist_ok=True)
        _open_path(str(input_dir))

    def _open_mutation(self):
        from config import MUTATION_DIR
        MUTATION_DIR.mkdir(exist_ok=True)
        _open_path(str(MUTATION_DIR))


if __name__ == "__main__":
    import sys
    import io

    # Force UTF-8 encoding for subprocesses so emojis don't crash Windows CP1252
    if len(sys.argv) > 1 and sys.argv[1] in ("--run-downloader", "--run-main", "--worker"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

    if len(sys.argv) > 1:
        if sys.argv[1] == "--run-downloader":
            import odoo_downloader
            sys.argv = [sys.argv[0]] + sys.argv[2:]
            odoo_downloader.run_downloader()
            sys.exit(0)
        elif sys.argv[1] == "--run-main":
            # main.py usually runs its logic on import or inside a main() check.
            # If main.py just runs on import, importing it is enough.
            # If it requires __main__, we can run it using runpy.
            import runpy
            runpy.run_module('main', run_name='__main__')
            sys.exit(0)
        elif sys.argv[1] == "--install-playwright":
            import runpy
            sys.argv = [sys.argv[0], "install", "chromium"]
            runpy.run_module('playwright', run_name='__main__')
            sys.exit(0)

    app = App()
    app.mainloop()
