import sys
import openpyxl
from pathlib import Path
from datetime import datetime
import subprocess
import argparse

from config import ODO_JOURNAL_EXCEL_PATH, ODOO_JOURNAL_EDC, ODOO_JOURNAL_AR

def extract_journal_date_range(excel_path: Path | str) -> tuple[str, str] | None:
    """Extract (date_from, date_to) ISO strings from a Reconciliation Excel file in-process."""
    path = Path(excel_path)
    if not path.exists():
        return None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Daily Summary" not in wb.sheetnames:
            wb.close()
            return None
        ws = wb["Daily Summary"]
        header_row = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
        col_map = {str(cell).strip().lower(): idx for idx, cell in enumerate(header_row) if cell}
        c_date = col_map.get("date", 1)
        c_pdate = col_map.get("payment date", 2)

        dates = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if c_date < len(row) and row[c_date] and row[c_date] != "-":
                dates.append(row[c_date])
            if c_pdate < len(row) and row[c_pdate] and row[c_pdate] != "-":
                dates.append(row[c_pdate])
        wb.close()

        parsed_dates = []
        for d in dates:
            if isinstance(d, datetime):
                parsed_dates.append(d)
                continue
            d_str = str(d).strip().split(" ")[0]
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d %b %y"):
                try:
                    parsed_dates.append(datetime.strptime(d_str, fmt))
                    break
                except Exception:
                    pass

        if parsed_dates:
            min_d = min(parsed_dates).strftime("%Y-%m-%d")
            max_d = max(parsed_dates).strftime("%Y-%m-%d")
            return (min_d, max_d)
    except Exception:
        pass
    return None


def get_distinct_dates_from_recap(excel_path: Path | str) -> list[str]:
    """Extract sorted distinct ISO date strings (e.g. ['2026-07-06', '2026-07-07']) from Daily Summary."""
    path = Path(excel_path)
    if not path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Daily Summary" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["Daily Summary"]
        header_row = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
        col_map = {str(cell).strip().lower(): idx for idx, cell in enumerate(header_row) if cell}
        c_date = col_map.get("date", 1)
        c_pdate = col_map.get("payment date", 2)

        dates = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if c_date < len(row) and row[c_date] and row[c_date] != "-":
                dates.append(row[c_date])
            if c_pdate < len(row) and row[c_pdate] and row[c_pdate] != "-":
                dates.append(row[c_pdate])
        wb.close()

        parsed_dates = set()
        for d in dates:
            if isinstance(d, datetime):
                parsed_dates.add(d.strftime("%Y-%m-%d"))
                continue
            d_str = str(d).strip().split(" ")[0]
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d %b %y"):
                try:
                    dt = datetime.strptime(d_str, fmt)
                    parsed_dates.add(dt.strftime("%Y-%m-%d"))
                    break
                except Exception:
                    pass

        return sorted(list(parsed_dates))
    except Exception:
        return []


def check_journals(excel_path: Path | str, skip_download: bool = False, get_dates: bool = False, debug: bool = False):
    path = Path(excel_path)
    if not path.exists():
        print(f"❌ File not found: {excel_path}")
        return False
        
    print(f"\n── Stage 2: Checking Journal Entries ──")
    print(f"[+] Membaca {path.name}...")
    
    wb = openpyxl.load_workbook(path)
    try:
        if "Daily Summary" not in wb.sheetnames:
            print("⚠️ Sheet 'Daily Summary' not found.")
            return False
            
        ws = wb["Daily Summary"]
        
        # Map columns
        col_map = {str(ws.cell(row=3, column=c).value).strip().lower(): c for c in range(1, ws.max_column + 1) if ws.cell(row=3, column=c).value}
        c_date = col_map.get("date", 2)
        c_pdate = col_map.get("payment date", 3)
        c_bank = col_map.get("bank", 4)
        c_journal = col_map.get("journal", 5)
        c_jstatus = col_map.get("journal information", 11)
        
        # Find earliest date and latest payment date
        dates = []
        pdates = []
        
        for row in range(4, ws.max_row + 1):
            d_val = ws.cell(row=row, column=c_date).value
            pd_val = ws.cell(row=row, column=c_pdate).value
            
            if d_val and d_val != "-":
                try:
                    # Actually, in daily summary it's '26 Jul 2026' style.
                    dates.append(d_val)
                except: pass
                
            if pd_val and pd_val != "-":
                pdates.append(pd_val)
                
        if not dates:
            print("⚠️ No dates found in Daily Summary.")
            return False
            
        # We need to parse these strings into datetime to find min/max
        parsed_dates = []
        for d in dates + pdates:
            if isinstance(d, datetime):
                parsed_dates.append(d)
                continue
                
            d_str = str(d).strip()
            if " " in d_str and ":" in d_str:
                d_str = d_str.split(" ")[0] # strip time
                
            try:
                parsed_dates.append(datetime.strptime(d_str, "%d/%m/%Y"))
            except ValueError:
                try:
                    parsed_dates.append(datetime.strptime(d_str, "%Y-%m-%d"))
                except ValueError:
                    try:
                        parsed_dates.append(datetime.strptime(d_str, "%d %b %y"))
                    except:
                        pass
                    
        if not parsed_dates:
            print("⚠️ Could not parse dates from Daily Summary.")
            return False
            
        min_date = min(parsed_dates)
        max_date = max(parsed_dates)
        
        date_from_str = min_date.strftime("%Y-%m-%d")
        date_to_str = max_date.strftime("%Y-%m-%d")
        
        if get_dates:
            print(f"[DATE_RANGE]|{date_from_str}|{date_to_str}")
            return True
            
        print(f"[+] Date Range for Journal Entries: {date_from_str} to {date_to_str}")
        
        if not skip_download:
            print("[+] Make sure Journal Entries are downloaded using odoo_downloader.py")
            
        # Load downloaded journals
        if not ODO_JOURNAL_EXCEL_PATH.exists():
            print(f"\n⚠️ File '{ODO_JOURNAL_EXCEL_PATH.name}' not found. Skipping Journal Entries check.\n")
            return True

        existing_journals = [] # list of dicts: {"journal": str, "date": str, "reference": str}
        try:
            j_wb = openpyxl.load_workbook(ODO_JOURNAL_EXCEL_PATH, read_only=True, data_only=True)
            try:
                j_ws = j_wb.active
                
                # Map columns for Journal Entries
                j_headers = {}
                for col_idx, cell in enumerate(next(j_ws.iter_rows(min_row=1, max_row=1, values_only=True))):
                    if cell:
                        j_headers[str(cell).strip().lower()] = col_idx
                        
                c_j_number = j_headers.get("number") or j_headers.get("name") or j_headers.get("no")
                c_j_journal = j_headers.get("journal")
                c_j_date = j_headers.get("date")
                c_j_ref = j_headers.get("reference") or j_headers.get("label")
                c_j_status = j_headers.get("status")
                c_j_total = j_headers.get("total signed") or j_headers.get("total")
                
                if c_j_journal is not None and c_j_date is not None:
                    for j_row in j_ws.iter_rows(min_row=2, values_only=True):
                        j_val = str(j_row[c_j_journal] or "").strip()
                        
                        d_val_raw = j_row[c_j_date]
                        if isinstance(d_val_raw, datetime):
                            d_val = d_val_raw.strftime("%Y-%m-%d")
                        else:
                            d_val = str(d_val_raw or "").strip()
                            if " " in d_val and ":" in d_val:
                                d_val = d_val.split(" ")[0]
                                
                        ref_val = str(j_row[c_j_ref] or "").strip() if c_j_ref is not None else ""
                        status_val = str(j_row[c_j_status] or "").strip() if c_j_status is not None else ""
                        num_val = str(j_row[c_j_number] or "").strip() if c_j_number is not None else ""
                        
                        try:
                            total_val = float(j_row[c_j_total]) if c_j_total is not None and j_row[c_j_total] is not None else 0.0
                        except:
                            total_val = 0.0
                        
                        if j_val and d_val:
                            existing_journals.append({
                                "number": num_val,
                                "journal": j_val,
                                "date": d_val,
                                "reference": ref_val.lower(),
                                "status": status_val,
                                "total": total_val
                            })
            finally:
                j_wb.close()
        except Exception as e:
            print(f"⚠️ Failed to read {ODO_JOURNAL_EXCEL_PATH.name}: {e}")
                
        # Update Daily Summary
        from journal_generator import format_date_indo
        from config import BANK_ACCOUNTS
        
        c_edc_num = col_map.get("edc number", 12)
        c_ar_num = col_map.get("ar number", 13)

        updated = 0
        for row in range(4, ws.max_row + 1):
            bank = str(ws.cell(row=row, column=c_bank).value or "").strip()
            d_val = ws.cell(row=row, column=c_date).value
            ds_journal = str(ws.cell(row=row, column=c_journal).value or "").strip()
            
            if not bank or bank == "None":
                continue
                
            bank_lower = bank.lower()
            
            # Find the alias by looking up the ds_journal (account name) in config
            alias = "main"
            for a, info in BANK_ACCOUNTS.get(bank_lower, {}).items():
                if info.get("group") == ds_journal or a == ds_journal.lower():
                    alias = a
                    break
                    
            # Construct the reference bank alias exactly as journal_generator.py does
            ref_bank_alias = BANK_ACCOUNTS.get(bank_lower, {}).get(alias, {}).get("alias", bank_lower.title() + " " + alias.title())
            
            pd_val = ws.cell(row=row, column=c_pdate).value
            
            if d_val and d_val != "-":
                try:
                    # Helper to convert to yyyy-mm-dd
                    def to_iso(val):
                        if isinstance(val, datetime): return val.strftime("%Y-%m-%d")
                        try:
                            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d %b %y"):
                                try: return datetime.strptime(str(val).strip(), fmt).strftime("%Y-%m-%d")
                                except: continue
                        except: pass
                        return str(val)

                    d_val_str = to_iso(d_val)
                    payment_date_str = to_iso(pd_val)
                    target_ref_lower = ref_bank_alias.lower()
                    
                    ar_status = None
                    edc_status = None
                    ar_diff = False
                    edc_diff = False
                    ar_number = ""
                    edc_number = ""
                    
                    c_odoo_amt = col_map.get("total odoo", 7)
                    try:
                        odoo_amt = float(ws.cell(row=row, column=c_odoo_amt).value) if ws.cell(row=row, column=c_odoo_amt).value else 0.0
                    except:
                        odoo_amt = 0.0

                    for j in existing_journals:
                        j_journal = j["journal"].lower()
                        j_ref = j["reference"].lower()
                        j_date = j["date"]
                        j_status = j.get("status", "").lower()
                        j_total = abs(j.get("total", 0.0))
                        
                        # Match AR
                        if j_journal == ODOO_JOURNAL_AR.lower() and j_date == payment_date_str and target_ref_lower in j_ref:
                            if j_status == "posted":
                                ar_status = "Posted"
                            else:
                                ar_status = "Draft"
                            ar_number = j.get("number", "")
                            ar_diff = abs(j_total - abs(odoo_amt)) > 1.0
                            
                        # Match EDC
                        if j_journal == ODOO_JOURNAL_EDC.lower() and j_date == d_val_str and target_ref_lower in j_ref:
                            if j_status == "posted":
                                edc_status = "Posted"
                            else:
                                edc_status = "Draft"
                            edc_number = j.get("number", "")
                            edc_diff = abs(j_total - abs(odoo_amt)) > 1.0

                    # Determine final status
                    if ar_status and edc_status:
                        if ar_status == edc_status:
                            if ar_diff or edc_diff:
                                final_status = f"{ar_status} (Both Difference)"
                            else:
                                final_status = f"{ar_status} (Both)"
                        else:
                            edc_part = f"{edc_status} (EDC Difference)" if edc_diff else f"{edc_status} (EDC)"
                            ar_part = f"{ar_status} (AR Difference)" if ar_diff else f"{ar_status} (AR)"
                            final_status = f"{edc_part} | {ar_part}"
                    elif edc_status:
                        if edc_diff:
                            final_status = f"{edc_status} (EDC Difference)"
                        else:
                            final_status = f"{edc_status} (EDC)"
                    elif ar_status:
                        if ar_diff:
                            final_status = f"{ar_status} (AR Difference)"
                        else:
                            final_status = f"{ar_status} (AR)"
                    else:
                        final_status = "Not Yet"
                        
                    if ds_journal != "Unknown":
                        ws.cell(row=row, column=c_jstatus).value = final_status
                        from openpyxl.styles import Font
                        if c_edc_num:
                            ce = ws.cell(row=row, column=c_edc_num)
                            ce.value = edc_number if edc_number else "-"
                            ce.font = Font(size=10, bold=True)
                        if c_ar_num:
                            ca = ws.cell(row=row, column=c_ar_num)
                            ca.value = ar_number if ar_number else "-"
                            ca.font = Font(size=10, bold=True)

                        if final_status != "Not Yet":
                            updated += 1
                            if debug:
                                print(f"      [DEBUG] Found {final_status} ({edc_number} / {ar_number}) for {target_ref_lower}")
                except Exception as e:
                    print(f"⚠️ Failed to parse row {row}: {e}")

                
        if updated > 0:
            from openpyxl.utils import get_column_letter
            if c_edc_num:
                ws.column_dimensions[get_column_letter(c_edc_num)].width = 18
            if c_ar_num:
                ws.column_dimensions[get_column_letter(c_ar_num)].width = 18
            if c_jstatus:
                ws.column_dimensions[get_column_letter(c_jstatus)].width = 22
            from odoo_journal_creator import safe_save_workbook
            safe_save_workbook(wb, path)
            print(f"\n✅ Successfully updated {updated} Journal Information rows in {path.name}")
        else:
            print("\n⚠️ Tidak ada baris yang diperbarui.")

            
        return True
    finally:
        wb.close()


check_journal_entries = check_journals


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Check Journal Entries against Daily Summary")
    parser.add_argument("excel_path", type=str, help="Path to Reconciliation Excel file")
    parser.add_argument("--skip-download", action="store_true", help="Skip downloading Journal Entries from Odoo")
    parser.add_argument("--debug", action="store_true", help="Print debug information about the matching process")
    parser.add_argument("--get-dates", action="store_true", help="Only extract and print the dates, then exit")
    
    args = parser.parse_args()
    check_journals(args.excel_path, args.skip_download, args.debug, args.get_dates)
