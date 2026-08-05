"""
journal_generator.py — Generate Odoo-compatible Excel import files based on reconciliation results.
"""

import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook

from config import (
    OUTPUT_DIR,
    BANK_ACCOUNTS,
    ODOO_COMPANY_NAME,
    ODOO_JOURNAL_EDC,
)

def get_alias_by_group(bank_name: str, group_name: str) -> str:
    """Find the bank alias based on the group name."""
    bank_dict = BANK_ACCOUNTS.get(bank_name.lower(), {})
    for alias, props in bank_dict.items():
        if props.get("group", "").lower() == group_name.lower():
            return alias
    return ""

def format_date_indo(date_str: str) -> str:
    """Format DD/MM/YYYY into 'D Month YY' for Indonesian (e.g. 8 Juli 26)."""
    if not date_str:
        return ""
    try:
        try:
            dt = datetime.strptime(str(date_str), "%d/%m/%Y")
        except ValueError:
            dt = datetime.strptime(str(date_str), "%Y-%m-%d")
            
        months = ["January", "February", "March", "April", "May", "June", 
                  "July", "August", "September", "October", "November", "December"]
        return f"{dt.day} {months[dt.month - 1]} {dt.strftime('%y')}"
    except Exception:
        return str(date_str)

def generate_journal_import(reconciliation_file: Path, config_path: Path | None = None, mode: str = "both", is_preview: bool = False) -> Path | None:
    """
    Reads the 'Daily Summary' from the reconciliation file and generates 
    an Odoo-compatible Excel file for Settlement EDC import based on config.
    """
    wb_source = load_workbook(reconciliation_file, data_only=True)
    if "Daily Summary" not in wb_source.sheetnames:
        print("❌ Sheet 'Daily Summary' not found.")
        return None
    ws_source = wb_source["Daily Summary"]
    
    # Extract Mutations
    mutations_data = []
    if "Mutation Summary" in wb_source.sheetnames:
        ws_mut = wb_source["Mutation Summary"]
        for r in range(4, ws_mut.max_row + 1):
            if not ws_mut.cell(row=r, column=4).value: continue
            mutations_data.append({
                "payment_date": ws_mut.cell(row=r, column=3).value,
                "bank": ws_mut.cell(row=r, column=4).value,
                "group": ws_mut.cell(row=r, column=5).value,
                "category": ws_mut.cell(row=r, column=7).value,
                "amount": ws_mut.cell(row=r, column=8).value
            })

    # Extract Admin Fees
    admin_fees_data = []
    if "Admin Fee" in wb_source.sheetnames:
        ws_adm = wb_source["Admin Fee"]
        for r in range(4, ws_adm.max_row + 1):
            if not ws_adm.cell(row=r, column=4).value: continue
            admin_fees_data.append({
                "payment_date": ws_adm.cell(row=r, column=3).value,
                "bank": ws_adm.cell(row=r, column=4).value,
                "group": ws_adm.cell(row=r, column=5).value,
                "category": ws_adm.cell(row=r, column=7).value,
                "amount": ws_adm.cell(row=r, column=8).value
            })

    
    # Extract "Sesuai" items
    items = []
    for row in range(4, ws_source.max_row + 1):
        odoo_date = ws_source.cell(row=row, column=2).value    # Odoo Date
        payment_date = ws_source.cell(row=row, column=3).value # Payment Date
        bank = ws_source.cell(row=row, column=4).value         # Bank
        group = ws_source.cell(row=row, column=5).value        # Journal
        total_odo = ws_source.cell(row=row, column=7).value    # Total Odoo
        selisih = ws_source.cell(row=row, column=8).value      # Difference
        status = ws_source.cell(row=row, column=9).value       # Status
        
        if not bank or not status:
            continue
            
        try:
            diff = abs(float(selisih)) if selisih is not None else 0
        except:
            diff = 0
            
        from config import JOURNAL_TOLERANCE
            
        if diff <= JOURNAL_TOLERANCE:
            items.append({
                "row": row,
                "bank": bank,
                "group": group,
                "odoo_date": odoo_date,
                "payment_date": payment_date,
                "amount": total_odo
            })

    # Filter items based on config
    if config_path and config_path.exists():
        import json
        try:
            config_data = json.loads(config_path.read_text())
            # config_data is a list of dicts: {"row": int, "edc": bool, "ar": bool}
            config_map = {item["row"]: item for item in config_data}
            filtered_items = []
            for item in items:
                if item["row"] in config_map:
                    item["create_edc"] = config_map[item["row"]].get("edc", True)
                    item["create_ar"] = config_map[item["row"]].get("ar", True)
                    filtered_items.append(item)
            items = filtered_items
        except Exception as e:
            print(f"⚠️ Failed to read config: {e}. All 'Match' items will be processed with EDC=True, AR=True.")
            for item in items:
                item["create_edc"] = True
                item["create_ar"] = True
    else:
        for item in items:
            item["create_edc"] = True
            item["create_ar"] = True

    if not items:
        print(f"ℹ️ No data with difference <= {JOURNAL_TOLERANCE} in Daily Summary.")
        return None

    headers = [
        "Company", "Date", "Journal", "Number", "Partner", "Reference", 
        "Journal Items/Account", "Journal Items/Credit", "Journal Items/Debit"
    ]
    
    from config import BASE_DIR
    template_path = BASE_DIR / "template" / "Journal Entry (account.move).xlsx"
    
    styles_row = []
    import copy
    
    if template_path.exists():
        wb_out = load_workbook(template_path)
        ws_out = wb_out.active
        # capture styles from row 2
        for col in range(1, ws_out.max_column + 1):
            cell = ws_out.cell(row=2, column=col)
            styles_row.append({
                "font": copy.copy(cell.font),
                "border": copy.copy(cell.border),
                "fill": copy.copy(cell.fill),
                "number_format": copy.copy(cell.number_format),
                "alignment": copy.copy(cell.alignment),
                "protection": copy.copy(cell.protection),
            })
        ws_out.delete_rows(2, ws_out.max_row)
    else:
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "account.move"
        ws_out.append(headers)
        
    def _apply_style(ws, row_idx):
        if not styles_row: return
        from openpyxl.styles import Protection
        for col_idx, styles in enumerate(styles_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = copy.copy(styles["font"])
            cell.border = copy.copy(styles["border"])
            cell.fill = copy.copy(styles["fill"])
            cell.number_format = copy.copy(styles["number_format"])
            cell.alignment = copy.copy(styles["alignment"])
            
            # Unlocked columns: Date(2), Reference(6), Account(7), Credit(8), Debit(9)
            if col_idx in (2, 6, 7, 8, 9):
                cell.protection = Protection(locked=False)
            else:
                cell.protection = Protection(locked=True)
    
    # Process each item
    for item in items:
        bank_name = item["bank"]
        group_name = item["group"]
        odoo_date = item["odoo_date"] or ""
        payment_date = item["payment_date"] or ""
        amount = item["amount"]
        
        alias = get_alias_by_group(bank_name, group_name)
        if not alias:
            print(f"⚠️ Peringatan: Alias tidak ditemukan untuk Bank={bank_name}, Account Name={group_name}. Lewati.")
            continue
            
        # Get Accounts from config
        props = BANK_ACCOUNTS[bank_name.lower()][alias]
        debit_acc = props.get("edc_debit")
        credit_acc = props.get("edc_credit")
        
        if not debit_acc or not credit_acc:
            print(f"⚠️ Peringatan: EDC Debit/Credit belum diset di .env untuk {bank_name.upper()} {alias}. Lewati.")
            continue
            
        # Format Reference (e.g. Settlement EDC BCA Main 8 Juli 26)
        formatted_date = format_date_indo(odoo_date)
        
        ref_bank_alias = f"{bank_name.upper()} {alias.capitalize()}"
        if bank_name.upper() == "BCA" and alias == "main":
            ref_bank_alias = "BCA Main"
        elif bank_name.upper() == "MANDIRI" and alias == "main":
            ref_bank_alias = "Mandiri"
        elif bank_name.upper() == "BRI" and alias == "lbf":
            ref_bank_alias = "BRI LBF"
        elif bank_name.upper() == "BRI" and alias == "frans":
            ref_bank_alias = "BRI Frans (Sanur)"
        elif bank_name.upper() == "BRI" and alias == "nara":
            ref_bank_alias = "BRI Nara"

        reference = f"Settlement EDC {ref_bank_alias} {formatted_date}"
        
        # EDC Date uses Odoo Date, AR Date uses Payment Date
        # Data is already in DD/MM/YYYY from excel_writer.py
        date_obj = str(odoo_date)
        ar_date_obj = str(payment_date)

        # EDC Journal
        if mode in ("both", "edc") and item.get("create_edc", True):
            # Row 1 (Debit)
            row1 = [
                ODOO_COMPANY_NAME,       # Company
                date_obj,                # Date
                ODOO_JOURNAL_EDC,        # Journal
                "/",                     # Number
                None,                    # Partner
                reference,               # Reference
                debit_acc,               # Journal Items/Account
                None,                    # Journal Items/Credit
                amount                   # Journal Items/Debit
            ]
            
            # Row 2 (Credit)
            row2 = [
                None,                    # Company
                None,                    # Date
                None,                    # Journal
                None,                    # Number
                None,                    # Partner
                None,                    # Reference
                credit_acc,              # Journal Items/Account
                amount,                  # Journal Items/Credit
                None                     # Journal Items/Debit
            ]
            
            ws_out.append(row1)
            _apply_style(ws_out, ws_out.max_row)
            
            ws_out.append(row2)
            _apply_style(ws_out, ws_out.max_row)

        # AR Journal (If needed)
        if mode in ("both", "ar") and item.get("create_ar", False):
            ar_reference = reference.replace("Settlement EDC", "Settlement AR")
            
            # Find matching Mutations
            matched_mutations = [m for m in mutations_data if m["payment_date"] == item["payment_date"] and m["group"] == item["group"]]
            # Find matching Admin Fees
            matched_admin = [a for a in admin_fees_data if a["payment_date"] == item["payment_date"] and a["group"] == item["group"]]
            
            from collections import defaultdict
            ar_debits = []
            
            # Process Mutations (Keep separate)
            for m in matched_mutations:
                cat = (m["category"] or "").replace(" ", "").lower()
                acc = props.get(f"ar_debit_{cat}") or props.get("ar_debit")
                if not acc: acc = f"MISSING ACCOUNT FOR AR DEBIT ({cat})"
                ar_debits.append({"account": acc, "amount": float(m["amount"] or 0)})
                
            # Process Admin Fees (Sum by account)
            admin_sums = defaultdict(float)
            for a in matched_admin:
                cat = (a["category"] or "").replace(" ", "").lower()
                acc = props.get(f"admin_debit_{cat}") or props.get("admin_debit")
                if not acc: acc = f"MISSING ACCOUNT FOR ADMIN DEBIT ({cat})"
                admin_sums[acc] += float(a["amount"] or 0)
                
            for acc, amt in admin_sums.items():
                if amt > 0:
                    ar_debits.append({"account": acc, "amount": amt})
                    
            total_debits = sum(d["amount"] for d in ar_debits)
            total_credits = float(amount)
            diff = total_debits - total_credits
            
            from config import ODOO_JOURNAL_AR
            
            first_row = True
            
            # 1. Debits
            for d in ar_debits:
                ws_out.append([
                    ODOO_COMPANY_NAME if first_row else None,
                    ar_date_obj if first_row else None,
                    ODOO_JOURNAL_AR if first_row else None,
                    "/" if first_row else None,
                    None,
                    ar_reference if first_row else None,
                    d["account"],
                    None,
                    d["amount"]
                ])
                first_row = False
                
            # 2. Credit (EDC Debit Account)
            ws_out.append([
                None, None, None, None, None, None,
                debit_acc,
                total_credits,
                None
            ])
            
            # 3. Balancing Difference
            if round(diff, 2) > 0:
                ws_out.append([
                    None, None, None, None, None, None,
                    "82005 Bank Difference Income",
                    abs(diff),
                    None
                ])
            elif round(diff, 2) < 0:
                ws_out.append([
                    None, None, None, None, None, None,
                    "8107 Bank Difference Loss",
                    None,
                    abs(diff)
                ])

    if ws_out.max_row <= 1:
        print("❌ Tidak ada jurnal yang berhasil dibuat. Periksa pengaturan .env")
        return None
    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col in ws_out.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_out.column_dimensions[col_letter].width = max_length + 2

    # Save the file to output/journal directory
    journal_dir = OUTPUT_DIR / "journal"
    journal_dir.mkdir(parents=True, exist_ok=True)
    
    ws_out.protection.sheet = True
    ws_out.freeze_panes = "A2"
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    mode_str = mode.upper() if mode != "both" else "EDC_AR"
    
    if is_preview:
        out_filename = f"Preview_{mode_str}_Journal.xlsx"
    else:
        out_filename = f"Journal_Import_{mode_str}_{timestamp}.xlsx"
        
    out_path = journal_dir / out_filename
    wb_out.save(out_path)
    print(f"✅ File Import Jurnal berhasil dibuat: {out_path.name}")
    return out_path

if __name__ == "__main__":
    # Test script directly
    latest = None
    files = list(OUTPUT_DIR.glob("[Rr]econciliation_*.xlsx"))
    if files:
        import os
        latest = max(files, key=os.path.getmtime)
        generate_journal_import(latest)
    else:
        print("No reconciliation file found.")
