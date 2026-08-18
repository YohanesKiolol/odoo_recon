"""
odoo_journal_creator.py — Automate Odoo Journal Entry creation directly via XML-RPC.
Directly creates balanced Draft Journal Entries (account.move) in Odoo with real-time logs.
Zero browser dependencies. 100% fast, reliable, and always in Draft state.
"""

import os
import sys
import argparse
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from config import (
    OUTPUT_DIR,
    BASE_DIR,
    ODOO_URL,
    ODOO_DB,
    ODOO_JOURNAL_EDC,
    ODOO_JOURNAL_AR,
)
from journal_generator import generate_journal_import
import odoo_inspector


def get_latest_excel_file() -> Path | None:
    """Find the most recently created reconciliation Excel file in output/."""
    excel_files = list(OUTPUT_DIR.glob("reconciliation_*.xlsx"))
    if not excel_files:
        excel_files = list(OUTPUT_DIR.glob("Reconciliation_*.xlsx"))
    if not excel_files:
        print(f"❌ No reconciliation Excel file found in '{OUTPUT_DIR}'.")
        return None
    return max(excel_files, key=os.path.getmtime)


def parse_journal_blocks_from_excel(import_excel_path: Path) -> list[dict]:
    """
    Parse grouped journal entries from the generated import Excel file.
    Returns list of dicts with: date, journal, ref, lines: [{account, debit, credit}].
    """
    wb = load_workbook(import_excel_path, data_only=True)
    ws = wb.active

    entries = []
    current_entry = None

    # Header is row 1
    # Columns:
    # 1: Company, 2: Date, 3: Journal, 4: Number, 5: Partner, 6: Reference,
    # 7: Journal Items/Account, 8: Journal Items/Credit, 9: Journal Items/Debit

    for r in range(2, ws.max_row + 1):
        c_date = ws.cell(row=r, column=2).value
        c_journal = ws.cell(row=r, column=3).value
        c_ref = ws.cell(row=r, column=6).value
        c_account = ws.cell(row=r, column=7).value
        c_credit = ws.cell(row=r, column=8).value
        c_debit = ws.cell(row=r, column=9).value

        # If Date & Ref are populated, start a new journal entry block
        if c_date and c_ref:
            if current_entry and current_entry["lines"]:
                entries.append(current_entry)

            # Format Date to YYYY-MM-DD
            d_str = str(c_date).strip()
            iso_date = ""
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d %B %Y", "%d %b %Y"):
                try:
                    iso_date = datetime.strptime(d_str, fmt).strftime("%Y-%m-%d")
                    break
                except ValueError:
                    continue
            if not iso_date:
                iso_date = d_str[:10]

            current_entry = {
                "date": iso_date,
                "journal_name": str(c_journal or ODOO_JOURNAL_EDC).strip(),
                "ref": str(c_ref).strip(),
                "lines": []
            }

        if current_entry and c_account:
            credit_val = float(c_credit) if c_credit not in (None, "", "-") else 0.0
            debit_val = float(c_debit) if c_debit not in (None, "", "-") else 0.0

            if credit_val > 0 or debit_val > 0:
                current_entry["lines"].append({
                    "account": str(c_account).strip(),
                    "debit": debit_val,
                    "credit": credit_val
                })

    if current_entry and current_entry["lines"]:
        entries.append(current_entry)

    try:
        wb.close()
    except Exception:
        pass

    return entries


def create_draft_journals_via_xmlrpc(import_excel_path: Path, config_path: Path | None = None) -> bool:
    """
    Directly create Draft Journal Entries (account.move) in Odoo via XML-RPC.
    """
    print(f"\n── Starting Direct Odoo Journal Creator (XML-RPC) ──")
    print(f"[+] Authenticating with Odoo: {odoo_inspector._get_base_url()} (DB: {odoo_inspector.ODOO_DB})...")

    uid, err = odoo_inspector.authenticate()
    if not uid:
        print(f"❌ Odoo Authentication Failed: {err}")
        return False

    print(f"✅ Authenticated successfully as UID {uid}!")

    # 1. Cache journals and accounts for fast ID resolution
    print("[+] Loading Odoo Chart of Accounts & Journals metadata...")
    try:
        all_journals = odoo_inspector._execute_kw(
            "account.journal", "search_read", [[]], {"fields": ["id", "name", "code", "type"]}
        )
        all_accounts = odoo_inspector._execute_kw(
            "account.account", "search_read", [[]], {"fields": ["id", "name", "code"]}
        )
    except Exception as e:
        print(f"❌ Failed to load metadata from Odoo: {e}")
        return False

    journal_by_name = {j["name"].strip().lower(): j["id"] for j in all_journals}
    journal_by_code = {j["code"].strip().lower(): j["id"] for j in all_journals}

    account_by_code = {str(a["code"]).strip(): a["id"] for a in all_accounts if a.get("code")}
    account_by_name = {a["name"].strip().lower(): a["id"] for a in all_accounts if a.get("name")}

    # 2. Parse entries from import file
    entries = parse_journal_blocks_from_excel(import_excel_path)
    if not entries:
        print("❌ No valid journal entry blocks found in import file.")
        return False

    print(f"[+] Found {len(entries)} Journal Entry block(s) ready to create.\n")

    created_count = 0
    skipped_count = 0
    error_count = 0

    for idx, entry in enumerate(entries, start=1):
        ref = entry["ref"]
        date_val = entry["date"]
        j_name = entry["journal_name"]
        lines = entry["lines"]

        total_debit = sum(l["debit"] for l in lines)
        total_credit = sum(l["credit"] for l in lines)

        print(f"[{idx}/{len(entries)}] Processing: '{ref}'")
        print(f"    • Date: {date_val} | Journal: {j_name}")
        print(f"    • Debit: Rp {total_debit:,.2f} | Credit: Rp {total_credit:,.2f}")

        # Check balance
        if round(total_debit, 2) != round(total_credit, 2):
            print(f"    ⚠️ SKIPPED: Unbalanced debit ({total_debit}) vs credit ({total_credit})")
            error_count += 1
            continue

        # Resolve journal_id
        journal_id = journal_by_name.get(j_name.lower()) or journal_by_code.get(j_name.lower())
        if not journal_id:
            # Fallback to general/miscellaneous journal
            journal_id = journal_by_name.get("miscellaneous operations") or journal_by_code.get("misc")
            if not journal_id and all_journals:
                journal_id = all_journals[0]["id"]

        # Check if already exists in Odoo
        try:
            existing = odoo_inspector._execute_kw(
                "account.move", "search_read",
                [[["ref", "=", ref], ["state", "!=", "cancel"]]],
                {"fields": ["id", "name", "state"], "limit": 1}
            )
            if existing:
                ex = existing[0]
                print(f"    ℹ️ Already exists in Odoo: {ex.get('name')} (ID: {ex.get('id')}, State: {ex.get('state')}) — Skipping.")
                skipped_count += 1
                continue
        except Exception as e:
            print(f"    [WARN] Duplicate check warning: {e}")

        # Build ORM line_ids
        move_lines = []
        unresolved_acc = None

        for l in lines:
            acc_str = l["account"]
            # Try code exact match first
            acc_id = account_by_code.get(acc_str)
            if not acc_id:
                # Try name exact match
                acc_id = account_by_name.get(acc_str.lower())
            if not acc_id:
                # Try finding code inside string (e.g. '0001 - BCA EDC')
                parts = acc_str.split(" ", 1)
                if parts and account_by_code.get(parts[0]):
                    acc_id = account_by_code[parts[0]]

            if not acc_id:
                unresolved_acc = acc_str
                break

            move_lines.append((0, 0, {
                "name": ref,
                "account_id": acc_id,
                "debit": l["debit"],
                "credit": l["credit"],
            }))

        if unresolved_acc:
            print(f"    ❌ FAILED: Account '{unresolved_acc}' not found in Odoo Chart of Accounts.")
            error_count += 1
            continue

        # Create Draft account.move
        move_vals = {
            "move_type": "entry",
            "date": date_val,
            "ref": ref,
            "journal_id": journal_id,
            "state": "draft",          # ALWAYS DRAFT
            "line_ids": move_lines
        }

        try:
            move_id = odoo_inspector._execute_kw("account.move", "create", [move_vals])
            created_count += 1
            print(f"    ✅ Successfully Created Draft Entry! (Odoo Move ID: {move_id})\n")
        except Exception as ex:
            print(f"    ❌ Failed to create journal entry: {ex}\n")
            error_count += 1

    print("── Journal Creation Summary ──")
    print(f"  • Created (Draft): {created_count}")
    print(f"  • Already Existed: {skipped_count}")
    print(f"  • Failed / Skipped: {error_count}")
    print("──────────────────────────────\n")

    if created_count > 0 or (skipped_count > 0 and error_count == 0):
        print("✅ Direct Odoo Journal Creation completed successfully!\n")
        return True
    return False


def main():
    import io as _io
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    elif hasattr(sys.stdout, "buffer"):
        sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

    parser = argparse.ArgumentParser(description="Direct Odoo Journal Creation via XML-RPC")
    parser.add_argument("--file", type=str, help="Path to reconciliation Excel file")
    parser.add_argument("--config", type=str, help="Path to journal config JSON")
    parser.add_argument("--import-file", type=str, help="Path to generated import Excel file")
    parser.add_argument("--email", type=str, default="", help="Odoo Email")
    parser.add_argument("--password", type=str, default="", help="Odoo Password")
    parser.add_argument("--headless", action="store_true", help="Legacy flag (ignored, runs direct API)")
    args = parser.parse_args()

    # 1. Get Recon File
    if args.file:
        file_path = Path(args.file)
        if not file_path.exists():
            print(f"❌ File not found: {file_path}")
            sys.exit(1)
    else:
        file_path = get_latest_excel_file()
        if not file_path:
            sys.exit(1)

    print(f"📁 Source Reconciliation Data: {file_path.name}")
    config_path = Path(args.config) if args.config else None

    # 2. Get or Generate Import File
    if args.import_file:
        import_file = Path(args.import_file)
        if not import_file.exists():
            print(f"❌ Import file not found: {import_file}")
            sys.exit(1)
    else:
        import_file = generate_journal_import(file_path, config_path)
        if not import_file:
            sys.exit(1)

    # 3. Create Draft Journals directly via XML-RPC
    success = create_draft_journals_via_xmlrpc(import_file, config_path)
    if not success:
        sys.exit(1)


def safe_save_workbook(wb, file_path: Path) -> bool:
    """Save openpyxl workbook safely without locking issues."""
    try:
        wb.save(str(file_path))
        return True
    except Exception as e:
        try:
            tmp_path = file_path.with_suffix(".tmp.xlsx")
            wb.save(str(tmp_path))
            if tmp_path.exists():
                os.replace(str(tmp_path), str(file_path))
                return True
        except Exception:
            pass
        print(f"Error saving workbook: {e}")
        return False


if __name__ == "__main__":
    main()
