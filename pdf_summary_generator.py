from __future__ import annotations

import os
import sys
import glob
from pathlib import Path
from datetime import datetime
from decimal import Decimal
import openpyxl

from config import OUTPUT_DIR, ODOO_COMPANY_NAME


def _format_rupiah(amount: float | Decimal | int | None) -> str:
    if amount is None or amount == "":
        return "Rp&nbsp;0"
    try:
        val = float(amount)
        if val == 0:
            return "Rp&nbsp;0"
        sign = "-" if val < 0 else ""
        abs_val = abs(val)
        formatted = f"{abs_val:,.0f}".replace(",", ".")
        return f"{sign}Rp&nbsp;{formatted}"
    except (ValueError, TypeError):
        return f"Rp&nbsp;{amount}"


def extract_reconciliation_summary(excel_path: Path | str | None = None) -> dict:
    """Extract comprehensive metrics, actual discrepancies, manual matches, and incomplete coverage notices."""
    if excel_path is None:
        files = glob.glob(str(OUTPUT_DIR / "[Rr]econciliation_*.xlsx"))
        if not files:
            raise FileNotFoundError("No Reconciliation report file found in output directory.")
        excel_path = Path(max(files, key=os.path.getmtime))
    else:
        excel_path = Path(excel_path)
        if not excel_path.exists():
            raise FileNotFoundError(f"Reconciliation file not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    
    # Metadata
    filename = excel_path.name
    generation_time = datetime.fromtimestamp(excel_path.stat().st_mtime).strftime("%d %B %Y, %H:%M:%S")
    company_name = ODOO_COMPANY_NAME or "Eyerizz Eyewear"

    # 1. Daily Summary Parsing
    daily_rows = []
    audited_bank_amt = 0.0
    audited_odoo_amt = 0.0
    audited_diff_amt = 0.0
    
    complete_batches = set()   # (date, journal)
    incomplete_batches = {}    # (date, journal) -> {bank, bank_amt, odoo_amt, diff_amt}
    
    banks_found = set()
    dates_found = []
    settlement_total = 0
    settlement_resolved = 0
    channel_stats = {} # journal -> {bank, bank_amt, odoo_amt, diff_amt, count, edc_nums: set(), ar_nums: set()}

    if "Daily Summary" in wb.sheetnames:
        ws_ds = wb["Daily Summary"]
        for row in ws_ds.iter_rows(min_row=4, values_only=True):
            if not row or row[0] is None:
                continue
            d_str = str(row[1] or "").strip()
            p_d_str = str(row[2] or "").strip()
            b_name = str(row[3] or "").strip()
            j_name = str(row[4] or b_name).strip()
            
            try: b_amt = float(row[5] or 0.0)
            except: b_amt = 0.0
            try: o_amt = float(row[6] or 0.0)
            except: o_amt = 0.0
            try: diff_amt = float(row[7] or 0.0)
            except: diff_amt = 0.0
            
            status = str(row[9] or "").strip()
            j_info = str(row[10] if len(row) > 10 else "").strip()
            edc_no = str(row[11] if len(row) > 11 else "").strip()
            ar_no = str(row[12] if len(row) > 12 else "").strip()

            if b_name: banks_found.add(b_name)
            if d_str: dates_found.append(d_str)

            key = (d_str, j_name)
            is_incomplete = "incomplete" in status.lower()

            if is_incomplete:
                incomplete_batches[key] = {
                    "date": d_str,
                    "bank": b_name,
                    "journal": j_name,
                    "bank_amount": b_amt,
                    "odoo_amount": o_amt,
                    "difference": diff_amt
                }
            else:
                complete_batches.add(key)
                audited_bank_amt += b_amt
                audited_odoo_amt += o_amt
                audited_diff_amt += diff_amt

            if j_info and j_info != "-":
                settlement_total += 1
                if "EDC" in j_info or "AR" in j_info or "SETTLEMENT" in j_info.upper():
                    settlement_resolved += 1

            # Channel stats aggregation
            if j_name not in channel_stats:
                channel_stats[j_name] = {
                    "bank": b_name,
                    "bank_amt": 0.0,
                    "odoo_amt": 0.0,
                    "diff_amt": 0.0,
                    "complete_days": 0,
                    "incomplete_days": 0,
                    "edc_nums": set(),
                    "ar_nums": set()
                }
            channel_stats[j_name]["bank_amt"] += b_amt
            channel_stats[j_name]["odoo_amt"] += o_amt
            channel_stats[j_name]["diff_amt"] += diff_amt
            if is_incomplete:
                channel_stats[j_name]["incomplete_days"] += 1
            else:
                channel_stats[j_name]["complete_days"] += 1

            if edc_no and edc_no != "-": channel_stats[j_name]["edc_nums"].add(edc_no)
            if ar_no and ar_no != "-": channel_stats[j_name]["ar_nums"].add(ar_no)

            daily_rows.append({
                "date": d_str,
                "payment_date": p_d_str,
                "bank": b_name,
                "journal": j_name,
                "bank_amount": b_amt,
                "odoo_amount": o_amt,
                "difference": diff_amt,
                "status": status,
                "is_incomplete": is_incomplete,
                "edc_number": edc_no,
                "ar_number": ar_no
            })

    # 2. Differences Parsing: Separate Actionable Unresolved vs Manual Matches vs Incomplete
    actual_discrepancies = []
    manual_matches = []
    incomplete_discrepancies = []
    
    act_bank_only_cnt = 0
    act_odoo_only_cnt = 0
    act_bank_only_amt = 0.0
    act_odoo_only_amt = 0.0

    if "Differences" in wb.sheetnames:
        ws_diff = wb["Differences"]
        for row in ws_diff.iter_rows(min_row=4, values_only=True):
            if not row or row[0] is None:
                continue
            d_str = str(row[1] or "").strip()
            b_name = str(row[2] or "").strip()
            j_name = str(row[3] or "").strip()
            o_num = str(row[4] or "").strip()
            ref = str(row[5] or "").strip()
            b_num = str(row[6] or "").strip()
            
            try: b_amt = float(row[8] or 0.0) if row[8] is not None else 0.0
            except: b_amt = 0.0
            try: o_amt = float(row[9] or 0.0) if row[9] is not None else 0.0
            except: o_amt = 0.0
            
            raw_status = str(row[12] or "").strip()
            key = (d_str, j_name)

            is_manual_match = "Match" in raw_status or "M0" in raw_status or "M1" in raw_status

            if "Bank" in raw_status or (b_amt > 0 and o_amt == 0):
                disc_type = "Bank Only"
                disc_amt = b_amt
            elif "Odoo" in raw_status or (o_amt > 0 and b_amt == 0):
                disc_type = "Odoo Only"
                disc_amt = o_amt
            else:
                disc_type = raw_status or "Amount Difference"
                disc_amt = abs(b_amt - o_amt)

            item = {
                "date": d_str,
                "bank": b_name,
                "journal": j_name,
                "odoo_number": o_num if o_num and o_num != "None" else "-",
                "reference": ref if ref and ref != "None" else "-",
                "bank_number": b_num if b_num and b_num != "None" else "-",
                "bank_amount": b_amt,
                "odoo_amount": o_amt,
                "amount": disc_amt,
                "type": disc_type,
                "status": raw_status
            }

            if is_manual_match:
                manual_matches.append(item)
            elif key in complete_batches:
                actual_discrepancies.append(item)
                if disc_type == "Bank Only":
                    act_bank_only_cnt += 1
                    act_bank_only_amt += disc_amt
                elif disc_type == "Odoo Only":
                    act_odoo_only_cnt += 1
                    act_odoo_only_amt += disc_amt
            else:
                incomplete_discrepancies.append(item)

    wb.close()

    # Parse date range
    parsed_dates = []
    for ds in dates_found:
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y", "%m/%d/%Y"):
            try:
                parsed_dates.append(datetime.strptime(str(ds).strip()[:10], fmt).date())
                break
            except Exception:
                continue
    
    if parsed_dates:
        min_date_str = min(parsed_dates).strftime("%d/%m/%Y")
        max_date_str = max(parsed_dates).strftime("%d/%m/%Y")
        date_range_display = f"{min_date_str} – {max_date_str}"
    else:
        date_range_display = "Current Period"

    # Incomplete stats summary
    incomplete_uncompared_odoo_amt = sum(b["odoo_amount"] for b in incomplete_batches.values() if b["bank_amount"] == 0)
    incomplete_uncompared_bank_amt = sum(b["bank_amount"] for b in incomplete_batches.values() if b["odoo_amount"] == 0)

    # Audited Match rate on complete batches
    if audited_bank_amt > 0:
        match_rate_pct = max(0.0, round(((audited_bank_amt - abs(audited_diff_amt)) / audited_bank_amt) * 100, 2))
    else:
        match_rate_pct = 100.0

    return {
        "company_name": company_name,
        "date_range": date_range_display,
        "filename": filename,
        "generation_time": generation_time,
        "audited_bank_amount": audited_bank_amt,
        "audited_odoo_amount": audited_odoo_amt,
        "audited_diff_amount": audited_diff_amt,
        "match_rate": match_rate_pct,
        "actual_discrepancies": actual_discrepancies,
        "act_bank_only_count": act_bank_only_cnt,
        "act_bank_only_amount": act_bank_only_amt,
        "act_odoo_only_count": act_odoo_only_cnt,
        "act_odoo_only_amount": act_odoo_only_amt,
        "total_actual_discrepancies": len(actual_discrepancies),
        "manual_matches": manual_matches,
        "total_manual_matches": len(manual_matches),
        "incomplete_batches": list(incomplete_batches.values()),
        "incomplete_discrepancies_count": len(incomplete_discrepancies),
        "incomplete_uncompared_odoo_amount": incomplete_uncompared_odoo_amt,
        "incomplete_uncompared_bank_amount": incomplete_uncompared_bank_amt,
        "settlement_total": settlement_total,
        "settlement_resolved": settlement_resolved,
        "settlement_pct": int((settlement_resolved / settlement_total * 100)) if settlement_total > 0 else 100,
        "banks": sorted(list(banks_found)),
        "channel_stats": channel_stats,
        "daily_rows": daily_rows
    }


def render_html_summary(data: dict) -> str:
    """Render clean, executive-ready HTML report with separated manual matches and non-wrapping currency."""
    company = data["company_name"]
    date_range = data["date_range"]
    gen_time = data["generation_time"]
    banks_list_str = ", ".join(data["banks"]) if data["banks"] else "All Banks"
    
    bank_amt_str = _format_rupiah(data["audited_bank_amount"])
    odoo_amt_str = _format_rupiah(data["audited_odoo_amount"])
    diff_amt_str = _format_rupiah(data["audited_diff_amount"])
    match_rate = data["match_rate"]
    
    total_act_disc = data["total_actual_discrepancies"]
    act_bank_only_cnt = data["act_bank_only_count"]
    act_bank_only_amt_str = _format_rupiah(data["act_bank_only_amount"])
    act_odoo_only_cnt = data["act_odoo_only_count"]
    act_odoo_only_amt_str = _format_rupiah(data["act_odoo_only_amount"])

    settle_pct = data["settlement_pct"]
    settle_res = data["settlement_resolved"]
    settle_tot = data["settlement_total"]

    # Channel stats table rows
    channel_rows_html = ""
    for ch_name, s in data["channel_stats"].items():
        diff_val = s["diff_amt"]
        diff_cls = "text-success" if diff_val == 0 else "text-warning"
        
        if s["incomplete_days"] > 0 and s["complete_days"] == 0:
            coverage_badge = f"<span class='badge badge-yellow'>⚠️ {s['incomplete_days']} Dates Pending Bank Upload</span>"
        elif s["incomplete_days"] > 0:
            coverage_badge = f"<span class='badge badge-blue'>✓ {s['complete_days']} Audited</span> <span class='badge badge-yellow'>⚠️ {s['incomplete_days']} Pending</span>"
        else:
            coverage_badge = f"<span class='badge badge-green'>✓ 100% Audited ({s['complete_days']} Days)</span>"

        edc_badge = f"<span class='badge badge-blue'>{len(s['edc_nums'])} EDC</span>" if s["edc_nums"] else "<span class='badge badge-gray'>-</span>"
        ar_badge = f"<span class='badge badge-purple'>{len(s['ar_nums'])} AR</span>" if s["ar_nums"] else "<span class='badge badge-gray'>-</span>"
        
        channel_rows_html += f"""
        <tr>
          <td class="font-bold">{ch_name}</td>
          <td><span class="bank-pill bank-{s['bank'].lower()}">{s['bank']}</span></td>
          <td class="text-right font-mono font-bold nowrap">{_format_rupiah(s['bank_amt'])}</td>
          <td class="text-right font-mono font-bold nowrap">{_format_rupiah(s['odoo_amt'])}</td>
          <td class="text-right font-mono font-bold nowrap {diff_cls}">{_format_rupiah(diff_val)}</td>
          <td>{coverage_badge}</td>
          <td class="text-center">{edc_badge} {ar_badge}</td>
        </tr>
        """

    # Discrepancies table rows (Pure Actionable Unresolved Only)
    actual_discrepancies = data["actual_discrepancies"]
    disc_rows_html = ""
    if actual_discrepancies:
        for i, d in enumerate(actual_discrepancies[:60], start=1):
            type_cls = "badge-bank" if "Bank" in d["type"] else "badge-odoo"
            disc_rows_html += f"""
            <tr>
              <td class="text-center text-muted">{i}</td>
              <td class="nowrap">{d['date']}</td>
              <td><span class="bank-pill bank-{d['bank'].lower()}">{d['bank']}</span></td>
              <td class="text-muted">{d['journal']}</td>
              <td><span class="badge {type_cls}">{d['type']}</span></td>
              <td class="font-mono">{d['bank_number']}</td>
              <td class="font-mono">{d['odoo_number']}</td>
              <td class="font-mono">{d['reference']}</td>
              <td class="text-right font-mono font-bold nowrap">{_format_rupiah(d['amount'])}</td>
            </tr>
            """
    else:
        disc_rows_html = """
        <tr>
          <td colspan="9" style="text-align:center; padding: 16px; color: #059669; font-weight: bold;">
            🎉 100% Reconciled! No unresolved discrepancies found in audited dates.
          </td>
        </tr>
        """

    # Manual Matches Table Rows
    manual_matches = data["manual_matches"]
    manual_rows_html = ""
    if manual_matches:
        for i, m in enumerate(manual_matches, start=1):
            variance = abs(m["bank_amount"] - m["odoo_amount"])
            manual_rows_html += f"""
            <tr>
              <td class="text-center text-muted">{i}</td>
              <td class="nowrap">{m['date']}</td>
              <td><span class="bank-pill bank-{m['bank'].lower()}">{m['bank']}</span></td>
              <td class="text-muted">{m['journal']}</td>
              <td><span class="badge badge-match">{m['status']}</span></td>
              <td class="font-mono">{m['bank_number']}</td>
              <td class="font-mono">{m['odoo_number']}</td>
              <td class="font-mono">{m['reference']}</td>
              <td class="text-right font-mono font-bold nowrap">{_format_rupiah(m['bank_amount'])}</td>
              <td class="text-right font-mono font-bold nowrap">{_format_rupiah(m['odoo_amount'])}</td>
              <td class="text-right font-mono font-bold nowrap text-muted">{_format_rupiah(variance)}</td>
            </tr>
            """
    else:
        manual_rows_html = """
        <tr>
          <td colspan="11" style="text-align:center; padding: 12px; color: #64748B;">
            No manual matches recorded for this reconciliation run.
          </td>
        </tr>
        """

    # Incomplete Coverage Warning Box
    incomplete_batches = data["incomplete_batches"]
    incomplete_box_html = ""
    if incomplete_batches:
        uncompared_odoo_str = _format_rupiah(data["incomplete_uncompared_odoo_amount"])
        incomplete_dates_summary = {}
        for b in incomplete_batches:
            incomplete_dates_summary.setdefault(b["journal"], []).append(b["date"])
        
        breakdown_items = ""
        for j_name, d_list in incomplete_dates_summary.items():
            breakdown_items += f"<li><strong>{j_name}</strong>: {len(d_list)} date(s) uncompared ({d_list[0]} … {d_list[-1]})</li>"

        incomplete_box_html = f"""
        <div class="incomplete-card avoid-break">
          <div class="incomplete-hdr">
            <span class="incomplete-icon">⚠️</span>
            <div>
              <div class="incomplete-title">Incomplete Source Data Notice ({len(incomplete_batches)} Date Batches)</div>
              <div class="incomplete-sub">Bank statements were not uploaded for the following dates/journals. These entries are excluded from active discrepancies until statement files are provided.</div>
            </div>
          </div>
          <div class="incomplete-body">
            <div style="margin-bottom: 6px;"><strong>Uncompared Odoo Payments:</strong> <span class="font-mono font-bold text-warning nowrap">{uncompared_odoo_str}</span> ({data['incomplete_discrepancies_count']:,} entries pending bank statement upload)</div>
            <ul style="margin: 4px 0 0 16px; padding: 0; font-size: 10px; color: #475569;">
              {breakdown_items}
            </ul>
          </div>
        </div>
        """

    html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Reconciliation Executive Summary - {company}</title>
<style>
  @page {{
    size: A4;
    margin: 14mm 14mm 16mm 14mm;
    @bottom-right {{
      content: "Page " counter(page) " of " counter(pages);
      font-size: 9px;
      color: #94A3B8;
      font-family: sans-serif;
    }}
  }}
  * {{ box-sizing: border-box; }}
  body {{
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    color: #0F172A;
    background: #FFFFFF;
    margin: 0;
    padding: 0;
    font-size: 11px;
    line-height: 1.4;
    -webkit-print-color-adjust: exact;
  }}
  
  /* Header */
  .header {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    border-bottom: 2px solid #E2E8F0;
    padding-bottom: 12px;
    margin-bottom: 14px;
  }}
  .header-left h1 {{
    margin: 0;
    font-size: 20px;
    font-weight: 800;
    color: #1E3A8A;
    letter-spacing: -0.5px;
  }}
  .header-left p {{
    margin: 3px 0 0 0;
    color: #64748B;
    font-size: 10.5px;
  }}
  .header-meta {{
    margin-top: 4px;
    font-size: 10px;
    color: #334155;
  }}
  .header-right {{
    text-align: right;
  }}
  .status-badge {{
    display: inline-block;
    background: #DCFCE7;
    color: #166534;
    font-weight: 700;
    font-size: 11px;
    padding: 4px 10px;
    border-radius: 20px;
    border: 1px solid #BBF7D0;
  }}
  .gen-time {{
    margin: 4px 0 0 0;
    font-size: 9.5px;
    color: #94A3B8;
  }}

  /* Scorecards Grid */
  .kpi-grid {{
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 10px;
    margin-bottom: 16px;
  }}
  .kpi-card {{
    background: #F8FAFC;
    border: 1px solid #E2E8F0;
    border-radius: 8px;
    padding: 10px 12px;
    position: relative;
  }}
  .kpi-card.primary {{ border-left: 4px solid #2563EB; }}
  .kpi-card.success {{ border-left: 4px solid #059669; }}
  .kpi-card.warning {{ border-left: 4px solid #D97706; }}
  .kpi-card.accent  {{ border-left: 4px solid #7C3AED; }}
  .kpi-title {{
    font-size: 9px;
    font-weight: 700;
    color: #64748B;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 4px;
  }}
  .kpi-value {{
    font-size: 14px;
    font-weight: 800;
    color: #0F172A;
    margin-bottom: 2px;
    font-family: Consolas, Menlo, monospace;
    white-space: nowrap;
  }}
  .kpi-sub {{
    font-size: 9.5px;
    color: #64748B;
  }}

  /* Section Titles */
  .section-title {{
    font-size: 12.5px;
    font-weight: 700;
    color: #1E293B;
    margin: 16px 0 8px 0;
    display: flex;
    align-items: center;
    gap: 6px;
  }}

  /* Tables */
  table {{
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 12px;
    font-size: 10px;
  }}
  th {{
    background: #F1F5F9;
    color: #475569;
    font-weight: 700;
    text-align: left;
    padding: 6px 8px;
    border-top: 1px solid #CBD5E1;
    border-bottom: 1px solid #CBD5E1;
    font-size: 9px;
    text-transform: uppercase;
    letter-spacing: 0.3px;
  }}
  td {{
    padding: 5px 8px;
    border-bottom: 1px solid #E2E8F0;
    color: #1E293B;
  }}
  tr:nth-child(even) {{ background: #FAFAFC; }}
  
  /* Utilities */
  .text-right {{ text-align: right; }}
  .text-center {{ text-align: center; }}
  .font-bold {{ font-weight: 700; }}
  .font-mono {{ font-family: Consolas, Menlo, monospace; font-size: 9.5px; }}
  .nowrap {{ white-space: nowrap; }}
  .text-success {{ color: #059669; }}
  .text-warning {{ color: #D97706; }}
  .text-danger {{ color: #DC2626; }}
  .text-muted {{ color: #64748B; }}

  /* Badges & Pills */
  .bank-pill {{
    display: inline-block;
    padding: 1px 6px;
    border-radius: 4px;
    font-size: 9px;
    font-weight: 700;
  }}
  .bank-bca {{ background: #DBEAFE; color: #1E40AF; }}
  .bank-mandiri {{ background: #FEF3C7; color: #92400E; }}
  .bank-bri {{ background: #D1FAE5; color: #065F46; }}

  .badge {{
    display: inline-block;
    padding: 2px 6px;
    border-radius: 4px;
    font-size: 9px;
    font-weight: 700;
    white-space: nowrap;
  }}
  .badge-bank {{ background: #EEF2FF; color: #4338CA; border: 1px solid #C7D2FE; }}
  .badge-odoo {{ background: #FEF3C7; color: #B45309; border: 1px solid #FDE68A; }}
  .badge-match {{ background: #F3E8FF; color: #6B21A8; border: 1px solid #E9D5FF; }}
  .badge-green {{ background: #DCFCE7; color: #166534; }}
  .badge-blue {{ background: #E0E7FF; color: #3730A3; }}
  .badge-yellow {{ background: #FEF3C7; color: #92400E; }}
  .badge-purple {{ background: #F3E8FF; color: #6B21A8; }}
  .badge-gray {{ background: #F1F5F9; color: #64748B; }}

  /* Incomplete Data Card */
  .incomplete-card {{
    background: #FFFBEB;
    border: 1px solid #FCD34D;
    border-left: 4px solid #F59E0B;
    border-radius: 8px;
    padding: 10px 14px;
    margin-bottom: 14px;
  }}
  .incomplete-hdr {{
    display: flex;
    align-items: flex-start;
    gap: 8px;
    margin-bottom: 6px;
  }}
  .incomplete-icon {{ font-size: 15px; }}
  .incomplete-title {{ font-size: 11px; font-weight: 700; color: #92400E; }}
  .incomplete-sub {{ font-size: 10px; color: #B45309; }}
  .incomplete-body {{ font-size: 10px; color: #78350F; }}

  /* Insights Box */
  .insights-card {{
    background: #F8FAFC;
    border: 1px solid #E2E8F0;
    border-radius: 8px;
    padding: 12px 14px;
    margin-bottom: 14px;
  }}
  .insights-grid {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 14px;
  }}
  .insights-item h4 {{
    margin: 0 0 4px 0;
    font-size: 11px;
    font-weight: 700;
    color: #1E293B;
  }}
  .insights-item p {{
    margin: 0;
    color: #64748B;
    font-size: 10px;
  }}

  /* Page break rules */
  .avoid-break {{ page-break-inside: avoid; }}
</style>
</head>
<body>

  <!-- ========================== PAGE 1: EXECUTIVE SUMMARY ========================== -->
  <div class="header">
    <div class="header-left">
      <h1>Reconciliation Executive Summary</h1>
      <p><strong>{company}</strong> &nbsp;•&nbsp; Date Range: <strong>{date_range}</strong></p>
      <div class="header-meta">
        🏦 <strong>Banks Included:</strong> {banks_list_str} &nbsp;|&nbsp; 📁 <strong>Source:</strong> {data['filename']}
      </div>
    </div>
    <div class="header-right">
      <span class="status-badge">● Settlement Audited</span>
      <p class="gen-time">Generated: {gen_time}</p>
    </div>
  </div>

  <!-- KPI Scorecards -->
  <div class="kpi-grid">
    <div class="kpi-card primary">
      <div class="kpi-title">Audited Bank Volume</div>
      <div class="kpi-value">{bank_amt_str}</div>
      <div class="kpi-sub">Complete Statements Audited</div>
    </div>
    <div class="kpi-card success">
      <div class="kpi-title">Audited Odoo Payments</div>
      <div class="kpi-value">{odoo_amt_str}</div>
      <div class="kpi-sub">Matching Date Entries</div>
    </div>
    <div class="kpi-card accent">
      <div class="kpi-title">Reconciliation Match</div>
      <div class="kpi-value">{match_rate}%</div>
      <div class="kpi-sub">{settle_pct}% Journals Linked ({settle_res}/{settle_tot})</div>
    </div>
    <div class="kpi-card warning">
      <div class="kpi-title">Active Discrepancies</div>
      <div class="kpi-value">{total_act_disc} Items</div>
      <div class="kpi-sub">Net Difference: <span class="nowrap">{diff_amt_str}</span></div>
    </div>
  </div>

  <!-- Channel Breakdown Table -->
  <div class="section-title">📊 Per-Channel & Bank Settlement Performance</div>
  <table>
    <thead>
      <tr>
        <th>Channel / Journal</th>
        <th>Bank</th>
        <th class="text-right">Bank Settlement</th>
        <th class="text-right">Odoo Recorded</th>
        <th class="text-right">Net Difference</th>
        <th>Data Coverage</th>
        <th class="text-center">Settlement Status</th>
      </tr>
    </thead>
    <tbody>
      {channel_rows_html}
    </tbody>
  </table>

  <!-- Incomplete Data Notice Box -->
  {incomplete_box_html}

  <!-- Accounting & Settlement Status -->
  <div class="insights-card avoid-break">
    <div class="insights-grid">
      <div class="insights-item">
        <h4>📑 EDC & Settlement Journal Coverage ({settle_pct}%)</h4>
        <p><strong>{settle_res} of {settle_tot}</strong> settlement journal numbers successfully matched to Odoo draft entries. {settle_tot - settle_res} pending accounting resolution.</p>
      </div>
      <div class="insights-item">
        <h4>🔍 Active Discrepancy Breakdown</h4>
        <p><strong>Bank Only:</strong> {act_bank_only_cnt} transactions (<span class="nowrap">{act_bank_only_amt_str}</span>) unrecorded in Odoo.<br>
        <strong>Odoo Only:</strong> {act_odoo_only_cnt} transactions (<span class="nowrap">{act_odoo_only_amt_str}</span>) unsettled in bank statements.<br>
        <strong>Manual Matches:</strong> {data['total_manual_matches']} matched pairs logged in audit trail.</p>
      </div>
    </div>
  </div>

  <!-- Page 1 Footer -->
  <div style="margin-top: 16px; padding-top: 8px; border-top: 1px solid #E2E8F0; display: flex; justify-content: space-between; font-size: 9.5px; color: #94A3B8;">
    <div>Report Source: {data['filename']}</div>
    <div>Executive Summary Overview • Page 1</div>
  </div>

  <!-- ========================== PAGE 2: ACTIVE DISCREPANCIES ========================== -->
  <div class="page-break" style="page-break-before: always; break-before: page;"></div>

  <div class="header" style="margin-top: 6px;">
    <div class="header-left">
      <h1 style="font-size: 17px; color: #B45309;">⚠️ Active Discrepancy Items (Unresolved)</h1>
      <p><strong>{company}</strong> &nbsp;•&nbsp; Date Range: <strong>{date_range}</strong> &nbsp;•&nbsp; <strong>{total_act_disc} Actionable Items</strong></p>
    </div>
    <div class="header-right">
      <span class="status-badge" style="background: #FEF3C7; color: #92400E; border-color: #FDE68A;">● Review Required</span>
      <p class="gen-time">Generated: {gen_time}</p>
    </div>
  </div>

  <!-- Actionable Discrepancy Detail Table -->
  <table>
    <thead>
      <tr>
        <th class="text-center" style="width: 25px;">#</th>
        <th>Date</th>
        <th>Bank</th>
        <th>Journal</th>
        <th>Type</th>
        <th>Bank Trace #</th>
        <th>Odoo Payment #</th>
        <th>Invoice Ref</th>
        <th class="text-right">Amount</th>
      </tr>
    </thead>
    <tbody>
      {disc_rows_html}
    </tbody>
  </table>

  <!-- ========================== MANUAL MATCHES AUDIT TRAIL ========================== -->
  <div class="section-title" style="margin-top: 24px; color: #6B21A8;">
    🤝 Manual Matches & Adjustments Audit Trail ({data['total_manual_matches']} Matched Pairs)
  </div>
  <table>
    <thead>
      <tr>
        <th class="text-center" style="width: 25px;">#</th>
        <th>Date</th>
        <th>Bank</th>
        <th>Journal</th>
        <th>Match Status</th>
        <th>Bank Trace #</th>
        <th>Odoo Payment #</th>
        <th>Invoice Ref</th>
        <th class="text-right">Bank Amt</th>
        <th class="text-right">Odoo Amt</th>
        <th class="text-right">Difference</th>
      </tr>
    </thead>
    <tbody>
      {manual_rows_html}
    </tbody>
  </table>

  <!-- Page 2/Final Footer -->
  <div style="margin-top: 24px; padding-top: 10px; border-top: 1px solid #E2E8F0; display: flex; justify-content: space-between; font-size: 9.5px; color: #94A3B8;">
    <div>Report Source: {data['filename']}</div>
    <div>System Generated • Automation Suite</div>
  </div>

</body>
</html>
    """
    return html


def generate_executive_summary_pdf(excel_path: Path | str | None = None, output_pdf_path: Path | str | None = None) -> Path:
    """
    Extracts data from the latest reconciliation Excel file, generates a clean HTML summary,
    and converts it to a professional PDF using Playwright headless Chromium.
    """
    data = extract_reconciliation_summary(excel_path)
    html_content = render_html_summary(data)

    if output_pdf_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        safe_dates = data["date_range"].replace(" ", "").replace("/", "").replace("–", "_to_").replace("-", "_to_")
        output_pdf_path = OUTPUT_DIR / f"Executive_Summary_{safe_dates}.pdf"
    else:
        output_pdf_path = Path(output_pdf_path)

    # Convert HTML to PDF via Playwright
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page()
        page.set_content(html_content, wait_until="networkidle")
        page.pdf(
            path=str(output_pdf_path),
            format="A4",
            print_background=True,
            margin={"top": "12mm", "bottom": "14mm", "left": "12mm", "right": "12mm"}
        )
        browser.close()

    print(f"✅ Executive Summary PDF successfully generated: {output_pdf_path.name}")
    return output_pdf_path


if __name__ == "__main__":
    pdf_file = generate_executive_summary_pdf()
    print(f"Generated PDF: {pdf_file}")
