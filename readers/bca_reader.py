"""
readers/bca_reader.py

Reads BCA bank transactions from a password-protected Excel file.

File discovery:
  - Searches BCA_EXCEL_DIR for an Excel file whose name contains BCA_EXCEL_PATTERN
  - Default pattern: 'ReportMerchantBCA_' (e.g. ReportMerchantBCA_20260720.xlsx)

Excel structure:
  - Password-protected (msoffcrypto-tool to decrypt)
  - Header at row 5 (rows 1-4 are preamble)
  - Column: 'Original Amount' — no thousand separator, ends with .00
  - Column: 'Transaction Date' — format DD/MM/YYYY (e.g. "17/06/2026")
  - Filter: only rows where Transaction Date matches the ODO date
"""

import io
from datetime import date
from pathlib import Path
from decimal import Decimal

import msoffcrypto
import openpyxl

from amount_utils import parse_amount, normalize_for_compare


def _parse_bca_date(raw) -> date | None:
    """Parse BCA date cell: '17/06/2026' → date(2026, 6, 17)"""
    if raw is None:
        return None
    # If Excel stored as actual date object
    if isinstance(raw, date):
        return raw
    from datetime import datetime
    s = str(raw).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def _find_bca_excels(excel_dir: Path, excel_pattern: str = "") -> list[Path]:
    """
    Find ALL BCA Excel files in excel_dir using template structure detection.
    """
    candidates = []
    if excel_pattern:
        candidates = sorted(
            p for p in excel_dir.iterdir()
            if p.suffix.lower() in (".xlsx", ".xls")
            and excel_pattern.lower() in p.name.lower()
            and not p.name.startswith("~$")
        )
        if candidates:
            return candidates

    # Template/Content-based discovery: probe every .xlsx by content
    try:
        from readers.file_detector import _probe_and_alias_bca_xlsx
        from config import BCA_EXCEL_PASSWORD
    except Exception:
        pass
    else:
        content_matches = []
        for p in sorted(excel_dir.iterdir()):
            if p.suffix.lower() not in (".xlsx", ".xls") or p.name.startswith("~$"):
                continue
            try:
                is_bca, _ = _probe_and_alias_bca_xlsx(p, BCA_EXCEL_PASSWORD, {})
                if is_bca:
                    content_matches.append(p)
            except Exception:
                continue
        if content_matches:
            return content_matches

    # If xlsx files exist in bca directory, return them directly
    plain_excels = sorted(p for p in excel_dir.iterdir() if p.suffix.lower() in (".xlsx", ".xls") and not p.name.startswith("~$"))
    if plain_excels:
        return plain_excels

    all_xlsx = [p.name for p in excel_dir.glob("*.xlsx")] + [p.name for p in excel_dir.glob("*.xls")]
    raise FileNotFoundError(
        f"No BCA Excel files found in: {excel_dir}\n"
        f"Available files: {all_xlsx}"
    )


def _read_one_bca(
    excel_path: Path,
    password: str,
    amount_col: str,
    date_col: str,
    number_col: str = "",
    filter_dates: set | None = None,   # if provided, only keep rows with matching date
) -> list[dict]:
    """Read transactions from a single BCA Excel file.
    If filter_dates is given, only rows whose date is in that set are returned.
    """
    print(f"  BCA file: {excel_path.name}")
    print(f"  Decrypting {excel_path.name}...")
    try:
        decrypted = io.BytesIO()
        with open(excel_path, "rb") as f:
            office_file = msoffcrypto.OfficeFile(f)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
        decrypted.seek(0)
        wb = openpyxl.load_workbook(decrypted, data_only=True)
    except Exception:
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception as e:
            raise RuntimeError(
                f"Cannot decrypt BCA Excel '{excel_path.name}': {e}\n"
                f"Check BCA_EXCEL_PASSWORD in your .env file."
            )

    try:
        ws = wb.active
        assert ws is not None

        header_row = None
        header_row_idx = 5
        for r_idx in range(1, min(12, ws.max_row + 1)):
            row_vals = [str(cell.value or "").strip().lower() for cell in ws[r_idx]]
            if any("date" in v or "tanggal" in v for v in row_vals) and any("amount" in v or "nilai" in v or "gross" in v or "original" in v for v in row_vals):
                header_row = [cell.value for cell in ws[r_idx]]
                header_row_idx = r_idx
                break

        if header_row is None:
            header_row = [cell.value for cell in ws[5]] if ws.max_row >= 5 else [cell.value for cell in ws[1]]
            header_row_idx = 5 if ws.max_row >= 5 else 1

        headers = [str(h).strip() if h is not None else "" for h in header_row]

        def _find_col(col_name: str, fallback_keywords: tuple = ()) -> int:
            col_name = col_name.strip()
            if col_name in headers:
                return headers.index(col_name)
            for i, h in enumerate(headers):
                if h.lower() == col_name.lower():
                    return i
            for kw in fallback_keywords:
                for i, h in enumerate(headers):
                    if kw.lower() in h.lower():
                        return i
            raise ValueError(
                f"Column '{col_name}' not found in BCA Excel '{excel_path.name}'.\n"
                f"Available columns: {[h for h in headers if h]}"
            )

        amount_idx = _find_col(amount_col, ("original amount", "gross amount", "total amount", "amount", "nilai"))
        date_idx   = _find_col(date_col, ("transaction date", "tanggal transaksi", "trans date", "date", "tgl"))

        number_idx: int | None = None
        if number_col:
            try:
                number_idx = _find_col(number_col.strip())
            except ValueError:
                print(f"  [WARN] BCA: number column '{number_col}' not found — skipping")

        ref_idx: int | None = None
        try:
            ref_idx = _find_col("Reference Number")
        except ValueError:
            pass

        void_idx: int | None = None
        try:
            void_idx = _find_col("Void")
        except ValueError:
            pass

        reversal_idx: int | None = None
        try:
            reversal_idx = _find_col("Reversal")
        except ValueError:
            pass

        refund_idx: int | None = None
        try:
            refund_idx = _find_col("Refund")
        except ValueError:
            pass

        mid_idx: int | None = None

        try:
            mid_idx = _find_col("Merchant ID")
        except ValueError:
            pass

        store_name = ""
        try:
            import config
            bank_accounts = getattr(config, "BANK_ACCOUNTS", {})
            bca_accs = bank_accounts.get("bca", {})

            # 1. Match MID from rows against configured account MIDs
            if mid_idx is not None:
                for row_sample in ws.iter_rows(min_row=6, max_row=12, values_only=True):
                    if mid_idx < len(row_sample) and row_sample[mid_idx]:
                        raw_mid = str(row_sample[mid_idx]).strip().lstrip("'")
                        for alias_k, acc_info in bca_accs.items():
                            if (acc_info.get("mid") or "").strip() == raw_mid:
                                store_name = acc_info.get("store") or alias_k.capitalize()
                                break
                        if store_name:
                            break

            # 2. Check folder alias
            if not store_name and excel_path.parent.name.lower() in bca_accs:
                store_name = bca_accs[excel_path.parent.name.lower()].get("store", "")

            # 3. Fallback to main store
            if not store_name:
                store_name = bca_accs.get("main", {}).get("store") or "Sanur"
        except Exception:
            store_name = "Sanur"

        txns = []
        skipped_empty = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):

            if not any(c is not None for c in row):
                continue

            raw_amount = row[amount_idx] if amount_idx < len(row) else None
            if raw_amount is None or str(raw_amount).strip() in ("", "0", "0.00"):
                skipped_empty += 1
                continue

            raw_date = row[date_idx] if date_idx < len(row) else None
            txn_date = _parse_bca_date(raw_date)

            if txn_date is None:
                skipped_empty += 1
                continue

            # Date filter: skip rows whose date is not in the allowed set
            if filter_dates is not None and txn_date not in filter_dates:
                continue

            try:
                amount = normalize_for_compare(parse_amount(raw_amount))
            except ValueError as e:
                print(f"  [WARN] BCA row {row_num}: {e} — skipped")
                skipped_empty += 1
                continue

            from datetime import date as dt_date, timedelta
            if isinstance(txn_date, dt_date):
                payment_date = txn_date + timedelta(days=1)
            else:
                payment_date = ""

            desc = ""
            for desc_col_name in ("Transaction Remark", "Keterangan", "Description", "Remark"):
                try:
                    di = _find_col(desc_col_name)
                    val = str(row[di]).strip() if row[di] is not None else ""
                    if val:
                        desc = val
                        break
                except (ValueError, IndexError):
                    continue

            if not desc:
                parts = []
                # 1. Card Number or Phone Number
                ident = ""
                for col in ("Card Number", "Phone Number"):
                    try:
                        idx = _find_col(col)
                        val = str(row[idx]).strip() if row[idx] is not None else ""
                        if val and val.lower() != "none":
                            ident = val
                            break
                    except (ValueError, IndexError):
                        pass
                if ident:
                    parts.append(ident)

                # 2. Date, Time, Method, Type
                for col_name in ("Transaction Date", "Transaction Time", "Payment Method", "Payment Type"):
                    try:
                        idx = _find_col(col_name)
                        raw_val = row[idx]
                        if raw_val is not None:
                            if col_name == "Transaction Date" and hasattr(raw_val, "strftime"):
                                val = raw_val.strftime("%Y-%m-%d")
                            elif col_name == "Transaction Time" and hasattr(raw_val, "strftime"):
                                val = raw_val.strftime("%H:%M:%S")
                            else:
                                val = str(raw_val).strip()

                            if val and val.lower() != "none":
                                parts.append(val)
                    except (ValueError, IndexError):
                        pass
                desc = ", ".join(parts)

            number = ""
            if number_idx is not None and number_idx < len(row):
                number = str(row[number_idx]).strip() if row[number_idx] is not None else ""

            ref_num = ""
            if ref_idx is not None and ref_idx < len(row):
                ref_num = str(row[ref_idx]).strip() if row[ref_idx] is not None else ""

            if not number and ref_num:
                number = ref_num


            is_void = False
            if void_idx is not None and void_idx < len(row):
                is_void = (str(row[void_idx]).strip().upper() == "Y")

            is_reversal = False
            if reversal_idx is not None and reversal_idx < len(row):
                is_reversal = (str(row[reversal_idx]).strip().upper() == "Y")

            is_refund = False
            if refund_idx is not None and refund_idx < len(row):
                is_refund = (str(row[refund_idx]).strip().upper() == "Y")
                if is_refund:
                    print(f"  [WARN] BCA row {row_num}: Refund detected! Manual check required to ensure correct netting.")

            txns.append({
                "amount":      amount,
                "amount_raw":  raw_amount,
                "date":        str(txn_date) if txn_date else str(raw_date or ""),
                "payment_date": str(payment_date) if payment_date else "",
                "description": desc,
                "number":      number,
                "ref_num":     ref_num,
                "is_void":     is_void,
                "is_reversal": is_reversal,
                "is_refund":   is_refund,
                "filename":    excel_path.name,
                "source":      "Bank (BCA)",
                "bank":        "BCA",
                "store":       store_name,
            })


        print(f"    → {len(txns)} transactions ({skipped_empty} empty skipped)")
        return txns
    finally:
        wb.close()
        decrypted.close()   # release in-memory BytesIO — Windows holds handles until explicit close



def read_bca(
    excel_dir: Path,
    excel_pattern: str,
    password: str,
    amount_col: str,
    date_col: str,
    number_col: str = "",
    filter_date=None,          # single date (legacy, converted to set)
    filter_dates: set | None = None,  # set of dates from ODO — primary filter
) -> tuple[list[dict], list[dict]]:
    """
    Read BCA transactions from ALL matching Excel files in excel_dir.

    Date filtering:
      - filter_dates (set[date]): only keep rows matching any date in the set.
        Derived from ODO BCA transactions → preserves original single-date
        behavior when ODO has 1 date, and extends to multi-date automatically.
      - filter_date (date, legacy): converted to a one-element set.
      - If neither is provided: all dates are included (no filter).
    """
    # Backwards-compat: convert legacy single filter_date to set
    if filter_dates is None and filter_date is not None:
        filter_dates = {filter_date}

    if not excel_dir.exists():
        raise FileNotFoundError(
            f"BCA Excel directory not found: {excel_dir}\n"
            f"Check BCA_EXCEL_DIR in your .env file."
        )

    excel_files = _find_bca_excels(excel_dir, excel_pattern)
    print(f"  Found {len(excel_files)} BCA file(s): {[p.name for p in excel_files]}")

    all_txns: list[dict] = []
    if len(excel_files) <= 1:
        # Single file — no parallelism overhead worth paying
        for path in excel_files:
            all_txns.extend(
                _read_one_bca(path, password, amount_col, date_col, number_col,
                              filter_dates=filter_dates)
            )
    else:
        # Multiple files — decrypt + parse in parallel (I/O-bound, threads bypass GIL for disk reads)
        from concurrent.futures import ThreadPoolExecutor, as_completed
        futures_map = {}
        with ThreadPoolExecutor(max_workers=min(len(excel_files), 4),
                                thread_name_prefix="bca_read") as pool:
            for path in excel_files:
                fut = pool.submit(
                    _read_one_bca, path, password, amount_col, date_col,
                    number_col, filter_dates
                )
                futures_map[fut] = path
            results_by_path = {}
            for fut in as_completed(futures_map):
                path = futures_map[fut]
                try:
                    results_by_path[path] = fut.result()
                except Exception as e:
                    print(f"  [WARN] BCA: failed reading {path.name}: {e}")
                    results_by_path[path] = []
        # Re-apply original file order so output is deterministic
        for path in excel_files:
            all_txns.extend(results_by_path.get(path, []))


    # ── Exclude Voided/Reversed Transactions ──────────────────────────────────
    # If a trace number is marked as Void/Reversal, we must drop both the void
    # transaction itself AND the original transaction.
    voided_keys = set()
    for t in all_txns:
        if t.get("is_void") or t.get("is_reversal"):
            num = (t.get("number") or "").strip()
            if num:
                voided_keys.add((t["date"], num))

    excluded_txns = []
    if voided_keys:
        filtered_txns = []
        for t in all_txns:
            num = (t.get("number") or "").strip()
            if num and (t["date"], num) in voided_keys:
                if t.get("is_void"):
                    t["exclusion_reason"] = "Void"
                elif t.get("is_reversal"):
                    t["exclusion_reason"] = "Reversal"
                elif t.get("is_refund"):
                    t["exclusion_reason"] = "Refund"
                else:
                    t["exclusion_reason"] = "Original transaction of Void/Reversal"
                excluded_txns.append(t)
                continue
            filtered_txns.append(t)
        
        print(f"  [INFO] BCA: Excluded {len(excluded_txns)} rows due to Void/Reversal")
        all_txns = filtered_txns

    # ── Deduplicate overlapping files ─────────────────────────────────────────
    # Key: (date, trace_number) if trace_number exists, else (date, amount, desc)
    seen: set = set()
    deduped: list[dict] = []
    for t in all_txns:
        num = (t.get("number") or "").strip()
        ref = (t.get("ref_num") or "").strip()
        if num:
            key = (t["date"], num, ref)
        else:
            key = (t["date"], str(t["amount"]), t.get("description", ""))
        if key not in seen:
            seen.add(key)
            deduped.append(t)

    dupes = len(all_txns) - len(deduped)
    if dupes:
        print(f"  [INFO] BCA: {dupes} duplicate row(s) removed (overlapping files)")
    if filter_dates is not None:
        final_dates = sorted({t["date"] for t in deduped})
        print(f"  ✅ BCA: {len(deduped)} transactions loaded ({len(excel_files)} file(s), dates: {', '.join(final_dates)})")
    else:
        print(f"  ✅ BCA: {len(deduped)} transactions loaded ({len(excel_files)} file(s))")

    return deduped, excluded_txns
