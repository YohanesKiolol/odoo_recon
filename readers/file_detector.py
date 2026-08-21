"""
readers/file_detector.py — content-based file type detection for uploads.

Identifies bank + category from file internals so renamed files are still
routed correctly. Filename is never used for routing decisions.

Detection order:
  .xlsx  -> BCA EDC (AES-encrypted + BCA headers) | Odoo (unencrypted)
  .zip   -> Mandiri EDC (AES-encrypted + CSV inside) | BRI EDC (plain ZIP + PDF inside)
  .csv   -> Mandiri EDC raw (MERCHANT STATEMENT header) |
            BCA mutation | BRI mutation | Mandiri mutation
  .pdf   -> BRI EDC raw (AMT_TRX column in PDF table)

Extracted-file support (#13):
  Mandiri EDC CSV uploaded directly -> AES-re-wrapped into .zip, placed in MANDIRI_ZIP_DIR
  BRI EDC PDF uploaded directly     -> plain-re-wrapped into .zip, placed in BRI_ZIP_DIR
"""

from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass, field
from pathlib import Path


@dataclass
class DetectionResult:
    bank: str         # "bca" | "mandiri" | "bri" | "odoo"
    category: str     # "edc" | "mutation" | "odoo"
    alias: str        # matched alias or "main"
    target_dir: Path  # final destination directory
    wrap_as_zip: bool = field(default=False)  # True = re-zip before copy
    zip_pdf_entry_name: str = field(default="")  # for BRI PDF re-wrap
    is_valid: bool = field(default=True)


# --------------------------------------------------------------------------- #
# helpers                                                                      #
# --------------------------------------------------------------------------- #

def _normalize_pdf_header(h: str) -> str:
    """Fix doubled PDF headers: 'MMIIDD' -> 'MID'."""
    if not h:
        return h
    h = h.replace("\n", " ").replace("\r", "").strip()
    if len(h) >= 2 and len(h) % 2 == 0:
        halved = "".join(h[i] for i in range(0, len(h), 2))
        if "".join(c * 2 for c in halved) == h:
            return halved
    return h


def _alias_from_mid(mid_value: str, accounts: dict) -> str:
    """Match a MID value read from file against all configured aliases."""
    mid_clean = mid_value.strip().lstrip("'")
    for alias, acc_info in accounts.items():
        configured_mid = (acc_info.get("mid") or "").strip()
        if configured_mid and configured_mid == mid_clean:
            return alias
    if "main" in accounts:
        return "main"
    return next(iter(accounts), "main")


def _best_alias_from_text(accounts: dict, candidate_texts: list[str], key: str = "acc") -> str:
    """Return alias whose configured key appears in any candidate text.
    Also tries stripping leading zeros from the identifier (e.g. BRI CSVs store
    '1701005774562' while .env has '001701005774562').
    """
    for alias, acc_info in accounts.items():
        ident = (acc_info.get(key) or "").strip()
        if ident:
            ident_up = ident.upper()
            ident_stripped = ident.lstrip("0").upper()  # zero-stripped variant
            for text in candidate_texts:
                text_up = text.upper()
                if ident_up in text_up:
                    return alias
                # Also match stripped version (avoids leading-zero mismatch)
                if ident_stripped and ident_stripped in text_up:
                    return alias
    if "main" in accounts:
        return "main"
    return next(iter(accounts), "main")


# --------------------------------------------------------------------------- #
# xlsx probes (single decrypt pass for BCA — fixes #3)                        #
# --------------------------------------------------------------------------- #

def _probe_and_alias_bca_xlsx(path: Path, password: str, accounts: dict) -> tuple[bool, str]:
    """
    Single decrypt pass: check BCA EDC headers AND read Merchant ID for alias.
    Returns (is_bca, alias).  Fixes #3 (was two separate decrypts).
    """
    try:
        import msoffcrypto
        import openpyxl
        buf = io.BytesIO()
        with open(path, "rb") as f:
            office = msoffcrypto.OfficeFile(f)
            office.load_key(password=password)
            office.decrypt(buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf, data_only=True, read_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return False, "main"

        headers = [str(c.value or "").strip() for c in ws[5]]
        headers_lower = [h.lower() for h in headers]

        # Must have both BCA-specific columns
        if not (any("original amount" in h for h in headers_lower) and
                any("transaction date" in h for h in headers_lower)):
            wb.close()
            return False, "main"

        # Find 'Merchant ID' column
        mid_idx = next(
            (i for i, h in enumerate(headers_lower) if h == "merchant id"),
            None
        )
        alias = "main"
        if mid_idx is not None:
            for row in ws.iter_rows(min_row=6, max_row=15, values_only=True):
                cell = row[mid_idx] if mid_idx < len(row) else None
                if cell is not None:
                    alias = _alias_from_mid(str(cell), accounts)
                    break

        wb.close()
        return True, alias
    except Exception:
        return False, "main"


def _probe_odoo_xlsx(path: Path) -> bool:
    """
    True if unencrypted xlsx that looks like an Odoo payment export.
    Tightened (#9): requires 'payment' AND at least one of journal/partner/amount
    in the first 3 rows to avoid false-positives.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb.active
            if ws is None:
                return False
            texts: list[str] = []
            for row in ws.iter_rows(max_row=3, values_only=True):
                for cell in row:
                    if cell is not None:
                        texts.append(str(cell).lower())
        finally:
            wb.close()
        joined = " ".join(texts)
        has_payment = "payment" in joined or "account.payment" in joined
        has_odoo_col = any(k in joined for k in ("journal", "partner", "reconcil"))
        return has_payment and has_odoo_col
    except Exception:
        return False



# --------------------------------------------------------------------------- #
# zip probes                                                                   #
# --------------------------------------------------------------------------- #

def _probe_mandiri_zip(path: Path, password: str) -> bool:
    """True if AES ZIP openable with Mandiri password and contains Mandiri EDC CSV."""
    try:
        import pyzipper
        pwd = password.encode("utf-8")
        with pyzipper.AESZipFile(path, "r") as zf:
            zf.setpassword(pwd)
            csv_names = [n for n in zf.namelist() if n.lower().endswith(".csv")]
            if not csv_names:
                return False
            data = zf.read(csv_names[0])
            lines = data.decode("utf-8", errors="replace").splitlines()
            if len(lines) < 6:
                return False
            header = lines[5].upper()
            return "AMOUNT" in header and ("MID" in header or "TRXDATE" in header)
    except Exception:
        return False


def _probe_bri_zip(path: Path, pdf_pattern: str = "detail") -> bool:
    """True if plain ZIP containing a PDF whose name starts with pdf_pattern."""
    try:
        with zipfile.ZipFile(path, "r") as zf:
            return any(
                Path(n).suffix.lower() == ".pdf" and
                Path(n).name.lower().startswith(pdf_pattern.lower())
                for n in zf.namelist()
            )
    except Exception:
        return False


# --------------------------------------------------------------------------- #
# csv probes                                                                   #
# --------------------------------------------------------------------------- #

def _csv_head(path: Path, n_lines: int = 15) -> tuple[str, list[str]]:
    try:
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            lines = [f.readline() for _ in range(n_lines)]
        return "".join(lines), lines
    except Exception:
        return "", []


def _probe_mandiri_edc_csv(raw: str) -> bool:
    """True if CSV is a Mandiri EDC report (extracted from ZIP). Unique first line."""
    return "MERCHANT STATEMENT" in raw


def _probe_bca_mutation(raw: str) -> bool:
    return "Periode" in raw and ("Tanggal" in raw or "Date" in raw)


def _probe_bri_mutation(raw: str) -> bool:
    return "MUTASI_KREDIT" in raw and "MUTASI_DEBET" in raw and "TGL_TRAN" in raw


def _probe_mandiri_mutation(lines: list[str]) -> bool:
    if not lines:
        return False
    try:
        header = next(csv.reader([lines[0]]))
        cols = {c.strip() for c in header}
        return {"Date", "Description", "Credit", "Debit"}.issubset(cols)
    except Exception:
        return False


# --------------------------------------------------------------------------- #
# pdf probe (BRI EDC raw — feature #13b)                                      #
# --------------------------------------------------------------------------- #

def _probe_bri_edc_pdf(path: Path) -> bool:
    """True if PDF contains a table with AMT_TRX column (BRI EDC report)."""
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages[:2]:
                tables = page.extract_tables()
                for table in tables:
                    if not table:
                        continue
                    headers = [_normalize_pdf_header(h or "") for h in table[0]]
                    if any(h.upper() == "AMT_TRX" for h in headers):
                        return True
        return False
    except Exception:
        return False


# --------------------------------------------------------------------------- #
# alias detection                                                              #
# --------------------------------------------------------------------------- #

def _alias_mandiri_zip(path: Path, password: str, accounts: dict) -> str:
    """Read MID from first non-empty data row in Mandiri EDC CSV (fixes #5)."""
    try:
        import pyzipper
        pwd = password.encode("utf-8")
        with pyzipper.AESZipFile(path, "r") as zf:
            zf.setpassword(pwd)
            csv_names = [n for n in zf.namelist() if n.lower().endswith(".csv")]
            if not csv_names:
                return _best_alias_from_text(accounts, [], key="mid")
            data = zf.read(csv_names[0])
            lines = data.decode("utf-8", errors="replace").splitlines()
        return _alias_from_mandiri_edc_lines(lines, accounts)
    except Exception:
        return _best_alias_from_text(accounts, [], key="mid")


def _alias_from_mandiri_edc_lines(lines: list[str], accounts: dict) -> str:
    """Scan up to 10 data rows for a non-empty MID value (fixes #5)."""
    if len(lines) < 7:
        return _best_alias_from_text(accounts, [], key="mid")
    try:
        header = next(csv.reader([lines[5]]))
        header = [h.strip() for h in header]
        mid_col_idx = next(
            (i for i, h in enumerate(header) if h.upper() == "MID"), None
        )
        if mid_col_idx is None:
            return _best_alias_from_text(accounts, [], key="mid")
        # Scan up to 10 data rows for first non-empty MID (#5)
        for line in lines[6:16]:
            try:
                row = next(csv.reader([line]))
                row = [v.strip() for v in row]
                if mid_col_idx < len(row) and row[mid_col_idx]:
                    return _alias_from_mid(row[mid_col_idx], accounts)
            except Exception:
                continue
    except Exception:
        pass
    return _best_alias_from_text(accounts, [], key="mid")


def _alias_bri_zip(path: Path, accounts: dict, pdf_pattern: str = "detail") -> tuple[str, list[str]]:
    """
    Read MID and Account from BRI ZIP file and member PDFs.
    Returns (alias, all_mids_found).
    """
    mids_found: list[str] = []
    try:
        import pdfplumber
        with zipfile.ZipFile(path, "r") as zf:
            namelist = zf.namelist()

            # 1. Match member filenames against configured MIDs and Account numbers
            for name in namelist:
                for alias, acc_info in accounts.items():
                    mid = (acc_info.get("mid") or "").strip()
                    acc = (acc_info.get("acc") or "").strip()
                    mid_s = mid.lstrip("0")
                    acc_s = acc.lstrip("0")
                    if (mid and mid in name) or (mid_s and len(mid_s) >= 5 and mid_s in name):
                        mids_found.append(mid)
                        return alias, mids_found
                    if (acc and acc in name) or (acc_s and len(acc_s) >= 5 and acc_s in name):
                        return alias, mids_found

            # 2. Inspect PDF text and tables
            for name in namelist:
                if name.lower().endswith(".pdf"):
                    try:
                        pdf_bytes = zf.read(name)
                        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                            for page in pdf.pages[:3]:
                                txt = page.extract_text() or ""
                                for alias, acc_info in accounts.items():
                                    mid = (acc_info.get("mid") or "").strip()
                                    acc = (acc_info.get("acc") or "").strip()
                                    mid_s = mid.lstrip("0")
                                    acc_s = acc.lstrip("0")
                                    if (mid and mid in txt) or (mid_s and len(mid_s) >= 5 and mid_s in txt):
                                        mids_found.append(mid)
                                        return alias, mids_found
                                    if (acc and acc in txt) or (acc_s and len(acc_s) >= 5 and acc_s in txt):
                                        return alias, mids_found

                                tables = page.extract_tables()
                                for table in tables:
                                    if not table or len(table) < 2: continue
                                    headers = [_normalize_pdf_header(h or "") for h in table[0]]
                                    mid_idx = next((i for i, h in enumerate(headers) if h.upper() == "MID"), None)
                                    if mid_idx is not None:
                                        for row in table[1:]:
                                            if row and mid_idx < len(row) and row[mid_idx]:
                                                m_val = str(row[mid_idx]).strip().lstrip("'")
                                                if m_val:
                                                    mids_found.append(m_val)
                                                    return _alias_from_mid(m_val, accounts), mids_found
                    except Exception:
                        pass
    except Exception:
        pass

    alias = _alias_from_mid(mids_found[0], accounts) if mids_found else \
            _best_alias_from_text(accounts, [], key="mid")
    return alias, mids_found



def _alias_csv(path: Path, accounts: dict, key: str = "acc") -> str:
    try:
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            content = f.read(32768)
        return _best_alias_from_text(accounts, [content], key=key)
    except Exception:
        return _best_alias_from_text(accounts, [], key=key)


# --------------------------------------------------------------------------- #
# re-wrap helpers (feature #13)                                                #
# --------------------------------------------------------------------------- #

def _wrap_in_aes_zip(src: Path, dest_zip: Path, password: str) -> None:
    """Wrap src file into an AES-encrypted ZIP at dest_zip."""
    import pyzipper
    pwd = password.encode("utf-8")
    with pyzipper.AESZipFile(dest_zip, "w",
                              compression=zipfile.ZIP_DEFLATED,
                              encryption=pyzipper.WZ_AES) as zf:
        zf.setpassword(pwd)
        zf.write(src, arcname=src.name)


def _wrap_in_plain_zip(src: Path, dest_zip: Path, entry_name: str) -> None:
    """Wrap src file into a plain ZIP at dest_zip, stored as entry_name."""
    with zipfile.ZipFile(dest_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(src, arcname=entry_name)


# --------------------------------------------------------------------------- #
# public API                                                                   #
# --------------------------------------------------------------------------- #

def detect_file(path: Path) -> DetectionResult | None:
    """
    Probe file content to determine bank, category, alias, target_dir.
    Returns None if file cannot be identified.
    """
    from config import (
        BCA_EXCEL_DIR, BCA_EXCEL_PASSWORD,
        MANDIRI_ZIP_DIR, MANDIRI_ZIP_PASSWORD,
        BRI_ZIP_DIR, BRI_PDF_PATTERN,
        ODO_EXCEL_PATH,
        MUTATION_DIR,
        BANK_ACCOUNTS,
    )

    ext = path.suffix.lower()

    # ── xlsx ─────────────────────────────────────────────────────────────────
    if ext in (".xlsx", ".xls"):
        accounts = BANK_ACCOUNTS.get("bca", {})
        is_bca, alias = _probe_and_alias_bca_xlsx(path, BCA_EXCEL_PASSWORD, accounts)
        if is_bca:
            return DetectionResult(
                bank="bca", category="edc", alias=alias,
                target_dir=BCA_EXCEL_DIR / alias,
            )
        if _probe_odoo_xlsx(path):
            return DetectionResult(
                bank="odoo", category="odoo", alias="",
                target_dir=ODO_EXCEL_PATH.parent,
            )
        return None

    # ── zip ───────────────────────────────────────────────────────────────────
    if ext == ".zip":
        if _probe_mandiri_zip(path, MANDIRI_ZIP_PASSWORD):
            accounts = BANK_ACCOUNTS.get("mandiri", {})
            alias = _alias_mandiri_zip(path, MANDIRI_ZIP_PASSWORD, accounts)
            return DetectionResult(
                bank="mandiri", category="edc", alias=alias,
                target_dir=MANDIRI_ZIP_DIR / alias,
            )
        if _probe_bri_zip(path, BRI_PDF_PATTERN):
            accounts = BANK_ACCOUNTS.get("bri", {})
            alias, all_mids = _alias_bri_zip(path, accounts, BRI_PDF_PATTERN)
            if len(all_mids) > 1:
                # #4: warn about multi-MID ZIPs (logged in gui.py)
                print(f"  [WARN] Multiple MIDs in BRI ZIP: {', '.join(all_mids)} — routed to alias '{alias}'")
            return DetectionResult(
                bank="bri", category="edc", alias=alias,
                target_dir=BRI_ZIP_DIR / alias,
            )
        return None

    # ── csv ───────────────────────────────────────────────────────────────────
    if ext == ".csv":
        raw, lines = _csv_head(path)
        # Mandiri EDC CSV (extracted from ZIP) — check FIRST, before mutation probes
        if _probe_mandiri_edc_csv(raw):
            accounts = BANK_ACCOUNTS.get("mandiri", {})
            _, lines30 = _csv_head(path, n_lines=30)  # need rows 6-16 for MID scan
            alias = _alias_from_mandiri_edc_lines(lines30, accounts)
            return DetectionResult(
                bank="mandiri", category="edc", alias=alias,
                target_dir=MANDIRI_ZIP_DIR / alias,
                wrap_as_zip=True,  # re-wrap into AES ZIP
            )
        if _probe_bca_mutation(raw):
            accounts = BANK_ACCOUNTS.get("bca", {})
            alias = _alias_csv(path, accounts, key="acc")
            return DetectionResult(
                bank="bca", category="mutation", alias=alias,
                target_dir=MUTATION_DIR / "bca" / alias,
            )
        if _probe_bri_mutation(raw):
            accounts = BANK_ACCOUNTS.get("bri", {})
            alias = _alias_csv(path, accounts, key="acc")
            return DetectionResult(
                bank="bri", category="mutation", alias=alias,
                target_dir=MUTATION_DIR / "bri" / alias,
            )
        if _probe_mandiri_mutation(lines):
            accounts = BANK_ACCOUNTS.get("mandiri", {})
            alias = _alias_csv(path, accounts, key="acc")
            return DetectionResult(
                bank="mandiri", category="mutation", alias=alias,
                target_dir=MUTATION_DIR / "mandiri" / alias,
            )
        return None

    # ── pdf (#12 + feature #13b) ─────────────────────────────────────────────
    if ext == ".pdf":
        if _probe_bri_edc_pdf(path):
            accounts = BANK_ACCOUNTS.get("bri", {})
            # Read PDF bytes for alias detection
            pdf_bytes = path.read_bytes()
            alias, all_mids = _alias_from_bri_pdf_bytes(pdf_bytes, accounts)
            if len(all_mids) > 1:
                print(f"  [WARN] Multiple MIDs in BRI PDF: {', '.join(all_mids)} — routed to alias '{alias}'")
            # Ensure PDF entry name inside ZIP starts with BRI_PDF_PATTERN
            entry_name = path.name
            if not entry_name.lower().startswith(BRI_PDF_PATTERN.lower()):
                entry_name = BRI_PDF_PATTERN + "_" + entry_name
            return DetectionResult(
                bank="bri", category="edc", alias=alias,
                target_dir=BRI_ZIP_DIR / alias,
                wrap_as_zip=True,
                zip_pdf_entry_name=entry_name,
            )
        # Not a recognised BRI EDC PDF — give helpful message via None + caller logs
        return None

    return None


def copy_file(result: DetectionResult, src: Path) -> Path:
    """
    Copy (or re-wrap) src into result.target_dir.
    Returns the final destination path.
    """
    from config import MANDIRI_ZIP_PASSWORD, BRI_PDF_PATTERN

    result.target_dir.mkdir(parents=True, exist_ok=True)

    if result.wrap_as_zip:
        dest_zip = result.target_dir / (src.name + ".zip")
        if result.zip_pdf_entry_name:
            # BRI PDF re-wrap into plain ZIP
            _wrap_in_plain_zip(src, dest_zip, result.zip_pdf_entry_name)
        else:
            # Mandiri EDC CSV re-wrap into AES ZIP
            _wrap_in_aes_zip(src, dest_zip, MANDIRI_ZIP_PASSWORD)
        return dest_zip
    else:
        import shutil
        dest = result.target_dir / src.name
        shutil.copy2(src, dest)
        return dest
