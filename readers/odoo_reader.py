"""
readers/odoo_reader.py

Reads the Odoo Payments export: Payments (account.payment).xlsx
- Header at row 1
- Date reference in cell A3 (format: "20 Jul 2026 (99)" → parsed to date)
- Column 'Amount Signed' — no thousand separator, no .00
- Data is grouped by Transaction Type rows; reads amounts under each group

IMPORTANT — group boundary logic:
  The ODO export structure is:
    "BCA EDC Sanur (20)"  col A = str  ← known group header row (has subtotal amount)
    2026-07-20 00:00:00   col A = date ← individual transaction row
    ...
    "Paypal (1)"          col A = str  ← unknown group → STOP collecting
    2026-07-20 00:00:00   col A = date ← PayPal row, skipped
    "Mandiri EDC (45)"    col A = str  ← known group header
    ...

  Key insight: col A is a datetime object for data rows and a plain string for group
  headers/separators. We use isinstance(col_a, (date, datetime)) to discriminate.
"""

import re
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl

from amount_utils import parse_amount, normalize_for_compare


def _parse_odo_date(raw: str) -> date:
    """
    Parse ODO date from A3: "20 Jul 2026 (99)" → date(2026, 7, 20)
    The "(99)" count suffix is stripped first.
    """
    cleaned = re.sub(r"\(.*?\)", "", raw).strip()
    for fmt in ("%d %b %Y", "%d %B %Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            from datetime import datetime
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue
    raise ValueError(
        f"Cannot parse ODO date from A3: {raw!r}\n"
        f"Expected format like '20 Jul 2026 (99)'"
    )


def _is_date_cell(val) -> bool:
    """Return True if the cell value is a date/datetime (= a data row in ODO)."""
    return isinstance(val, (date, datetime))


def read_odoo(
    excel_path: Path,
    amount_col: str,
    group_map: dict[str, str],      # {odo_group_name: bank_key}
    number_col: str = "",           # e.g. 'Number'
) -> tuple[date, dict[str, list[dict]]]:
    """
    Read ODO Payments Excel.

    Returns:
        odo_date  : date parsed from A3
        bank_txns : {bank_key: [transaction dicts]}

    Each transaction dict:
        amount      : Decimal (normalized)
        amount_raw  : original cell value
        description : str
        number      : str (from number_col, empty if not configured)
        date        : str
        source      : "Odoo"
    """
    if not excel_path.exists():
        raise FileNotFoundError(
            f"ODO Excel not found: {excel_path}\n"
            f"Check ODO_EXCEL_PATH in your .env file."
        )

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    assert ws is not None, "No active worksheet in ODO Excel"

    # ── Date from A3 ──────────────────────────────────────────────────────────
    a3_val = ws["A3"].value
    if a3_val is None:
        raise ValueError("Cell A3 in ODO file is empty — cannot determine date.")
    odo_date = _parse_odo_date(str(a3_val))

    # ── Resolve column indices from header row 1 ───────────────────────────────
    header_row = [cell.value for cell in ws[1]]
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    def _col_idx(col_name: str, required: bool = True) -> int | None:
        if not col_name:
            return None
        col_name = col_name.strip()
        for i, h in enumerate(headers):
            if h == col_name or h.lower() == col_name.lower():
                return i
        if required:
            raise ValueError(
                f"Column '{col_name}' not found in ODO file.\n"
                f"Available columns: {[h for h in headers if h]}\n"
                f"Check config in your .env file."
            )
        print(f"  [WARN] ODO: column '{col_name}' not found — skipping")
        return None

    amount_idx = _col_idx(amount_col, required=True)
    number_idx = _col_idx(number_col, required=False)

    assert amount_idx is not None  # guaranteed by required=True above

    group_names = list(group_map.keys())

    # ── Parse rows ────────────────────────────────────────────────────────────
    bank_txns: dict[str, list[dict]] = {v: [] for v in group_map.values()}
    current_bank_key: str | None = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row):
            continue  # fully blank row

        col_a = row[0]
        raw_amount = row[amount_idx] if amount_idx < len(row) else None

        # ── Group header detection ─────────────────────────────────────────────
        # col A is a datetime for data rows; a string for group headers.
        if not _is_date_cell(col_a):
            val = str(col_a).strip() if col_a is not None else ""
            if not val:
                continue  # empty col A, blank-ish row

            # Check if it matches a known bank group
            matched = None
            for name in group_names:
                if name.lower() in val.lower():
                    matched = name
                    break

            if matched:
                current_bank_key = group_map[matched]
                print(f"  [ODO] Group → {current_bank_key}: '{val}'")
            else:
                # Unknown group (PayPal, Petty Cash, etc.) — stop collecting
                if current_bank_key is not None:
                    print(f"  [ODO] Unknown group '{val}' — stopping {current_bank_key} collection")
                current_bank_key = None
            continue  # group header row itself is never a data row

        # ── Data row (col A is a date) ─────────────────────────────────────────
        if current_bank_key is None:
            continue

        if raw_amount is None or str(raw_amount).strip() in ("", "0", "0.0", "0.00"):
            continue

        try:
            amount = normalize_for_compare(parse_amount(raw_amount))
        except ValueError as e:
            print(f"  [WARN] ODO row skip: {e}")
            continue

        # Actual transaction date from col A (datetime → date string)
        txn_date = col_a.date() if isinstance(col_a, datetime) else col_a
        desc = str(txn_date).split()[0] if txn_date else ""   # keep date as desc for compat
        number = ""
        if number_idx is not None and number_idx < len(row):
            number = str(row[number_idx]).strip() if row[number_idx] is not None else ""

        bank_txns[current_bank_key].append({
            "amount":      amount,
            "amount_raw":  raw_amount,
            "description": desc,
            "number":      number,
            "date":        str(txn_date).split()[0],   # "2026-07-20" (drop time part)
            "source":      "Odoo",
        })

    return odo_date, bank_txns
