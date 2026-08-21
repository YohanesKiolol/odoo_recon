"""Audited Financial Volume Breakdown modal dialog."""
import tkinter as tk
import customtkinter as ctk
from ui.theme import (
    BG, PANEL, PREVIEW_BG, BORDER, BORDER_DARK, TEXT, MUTED,
    WHITE, ACCENT, ACCENT_DARK, SUCCESS, ERROR, FONT_FAMILY, FONT_MONO
)

def open_breakdown_modal(parent, stats: dict = None, on_open_output = None):
    """Show detailed modal with financial breakdown per bank from latest reconciliation."""
    if not stats:
        try:
            import glob, os
            from openpyxl import load_workbook
            from config import OUTPUT_DIR
            out_files = glob.glob(str(OUTPUT_DIR / "[Rr]econciliation_*.xlsx"))
            if out_files:
                latest = max(out_files, key=os.path.getmtime)
                wb = load_workbook(latest, read_only=True, data_only=True)
                bank_stats = {}
                if "Daily Summary" in wb.sheetnames:
                    ws_ds = wb["Daily Summary"]
                    for row in ws_ds.iter_rows(min_row=4, values_only=True):
                        if not row or not row[0]: continue
                        d_str = str(row[1] or "").strip()
                        b_n = str(row[3] or "Other").strip()
                        tb = float(row[5] or 0.0)
                        to = float(row[6] or 0.0)
                        df = abs(float(row[7] or 0.0))
                        st = bank_stats.setdefault(b_n, {"dates": set(), "tot_bank": 0.0, "tot_odoo": 0.0, "tot_diff": 0.0})
                        if d_str: st["dates"].add(d_str)
                        st["tot_bank"] += tb
                        st["tot_odoo"] += to
                        st["tot_diff"] += df
                wb.close()
                if bank_stats:
                    stats = bank_stats
        except Exception:
            pass

    if not stats:
        if on_open_output:
            on_open_output()
        return

    top = ctk.CTkToplevel(parent)
    top.withdraw()
    top.title("Audited Financial Volume Breakdown")
    top.geometry("560x420")
    top.resizable(False, False)
    top.transient(parent)


    sw, sh = top.winfo_screenwidth(), top.winfo_screenheight()
    cx, cy = max(0, int(sw / 2 - 560 / 2)), max(0, int(sh / 2 - 420 / 2))
    top.geometry(f"560x420+{cx}+{cy}")
    top.configure(fg_color=BG)

    content = ctk.CTkFrame(top, fg_color=PANEL, corner_radius=10, border_color=BORDER, border_width=1)
    content.pack(fill="both", expand=True, padx=16, pady=16)

    hdr = ctk.CTkFrame(content, fg_color="transparent")
    hdr.pack(fill="x", padx=16, pady=(16, 8))
    ctk.CTkLabel(hdr, text="💰 Reconciled Volume & Net Variance", font=(FONT_FAMILY, 14, "bold"), text_color=TEXT).pack(anchor="w")
    ctk.CTkLabel(hdr, text="Financial audit breakdown by bank from the latest reconciliation report.", font=(FONT_FAMILY, 10), text_color=MUTED).pack(anchor="w", pady=(2, 0))

    # Table container
    tbl_frame = ctk.CTkFrame(content, fg_color=PREVIEW_BG, corner_radius=8, border_color=BORDER, border_width=1)
    tbl_frame.pack(fill="both", expand=True, padx=16, pady=10)

    # Header Row
    h_row = ctk.CTkFrame(tbl_frame, fg_color="transparent", height=28)
    h_row.pack(fill="x", padx=8, pady=(8, 4))
    ctk.CTkLabel(h_row, text="Bank", font=(FONT_FAMILY, 10, "bold"), text_color=MUTED, width=70, anchor="w").pack(side="left")
    ctk.CTkLabel(h_row, text="Dates", font=(FONT_FAMILY, 10, "bold"), text_color=MUTED, width=50, anchor="center").pack(side="left")
    ctk.CTkLabel(h_row, text="Total Bank", font=(FONT_FAMILY, 10, "bold"), text_color=MUTED, width=110, anchor="e").pack(side="left")
    ctk.CTkLabel(h_row, text="Total Odoo", font=(FONT_FAMILY, 10, "bold"), text_color=MUTED, width=110, anchor="e").pack(side="left")
    ctk.CTkLabel(h_row, text="Variance", font=(FONT_FAMILY, 10, "bold"), text_color=MUTED, width=100, anchor="e").pack(side="left")

    # Separator
    tk.Frame(tbl_frame, bg=BORDER, height=1).pack(fill="x", padx=8, pady=2)

    tot_b = 0.0
    tot_o = 0.0
    tot_d = 0.0
    tot_dt_cnt = 0

    for b_name, s in sorted(stats.items()):
        b_val = s["tot_bank"]
        o_val = s["tot_odoo"]
        d_val = s["tot_diff"]
        d_cnt = len(s["dates"])
        tot_b += b_val
        tot_o += o_val
        tot_d += d_val
        tot_dt_cnt += d_cnt

        row_f = ctk.CTkFrame(tbl_frame, fg_color="transparent", height=24)
        row_f.pack(fill="x", padx=8, pady=2)
        ctk.CTkLabel(row_f, text=b_name.upper(), font=(FONT_FAMILY, 10, "bold"), text_color=TEXT, width=70, anchor="w").pack(side="left")
        ctk.CTkLabel(row_f, text=f"{d_cnt} d", font=(FONT_FAMILY, 10), text_color=MUTED, width=50, anchor="center").pack(side="left")
        ctk.CTkLabel(row_f, text=f"Rp {b_val:,.0f}", font=(FONT_MONO, 10), text_color=TEXT, width=110, anchor="e").pack(side="left")
        ctk.CTkLabel(row_f, text=f"Rp {o_val:,.0f}", font=(FONT_MONO, 10), text_color=TEXT, width=110, anchor="e").pack(side="left")
        d_color = SUCCESS if d_val < 1.0 else ERROR
        d_str = "Rp 0" if d_val < 1.0 else f"Rp {d_val:,.0f}"
        ctk.CTkLabel(row_f, text=d_str, font=(FONT_MONO, 10, "bold"), text_color=d_color, width=100, anchor="e").pack(side="left")

    # Total Row
    tk.Frame(tbl_frame, bg=BORDER, height=1).pack(fill="x", padx=8, pady=4)
    tot_row = ctk.CTkFrame(tbl_frame, fg_color="transparent", height=28)
    tot_row.pack(fill="x", padx=8, pady=(2, 6))
    ctk.CTkLabel(tot_row, text="TOTAL", font=(FONT_FAMILY, 10, "bold"), text_color=TEXT, width=70, anchor="w").pack(side="left")
    ctk.CTkLabel(tot_row, text=f"{tot_dt_cnt} d", font=(FONT_FAMILY, 10, "bold"), text_color=TEXT, width=50, anchor="center").pack(side="left")
    ctk.CTkLabel(tot_row, text=f"Rp {tot_b:,.0f}", font=(FONT_MONO, 10), text_color=TEXT, width=110, anchor="e").pack(side="left")
    ctk.CTkLabel(tot_row, text=f"Rp {tot_o:,.0f}", font=(FONT_MONO, 10), text_color=TEXT, width=110, anchor="e").pack(side="left")
    tot_d_color = SUCCESS if tot_d < 1.0 else ERROR
    tot_d_str = "Rp 0" if tot_d < 1.0 else f"Rp {tot_d:,.0f}"
    ctk.CTkLabel(tot_row, text=tot_d_str, font=(FONT_MONO, 10, "bold"), text_color=tot_d_color, width=100, anchor="e").pack(side="left")

    # Footer Buttons
    btn_box = ctk.CTkFrame(content, fg_color="transparent")
    btn_box.pack(fill="x", side="bottom", padx=16, pady=(8, 16))
    ctk.CTkButton(
        btn_box, text="Open Reconciliation Excel", height=32,
        fg_color=ACCENT, hover_color=ACCENT_DARK, font=(FONT_FAMILY, 10, "bold"),
        command=lambda: [top.destroy(), on_open_output() if on_open_output else None]
    ).pack(side="left")
    ctk.CTkButton(
        btn_box, text="Close", height=32, width=80,
        fg_color=WHITE, text_color=TEXT, border_color=BORDER, border_width=1,
        hover_color=PREVIEW_BG, font=(FONT_FAMILY, 10, "bold"),
        command=top.destroy
    ).pack(side="right")

    top.update_idletasks()
    top.deiconify()
    top.grab_set()
    return top
