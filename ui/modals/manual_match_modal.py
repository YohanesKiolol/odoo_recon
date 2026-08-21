"""Manual Match and Discrepancy Pair Reconciliation modal dialog."""
import glob
import os
import tkinter as tk
from tkinter import messagebox
import customtkinter as ctk
from pathlib import Path
from collections import defaultdict
from excel_writer import safe_load_workbook
from ui.theme import (
    BG, PANEL, SIDEBAR_BG, PREVIEW_BG, BORDER, BORDER_DARK,
    ACCENT, ACCENT_DARK, SUCCESS, ERROR, WARN, TEXT, MUTED, WHITE,
    FONT_FAMILY, _EMOJI_FONT
)
from ui.widgets.window_utils import _center_modal_on_parent, _open_path

def open_manual_match_modal(parent, output_dir: Path, log_write_fn=None, set_status_fn=None, on_open_journal=None):
    """Open the Manual Match & Discrepancy Reconciler modal dialog."""
    output_files = glob.glob(str(output_dir / "[Rr]econciliation_*.xlsx"))
    if not output_files:
        if set_status_fn:
            set_status_fn("No reconciliation file found. Run scan/recon first!", ERROR)
        return
    latest_file = max(output_files, key=os.path.getctime)

    try:
        wb = safe_load_workbook(latest_file, data_only=True)
        if "Differences" not in wb.sheetnames:
            if set_status_fn:
                set_status_fn("No 'Differences' sheet found in recon file", ERROR)
            wb.close()
            return

        ws = wb["Differences"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        if set_status_fn:
            set_status_fn(f"Error loading recon file: {e}", ERROR)
        return

    if len(rows) < 3:
        if set_status_fn:
            set_status_fn("Differences sheet is empty!", WARN)
        messagebox.showinfo("Manual Match", "Differences sheet is empty. No discrepancy items to reconcile.")
        return

    hdr_row = [str(v or "").strip().lower() for v in rows[2]]
    def _find_c(name, fallback):
        for i, h in enumerate(hdr_row):
            if name in h: return i
        return fallback

    c_date     = _find_c("date", 1)
    c_bank     = _find_c("bank", 2)
    c_journal  = _find_c("journal", 3)
    c_odo_num  = _find_c("odoo number", 4)
    c_ref      = _find_c("reference", 5)
    c_bank_num = _find_c("bank number", 6)
    c_file     = _find_c("filename", 7)
    c_bank_amt = _find_c("bank amount", 8)
    c_odo_amt  = _find_c("odoo amount", 9 if len(hdr_row) > 13 else 8)
    c_src      = _find_c("source", 10 if len(hdr_row) > 13 else 9)
    c_recon    = _find_c("reconciled", 11 if len(hdr_row) > 13 else 10)
    c_status   = _find_c("status", 12 if len(hdr_row) > 13 else 11)

    bank_items = []
    odo_items = []
    odo_by_key = defaultdict(list)

    for r_idx, r in enumerate(rows[3:], 4):
        if not r or len(r) <= c_status:
            continue
        st = str(r[c_status] or "").strip()
        if not st or st.startswith("DIFFERENCES") or st.startswith("Date:"):
            continue

        src = str(r[c_src] or "").strip()
        if src == "Bank":
            amt = float(r[c_bank_amt] or r[c_odo_amt] or 0.0)
        else:
            amt = float(r[c_odo_amt] or r[c_bank_amt] or 0.0)

        row_dict = {
            "row_idx": r_idx,
            "no": r[0],
            "date": str(r[c_date] or "").strip(),
            "bank": str(r[c_bank] or "").strip(),
            "journal": str(r[c_journal] or "").strip(),
            "number_odo": str(r[c_odo_num] or "").strip(),
            "reference": str(r[c_ref] or "").strip(),
            "number_bank": str(r[c_bank_num] or "").strip(),
            "filename": str(r[c_file] or "").strip(),
            "amount": amt,
            "source": src,
            "reconciled": str(r[c_recon] or "").strip(),
            "status": st,
        }
        if st == "Only in Bank":
            bank_items.append(row_dict)
        elif st == "Only in Odoo":
            odo_items.append(row_dict)
            odo_by_key[(row_dict["date"], row_dict["journal"].lower())].append(row_dict)

    if not bank_items and not odo_items:
        if set_status_fn:
            set_status_fn("No discrepancy items found in Differences sheet!", SUCCESS)
        messagebox.showinfo("Manual Match", "No unmatched discrepancy items found in the Differences sheet! All items are already reconciled or matched.")
        return

    candidate_pairs = []
    for b in bank_items:
        key = (b["date"], b["journal"].lower())
        for o in odo_by_key.get(key, []):
            diff = b["amount"] - o["amount"]
            abs_diff = abs(diff)
            max_val = max(b["amount"], o["amount"], 1.0)
            pct = (abs_diff / max_val) * 100.0
            
            s1 = str(int(round(b["amount"])))
            s2 = str(int(round(o["amount"])))
            is_transposition = (len(s1) == len(s2) and sorted(s1) == sorted(s2))

            if pct <= 3.0 or (pct <= 8.0 and abs_diff <= 50000) or is_transposition:
                candidate_pairs.append({
                    "bank": b,
                    "odoo": o,
                    "diff": diff,
                    "abs_diff": abs_diff,
                    "pct": pct,
                    "date": b["date"],
                    "journal": b["journal"],
                })

    def _parse_date_key(d_str: str):
        try:
            if "/" in d_str:
                parts = d_str.split("/")
                if len(parts) == 3:
                    return (int(parts[2]), int(parts[1]), int(parts[0]))
            elif "-" in d_str:
                parts = d_str.split("-")
                if len(parts) == 3:
                    return (int(parts[0]), int(parts[1]), int(parts[2]))
        except Exception:
            pass
        return (9999, 12, 31)

    bank_items.sort(key=lambda r: (_parse_date_key(r["date"]), r["row_idx"]))
    odo_items.sort(key=lambda r: (_parse_date_key(r["date"]), r["row_idx"]))
    candidate_pairs.sort(key=lambda x: (_parse_date_key(x["date"]), x["pct"], x["abs_diff"]))

    top = ctk.CTkToplevel(parent)
    top.withdraw()
    top.title("Manual Match — Reconcile Differences")
    top.minsize(860, 480)
    top.configure(fg_color=BG)
    top.transient(parent)
    _center_modal_on_parent(top, parent)

    active_matched = []
    auto_page = [0]
    CANDIDATES_PER_PAGE = 10

    hdr = ctk.CTkFrame(top, fg_color=PANEL, height=76, corner_radius=0, border_color=BORDER, border_width=1)
    hdr.pack(fill="x", side="top")
    hdr.pack_propagate(False)

    hdr_in = ctk.CTkFrame(hdr, fg_color="transparent")
    hdr_in.pack(fill="both", expand=True, padx=24, pady=14)

    ctk.CTkLabel(
        hdr_in, text="🧩 Manual Reconciliation & Pair Matching",
        font=(FONT_FAMILY, 16, "bold"), text_color=ACCENT
    ).pack(anchor="w")
    ctk.CTkLabel(
        hdr_in, text=f"Review candidate pairs or manually match Bank vs Odoo discrepancies. Loaded {len(bank_items)} Bank & {len(odo_items)} Odoo items.",
        font=(FONT_FAMILY, 10, "bold"), text_color=MUTED
    ).pack(anchor="w", pady=(2, 0))

    lbl_modal_msg = ctk.CTkLabel(
        hdr_in, text="",
        font=(FONT_FAMILY, 10, "bold"), text_color=ERROR
    )
    lbl_modal_msg.pack(anchor="w", pady=(2, 0))

    def show_modal_msg(msg: str, is_err=True):
        lbl_modal_msg.configure(text=msg, text_color=ERROR if is_err else SUCCESS)
        top.after(5000, lambda: lbl_modal_msg.configure(text=""))

    footer = ctk.CTkFrame(top, fg_color=SIDEBAR_BG, height=60, corner_radius=0, border_color=BORDER, border_width=1)
    footer.pack(fill="x", side="bottom")

    body_frame = ctk.CTkFrame(top, fg_color=BG, corner_radius=0)
    body_frame.pack(fill="both", expand=True, padx=20, pady=12)

    tab_bar = ctk.CTkFrame(body_frame, fg_color="transparent")
    tab_bar.pack(fill="x", pady=(0, 8))

    active_tab = {"mode": "auto"}

    btn_tab_auto = ctk.CTkButton(
        tab_bar, text=f"⚡ Smart Candidates ({len(candidate_pairs)})", height=32,
        fg_color=WHITE, hover_color=PREVIEW_BG, border_color=ACCENT, border_width=2,
        text_color=ACCENT, font=(FONT_FAMILY, 10, "bold")
    )
    btn_tab_auto.pack(side="left", padx=(0, 6))

    btn_tab_manual = ctk.CTkButton(
        tab_bar, text=f"🔍 Free-Form Selector (Bank: {len(bank_items)} | Odoo: {len(odo_items)})", height=32,
        fg_color=PANEL, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1,
        text_color=TEXT, font=(FONT_FAMILY, 10, "bold")
    )
    btn_tab_manual.pack(side="left")

    content_box = ctk.CTkFrame(body_frame, fg_color=PANEL, corner_radius=10, border_color=BORDER, border_width=1)
    content_box.pack(fill="both", expand=True)

    queue_card = ctk.CTkFrame(body_frame, fg_color=PANEL, corner_radius=10, border_color=BORDER, border_width=1, height=230)
    queue_card.pack(fill="x", pady=(10, 0))
    queue_card.pack_propagate(False)

    q_hdr = ctk.CTkFrame(queue_card, fg_color="transparent")
    q_hdr.pack(fill="x", padx=16, pady=(8, 2))

    lbl_q_title = ctk.CTkLabel(q_hdr, text="📋 Confirmed Matched Pairs Queue (0 pairs)", font=(FONT_FAMILY, 10, "bold"), text_color=TEXT)
    lbl_q_title.pack(side="left")

    q_scroll = ctk.CTkScrollableFrame(queue_card, fg_color=PREVIEW_BG, corner_radius=6, border_color=BORDER, border_width=1)
    q_scroll.pack(fill="both", expand=True, padx=12, pady=(0, 8))

    def render_queue():
        for w in q_scroll.winfo_children():
            w.destroy()
        lbl_q_title.configure(text=f"📋 Confirmed Matched Pairs Queue ({len(active_matched)} pair{'s' if len(active_matched)!=1 else ''})")
        btn_submit_matched.configure(text=f"💾 Update Recon & Open Journal ({len(active_matched)})")

        col_widths = [110, 220, 160, 160, 160, 90]
        for col, w in enumerate(col_widths):
            q_scroll.grid_columnconfigure(col, weight=1, minsize=w)

        if not active_matched:
            tk.Label(q_scroll, text="No matched pairs queued yet. Click '🔗 Match Pair' above to add items here.", bg=PREVIEW_BG, fg=MUTED, font=(FONT_FAMILY, 9, "bold")).pack(pady=10)
            return

        q_headers = ["Date", "Journal", "Bank Amount", "Odoo Amount", "Net Difference", "Action"]
        for col, h in enumerate(q_headers):
            tk.Label(q_scroll, text=h, bg=PREVIEW_BG, fg=MUTED, font=(FONT_FAMILY, 9, "bold")).grid(row=0, column=col, padx=12, pady=(4, 6), sticky="w")
        
        tk.Frame(q_scroll, bg=BORDER_DARK, height=1).grid(row=1, column=0, columnspan=len(q_headers), sticky="ew")

        for idx, pair in enumerate(active_matched, 2):
            b, o, diff = pair["bank"], pair["odoo"], pair["diff"]
            impact_color = ERROR if diff < 0 else WARN if diff > 0 else SUCCESS

            tk.Label(q_scroll, text=f"📅 {b['date']}", bg=PREVIEW_BG, fg=TEXT, font=(_EMOJI_FONT, 9, "bold")).grid(row=idx, column=0, padx=12, pady=2, sticky="w")
            tk.Label(q_scroll, text=f"🏦 {b['journal']}", bg=PREVIEW_BG, fg=ACCENT, font=(_EMOJI_FONT, 9, "bold")).grid(row=idx, column=1, padx=12, pady=2, sticky="w")
            tk.Label(q_scroll, text=f"Rp {b['amount']:,.0f}", bg=PREVIEW_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold")).grid(row=idx, column=2, padx=12, pady=2, sticky="w")
            tk.Label(q_scroll, text=f"Rp {o['amount']:,.0f}", bg=PREVIEW_BG, fg=ACCENT, font=(FONT_FAMILY, 10, "bold")).grid(row=idx, column=3, padx=12, pady=2, sticky="w")
            tk.Label(q_scroll, text=f"Rp {diff:,.0f}", bg=PREVIEW_BG, fg=impact_color, font=(FONT_FAMILY, 10, "bold")).grid(row=idx, column=4, padx=12, pady=2, sticky="w")

            def _unpair(p=pair):
                active_matched.remove(p)
                render_queue()
                if active_tab["mode"] == "auto":
                    render_auto_tab()
                else:
                    render_manual_tab()

            btn_unp = tk.Label(q_scroll, text="❌ Unpair", bg=PREVIEW_BG, fg=ERROR, cursor="hand2", font=(_EMOJI_FONT, 9, "bold"))
            btn_unp.bind("<Button-1>", lambda e, func=_unpair: func())
            btn_unp.grid(row=idx, column=5, padx=12, pady=2, sticky="w")

    tab_auto_frame = ctk.CTkFrame(content_box, fg_color="transparent")
    tab_manual_frame = ctk.CTkFrame(content_box, fg_color="transparent")
    tab_auto_frame.pack(fill="both", expand=True)

    def switch_to_auto():
        active_tab["mode"] = "auto"
        btn_tab_auto.configure(fg_color=WHITE, border_color=ACCENT, border_width=2, text_color=ACCENT)
        btn_tab_manual.configure(fg_color=PANEL, border_color=BORDER_DARK, border_width=1, text_color=TEXT)
        tab_manual_frame.pack_forget()
        tab_auto_frame.pack(fill="both", expand=True)

    def switch_to_manual():
        active_tab["mode"] = "manual"
        btn_tab_auto.configure(fg_color=PANEL, border_color=BORDER_DARK, border_width=1, text_color=TEXT)
        btn_tab_manual.configure(fg_color=WHITE, border_color=ACCENT, border_width=2, text_color=ACCENT)
        tab_auto_frame.pack_forget()
        tab_manual_frame.pack(fill="both", expand=True)
        render_manual_tab()

    btn_tab_auto.configure(command=switch_to_auto)
    btn_tab_manual.configure(command=switch_to_manual)

    def render_auto_tab():
        matched_bank_rows = {p["bank"]["row_idx"] for p in active_matched}
        matched_odo_rows = {p["odoo"]["row_idx"] for p in active_matched}

        avail_candidates = [
            c for c in candidate_pairs
            if c["bank"]["row_idx"] not in matched_bank_rows and c["odoo"]["row_idx"] not in matched_odo_rows
        ]

        for w in tab_auto_frame.winfo_children():
            w.destroy()

        if not avail_candidates:
            ctk.CTkLabel(tab_auto_frame, text="No candidate pairs remaining to match.", font=(FONT_FAMILY, 11, "bold"), text_color=MUTED).pack(pady=40)
            return

        total_pages = max(1, (len(avail_candidates) + CANDIDATES_PER_PAGE - 1) // CANDIDATES_PER_PAGE)
        if auto_page[0] >= total_pages:
            auto_page[0] = total_pages - 1

        p_bar = ctk.CTkFrame(tab_auto_frame, fg_color="transparent")
        p_bar.pack(fill="x", padx=14, pady=(8, 4))

        ctk.CTkLabel(p_bar, text=f"Showing candidates {auto_page[0]*CANDIDATES_PER_PAGE + 1}–{min((auto_page[0]+1)*CANDIDATES_PER_PAGE, len(avail_candidates))} of {len(avail_candidates)}", font=(FONT_FAMILY, 9, "bold"), text_color=MUTED).pack(side="left")

        p_controls = ctk.CTkFrame(p_bar, fg_color="transparent")
        p_controls.pack(side="right")

        def _prev_p():
            if auto_page[0] > 0:
                auto_page[0] -= 1
                render_auto_tab()

        def _next_p():
            if auto_page[0] < total_pages - 1:
                auto_page[0] += 1
                render_auto_tab()

        ctk.CTkButton(p_controls, text="◀ Prev", width=65, height=26, fg_color=PANEL, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1, text_color=TEXT, font=(FONT_FAMILY, 9, "bold"), state="normal" if auto_page[0]>0 else "disabled", command=_prev_p).pack(side="left", padx=2)
        ctk.CTkLabel(p_controls, text=f"Page {auto_page[0]+1} of {total_pages}", font=(FONT_FAMILY, 9, "bold"), text_color=TEXT).pack(side="left", padx=8)
        ctk.CTkButton(p_controls, text="Next ▶", width=65, height=26, fg_color=PANEL, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1, text_color=TEXT, font=(FONT_FAMILY, 9, "bold"), state="normal" if auto_page[0]<total_pages-1 else "disabled", command=_next_p).pack(side="left", padx=2)

        c_scroll = ctk.CTkScrollableFrame(tab_auto_frame, fg_color="transparent")
        c_scroll.pack(fill="both", expand=True, padx=12, pady=(0, 8))

        def _reset_scroll(scroll_frame):
            def _apply():
                try:
                    scroll_frame._parent_canvas.yview_moveto(0.0)
                except Exception:
                    pass
            _apply()
            top.after(10, _apply)
            top.after(50, _apply)

        _reset_scroll(c_scroll)

        page_items = avail_candidates[auto_page[0]*CANDIDATES_PER_PAGE : (auto_page[0]+1)*CANDIDATES_PER_PAGE]

        for c in page_items:
            b, o, diff = c["bank"], c["odoo"], c["diff"]
            
            card = tk.Frame(c_scroll, bg=WHITE, highlightbackground=BORDER_DARK, highlightcolor=BORDER_DARK, highlightthickness=1)
            card.pack(fill="x", pady=3, padx=4)

            c_inner = tk.Frame(card, bg=WHITE)
            c_inner.pack(fill="x", padx=12, pady=6)

            info_left = tk.Frame(c_inner, bg=WHITE)
            info_left.pack(side="left", fill="both", expand=True)

            txt_hdr = f"📅 {b['date']}  |  🏦 {b['journal']}"
            txt_dtl = f"Bank: Rp {b['amount']:,.0f} (Ref: {b['number_bank'] or 'N/A'})   •   Odoo: Rp {o['amount']:,.0f} (Doc: {o['number_odo'] or 'N/A'})"
            
            lbl_h = tk.Label(info_left, text=txt_hdr, font=(_EMOJI_FONT, 10, "bold"), fg=ACCENT, bg=WHITE, anchor="w")
            lbl_h.pack(fill="x", pady=(0, 2))

            lbl_d = tk.Label(info_left, text=txt_dtl, font=(FONT_FAMILY, 11, "bold"), fg=TEXT, bg=WHITE, anchor="w")
            lbl_d.pack(fill="x", pady=(0, 0))

            info_right = tk.Frame(c_inner, bg=WHITE)
            info_right.pack(side="right", anchor="e")

            var_color = SUCCESS if diff == 0 else WARN
            var_text = f"Variance: Rp {diff:,.0f}" if diff != 0 else "Exact Match"
            lbl_v = tk.Label(info_right, text=var_text, font=(FONT_FAMILY, 10, "bold"), fg=var_color, bg=WHITE, anchor="e")
            lbl_v.pack(side="left", padx=(0, 14))

            def _pair_auto(cand=c):
                o_r = cand["odoo"]
                o_recon = str(o_r.get("reconciled", "")).strip().lower()
                if o_recon in ("no", "false"):
                    show_modal_msg(
                        f"⚠️ Cannot match: Odoo payment ({o_r.get('number_odo') or o_r.get('amount')}) is not reconciled with invoice in Odoo (Reconciled: No)."
                    )
                    return
                active_matched.append(cand)
                show_modal_msg("✅ Matched pair added to queue.", is_err=False)
                render_queue()
                render_auto_tab()

            ctk.CTkButton(
                info_right, text="🔗 Match Pair", height=28, width=105,
                fg_color=ACCENT, hover_color=ACCENT_DARK, text_color=WHITE,
                font=(FONT_FAMILY, 10, "bold"), corner_radius=5,
                command=_pair_auto
            ).pack(side="left")

            for w in (card, c_inner, info_left, lbl_h, lbl_d, info_right, lbl_v):
                w.bind("<Enter>", lambda e, c=card, ci=c_inner, il=info_left, lh=lbl_h, ld=lbl_d, ir=info_right, lv=lbl_v: [
                    c.config(bg=PREVIEW_BG), ci.config(bg=PREVIEW_BG), il.config(bg=PREVIEW_BG), lh.config(bg=PREVIEW_BG), ld.config(bg=PREVIEW_BG), ir.config(bg=PREVIEW_BG), lv.config(bg=PREVIEW_BG)
                ])
                w.bind("<Leave>", lambda e, c=card, ci=c_inner, il=info_left, lh=lbl_h, ld=lbl_d, ir=info_right, lv=lbl_v: [
                    c.config(bg=WHITE), ci.config(bg=WHITE), il.config(bg=WHITE), lh.config(bg=WHITE), ld.config(bg=WHITE), ir.config(bg=WHITE), lv.config(bg=WHITE)
                ])

    def render_manual_tab():
        for w in tab_manual_frame.winfo_children():
            w.destroy()

        bar_sel = ctk.CTkFrame(tab_manual_frame, fg_color=PREVIEW_BG, height=52, corner_radius=0, border_color=BORDER, border_width=1)
        bar_sel.pack(fill="x", side="bottom")
        bar_sel.pack_propagate(False)

        split = ctk.CTkFrame(tab_manual_frame, fg_color="transparent")
        split.pack(fill="both", expand=True, padx=10, pady=10)
        split.rowconfigure(0, weight=1)
        split.columnconfigure(0, weight=1, uniform="equal_cols")
        split.columnconfigure(1, weight=1, uniform="equal_cols")

        matched_b_ids = {p["bank"]["row_idx"] for p in active_matched}
        matched_o_ids = {p["odoo"]["row_idx"] for p in active_matched}

        avail_b = [b for b in bank_items if b["row_idx"] not in matched_b_ids]
        avail_o = [o for o in odo_items if o["row_idx"] not in matched_o_ids]

        sel_b = [None]
        sel_o = [None]
        ITEMS_PAGE_SIZE = 25
        page_b = [0]
        page_o = [0]

        box_b = ctk.CTkFrame(split, fg_color=PREVIEW_BG, corner_radius=8, border_color=BORDER, border_width=1)
        box_b.grid(row=0, column=0, sticky="nsew", padx=(0, 4))

        hdr_box_b = ctk.CTkFrame(box_b, fg_color="transparent")
        hdr_box_b.pack(fill="x", padx=10, pady=(6, 2))
        lbl_title_b = ctk.CTkLabel(hdr_box_b, text=f"Bank Items ({len(avail_b)})", font=(FONT_FAMILY, 10, "bold"), text_color=TEXT)
        lbl_title_b.pack(side="left")

        p_bar_b = ctk.CTkFrame(hdr_box_b, fg_color="transparent")
        p_bar_b.pack(side="right")
        lbl_page_b = ctk.CTkLabel(p_bar_b, text="Page 1/1", font=(FONT_FAMILY, 9), text_color=MUTED)
        lbl_page_b.pack(side="left", padx=4)
        btn_prev_b = ctk.CTkButton(p_bar_b, text="◀", width=26, height=22, font=(FONT_FAMILY, 9, "bold"), fg_color=PANEL, hover_color=PREVIEW_BG, text_color=TEXT, border_width=1, border_color=BORDER_DARK)
        btn_prev_b.pack(side="left", padx=1)
        btn_next_b = ctk.CTkButton(p_bar_b, text="▶", width=26, height=22, font=(FONT_FAMILY, 9, "bold"), fg_color=PANEL, hover_color=PREVIEW_BG, text_color=TEXT, border_width=1, border_color=BORDER_DARK)
        btn_next_b.pack(side="left", padx=1)

        entry_search_b = ctk.CTkEntry(box_b, placeholder_text="🔍 Filter Bank by date/amount/ref...", height=28, font=(FONT_FAMILY, 9.5))
        entry_search_b.pack(fill="x", padx=8, pady=(0, 4))

        s_b = ctk.CTkScrollableFrame(box_b, fg_color="transparent")
        s_b.pack(fill="both", expand=True, padx=4, pady=4)

        box_o = ctk.CTkFrame(split, fg_color=PREVIEW_BG, corner_radius=8, border_color=BORDER, border_width=1)
        box_o.grid(row=0, column=1, sticky="nsew", padx=(4, 0))

        hdr_box_o = ctk.CTkFrame(box_o, fg_color="transparent")
        hdr_box_o.pack(fill="x", padx=10, pady=(6, 2))
        lbl_title_o = ctk.CTkLabel(hdr_box_o, text=f"Odoo Items ({len(avail_o)})", font=(FONT_FAMILY, 10, "bold"), text_color=TEXT)
        lbl_title_o.pack(side="left")

        p_bar_o = ctk.CTkFrame(hdr_box_o, fg_color="transparent")
        p_bar_o.pack(side="right")
        lbl_page_o = ctk.CTkLabel(p_bar_o, text="Page 1/1", font=(FONT_FAMILY, 9), text_color=MUTED)
        lbl_page_o.pack(side="left", padx=4)
        btn_prev_o = ctk.CTkButton(p_bar_o, text="◀", width=26, height=22, font=(FONT_FAMILY, 9, "bold"), fg_color=PANEL, hover_color=PREVIEW_BG, text_color=TEXT, border_width=1, border_color=BORDER_DARK)
        btn_prev_o.pack(side="left", padx=1)
        btn_next_o = ctk.CTkButton(p_bar_o, text="▶", width=26, height=22, font=(FONT_FAMILY, 9, "bold"), fg_color=PANEL, hover_color=PREVIEW_BG, text_color=TEXT, border_width=1, border_color=BORDER_DARK)
        btn_next_o.pack(side="left", padx=1)

        entry_search_o = ctk.CTkEntry(box_o, placeholder_text="🔍 Filter Odoo by date/amount/doc...", height=28, font=(FONT_FAMILY, 9.5))
        entry_search_o.pack(fill="x", padx=8, pady=(0, 4))

        s_o = ctk.CTkScrollableFrame(box_o, fg_color="transparent")
        s_o.pack(fill="both", expand=True, padx=4, pady=4)

        bar_in = ctk.CTkFrame(bar_sel, fg_color="transparent")
        bar_in.pack(fill="both", expand=True, padx=14, pady=6)

        lbl_sel_info = ctk.CTkLabel(bar_in, text="Select 1 Bank item and 1 Odoo item to calculate variance and pair.", font=(FONT_FAMILY, 10.5, "bold"), text_color=MUTED)
        lbl_sel_info.pack(side="left")

        def _update_sel_bar():
            if sel_b[0] and sel_o[0]:
                b_amt = sel_b[0]["amount"]
                o_amt = sel_o[0]["amount"]
                diff = b_amt - o_amt
                lbl_sel_info.configure(
                    text=f"Selected: Bank Rp {b_amt:,.0f}  |  Odoo Rp {o_amt:,.0f}   →   Net Difference: Rp {diff:,.0f}",
                    font=(FONT_FAMILY, 11.5, "bold"),
                    text_color=ACCENT
                )
                btn_pair_sel.configure(state="normal", fg_color=ACCENT)
            elif sel_b[0]:
                lbl_sel_info.configure(
                    text=f"Selected Bank: Rp {sel_b[0]['amount']:,.0f} ({sel_b[0]['date']} | {sel_b[0]['journal']}) — Select matching Odoo item",
                    font=(FONT_FAMILY, 10.5, "bold"),
                    text_color=ACCENT
                )
                btn_pair_sel.configure(state="disabled", fg_color=BORDER_DARK)
            elif sel_o[0]:
                lbl_sel_info.configure(
                    text=f"Selected Odoo: Rp {sel_o[0]['amount']:,.0f} ({sel_o[0]['date']} | {sel_o[0]['journal']}) — Select matching Bank item",
                    font=(FONT_FAMILY, 10.5, "bold"),
                    text_color=ACCENT
                )
                btn_pair_sel.configure(state="disabled", fg_color=BORDER_DARK)
            else:
                lbl_sel_info.configure(
                    text="Select 1 Bank item and 1 Odoo item to calculate variance and pair.",
                    font=(FONT_FAMILY, 10.5, "bold"),
                    text_color=MUTED
                )
                btn_pair_sel.configure(state="disabled", fg_color=BORDER_DARK)

        def _confirm_manual_pair():
            if sel_b[0] and sel_o[0]:
                b, o = sel_b[0], sel_o[0]
                o_recon = str(o.get("reconciled", "")).strip().lower()
                if o_recon in ("no", "false"):
                    show_modal_msg(
                        f"⚠️ Cannot match: Selected Odoo payment ({o.get('number_odo') or o.get('amount')}) is not reconciled with invoice in Odoo (Reconciled: No)."
                    )
                    return
                diff = b["amount"] - o["amount"]
                active_matched.append({
                    "bank": b,
                    "odoo": o,
                    "diff": diff,
                    "abs_diff": abs(diff),
                    "pct": (abs(diff)/max(b["amount"], 1.0))*100.0,
                    "date": b["date"],
                    "journal": b["journal"],
                })
                show_modal_msg("✅ Matched pair added to queue.", is_err=False)
                sel_b[0] = None
                sel_o[0] = None
                render_queue()
                render_auto_tab()
                render_manual_tab()

        btn_pair_sel = ctk.CTkButton(
            bar_in, text="🔗 Pair Selected Items", height=36,
            fg_color=BORDER_DARK, state="disabled", text_color=WHITE,
            font=(FONT_FAMILY, 10, "bold"), command=_confirm_manual_pair
        )
        btn_pair_sel.pack(side="right")

        def _toggle_b(bi):
            if sel_b[0] == bi:
                sel_b[0] = None
            else:
                sel_b[0] = bi
            _rebuild_lists()
            _update_sel_bar()

        def _toggle_o(oi):
            if sel_o[0] == oi:
                sel_o[0] = None
            else:
                sel_o[0] = oi
            _rebuild_lists()
            _update_sel_bar()

        def _render_native_card(parent_f, item, is_sel, is_bank, on_click):
            bg_col = ACCENT if is_sel else WHITE
            fg_hdr = WHITE if is_sel else ACCENT
            fg_txt = WHITE if is_sel else TEXT
            border_col = ACCENT_DARK if is_sel else BORDER_DARK

            card = tk.Frame(parent_f, bg=bg_col, highlightbackground=border_col, highlightcolor=border_col, highlightthickness=1, cursor="hand2")
            card.pack(fill="x", pady=2, padx=2)

            top_txt = f"📅 {item['date']}  |  🏦 {item['journal']}"
            if is_bank:
                bot_txt = f"Rp {item['amount']:,.0f}  (Ref: {item.get('number_bank') or 'N/A'})"
            else:
                recon_st = str(item.get("reconciled") or "Yes")
                recon_tag = f"  •  Reconciled: {recon_st}"
                bot_txt = f"Rp {item['amount']:,.0f}  (Doc: {item.get('number_odo') or 'N/A'}{recon_tag})"

            l_top = tk.Label(card, text=top_txt, font=(_EMOJI_FONT, 10, "bold"), fg=fg_hdr, bg=bg_col, anchor="w", cursor="hand2")
            l_top.pack(fill="x", padx=8, pady=(4, 1))

            l_bot = tk.Label(card, text=bot_txt, font=(FONT_FAMILY, 11, "bold"), fg=fg_txt, bg=bg_col, anchor="w", cursor="hand2")
            l_bot.pack(fill="x", padx=8, pady=(0, 4))

            for w in (card, l_top, l_bot):
                w.bind("<Button-1>", lambda e, func=on_click: func())
                if not is_sel:
                    w.bind("<Enter>", lambda e, c=card, t=l_top, b=l_bot: (c.config(bg=PREVIEW_BG), t.config(bg=PREVIEW_BG), b.config(bg=PREVIEW_BG)))
                    w.bind("<Leave>", lambda e, c=card, t=l_top, b=l_bot: (c.config(bg=WHITE), t.config(bg=WHITE), b.config(bg=WHITE)))

        def _rebuild_lists():
            q_b = entry_search_b.get().strip().lower()
            q_o = entry_search_o.get().strip().lower()

            if sel_b[0]:
                target_d = sel_b[0]["date"]
                target_j = sel_b[0]["journal"].strip().lower()
                base_o = [o for o in avail_o if o["date"] == target_d and o["journal"].strip().lower() == target_j]
            else:
                base_o = avail_o

            if sel_o[0]:
                target_d = sel_o[0]["date"]
                target_j = sel_o[0]["journal"].strip().lower()
                base_b = [b for b in avail_b if b["date"] == target_d and b["journal"].strip().lower() == target_j]
            else:
                base_b = avail_b

            if q_b:
                filtered_b = [b for b in base_b if q_b in f"{b['date']} {b['journal']} {b['amount']} {b.get('number_bank','')}".lower()]
            else:
                filtered_b = base_b

            if q_o:
                filtered_o = [o for o in base_o if q_o in f"{o['date']} {o['journal']} {o['amount']} {o.get('number_odo','')}".lower()]
            else:
                filtered_o = base_o

            tot_p_b = max(1, (len(filtered_b) + ITEMS_PAGE_SIZE - 1) // ITEMS_PAGE_SIZE)
            if page_b[0] >= tot_p_b: page_b[0] = tot_p_b - 1
            if page_b[0] < 0: page_b[0] = 0

            tot_p_o = max(1, (len(filtered_o) + ITEMS_PAGE_SIZE - 1) // ITEMS_PAGE_SIZE)
            if page_o[0] >= tot_p_o: page_o[0] = tot_p_o - 1
            if page_o[0] < 0: page_o[0] = 0

            lbl_title_b.configure(text=f"Bank Items ({len(filtered_b)} shown of {len(avail_b)})")
            lbl_page_b.configure(text=f"P {page_b[0]+1}/{tot_p_b}")
            btn_prev_b.configure(state="normal" if page_b[0] > 0 else "disabled", command=lambda: [_change_page_b(-1)])
            btn_next_b.configure(state="normal" if page_b[0] < tot_p_b - 1 else "disabled", command=lambda: [_change_page_b(1)])

            lbl_title_o.configure(text=f"Odoo Items ({len(filtered_o)} shown of {len(avail_o)})")
            lbl_page_o.configure(text=f"P {page_o[0]+1}/{tot_p_o}")
            btn_prev_o.configure(state="normal" if page_o[0] > 0 else "disabled", command=lambda: [_change_page_o(-1)])
            btn_next_o.configure(state="normal" if page_o[0] < tot_p_o - 1 else "disabled", command=lambda: [_change_page_o(1)])

            for w in s_b.winfo_children(): w.destroy()
            slice_b = filtered_b[page_b[0]*ITEMS_PAGE_SIZE : (page_b[0]+1)*ITEMS_PAGE_SIZE]
            if not slice_b:
                tk.Label(s_b, text="No matching Bank items", bg=PANEL, fg=MUTED, font=(FONT_FAMILY, 9)).pack(pady=12)
            else:
                for b_item in slice_b:
                    is_sel = (sel_b[0] == b_item)
                    _render_native_card(s_b, b_item, is_sel, True, lambda bi=b_item: _toggle_b(bi))

            for w in s_o.winfo_children(): w.destroy()
            slice_o = filtered_o[page_o[0]*ITEMS_PAGE_SIZE : (page_o[0]+1)*ITEMS_PAGE_SIZE]
            if not slice_o:
                tk.Label(s_o, text="No matching Odoo items", bg=PANEL, fg=MUTED, font=(FONT_FAMILY, 9)).pack(pady=12)
            else:
                for o_item in slice_o:
                    is_sel = (sel_o[0] == o_item)
                    _render_native_card(s_o, o_item, is_sel, False, lambda oi=o_item: _toggle_o(oi))

        def _reset_scroll_frame(scroll_frame):
            def _apply():
                try:
                    scroll_frame._parent_canvas.yview_moveto(0.0)
                except Exception:
                    pass
            _apply()
            top.after(10, _apply)
            top.after(50, _apply)

        def _change_page_b(delta):
            page_b[0] += delta
            _rebuild_lists()
            _reset_scroll_frame(s_b)

        def _change_page_o(delta):
            page_o[0] += delta
            _rebuild_lists()
            _reset_scroll_frame(s_o)

        entry_search_b.bind("<KeyRelease>", lambda e: [setattr(page_b, '__setitem__', (0, 0)) if hasattr(page_b, '__setitem__') else None, _rebuild_lists(), _reset_scroll_frame(s_b)])
        entry_search_o.bind("<KeyRelease>", lambda e: [setattr(page_o, '__setitem__', (0, 0)) if hasattr(page_o, '__setitem__') else None, _rebuild_lists(), _reset_scroll_frame(s_o)])

        _rebuild_lists()

    f_in = ctk.CTkFrame(footer, fg_color="transparent")
    f_in.pack(fill="both", expand=True, padx=24, pady=12)

    ctk.CTkButton(
        f_in, text="Close", height=36, width=90,
        fg_color=WHITE, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1,
        text_color=TEXT, font=(FONT_FAMILY, 10, "bold"), command=top.destroy
    ).pack(side="left")

    right_btn_frame = ctk.CTkFrame(f_in, fg_color="transparent")
    right_btn_frame.pack(side="right")

    def _update_recon_file(open_journal_modal=False):
        if not active_matched:
            show_modal_msg("⚠️ No matched pairs in queue! Click '🔗 Match Pair' or '🔗 Pair Selected Items' first.")
            return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font as XLFont
            from excel_writer import _normalize_date_str, safe_load_workbook
            from reconciler import STATUS_DONE, STATUS_BANK_ONLY, STATUS_ODO_ONLY

            wb = safe_load_workbook(latest_file, data_only=False)
            GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")

            next_pair_idx = 1
            if "Differences" in wb.sheetnames:
                import re as _re
                ws_scan = wb["Differences"]
                for r_scan in range(4, ws_scan.max_row + 1):
                    v = ws_scan.cell(r_scan, 13).value or ws_scan.cell(r_scan, 12).value
                    if v:
                        m = _re.search(r"M(\d+)", str(v))
                        if m:
                            next_pair_idx = max(next_pair_idx, int(m.group(1)) + 1)

            SKIP_SHEETS = {"Daily Summary", "Differences", "Mutation Summary",
                           "Admin Fee", "Excluded Payment", "Other Payment",
                           "Other Mutation", "Legend"}

            def _ensure_diff_header(ws, col_no, label="Difference"):
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter as _gcl
                hdr_cell = ws.cell(3, col_no)
                if not hdr_cell.value:
                    hdr_cell.value = label
                    hdr_cell.font = Font(bold=True, color="FFFFFF", size=10)
                    hdr_cell.fill = PatternFill("solid", fgColor="1F4E79")
                    hdr_cell.alignment = Alignment(horizontal="center", vertical="center")
                    _thin = Side(style="thin", color="CCCCCC")
                    hdr_cell.border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
                    ws.column_dimensions[_gcl(col_no)].width = 18

            diff_rows_to_delete = []

            for pair_idx, pair in enumerate(active_matched, next_pair_idx):
                pair_tag = f"Match (M{pair_idx:02d})"
                b_item = pair["bank"]
                o_item = pair["odoo"]
                diff   = pair["diff"]
                b_date_norm = _normalize_date_str(b_item["date"])
                o_date_norm = _normalize_date_str(o_item["date"])

                if "Differences" in wb.sheetnames:
                    ws_diff = wb["Differences"]
                    _ensure_diff_header(ws_diff, 14)
                    b_row_d = None
                    o_row_d = None
                    for r_idx in range(4, ws_diff.max_row + 1):
                        src_v = str(ws_diff.cell(r_idx, 11).value or ws_diff.cell(r_idx, 10).value or "").strip()
                        d_v   = _normalize_date_str(ws_diff.cell(r_idx, 2).value)
                        a_v9  = float(ws_diff.cell(r_idx, 9).value or 0)
                        a_v10 = float(ws_diff.cell(r_idx, 10).value or 0)
                        if src_v == "Bank" and d_v == b_date_norm and (abs(a_v9 - b_item["amount"]) < 0.01 or abs(a_v10 - b_item["amount"]) < 0.01):
                            b_row_d = r_idx
                        elif src_v == "Odoo" and d_v == o_date_norm and (abs(a_v9 - o_item["amount"]) < 0.01 or abs(a_v10 - o_item["amount"]) < 0.01):
                            o_row_d = r_idx

                    if b_row_d:
                        ws_diff.cell(b_row_d, 5).value = o_item.get("number_odo", "")
                        ws_diff.cell(b_row_d, 6).value = o_item.get("reference", "")
                        c_o = ws_diff.cell(b_row_d, 10)
                        c_o.value = o_item["amount"]
                        c_o.number_format = '#,##0.00'
                        ws_diff.cell(b_row_d, 11).value = "Manual"
                        ws_diff.cell(b_row_d, 12).value = o_item.get("reconciled", "Yes")
                        ws_diff.cell(b_row_d, 13).value = pair_tag
                        ws_diff.cell(b_row_d, 13).font = XLFont(bold=True, size=10)
                        c_diff = ws_diff.cell(b_row_d, 14)
                        c_diff.value = diff
                        c_diff.number_format = '#,##0.00'
                        for c in range(1, 15): ws_diff.cell(b_row_d, c).fill = GREEN_FILL
                    if o_row_d:
                        diff_rows_to_delete.append(o_row_d)

                target_j = b_item["journal"].strip().lower()
                for sname in wb.sheetnames:
                    if sname in SKIP_SHEETS:
                        continue
                    if target_j not in sname.lower() and sname.lower() not in target_j:
                        continue
                    ws_b = wb[sname]
                    _ensure_diff_header(ws_b, 13)
                    b_row_b = None
                    o_row_b = None
                    bank_rows_to_delete = []
                    for r in range(4, ws_b.max_row + 1):
                        d_v = _normalize_date_str(ws_b.cell(r, 2).value)
                        src_v = str(ws_b.cell(r, 10).value or ws_b.cell(r, 9).value or "").strip()
                        a_v8 = float(ws_b.cell(r, 8).value or 0.0)
                        a_v9 = float(ws_b.cell(r, 9).value or 0.0)
                        if (src_v in ("Bank", "Both", "Manual") or ws_b.cell(r, 12).value == STATUS_DONE) and d_v == b_date_norm and (abs(a_v8 - b_item["amount"]) < 0.01 or abs(a_v9 - b_item["amount"]) < 0.01):
                            b_row_b = r
                        elif src_v == "Odoo" and d_v == o_date_norm and (abs(a_v8 - o_item["amount"]) < 0.01 or abs(a_v9 - o_item["amount"]) < 0.01):
                            o_row_b = r

                    if b_row_b:
                        ws_b.cell(b_row_b, 4).value = o_item.get("number_odo", "")
                        ws_b.cell(b_row_b, 5).value = o_item.get("reference", "")
                        c_o = ws_b.cell(b_row_b, 9)
                        c_o.value = o_item["amount"]
                        c_o.number_format = '#,##0.00'
                        ws_b.cell(b_row_b, 10).value = "Manual"
                        ws_b.cell(b_row_b, 11).value = o_item.get("reconciled", "Yes")
                        ws_b.cell(b_row_b, 12).value = pair_tag
                        ws_b.cell(b_row_b, 12).font = XLFont(bold=True, size=10)
                        c_diff = ws_b.cell(b_row_b, 13)
                        c_diff.value = diff
                        c_diff.number_format = '#,##0.00'
                        for c in range(1, 14): ws_b.cell(b_row_b, c).fill = GREEN_FILL
                    if o_row_b:
                        bank_rows_to_delete.append(o_row_b)

                    for r_del in sorted(bank_rows_to_delete, reverse=True):
                        ws_b.delete_rows(r_del)

            for r_del in sorted(diff_rows_to_delete, reverse=True):
                ws_diff.delete_rows(r_del)

            try:
                from odoo_journal_creator import close_excel_window_for_file
                close_excel_window_for_file(os.path.basename(latest_file))
            except Exception:
                pass

            from odoo_journal_creator import safe_save_workbook
            saved_path = safe_save_workbook(wb, Path(latest_file))
            if saved_path:
                if log_write_fn:
                    log_write_fn(f"\n✅ Updated reconciliation report ({saved_path.name}) with {len(active_matched)} matched pairs!\n", "ok")
                if set_status_fn:
                    set_status_fn(f"Updated recon report with {len(active_matched)} matched pairs", SUCCESS)
                try:
                    from excel_writer import save_manual_matches
                    output_d = saved_path.parent
                    sidecar_entries = []
                    for pi, pair in enumerate(active_matched, next_pair_idx):
                        sidecar_entries.append({
                            "pair_tag":        f"Match (M{pi:02d})",
                            "bank_date":       pair["bank"]["date"],
                            "bank_journal":    pair["bank"]["journal"],
                            "bank_amount":     pair["bank"]["amount"],
                            "bank_number":     pair["bank"].get("number_bank", ""),
                            "odoo_date":       pair["odoo"]["date"],
                            "odoo_journal":    pair["odoo"]["journal"],
                            "odoo_amount":     pair["odoo"]["amount"],
                            "odoo_number":     pair["odoo"].get("number_odo", ""),
                            "odoo_reference":  pair["odoo"].get("reference", ""),
                            "odoo_reconciled": pair["odoo"].get("reconciled", "Yes"),
                            "diff":            pair["diff"],
                        })

                    save_manual_matches(output_d, sidecar_entries)
                    if log_write_fn:
                        log_write_fn(f"✅ Saved {len(sidecar_entries)} match(es) to manual_matches.json\n", "ok")
                except Exception as _se:
                    if log_write_fn:
                        log_write_fn(f"⚠️ Could not save manual_matches.json: {_se}\n", "warn")
                top.destroy()
                if open_journal_modal and on_open_journal:
                    on_open_journal()
                else:
                    _open_path(str(saved_path))
            else:
                if log_write_fn:
                    log_write_fn(f"\n⚠️ Could not save reconciliation report ({os.path.basename(latest_file)}). Please close Excel and try again.\n", "warn")
                show_modal_msg("⚠️ Could not save Excel file. Please close Excel and try again.")

        except Exception as e:
            if set_status_fn:
                set_status_fn(f"Failed to update recon report: {e}", ERROR)
            if log_write_fn:
                log_write_fn(f"\n❌ Error updating recon report: {e}\n", "err")
            show_modal_msg(f"❌ Error updating recon report: {e}")

    btn_update_recon = ctk.CTkButton(
        right_btn_frame, text=f"💾 Update Recon Only", height=38,
        fg_color=WHITE, hover_color=PREVIEW_BG, border_color=ACCENT, border_width=2,
        text_color=ACCENT, font=(FONT_FAMILY, 11, "bold"), command=lambda: _update_recon_file(open_journal_modal=False)
    )
    btn_update_recon.pack(side="left", padx=(0, 8))

    btn_submit_matched = ctk.CTkButton(
        right_btn_frame, text=f"💾 Update Recon & Open Journal (0)", height=38,
        fg_color=ACCENT, hover_color=ACCENT_DARK, text_color=WHITE,
        font=(FONT_FAMILY, 11, "bold"), command=lambda: _update_recon_file(open_journal_modal=True)
    )
    btn_submit_matched.pack(side="left")

    render_queue()
    render_auto_tab()
    switch_to_auto()
    top.update_idletasks()
    top.deiconify()
    top.grab_set()
    return top
