"""Sales Portal Discrepancies and Odoo Diagnostics modal dialogs."""
import glob
import os
import threading
import tkinter as tk
import customtkinter as ctk
from pathlib import Path
from openpyxl import load_workbook
from cloud_sync import is_cloud_configured, push_bank_discrepancies, fetch_discrepancies
import odoo_inspector
from ui.theme import (
    BG, PANEL, SIDEBAR_BG, PREVIEW_BG, BORDER, BORDER_DARK,
    ACCENT, ACCENT_DARK, SUCCESS, SUCCESS_DARK, ERROR, ERROR_LIGHT,
    WARN, TEXT, MUTED, WHITE, FONT_FAMILY, FONT_MONO,
    BANK_BADGE_COLS, TYPE_BADGE_INFO
)
from ui.widgets.window_utils import _center_modal_on_parent

def open_discrepancy_inspection_modal(item: dict, parent_win=None):
    """Open live Odoo inspection diagnostics modal for a discrepancy."""
    parent = parent_win if (parent_win and parent_win.winfo_exists()) else None
    dlg = ctk.CTkToplevel(parent)
    dlg.withdraw()
    dlg.title("Odoo Discrepancy Diagnostics (Finance)")
    dlg.minsize(760, 400)
    dlg.resizable(True, True)
    if parent:
        dlg.transient(parent)

    sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
    cx, cy = max(0, int(sw / 2 - 820 / 2)), max(0, int(sh / 2 - 460 / 2))
    dlg.geometry(f"820x460+{cx}+{cy}")
    dlg.configure(fg_color=BG)

    scroll_root = ctk.CTkScrollableFrame(dlg, fg_color="transparent")
    scroll_root.pack(fill="both", expand=True, padx=16, pady=16)

    content = ctk.CTkFrame(scroll_root, fg_color=PANEL, corner_radius=10, border_color=BORDER, border_width=1)
    content.pack(fill="both", expand=True)

    # 1. Header Info
    hdr = ctk.CTkFrame(content, fg_color=PREVIEW_BG, corner_radius=8, border_color=BORDER, border_width=1)
    hdr.pack(fill="x", padx=16, pady=(16, 12))

    amt = float(item.get("amount", 0.0))
    t_date = str(item.get("date", item.get("transaction_date", ""))).strip()
    b_name = str(item.get("bank", item.get("bank_name", ""))).upper()
    journal = str(item.get("journal", "")).strip()
    dtype = str(item.get("discrepancy_type") or "bank_only").strip()
    badge_map = {
        "bank_only":         ("🏦 Bank Only", "#4338CA", "#EEF2FF"),
        "odoo_only":         ("📦 Odoo Only", "#B45309", "#FEF3C7"),
        "unreconciled_odoo": ("⚠️ Unreconciled", "#BE123C", "#FFE4E6"),
    }
    t_lbl, t_fg, t_bg = badge_map.get(dtype, ("Discrepancy", TEXT, BG))

    ctk.CTkLabel(hdr, text=f"Rp {amt:,.2f}  •  {t_lbl}", font=(FONT_FAMILY, 16, "bold"), text_color=TEXT).pack(anchor="w", padx=14, pady=(10, 3))
    ctk.CTkLabel(hdr, text=f"{b_name} • {journal} • Date: {t_date}", font=(FONT_FAMILY, 11, "bold"), text_color=ACCENT).pack(anchor="w", padx=14)

    b_num = str(item.get("number_bank") or item.get("bank_number") or "-").strip()
    o_num = str(item.get("number_odo") or item.get("odoo_number") or item.get("invoice_no") or "").strip()
    ref_txt = f"Bank Ref: {b_num}"
    if o_num:
        ref_txt += f"  •  Odoo Doc: {o_num}"
    ctk.CTkLabel(hdr, text=ref_txt, font=(FONT_MONO, 11), text_color=MUTED).pack(anchor="w", padx=14, pady=(3, 10))

    # 2. Odoo Live Diagnostics Assistant
    diag_frame = ctk.CTkFrame(content, fg_color="#F8FAFC", corner_radius=8, border_color=BORDER_DARK, border_width=1)
    diag_frame.pack(fill="x", padx=16, pady=(0, 16))

    diag_hdr = ctk.CTkFrame(diag_frame, fg_color="transparent")
    diag_hdr.pack(fill="x", padx=14, pady=(10, 6))

    ctk.CTkLabel(diag_hdr, text="🔍 Odoo Live Diagnostics (Read-Only)", font=(FONT_FAMILY, 12, "bold"), text_color=ACCENT_DARK).pack(side="left")
    lbl_status = ctk.CTkLabel(diag_hdr, text="⏳ Checking Odoo...", font=(FONT_FAMILY, 11, "bold"), text_color=MUTED)
    lbl_status.pack(side="right")

    diag_body = ctk.CTkFrame(diag_frame, fg_color="transparent")
    diag_body.pack(fill="x", padx=14, pady=(0, 14))

    btn_bar = ctk.CTkFrame(content, fg_color="transparent")
    btn_bar.pack(fill="x", side="bottom", padx=16, pady=(0, 16))

    btn_close = ctk.CTkButton(
        btn_bar, text="Close", height=32, width=90,
        fg_color=WHITE, hover_color=PREVIEW_BG,
        border_color=BORDER_DARK, border_width=1,
        text_color=TEXT, font=(FONT_FAMILY, 10, "bold"),
        command=dlg.destroy
    )
    btn_close.pack(side="right")

    inspect_item = {
        "discrepancy_type": dtype,
        "amount": amt,
        "transaction_date": t_date,
        "odoo_number": o_num,
        "bank_number": b_num,
        "bank_name": b_name,
        "journal": journal
    }

    def _apply_diag_ui(res: dict):
        if not dlg.winfo_exists():
            return
        for w in diag_body.winfo_children():
            w.destroy()

        if not res.get("success"):
            lbl_status.configure(text="⚠️ Odoo Check Offline", text_color=WARN)
            err_msg = res.get("error") or "Could not connect to Odoo server."
            ctk.CTkLabel(diag_body, text=err_msg, font=(FONT_FAMILY, 11), text_color=MUTED, wraplength=640, justify="left").pack(anchor="w", pady=4)
            return

        lbl_status.configure(text="✅ Checked Live", text_color=SUCCESS)

        if dtype == "bank_only":
            summary = res.get("summary", "")
            cross_payments = res.get("cross_journal_payments", [])
            same_payments = res.get("same_journal_payments", [])
            invoices = res.get("invoices", [])

            ctk.CTkLabel(diag_body, text=summary, font=(FONT_FAMILY, 12, "bold"), text_color=TEXT, wraplength=640, justify="left").pack(anchor="w", pady=(0, 8))

            if cross_payments:
                cj_hdr = ctk.CTkFrame(diag_body, fg_color="transparent")
                cj_hdr.pack(fill="x", pady=(6, 2))
                ctk.CTkLabel(cj_hdr, text="🚨 Possible Wrong Journal Payment in Odoo:", font=(FONT_FAMILY, 11, "bold"), text_color=ERROR).pack(side="left")

                for cp in cross_payments:
                    cp_card = ctk.CTkFrame(diag_body, fg_color="#FEF2F2", corner_radius=8, border_color="#FECACA", border_width=1)
                    cp_card.pack(fill="x", pady=4)

                    top_row = ctk.CTkFrame(cp_card, fg_color="transparent")
                    top_row.pack(fill="x", padx=12, pady=(8, 2))
                    ctk.CTkLabel(top_row, text=f" 🔄 Recorded in {cp['actual_journal']} ", font=(FONT_FAMILY, 10, "bold"), text_color=WHITE, fg_color="#DC2626", corner_radius=4).pack(side="left", padx=(0, 8))
                    ctk.CTkLabel(top_row, text=cp['name'], font=(FONT_MONO, 13, "bold"), text_color=TEXT).pack(side="left")
                    ctk.CTkLabel(top_row, text=f"Rp {cp['amount']:,.2f}", font=(FONT_FAMILY, 13, "bold"), text_color=ERROR).pack(side="right")

                    desc = f"Customer: {cp['customer']}  •  Date: {cp['date']}  •  Ref: {cp['ref']}  •  State: {cp['state'].title()}"
                    ctk.CTkLabel(cp_card, text=desc, font=(FONT_FAMILY, 10, "bold"), text_color=MUTED).pack(anchor="w", padx=12, pady=(2, 4))
                    sugg = f"↳ 💡 Hint: Reassign journal from '{cp['actual_journal']}' to '{cp['expected_journal']}' in Odoo, or match via Cross-Journal."
                    ctk.CTkLabel(cp_card, text=sugg, font=(FONT_FAMILY, 10), text_color=TEXT, wraplength=620, justify="left").pack(anchor="w", padx=12, pady=(0, 8))

            if same_payments and not cross_payments:
                for sp in same_payments:
                    sp_card = ctk.CTkFrame(diag_body, fg_color=WHITE, corner_radius=8, border_color=BORDER, border_width=1)
                    sp_card.pack(fill="x", pady=4)
                    top_row = ctk.CTkFrame(sp_card, fg_color="transparent")
                    top_row.pack(fill="x", padx=12, pady=(8, 2))
                    ctk.CTkLabel(top_row, text=" 💳 Existing Payment ", font=(FONT_FAMILY, 10, "bold"), text_color=WHITE, fg_color=WARN, corner_radius=4).pack(side="left", padx=(0, 8))
                    ctk.CTkLabel(top_row, text=sp['name'], font=(FONT_MONO, 13, "bold"), text_color=TEXT).pack(side="left")
                    ctk.CTkLabel(top_row, text=f"Rp {sp['amount']:,.2f}", font=(FONT_FAMILY, 13, "bold"), text_color=TEXT).pack(side="right")
                    desc = f"Journal: {sp['actual_journal']}  •  Customer: {sp['customer']}  •  Date: {sp['date']}  •  State: {sp['state'].title()}"
                    ctk.CTkLabel(sp_card, text=desc, font=(FONT_FAMILY, 10, "bold"), text_color=MUTED).pack(anchor="w", padx=12, pady=(2, 8))

            if invoices:
                inv_hdr = ctk.CTkFrame(diag_body, fg_color="transparent")
                inv_hdr.pack(fill="x", pady=(8, 2))
                ctk.CTkLabel(inv_hdr, text="📋 Candidate Draft or Unpaid Invoices:", font=(FONT_FAMILY, 11, "bold"), text_color=ACCENT_DARK).pack(side="left")

                for inv in invoices:
                    inv_card = ctk.CTkFrame(diag_body, fg_color=WHITE, corner_radius=8, border_color=BORDER, border_width=1)
                    inv_card.pack(fill="x", pady=4)

                    top_row = ctk.CTkFrame(inv_card, fg_color="transparent")
                    top_row.pack(fill="x", padx=12, pady=(10, 4))

                    badge_color = SUCCESS if inv["status_code"] == "AVAILABLE_OPEN" else (WARN if inv["status_code"] == "DRAFT_INVOICE" else ERROR)
                    ctk.CTkLabel(top_row, text=f" {inv['badge']} ", font=(FONT_FAMILY, 10, "bold"), text_color=WHITE, fg_color=badge_color, corner_radius=4).pack(side="left", padx=(0, 8))
                    ctk.CTkLabel(top_row, text=inv['name'], font=(FONT_MONO, 13, "bold"), text_color=TEXT).pack(side="left")
                    ctk.CTkLabel(top_row, text=f"Rp {inv['amount']:,.2f}", font=(FONT_FAMILY, 13, "bold"), text_color=SUCCESS_DARK).pack(side="right")

                    desc = f"Customer: {inv['customer']}  •  Date: {inv['date']}"
                    ctk.CTkLabel(inv_card, text=desc, font=(FONT_FAMILY, 11, "bold"), text_color=MUTED).pack(anchor="w", padx=12, pady=(0, 4))
                    if inv.get("detail"):
                        ctk.CTkLabel(inv_card, text=f"↳ {inv['detail']}", font=(FONT_FAMILY, 11), text_color=TEXT, wraplength=620, justify="left").pack(anchor="w", padx=12, pady=(0, 10))

        elif dtype == "odoo_only":
            summary = res.get("summary", "")
            linked_invoices = res.get("linked_invoices", [])
            ctk.CTkLabel(diag_body, text=summary, font=(FONT_FAMILY, 12, "bold"), text_color=TEXT, wraplength=640, justify="left").pack(anchor="w", pady=(0, 8))

            if linked_invoices:
                for li in linked_invoices:
                    li_box = ctk.CTkFrame(diag_body, fg_color=WHITE, corner_radius=8, border_color=BORDER, border_width=1)
                    li_box.pack(fill="x", pady=4)

                    r1 = ctk.CTkFrame(li_box, fg_color="transparent")
                    r1.pack(fill="x", padx=12, pady=(10, 4))
                    ctk.CTkLabel(r1, text="🔗 Linked Invoice:", font=(FONT_FAMILY, 11, "bold"), text_color=ACCENT).pack(side="left", padx=(0, 6))
                    ctk.CTkLabel(r1, text=li['name'], font=(FONT_MONO, 13, "bold"), text_color=TEXT).pack(side="left")
                    ctk.CTkLabel(r1, text=f"Rp {li['amount']:,.2f}", font=(FONT_FAMILY, 13, "bold"), text_color=TEXT).pack(side="right")

                    desc = f"Customer: {li['customer']}  •  Invoice Date: {li['date']}  •  State: {li['state']}"
                    ctk.CTkLabel(li_box, text=desc, font=(FONT_FAMILY, 11, "bold"), text_color=MUTED).pack(anchor="w", padx=12, pady=(0, 10))
            else:
                no_inv_box = ctk.CTkFrame(diag_body, fg_color=WHITE, corner_radius=6, border_color=BORDER, border_width=1)
                no_inv_box.pack(fill="x", pady=3)
                ctk.CTkLabel(no_inv_box, text="⚪ No linked invoice found on this payment document in Odoo.", font=(FONT_FAMILY, 11, "bold"), text_color=MUTED).pack(anchor="w", padx=12, pady=10)

        elif dtype == "unreconciled_odoo":
            summary = res.get("summary", "")
            invoices = res.get("invoices", [])
            ctk.CTkLabel(diag_body, text=summary, font=(FONT_FAMILY, 12, "bold"), text_color=TEXT, wraplength=640, justify="left").pack(anchor="w", pady=(0, 8))

            if invoices:
                for inv in invoices:
                    inv_card = ctk.CTkFrame(diag_body, fg_color=WHITE, corner_radius=8, border_color=BORDER, border_width=1)
                    inv_card.pack(fill="x", pady=4)

                    r1 = ctk.CTkFrame(inv_card, fg_color="transparent")
                    r1.pack(fill="x", padx=12, pady=(10, 4))
                    ctk.CTkLabel(r1, text=f" {inv['badge']} ", font=(FONT_FAMILY, 10, "bold"), text_color=WHITE, fg_color=ERROR if inv.get("other_payments") else SUCCESS, corner_radius=4).pack(side="left", padx=(0, 8))
                    ctk.CTkLabel(r1, text=f"Invoice: {inv['name']}", font=(FONT_MONO, 13, "bold"), text_color=TEXT).pack(side="left")
                    ctk.CTkLabel(r1, text=f"Rp {inv['amount']:,.2f}", font=(FONT_FAMILY, 13, "bold"), text_color=TEXT).pack(side="right")

                    desc = f"Customer: {inv['customer']}  •  Date: {inv['date']}"
                    ctk.CTkLabel(inv_card, text=desc, font=(FONT_FAMILY, 11, "bold"), text_color=MUTED).pack(anchor="w", padx=12, pady=(0, 4))

                    if inv.get("other_payments"):
                        other_box = ctk.CTkFrame(inv_card, fg_color=ERROR_LIGHT, corner_radius=6, border_color=ERROR, border_width=1)
                        other_box.pack(fill="x", padx=12, pady=(4, 10))
                        ctk.CTkLabel(other_box, text="⚠️ Already linked to other payment(s):", font=(FONT_FAMILY, 11, "bold"), text_color=ERROR).pack(anchor="w", padx=8, pady=(6, 2))
                        for op in inv["other_payments"]:
                            op_txt = f"• {op['ref']}  |  Date: {op['date']}  |  Journal: {op['journal']}  |  Rp {op['amount']:,.2f}"
                            ctk.CTkLabel(other_box, text=op_txt, font=(FONT_FAMILY, 11, "bold"), text_color=TEXT, justify="left").pack(anchor="w", padx=12, pady=1)
                    else:
                        ctk.CTkLabel(inv_card, text="🟢 Invoice is not linked to any other customer payment.", font=(FONT_FAMILY, 11, "bold"), text_color=SUCCESS).pack(anchor="w", padx=12, pady=(2, 10))

        else:
            summary = res.get("summary", "")
            ctk.CTkLabel(diag_body, text=summary, font=(FONT_FAMILY, 9), text_color=MUTED).pack(anchor="w")

    dlg.update_idletasks()
    dlg.deiconify()
    dlg.grab_set()

    def _bg_inspect():
        try:
            res = odoo_inspector.inspect_discrepancy(inspect_item)
        except Exception as e:
            res = {"success": False, "error": str(e)}
        if dlg.winfo_exists():
            dlg.after(0, lambda: _apply_diag_ui(res))

    threading.Thread(target=_bg_inspect, daemon=True).start()
    return dlg


def open_sales_portal_modal(parent, output_dir: Path, email_var: tk.StringVar, log_write_fn=None, set_status_fn=None):
    """Open Share Discrepancies to Sales Portal modal."""
    if not is_cloud_configured():
        if log_write_fn:
            log_write_fn("\n⚠️ Supabase is not configured in .env. Please set SUPABASE_URL and SUPABASE_KEY first.\n", "warn")
        if set_status_fn:
            set_status_fn("Cloud sync not configured", WARN)
        return

    output_files = glob.glob(str(output_dir / "[Rr]econciliation_*.xlsx"))
    if not output_files:
        if set_status_fn:
            set_status_fn("No reconciliation file found. Run recon first!", ERROR)
        return
    latest_file = max(output_files, key=os.path.getctime)

    try:
        wb = load_workbook(latest_file, read_only=True, data_only=True)
        if "Differences" not in wb.sheetnames:
            if set_status_fn:
                set_status_fn("No 'Differences' sheet found", ERROR)
            wb.close()
            return

        valid_recon_days = set()
        if "Daily Summary" in wb.sheetnames:
            ws_sum = wb["Daily Summary"]
            for r_sum in ws_sum.iter_rows(values_only=True):
                if not r_sum or len(r_sum) < 7:
                    continue
                d_s = str(r_sum[1] or "").strip()
                b_s = str(r_sum[3] or "").strip().upper()
                j_s = str(r_sum[4] or "").strip()
                try:
                    tot_b = float(r_sum[5] or 0.0)
                    tot_o = float(r_sum[6] or 0.0)
                except (ValueError, TypeError):
                    continue
                if tot_b > 0 and tot_o > 0:
                    valid_recon_days.add((d_s, b_s, j_s))
                    valid_recon_days.add((d_s, j_s))

        ws = wb["Differences"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        if set_status_fn:
            set_status_fn(f"Error loading recon file: {e}", ERROR)
        return

    hdr_row = [str(cell or "").strip().lower() for cell in (rows[2] if len(rows) > 2 else [])]
    def _find_c(name, fallback):
        for i, h in enumerate(hdr_row):
            if name in h: return i
        return fallback

    c_date     = _find_c("date", 1)
    c_bank     = _find_c("bank", 2)
    c_journal  = _find_c("journal", 3)
    c_odo_num  = _find_c("odoo number", 4)
    c_inv_ref  = _find_c("reference", 5)
    c_bank_num = _find_c("bank number", 6)
    c_file     = _find_c("filename", 7)
    c_bank_amt = _find_c("bank amount", 8)
    c_odo_amt  = _find_c("odoo amount", 9)
    c_src      = _find_c("source", 10)
    c_recon    = _find_c("reconciled", 11)
    c_status   = _find_c("status", 12)

    discrepancy_items = []
    for r_idx, r in enumerate(rows[3:], 4):
        if not r or len(r) <= c_status:
            continue

        r_date = str(r[c_date] or "").strip()
        r_bank = str(r[c_bank] or "").strip().upper()
        r_journal = str(r[c_journal] or "").strip()

        if valid_recon_days and (r_date, r_bank, r_journal) not in valid_recon_days and (r_date, r_journal) not in valid_recon_days:
            continue

        st = str(r[c_status] or "").strip()
        src = str(r[c_src] or "").strip()
        recon = str(r[c_recon] or "").strip().lower() if len(r) > c_recon else ""
        b_amt = float(r[c_bank_amt] or 0.0) if len(r) > c_bank_amt and r[c_bank_amt] else 0.0
        o_amt = float(r[c_odo_amt] or 0.0) if len(r) > c_odo_amt and r[c_odo_amt] else 0.0

        if st == "Only in Bank" or src == "Bank":
            disc_type = "bank_only"
            amt = b_amt or o_amt
        elif recon in ("no", "false") and st != "Only in Bank" and src != "Bank":
            disc_type = "unreconciled_odoo"
            amt = o_amt or b_amt
        elif st == "Only in Odoo" or src == "Odoo":
            disc_type = "odoo_only"
            amt = o_amt or b_amt
        else:
            continue

        if amt <= 0.0:
            continue

        odo_num = str(r[c_odo_num] or "").strip() if len(r) > c_odo_num else ""
        inv_ref = str(r[c_inv_ref] or "").strip() if len(r) > c_inv_ref else ""
        bank_num = str(r[c_bank_num] or "").strip() if len(r) > c_bank_num else ""

        discrepancy_items.append({
            "row_idx": r_idx,
            "date": str(r[c_date] or "").strip(),
            "bank": str(r[c_bank] or "").strip(),
            "journal": str(r[c_journal] or "").strip(),
            "number_bank": bank_num,
            "number_odo": odo_num,
            "invoice_no": inv_ref,
            "is_reconciled": str(r[c_recon] or "Yes").strip() if len(r) > c_recon else "Yes",
            "discrepancy_type": disc_type,
            "filename": str(r[c_file] or "").strip() if len(r) > c_file else "",
            "amount": amt,
        })

    def _disc_sort_key(it):
        b = str(it.get("bank", "")).strip().upper()
        d_raw = str(it.get("date", "")).strip()
        if "/" in d_raw:
            pts = d_raw.split("/")
            if len(pts) == 3:
                d_norm = f"{pts[2]}{pts[1].zfill(2)}{pts[0].zfill(2)}"
            else:
                d_norm = d_raw
        else:
            d_norm = d_raw.replace("-", "")
        return (b, d_norm)

    discrepancy_items.sort(key=_disc_sort_key)
    for item in discrepancy_items:
        item["is_synced"] = False

    top = ctk.CTkToplevel(parent)
    top.withdraw()
    top.title("Share Discrepancies to Sales Portal")
    top.minsize(860, 480)
    top.resizable(True, True)
    top.configure(fg_color=BG)
    top.transient(parent)
    _center_modal_on_parent(top, parent)

    hdr_frame = ctk.CTkFrame(top, fg_color=PANEL, corner_radius=0, height=76, border_color=BORDER, border_width=1)
    hdr_frame.pack(fill="x")
    hdr_frame.pack_propagate(False)

    hi = ctk.CTkFrame(hdr_frame, fg_color="transparent")
    hi.pack(fill="both", expand=True, padx=24, pady=12)

    ctk.CTkLabel(hi, text="Share Discrepancies to Sales Portal", font=(FONT_FAMILY, 16, "bold"), text_color=ACCENT).pack(anchor="w")
    ctk.CTkLabel(hi, text="Review Bank Only, Odoo Only, and Unreconciled discrepancies to upload to the Sales Portal.", font=(FONT_FAMILY, 11, "bold"), text_color=MUTED).pack(anchor="w")

    filt_frame = ctk.CTkFrame(top, fg_color=SIDEBAR_BG, corner_radius=0, height=62, border_color=BORDER, border_width=1)
    filt_frame.pack(fill="x")
    filt_frame.pack_propagate(False)

    fi = ctk.CTkFrame(filt_frame, fg_color="transparent")
    fi.pack(fill="both", expand=True, padx=16, pady=10)

    f_left = ctk.CTkFrame(fi, fg_color="transparent")
    f_left.pack(side="left", fill="y")

    cur_type = ["ALL"]
    cur_bank = ["ALL"]
    cur_status = ["ALL"]
    search_q = [""]

    type_btn_map = {}
    for t_key, t_lbl in [
        ("ALL", "All Types"),
        ("bank_only", "🏦 Bank Only"),
        ("odoo_only", "📦 Odoo Only"),
        ("unreconciled_odoo", "⚠️ Unreconciled"),
    ]:
        tb = ctk.CTkButton(
            f_left, text=t_lbl, height=36, width=135 if t_key == "unreconciled_odoo" else (125 if t_key == "bank_only" else (120 if t_key == "odoo_only" else 95)),
            fg_color=ACCENT if t_key == "ALL" else WHITE,
            text_color=WHITE if t_key == "ALL" else TEXT,
            border_color=ACCENT if t_key == "ALL" else BORDER_DARK,
            border_width=1,
            hover_color=ACCENT_DARK if t_key == "ALL" else PREVIEW_BG,
            font=(FONT_FAMILY, 11, "bold"), corner_radius=6
        )
        tb.pack(side="left", padx=(0, 6))
        type_btn_map[t_key] = tb

    sep_m = tk.Frame(f_left, bg=BORDER_DARK, width=1, height=24)
    sep_m.pack(side="left", padx=8, pady=4)

    bank_btn_map = {}
    for bname in ["ALL", "BCA", "MANDIRI", "BRI"]:
        bb = ctk.CTkButton(
            f_left, text=bname, height=36, width=55 if bname != "MANDIRI" else 80,
            fg_color=ACCENT if bname == "ALL" else WHITE,
            text_color=WHITE if bname == "ALL" else TEXT,
            border_color=ACCENT if bname == "ALL" else BORDER_DARK,
            border_width=1, hover_color=ACCENT_DARK if bname == "ALL" else PREVIEW_BG,
            font=(FONT_FAMILY, 11, "bold"), corner_radius=6
        )
        bb.pack(side="left", padx=(0, 4))
        bank_btn_map[bname] = bb

    f_right = ctk.CTkFrame(fi, fg_color="transparent")
    f_right.pack(side="right", fill="y")

    search_v = tk.StringVar()
    search_ent = ctk.CTkEntry(
        f_right, textvariable=search_v, placeholder_text="🔍 Search...",
        height=36, width=180, corner_radius=6, border_color=BORDER_DARK, fg_color=WHITE, text_color=TEXT,
        font=(FONT_FAMILY, 11, "bold")
    )
    search_ent.pack(side="right", padx=(8, 0))

    btn_st_all = ctk.CTkButton(
        f_right, text="All", height=36, width=65,
        fg_color=ACCENT, text_color=WHITE, border_color=ACCENT, border_width=1,
        hover_color=ACCENT_DARK, font=(FONT_FAMILY, 11, "bold"), corner_radius=6
    )
    btn_st_all.pack(side="left", padx=(0, 4))
    btn_st_ready = ctk.CTkButton(
        f_right, text="Ready to Send", height=36, width=120,
        fg_color=WHITE, text_color=TEXT, border_color=BORDER_DARK, border_width=1,
        hover_color=PREVIEW_BG, font=(FONT_FAMILY, 11, "bold"), corner_radius=6
    )
    btn_st_ready.pack(side="left", padx=(0, 4))
    btn_st_synced = ctk.CTkButton(
        f_right, text="Synced", height=36, width=85,
        fg_color=WHITE, text_color=TEXT, border_color=BORDER_DARK, border_width=1,
        hover_color=PREVIEW_BG, font=(FONT_FAMILY, 11, "bold"), corner_radius=6
    )
    btn_st_synced.pack(side="left", padx=(0, 4))

    scroll_cards = ctk.CTkScrollableFrame(top, fg_color="transparent", scrollbar_button_color=BORDER, scrollbar_button_hover_color=MUTED)
    scroll_cards.pack(fill="both", expand=True, padx=20, pady=8)

    footer = ctk.CTkFrame(top, fg_color=PANEL, corner_radius=0, height=64, border_color=BORDER, border_width=1)
    footer.pack(fill="x", side="bottom")
    footer.pack_propagate(False)

    foot_inner = ctk.CTkFrame(footer, fg_color="transparent")
    foot_inner.pack(fill="both", expand=True, padx=20, pady=10)

    btn_select_all = ctk.CTkButton(foot_inner, text="Select All Available", height=38, width=175, fg_color=WHITE, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1, text_color=TEXT, font=(FONT_FAMILY, 11, "bold"), corner_radius=6)
    btn_select_all.pack(side="left")

    lbl_summary = ctk.CTkLabel(foot_inner, text="Selected 0 items", font=(FONT_FAMILY, 12, "bold"), text_color=TEXT)
    lbl_summary.pack(side="left", padx=(16, 0))

    def _close_modal():
        try:
            top.grab_release()
        except Exception:
            pass
        try:
            top.destroy()
        except Exception:
            pass

    top.protocol("WM_DELETE_WINDOW", _close_modal)
    btn_close = ctk.CTkButton(foot_inner, text="Close", height=38, width=90, fg_color=WHITE, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1, text_color=TEXT, font=(FONT_FAMILY, 11, "bold"), corner_radius=6, command=_close_modal)
    btn_close.pack(side="right")

    btn_submit = ctk.CTkButton(foot_inner, text="Send Selected to Sales Portal (0)", height=38, fg_color=BORDER_DARK, hover_color=ACCENT_DARK, text_color=WHITE, font=(FONT_FAMILY, 12, "bold"), corner_radius=6, state="disabled")
    btn_submit.pack(side="right", padx=(0, 12))

    item_checks = {}
    for idx, item in enumerate(discrepancy_items):
        item_checks[idx] = tk.BooleanVar(value=False)

    def _update_counts():
        selected_count = 0
        selected_sum = 0.0
        available_count = 0
        for idx, item in enumerate(discrepancy_items):
            if not item["is_synced"]:
                available_count += 1
                if item_checks[idx].get():
                    selected_count += 1
                    selected_sum += item["amount"]
        lbl_summary.configure(text=f"Selected {selected_count} item(s) • Total: Rp {selected_sum:,.2f}")
        btn_submit.configure(text=f"Send Selected to Sales Portal ({selected_count})")
        if selected_count == 0:
            btn_submit.configure(state="disabled", fg_color=BORDER_DARK)
        else:
            btn_submit.configure(state="normal", fg_color=ACCENT)

        if available_count > 0 and selected_count == available_count:
            btn_select_all.configure(text="Deselect All")
        else:
            btn_select_all.configure(text="Select All Available")

    def _toggle_select_all():
        available = [i for i, item in enumerate(discrepancy_items) if not item["is_synced"]]
        all_on = all(item_checks[i].get() for i in available) if available else False
        new_val = not all_on
        for i in available:
            item_checks[i].set(new_val)
        _update_counts()

    btn_select_all.configure(command=_toggle_select_all)

    def _make_cb(parent, variable, bg_color, command=None):
        s = 18
        r = 3
        m = 1
        x, y = m, m
        w, h = s - 2*m, s - 2*m
        d = 2 * r

        cv = tk.Canvas(parent, width=s, height=s, bg=bg_color, highlightthickness=0, cursor="hand2")

        def _rrect(fill, outline, bw):
            cv.delete("all")
            if fill and fill != bg_color:
                cv.create_rectangle(x+r, y, x+w-r, y+h, fill=fill, outline="")
                cv.create_rectangle(x, y+r, x+w, y+h-r, fill=fill, outline="")
                cv.create_oval(x, y, x+d, y+d, fill=fill, outline="")
                cv.create_oval(x+w-d, y, x+w, y+d, fill=fill, outline="")
                cv.create_oval(x, y+h-d, x+d, y+h, fill=fill, outline="")
                cv.create_oval(x+w-d, y+h-d, x+w, y+h, fill=fill, outline="")
            cv.create_arc(x, y, x+d, y+d, start=90, extent=90, style="arc", outline=outline, width=bw)
            cv.create_arc(x+w-d, y, x+w, y+d, start=0, extent=90, style="arc", outline=outline, width=bw)
            cv.create_arc(x, y+h-d, x+d, y+h, start=180, extent=90, style="arc", outline=outline, width=bw)
            cv.create_arc(x+w-d, y+h-d, x+w, y+h, start=270, extent=90, style="arc", outline=outline, width=bw)
            cv.create_line(x+r, y, x+w-r, y, fill=outline, width=bw)
            cv.create_line(x+r, y+h, x+w-r, y+h, fill=outline, width=bw)
            cv.create_line(x, y+r, x, y+h-r, fill=outline, width=bw)
            cv.create_line(x+w, y+r, x+w, y+h-r, fill=outline, width=bw)

        def _draw(*_):
            cv.delete("all")
            if variable.get():
                _rrect(fill=ACCENT, outline=ACCENT_DARK, bw=1)
                cv.create_line(4, 9, 7, 13, fill=WHITE, width=2, capstyle="round", joinstyle="round")
                cv.create_line(7, 13, 14, 5, fill=WHITE, width=2, capstyle="round", joinstyle="round")
            else:
                _rrect(fill=bg_color, outline="#94A3B8", bw=2)

        def _toggle(e):
            variable.set(not variable.get())
            _draw()
            if command:
                command()

        cv.bind("<Button-1>", _toggle)
        variable.trace_add("write", _draw)
        _draw()
        return cv

    display_limit = [100]

    def _render_modal_cards():
        for w in scroll_cards.winfo_children():
            w.destroy()

        visible = []
        for idx, item in enumerate(discrepancy_items):
            if cur_type[0] != "ALL" and item["discrepancy_type"] != cur_type[0]:
                continue
            if cur_status[0] == "READY" and item["is_synced"]:
                continue
            if cur_status[0] == "SYNCED" and not item["is_synced"]:
                continue

            bname = item["bank"].upper()
            if cur_bank[0] != "ALL" and bname != cur_bank[0]:
                continue

            if search_q[0]:
                s_str = f"{item['bank']} {item['journal']} {item['number_bank']} {item['number_odo']} {item['invoice_no']} {item['discrepancy_type']} {item['amount']}".lower()
                if search_q[0] not in s_str:
                    continue

            visible.append((idx, item))

        ready_cnt = sum(1 for it in discrepancy_items if not it["is_synced"])
        synced_cnt = sum(1 for it in discrepancy_items if it["is_synced"])
        btn_st_ready.configure(text=f"Ready ({ready_cnt})")
        btn_st_synced.configure(text=f"Synced ({synced_cnt})")
        btn_st_all.configure(text=f"All ({len(discrepancy_items)})")

        if not visible:
            emp = tk.Frame(scroll_cards, bg=PANEL, highlightbackground=BORDER, highlightthickness=1)
            emp.pack(fill="x", padx=16, pady=20)
            tk.Label(emp, text="No items match current filter criteria.", bg=PANEL, fg=MUTED, font=(FONT_FAMILY, 10, "bold")).pack(pady=16)
            return

        slice_items = visible[:display_limit[0]]

        for idx, item in slice_items:
            card = tk.Frame(scroll_cards, bg=PANEL, highlightbackground=BORDER, highlightthickness=1, bd=0)
            card.pack(fill="x", padx=4, pady=3)
            
            card.columnconfigure(0, minsize=34)
            card.columnconfigure(1, minsize=118)
            card.columnconfigure(2, minsize=120)
            card.columnconfigure(3, minsize=80)
            card.columnconfigure(4, minsize=95)
            card.columnconfigure(5, minsize=140)
            card.columnconfigure(6, weight=1)
            card.columnconfigure(7, minsize=155)
            card.columnconfigure(8, minsize=120)

            if not item["is_synced"]:
                cb = _make_cb(card, item_checks[idx], PANEL, command=_update_counts)
                cb.grid(row=0, column=0, padx=(10, 4), pady=7, sticky="w")
            else:
                lbl_lock = tk.Label(card, text="Synced", bg=PANEL, fg=MUTED, font=(FONT_FAMILY, 9, "bold"))
                lbl_lock.grid(row=0, column=0, padx=(8, 4), pady=7, sticky="w")

            btn_chk = ctk.CTkButton(
                card, text="🔍 Check Odoo", width=104, height=28,
                fg_color=WHITE, hover_color=PREVIEW_BG,
                border_color=BORDER_DARK, border_width=1,
                text_color=TEXT, font=(FONT_FAMILY, 9, "bold"), corner_radius=6,
                command=lambda it=item: open_discrepancy_inspection_modal(it, parent_win=top)
            )
            btn_chk.grid(row=0, column=1, padx=(0, 8), pady=4, sticky="w")

            t_lbl, t_fg, t_bg = TYPE_BADGE_INFO.get(item["discrepancy_type"], ("Discrepancy", TEXT, BG))
            lbl_tbadge = tk.Label(card, text=f" {t_lbl} ", bg=t_bg, fg=t_fg, font=(FONT_FAMILY, 9, "bold"), padx=5, pady=3, width=14, anchor="center")
            lbl_tbadge.grid(row=0, column=2, padx=(0, 6), pady=7, sticky="w")

            bname = item["bank"].upper()
            fg_b, bg_b = BANK_BADGE_COLS.get(bname, BANK_BADGE_COLS["OTHER"])
            lbl_badge = tk.Label(card, text=f" {bname} ", bg=bg_b, fg=fg_b, font=(FONT_FAMILY, 9, "bold"), padx=5, pady=3, width=7, anchor="center")
            lbl_badge.grid(row=0, column=3, padx=(0, 6), pady=7, sticky="w")

            lbl_d = tk.Label(card, text=item["date"], bg=PANEL, fg=TEXT, font=(FONT_FAMILY, 10, "bold"), width=10, anchor="w")
            lbl_d.grid(row=0, column=4, padx=(0, 6), pady=7, sticky="w")

            lbl_j = tk.Label(card, text=item["journal"], bg=PANEL, fg=MUTED, font=(FONT_FAMILY, 9), width=16, anchor="w")
            lbl_j.grid(row=0, column=5, padx=(0, 6), pady=7, sticky="w")

            if item["discrepancy_type"] == "bank_only":
                b_ref = item["number_bank"] or "-"
                ref_text = f"Ref: {b_ref}"
            elif item["discrepancy_type"] == "unreconciled_odoo":
                o_doc = item["number_odo"] or item["invoice_no"] or "-"
                ref_text = f"Unrecon: {o_doc}"
            else:
                o_doc = item["number_odo"] or item["invoice_no"] or "-"
                ref_text = f"Doc: {o_doc}"

            lbl_ref = tk.Label(card, text=ref_text, bg=PANEL, fg=TEXT, font=(FONT_MONO, 10, "bold"), anchor="w")
            lbl_ref.grid(row=0, column=6, padx=(0, 10), pady=7, sticky="ew")

            lbl_amt = tk.Label(card, text=f"Rp {item['amount']:,.2f}", bg=PANEL, fg=TEXT, font=(FONT_FAMILY, 11, "bold"), width=16, anchor="e")
            lbl_amt.grid(row=0, column=7, padx=(0, 10), pady=7, sticky="e")

            if item["is_synced"]:
                lbl_st = tk.Label(card, text="● Synced", bg=PANEL, fg="#059669", font=(FONT_FAMILY, 9, "bold"), width=12, anchor="w")
            else:
                lbl_st = tk.Label(card, text="● Ready to Sync", bg=PANEL, fg="#D97706", font=(FONT_FAMILY, 9, "bold"), width=12, anchor="w")
            lbl_st.grid(row=0, column=8, padx=(0, 10), pady=7, sticky="w")

            labels_hover = [lbl_d, lbl_j, lbl_ref, lbl_amt, lbl_st]
            if item["is_synced"]:
                labels_hover.append(lbl_lock)

            def _make_hover(c_widget=card, labels=labels_hover, it=item):
                def _enter(e):
                    c_widget.config(bg=PREVIEW_BG, highlightbackground=ACCENT)
                    for l in labels: l.config(bg=PREVIEW_BG)
                def _leave(e):
                    c_widget.config(bg=PANEL, highlightbackground=BORDER)
                    for l in labels: l.config(bg=PANEL)
                def _on_double_click(e):
                    open_discrepancy_inspection_modal(it, parent_win=top)
                c_widget.bind("<Enter>", _enter)
                c_widget.bind("<Leave>", _leave)
                c_widget.bind("<Double-Button-1>", _on_double_click)
                for l in labels:
                    l.bind("<Double-Button-1>", _on_double_click)
            _make_hover()

        if len(visible) > len(slice_items):
            remaining = len(visible) - len(slice_items)
            more_frame = tk.Frame(scroll_cards, bg=BG)
            more_frame.pack(fill="x", pady=10)
            def _show_more():
                display_limit[0] += 100
                _render_modal_cards()
            btn_more = ctk.CTkButton(
                more_frame, text=f"Load Next 100 Items (Showing {len(slice_items)} of {len(visible)} items — {remaining} remaining)",
                height=32, fg_color=WHITE, hover_color=PREVIEW_BG, border_color=BORDER_DARK, border_width=1,
                text_color=TEXT, font=(FONT_FAMILY, 9, "bold"), command=_show_more
            )
            btn_more.pack(pady=4)

        _update_counts()

        def _reset_scroll():
            try:
                scroll_cards._parent_canvas.yview_moveto(0.0)
            except Exception:
                pass
        _reset_scroll()
        top.after(10, _reset_scroll)
        top.after(50, _reset_scroll)

    def _switch_type(tk_val):
        display_limit[0] = 100
        cur_type[0] = tk_val
        for k, btn in type_btn_map.items():
            if k == tk_val:
                btn.configure(fg_color=ACCENT, text_color=WHITE, border_color=ACCENT, hover_color=ACCENT_DARK)
            else:
                btn.configure(fg_color=WHITE, text_color=TEXT, border_color=BORDER_DARK, hover_color=PREVIEW_BG)
        _render_modal_cards()

    for tk_val, btn in type_btn_map.items():
        btn.configure(command=lambda t=tk_val: _switch_type(t))

    def _switch_status(st):
        display_limit[0] = 100
        cur_status[0] = st
        btn_st_ready.configure(
            fg_color=ACCENT if st == "READY" else WHITE,
            text_color=WHITE if st == "READY" else TEXT,
            border_color=ACCENT if st == "READY" else BORDER_DARK,
            hover_color=ACCENT_DARK if st == "READY" else PREVIEW_BG
        )
        btn_st_synced.configure(
            fg_color=ACCENT if st == "SYNCED" else WHITE,
            text_color=WHITE if st == "SYNCED" else TEXT,
            border_color=ACCENT if st == "SYNCED" else BORDER_DARK,
            hover_color=ACCENT_DARK if st == "SYNCED" else PREVIEW_BG
        )
        btn_st_all.configure(
            fg_color=ACCENT if st == "ALL" else WHITE,
            text_color=WHITE if st == "ALL" else TEXT,
            border_color=ACCENT if st == "ALL" else BORDER_DARK,
            hover_color=ACCENT_DARK if st == "ALL" else PREVIEW_BG
        )
        _render_modal_cards()

    btn_st_ready.configure(command=lambda: _switch_status("READY"))
    btn_st_synced.configure(command=lambda: _switch_status("SYNCED"))
    btn_st_all.configure(command=lambda: _switch_status("ALL"))

    def _switch_bank(bn):
        display_limit[0] = 100
        cur_bank[0] = bn
        for k, btn in bank_btn_map.items():
            if k == bn:
                btn.configure(fg_color=ACCENT, text_color=WHITE, border_color=ACCENT)
            else:
                btn.configure(fg_color=WHITE, text_color=TEXT, border_color=BORDER_DARK)
        _render_modal_cards()

    for bn, btn in bank_btn_map.items():
        btn.configure(command=lambda b=bn: _switch_bank(b))

    def _on_search_change(*args):
        display_limit[0] = 100
        search_q[0] = search_v.get().strip().lower()
        _render_modal_cards()

    search_v.trace_add("write", _on_search_change)

    def _do_push_to_sales():
        to_send = []
        for idx, item in enumerate(discrepancy_items):
            if not item["is_synced"] and item_checks[idx].get():
                to_send.append(item)

        if not to_send:
            return

        btn_submit.configure(state="disabled", text="⏳ Sending to Cloud...")
        top.update_idletasks()

        recon_date = to_send[0]["date"] if to_send else ""
        user_prof = email_var.get() or "Desktop-User"
        res = push_bank_discrepancies(to_send, recon_date=recon_date, user_profile=user_prof)

        if res.get("success"):
            if log_write_fn:
                log_write_fn(f"\n✅ Uploaded {res.get('count', len(to_send))} discrepancies to Cloud Portal!\n", "ok")
            if set_status_fn:
                set_status_fn(f"Sent {res.get('count')} items to Cloud Portal", SUCCESS)
            for item in to_send:
                item["is_synced"] = True
            _switch_status("SYNCED")
            _render_modal_cards()
        else:
            if log_write_fn:
                log_write_fn(f"\n❌ Failed to send to cloud: {res.get('error')}\n", "err")
            if set_status_fn:
                set_status_fn(f"Upload failed: {res.get('error')}", ERROR)
            btn_submit.configure(state="normal", text="📤 Send Selected to Sales Portal")

    btn_submit.configure(command=_do_push_to_sales)
    _render_modal_cards()
    _update_counts()

    top.update_idletasks()
    top.deiconify()
    top.grab_set()

    def _bg_check_synced():
        try:
            existing = fetch_discrepancies(limit=1000)
            if not existing or not top.winfo_exists():
                return
            cloud_keys = set()
            for ci in existing:
                d = str(ci.get("transaction_date", "")).strip()
                b = str(ci.get("bank_name", "")).strip().upper()
                num = str(ci.get("bank_number", "")).strip()
                onum = str(ci.get("odoo_number", "")).strip()
                dtype = str(ci.get("discrepancy_type", "bank_only")).strip()
                a = round(float(ci.get("amount", 0.0)), 2)
                if num or onum:
                    cloud_keys.add((d, b, num, onum, dtype, a))
                    if num: cloud_keys.add((d, b, num, "", dtype, a))
                    if onum: cloud_keys.add((d, b, "", onum, dtype, a))
                else:
                    cloud_keys.add((d, b, "", "", dtype, a))

            changed = False
            for item in discrepancy_items:
                d = item["date"]
                b = item["bank"].upper()
                num = item["number_bank"]
                onum = item["number_odo"]
                dtype = item["discrepancy_type"]
                a = round(item["amount"], 2)
                is_match = False
                if (d, b, num, onum, dtype, a) in cloud_keys:
                    is_match = True
                elif num and (d, b, num, "", dtype, a) in cloud_keys:
                    is_match = True
                elif onum and (d, b, "", onum, dtype, a) in cloud_keys:
                    is_match = True

                if is_match:
                    item["is_synced"] = True
                    changed = True

            if changed and top.winfo_exists():
                top.after(0, _render_modal_cards)
                top.after(0, _update_counts)
        except Exception:
            pass

    threading.Thread(target=_bg_check_synced, daemon=True).start()
    return top
