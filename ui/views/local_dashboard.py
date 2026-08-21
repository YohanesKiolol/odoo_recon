"""
Local Batch Session Reconciliation Dashboard View.
Renders KPI Summary cards, Accounting Insights cards, and account-level drilldown tables.
"""
import os
import glob
import re
import zipfile
import io
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path

import tkinter as tk
import customtkinter as ctk

from ui.theme import (
    PANEL, PREVIEW_BG, BORDER, BORDER_DARK, ACCENT, SUCCESS, ERROR, WARN, TEXT, MUTED, WHITE,
    FONT_FAMILY, FONT_MONO, IS_WINDOWS
)




class LocalDashboardView(ctk.CTkFrame):
    def __init__(self, master, click_handlers=None, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.clicks = click_handlers or {}
        self._drill_thread = None
        self._drill_seq = 0
        self._drill_cache = None
        self._build_ui()

    def _build_ui(self):
        # 1. Main Input KPI Grid (4 Cards)
        kpi_grid = ctk.CTkFrame(self, fg_color="transparent")
        kpi_grid.pack(fill="x", padx=12, pady=(8, 8))
        kpi_grid.rowconfigure(0, weight=1)
        for c in range(4):
            kpi_grid.columnconfigure(c, weight=1, uniform="kpi_col")

        def _make_kpi_card(parent, col, icon, title, click_key=None):
            card = ctk.CTkFrame(parent, fg_color=PREVIEW_BG, corner_radius=8, border_color=BORDER_DARK, border_width=1, height=96)
            card.grid(row=0, column=col, padx=4, pady=2, sticky="nsew")
            card.pack_propagate(False)

            top_f = ctk.CTkFrame(card, fg_color="transparent")
            top_f.pack(fill="x", padx=10, pady=(8, 2))

            ico_font = ("Segoe UI Emoji", 13) if IS_WINDOWS else (FONT_FAMILY, 13)
            ico_lbl = ctk.CTkLabel(top_f, text=icon, font=ico_font, text_color=ACCENT, fg_color="transparent")
            ico_lbl.pack(side="left")

            t_lbl = ctk.CTkLabel(top_f, text=title, font=(FONT_FAMILY, 10, "bold"), text_color=MUTED, fg_color="transparent")
            t_lbl.pack(side="left", padx=(4, 0))

            val_lbl = ctk.CTkLabel(card, text="0 Files", font=(FONT_FAMILY, 13, "bold"), text_color=TEXT, fg_color="transparent", anchor="w")
            val_lbl.pack(fill="x", padx=10, pady=(2, 1))

            sub_lbl = ctk.CTkLabel(card, text="—", font=(FONT_FAMILY, 10, "bold"), text_color=MUTED, fg_color="transparent", anchor="w")
            sub_lbl.pack(fill="x", padx=10, pady=(1, 6))

            cmd = self.clicks.get(click_key)
            if cmd:
                for w in (card, top_f, ico_lbl, t_lbl, val_lbl, sub_lbl):
                    try:
                        w.bind("<Button-1>", lambda e, c=cmd: c())
                        w.configure(cursor="hand2")
                    except Exception:
                        pass

            return val_lbl, sub_lbl

        self.kpi_bank_val, self.kpi_bank_sub = _make_kpi_card(kpi_grid, 0, "🏦", "MERCHANT REPORT", "open_input")
        self.kpi_odoo_val, self.kpi_odoo_sub = _make_kpi_card(kpi_grid, 1, "💳", "ODOO DATA", "open_odoo_file")
        self.kpi_mut_val,  self.kpi_mut_sub  = _make_kpi_card(kpi_grid, 2, "📊", "MUTATIONS & FEES", "open_mutation")
        self.kpi_eng_val,  self.kpi_eng_sub  = _make_kpi_card(kpi_grid, 3, "⚡", "ENGINE STATUS", "open_output")

        # 2. Accounting Insights Grid (4 Cards)
        ctk.CTkFrame(self, height=1, fg_color=BORDER).pack(fill="x", padx=16, pady=(4, 6))

        ins_hdr = ctk.CTkFrame(self, fg_color="transparent")
        ins_hdr.pack(fill="x", padx=16, pady=(0, 4))
        ctk.CTkLabel(
            ins_hdr, text="🎯 Reconciliation & Accounting Insights",
            font=(FONT_FAMILY, 12, "bold"), text_color=TEXT, fg_color="transparent"
        ).pack(side="left")

        ins_grid = ctk.CTkFrame(self, fg_color="transparent")
        ins_grid.pack(fill="x", padx=12, pady=(0, 10))
        ins_grid.rowconfigure(0, weight=1)
        for c in range(4):
            ins_grid.columnconfigure(c, weight=1, uniform="kpi_col")

        def _make_insight_card(parent, col, icon, title, click_key=None):
            card = ctk.CTkFrame(parent, fg_color=PREVIEW_BG, corner_radius=8, border_color=BORDER_DARK, border_width=1, height=96)
            card.grid(row=0, column=col, padx=4, pady=2, sticky="nsew")
            card.pack_propagate(False)

            top_f = ctk.CTkFrame(card, fg_color="transparent")
            top_f.pack(fill="x", padx=10, pady=(8, 2))

            ico_font = ("Segoe UI Emoji", 13) if IS_WINDOWS else (FONT_FAMILY, 13)
            ico_lbl = ctk.CTkLabel(top_f, text=icon, font=ico_font, text_color=ACCENT, fg_color="transparent")
            ico_lbl.pack(side="left")

            t_lbl = ctk.CTkLabel(top_f, text=title, font=(FONT_FAMILY, 10, "bold"), text_color=MUTED, fg_color="transparent")
            t_lbl.pack(side="left", padx=(4, 0))

            val_lbl = ctk.CTkLabel(card, text="Checking...", font=(FONT_FAMILY, 13, "bold"), text_color=TEXT, fg_color="transparent", anchor="w")
            val_lbl.pack(fill="x", padx=10, pady=(2, 1))

            sub_lbl = ctk.CTkLabel(card, text="—", font=(FONT_FAMILY, 10, "bold"), text_color=MUTED, fg_color="transparent", anchor="w")
            sub_lbl.pack(fill="x", padx=10, pady=(1, 6))

            cmd = self.clicks.get(click_key)
            if cmd:
                for w in (card, top_f, ico_lbl, t_lbl, val_lbl, sub_lbl):
                    try:
                        w.bind("<Button-1>", lambda e, c=cmd: c())
                        w.configure(cursor="hand2")
                    except Exception:
                        pass

            return val_lbl, sub_lbl

        self.ins_health_val, self.ins_health_sub = _make_insight_card(ins_grid, 0, "🎯", "RECON MATCH HEALTH", "open_output")
        self.ins_jrn_val,    self.ins_jrn_sub    = _make_insight_card(ins_grid, 1, "📑", "SETTLEMENT JOURNALS", "on_journal")
        self.ins_sales_val,  self.ins_sales_sub  = _make_insight_card(ins_grid, 2, "🏪", "SALES PORTAL TICKETS", "on_sync_cloud")
        self.ins_cov_val,    self.ins_cov_sub    = _make_insight_card(ins_grid, 3, "💰", "RECONCILED AMOUNT", "show_breakdown")

        # 3. Account-level Drilldown Table
        ctk.CTkFrame(self, height=1, fg_color=BORDER).pack(fill="x", padx=16, pady=(4, 6))

        hdr_row = ctk.CTkFrame(self, fg_color="transparent")
        hdr_row.pack(fill="x", padx=16, pady=(0, 4))
        ctk.CTkLabel(
            hdr_row, text="📅 Date Coverage by Account & Mutation",
            font=(FONT_FAMILY, 11, "bold"), text_color=TEXT, fg_color="transparent"
        ).pack(side="left")

        self.drill_grid = ctk.CTkFrame(self, fg_color=PANEL, corner_radius=6, border_color=BORDER_DARK, border_width=1)
        self.drill_grid.pack(fill="x", padx=16, pady=(0, 10))


    def update_summary(self, input_dir: Path, mutation_dir: Path, output_dir: Path, is_running: bool = False, skip_drill: bool = False):
        """Update KPI metrics and trigger background insight extractions."""
        def _set_lbl(lbl, text, color):
            try:
                lbl.configure(text=text, text_color=color)
            except Exception:
                try:
                    lbl.config(text=text, fg=color)
                except Exception:
                    pass

        try:
            # 1. Merchant files count
            def _bfiles(bdir):
                return [f for f in bdir.rglob("*") if f.is_file() and not f.name.startswith(".") and not f.name.startswith("~$")] if bdir.exists() else []

            c_bca = _bfiles(input_dir / "bca")
            c_man = _bfiles(input_dir / "mandiri")
            c_bri = _bfiles(input_dir / "bri")
            tot_b = len(c_bca) + len(c_man) + len(c_bri)
            _set_lbl(self.kpi_bank_val, f"{tot_b} File{'s' if tot_b!=1 else ''}", SUCCESS if tot_b > 0 else TEXT)
            p = []
            if c_bca: p.append(f"BCA: {len(c_bca)}")
            if c_man: p.append(f"Mandiri: {len(c_man)}")
            if c_bri: p.append(f"BRI: {len(c_bri)}")
            _set_lbl(self.kpi_bank_sub, " | ".join(p) if p else "No merchant files", MUTED)

            # 2. Odoo files
            from config import ODO_EXCEL_PATH, ODO_JOURNAL_EXCEL_PATH
            has_p = ODO_EXCEL_PATH.exists()
            has_j = ODO_JOURNAL_EXCEL_PATH.exists()
            if has_p and has_j:
                _set_lbl(self.kpi_odoo_val, "2/2 Ready", SUCCESS)
                _set_lbl(self.kpi_odoo_sub, "Payments & Journals loaded", MUTED)
            elif has_p or has_j:
                _set_lbl(self.kpi_odoo_val, "1/2 Ready", WARN)
                _set_lbl(self.kpi_odoo_sub, "Payments ready" if has_p else "Journals ready", MUTED)
            else:
                _set_lbl(self.kpi_odoo_val, "Not Loaded", MUTED)
                _set_lbl(self.kpi_odoo_sub, "Auto-downloads via XML-RPC", MUTED)

            # 3. Mutations
            mut_cnt = sum(len(list((mutation_dir / b).rglob("*.csv"))) for b in ["bca", "mandiri", "bri"] if (mutation_dir / b).exists())
            _set_lbl(self.kpi_mut_val, f"{mut_cnt} CSV Files" if mut_cnt > 0 else "None Loaded", SUCCESS if mut_cnt > 0 else MUTED)
            _set_lbl(self.kpi_mut_sub, "Bank Statement Mutations" if mut_cnt > 0 else "No mutation CSVs", MUTED)

            # 4. Engine Status
            if is_running:
                _set_lbl(self.kpi_eng_val, "Running...", ACCENT)
                _set_lbl(self.kpi_eng_sub, "Reconciling data...", MUTED)
            else:
                _set_lbl(self.kpi_eng_val, "Ready", SUCCESS)
                out_files = glob.glob(str(output_dir / "[Rr]econciliation_*.xlsx"))
                if out_files:
                    latest = max(out_files, key=os.path.getmtime)
                    mtime = datetime.fromtimestamp(os.path.getmtime(latest)).strftime("%d/%m %H:%M")
                    _set_lbl(self.kpi_eng_sub, f"Last: {mtime}", MUTED)
                else:
                    _set_lbl(self.kpi_eng_sub, "Engine idle", MUTED)

            # 5. Async Insights Extraction
            def _fetch_insights():
                out_files = glob.glob(str(output_dir / "[Rr]econciliation_*.xlsx"))
                if not out_files:
                    def _idle():
                        _set_lbl(self.ins_health_val, "No report yet", MUTED)
                        _set_lbl(self.ins_health_sub, "Run reconciliation first", MUTED)
                        _set_lbl(self.ins_jrn_val, "—", MUTED)
                        _set_lbl(self.ins_jrn_sub, "No settlement drafts", MUTED)
                        _set_lbl(self.ins_sales_val, "—", MUTED)
                        _set_lbl(self.ins_sales_sub, "No discrepancies", MUTED)
                        _set_lbl(self.ins_cov_val, "—", MUTED)
                        _set_lbl(self.ins_cov_sub, "No audited volumes", MUTED)
                    self.after(0, _idle)
                    return

                try:
                    from openpyxl import load_workbook
                    latest = max(out_files, key=os.path.getmtime)
                    wb = load_workbook(latest, read_only=True, data_only=True)
                    tot_reconciled = 0.0
                    tot_diff = 0.0
                    unmatched_cnt = 0
                    jrn_ready = 0
                    jrn_posted = 0
                    bank_stats = {}

                    if "Daily Summary" in wb.sheetnames:
                        ws_sum = wb["Daily Summary"]
                        for r in ws_sum.iter_rows(min_row=4, values_only=True):
                            if not r or not r[0]: continue
                            d_str = str(r[1] or "").strip()
                            b_n = str(r[3] or "Other").strip()
                            tb = float(r[5] or 0.0)
                            to = float(r[6] or 0.0)
                            df = abs(float(r[7] or 0.0))
                            
                            st = bank_stats.setdefault(b_n, {"dates": set(), "tot_bank": 0.0, "tot_odoo": 0.0, "tot_diff": 0.0})
                            if d_str: st["dates"].add(d_str)
                            st["tot_bank"] += tb
                            st["tot_odoo"] += to
                            st["tot_diff"] += df

                            tot_reconciled += to
                            if df > 1.0: tot_diff += df
                            if len(r) > 10:
                                js = str(r[10] or "")
                                if "Draft" in js: jrn_ready += 1
                                if "Posted" in js: jrn_posted += 1

                    self._latest_bank_stats = bank_stats
                    if hasattr(self.master, "_latest_bank_stats"):
                        self.master._latest_bank_stats = bank_stats

                    if "Differences" in wb.sheetnames:
                        ws_diff = wb["Differences"]
                        for r in ws_diff.iter_rows(min_row=4, values_only=True):
                            if not r or len(r) < 12: continue
                            st = str(r[11] or "")
                            if "Only in" in st: unmatched_cnt += 1

                    wb.close()

                    def _apply_insights():
                        rate = "100% Balanced" if tot_diff < 1.0 and tot_reconciled > 0 else (f"Rp {tot_diff:,.0f} Diff" if tot_diff > 0 else "Ready")
                        _set_lbl(self.ins_health_val, rate, SUCCESS if tot_diff < 1.0 else ERROR)
                        _set_lbl(self.ins_health_sub, f"Audited vs Odoo", MUTED)
                        _set_lbl(self.ins_jrn_val, f"{jrn_ready} Ready / {jrn_posted} Posted", ACCENT if jrn_ready > 0 else SUCCESS)
                        _set_lbl(self.ins_jrn_sub, "Settlement Journals", MUTED)
                        _set_lbl(self.ins_sales_val, f"{unmatched_cnt} Items", WARN if unmatched_cnt > 0 else SUCCESS)
                        _set_lbl(self.ins_sales_sub, "Discrepancy Tickets", MUTED)
                        _set_lbl(self.ins_cov_val, f"Rp {tot_reconciled:,.0f}".replace(",", "."), SUCCESS if tot_reconciled > 0 else TEXT)
                        _set_lbl(self.ins_cov_sub, "Reconciled Volume", MUTED)

                    self.after(0, _apply_insights)
                except Exception:
                    pass


            threading.Thread(target=_fetch_insights, daemon=True).start()

            # 6. Date Drilldown
            if not skip_drill:
                self._start_drill_update(input_dir, mutation_dir)

        except Exception as e:
            print(f"[LocalDashboard] Summary update error: {e}")

    def _start_drill_update(self, input_dir: Path, mutation_dir: Path):
        """Update account-level date drilldown table in background thread."""
        self._drill_seq += 1
        seq = self._drill_seq

        def _bg():
            rows = self._compute_drill_rows(input_dir, mutation_dir, seq)
            if seq == self._drill_seq:
                self.after(0, lambda: self._render_drill_grid(rows))

        t = threading.Thread(target=_bg, daemon=True)
        t.start()

    def _compute_drill_rows(self, input_dir: Path, mutation_dir: Path, seq: int):
        def _parse_d(s):
            s = str(s).strip()[:10]
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y", "%Y%m%d"):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    pass
            return None

        def _fmt_span(dates_set):
            clean = sorted(dates_set)
            if not clean: return "—"
            if len(clean) == 1: return clean[0].strftime("%d/%m/%y")
            is_contiguous = all((clean[i+1]-clean[i]).days == 1 for i in range(len(clean)-1))
            if is_contiguous:
                return f"{clean[0].strftime('%d/%m/%y')} – {clean[-1].strftime('%d/%m/%y')} ({len(clean)} days)"
            if len(clean) <= 4:
                return " | ".join(d.strftime("%d/%m/%y") for d in clean)
            return f"{clean[0].strftime('%d/%m/%y')} … {clean[-1].strftime('%d/%m/%y')} ({len(clean)} dates)"

        def _get_bca_dates(folder: Path) -> set:
            dates = set()
            import msoffcrypto, openpyxl
            from config import BCA_EXCEL_PASSWORD
            for f in folder.glob("*.xlsx"):
                if f.name.startswith("~$") or f.name.startswith("."): continue
                try:
                    try:
                        buf = io.BytesIO()
                        with open(f, "rb") as fp:
                            office = msoffcrypto.OfficeFile(fp)
                            office.load_key(password=BCA_EXCEL_PASSWORD)
                            office.decrypt(buf)
                        buf.seek(0)
                        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
                    except Exception:
                        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                    ws = wb.active
                    if ws is not None:
                        headers = [str(c or "").strip().lower() for c in next(ws.iter_rows(min_row=5, max_row=5, values_only=True))]
                        dt_idx = next((i for i, h in enumerate(headers) if "transaction date" in h or "date" in h), None)
                        if dt_idx is not None:
                            for r in ws.iter_rows(min_row=6, values_only=True):
                                if dt_idx < len(r) and r[dt_idx]:
                                    d = _parse_d(r[dt_idx])
                                    if d: dates.add(d)
                    wb.close()
                except Exception:
                    pass
            return dates

        def _get_mandiri_dates(folder: Path) -> set:
            dates = set()
            from config import MANDIRI_ZIP_PASSWORD, MANDIRI_AMOUNT_COLUMN, MANDIRI_NUMBER_COLUMN
            from readers.mandiri_reader import _read_csv_from_bytes
            for f in folder.glob("*"):
                if f.name.startswith(".") or f.name.startswith("~$"): continue
                if f.suffix.lower() == ".csv":
                    try:
                        txns = _read_csv_from_bytes(f.read_bytes(), MANDIRI_AMOUNT_COLUMN, MANDIRI_NUMBER_COLUMN)
                        for t in txns:
                            d = _parse_d(t.get("date"))
                            if d: dates.add(d)
                    except Exception:
                        pass
                elif f.suffix.lower() == ".zip":
                    try:
                        import pyzipper
                        pwd = MANDIRI_ZIP_PASSWORD.encode("utf-8") if MANDIRI_ZIP_PASSWORD else b""
                        with pyzipper.AESZipFile(f, "r") as zf:
                            if pwd: zf.setpassword(pwd)
                            for name in zf.namelist():
                                txns = _read_csv_from_bytes(zf.read(name), MANDIRI_AMOUNT_COLUMN, MANDIRI_NUMBER_COLUMN)
                                for t in txns:
                                    d = _parse_d(t.get("date"))
                                    if d: dates.add(d)
                    except Exception:
                        pass
            return dates

        def _get_bri_dates(folder: Path) -> set:
            dates = set()
            from config import BRI_PDF_PATTERN, BRI_AMOUNT_COLUMN, BRI_NUMBER_COLUMN
            from readers.bri_reader import _extract_detail_pdf, _parse_pdf_table
            for f in folder.glob("*.zip"):
                if f.name.startswith(".") or f.name.startswith("~$"): continue
                try:
                    pdf_bytes = _extract_detail_pdf(f, BRI_PDF_PATTERN)
                    txns = _parse_pdf_table(pdf_bytes, BRI_AMOUNT_COLUMN, BRI_NUMBER_COLUMN)
                    for t in txns:
                        d = _parse_d(t.get("date"))
                        if d: dates.add(d)
                except Exception:
                    pass
            return dates

        def _get_mut_dates(folder: Path, bank_key: str, alias_key: str) -> set:
            dates = set()
            if not folder.exists(): return dates
            import readers.mutation_reader as mr
            reader_fn = getattr(mr, f"read_mutation_{bank_key.lower()}", None)
            for csv_f in folder.glob("*.csv"):
                if csv_f.name.startswith("."): continue
                try:
                    if reader_fn:
                        rows2, unks2 = reader_fn(csv_f, alias_key)
                        for r in (rows2 + unks2):
                            d = _parse_d(r.get("date"))
                            if d: dates.add(d)
                    else:
                        with open(csv_f, "r", encoding="utf-8-sig", errors="ignore") as fp:
                            for line in fp:
                                for m in re.finditer(r'(\d{2}[-/]\d{2}[-/]\d{4}|\d{4}[-/]\d{2}[-/]\d{2})', line):
                                    d = _parse_d(m.group(1))
                                    if d: dates.add(d)
                except Exception:
                    pass
            return dates

        rows = []
        try:
            import config
            bank_accounts = getattr(config, "BANK_ACCOUNTS", {})

            for b_key, aliases in bank_accounts.items():
                if not aliases: continue
                bank_name = b_key.upper()

                for alias_k, acc_data in aliases.items():
                    alias_label = acc_data.get("alias") or (alias_k.capitalize() if alias_k != "main" else bank_name)
                    acc_title = f"{bank_name} ({alias_label})" if alias_label.lower() != bank_name.lower() else bank_name

                    # 1. Statement dates (isolated to specific alias folder)
                    stmt_dates = set()
                    alias_dir = input_dir / b_key / alias_k
                    if alias_dir.exists():
                        if b_key.lower() == "bca": stmt_dates.update(_get_bca_dates(alias_dir))
                        elif b_key.lower() == "mandiri": stmt_dates.update(_get_mandiri_dates(alias_dir))
                        elif b_key.lower() == "bri": stmt_dates.update(_get_bri_dates(alias_dir))
                    elif alias_k == "main" and (input_dir / b_key).exists():
                        if b_key.lower() == "bca": stmt_dates.update(_get_bca_dates(input_dir / b_key))
                        elif b_key.lower() == "mandiri": stmt_dates.update(_get_mandiri_dates(input_dir / b_key))
                        elif b_key.lower() == "bri": stmt_dates.update(_get_bri_dates(input_dir / b_key))

                    # 2. Mutation dates (isolated to specific alias folder)
                    mut_dates = set()
                    mut_dir = mutation_dir / b_key / alias_k
                    if mut_dir.exists():
                        mut_dates.update(_get_mut_dates(mut_dir, b_key, alias_k))
                    elif alias_k == "main" and (mutation_dir / b_key).exists():
                        mut_dates.update(_get_mut_dates(mutation_dir / b_key, b_key, alias_k))

                    rows.append((acc_title, _fmt_span(stmt_dates), _fmt_span(mut_dates)))


        except Exception as e:
            print(f"[LocalDashboard] Drill rows error: {e}")

        return rows

    def _render_drill_grid(self, rows):
        for child in self.drill_grid.winfo_children():
            child.destroy()

        self.drill_grid.columnconfigure(0, weight=2)
        self.drill_grid.columnconfigure(1, weight=3)
        self.drill_grid.columnconfigure(2, weight=3)

        # Header
        tk.Label(self.drill_grid, text="Bank Account", bg=PREVIEW_BG, fg=MUTED, font=(FONT_FAMILY, 9, "bold"), anchor="w").grid(row=0, column=0, padx=8, pady=4, sticky="ew")
        tk.Label(self.drill_grid, text="Merchant Date Range", bg=PREVIEW_BG, fg=MUTED, font=(FONT_FAMILY, 9, "bold"), anchor="w").grid(row=0, column=1, padx=8, pady=4, sticky="ew")
        tk.Label(self.drill_grid, text="Mutation Date Range", bg=PREVIEW_BG, fg=MUTED, font=(FONT_FAMILY, 9, "bold"), anchor="w").grid(row=0, column=2, padx=8, pady=4, sticky="ew")

        for idx, (acc, d_span, mut_st) in enumerate(rows, start=1):
            bg_col = WHITE if idx % 2 == 1 else PREVIEW_BG
            tk.Label(self.drill_grid, text=acc, bg=bg_col, fg=TEXT, font=(FONT_FAMILY, 9, "bold"), anchor="w").grid(row=idx, column=0, padx=8, pady=3, sticky="ew")
            tk.Label(self.drill_grid, text=d_span, bg=bg_col, fg=SUCCESS if d_span != "—" else MUTED, font=(FONT_MONO, 9), anchor="w").grid(row=idx, column=1, padx=8, pady=3, sticky="ew")
            tk.Label(self.drill_grid, text=mut_st, bg=bg_col, fg=SUCCESS if mut_st != "—" else MUTED, font=(FONT_MONO, 9), anchor="w").grid(row=idx, column=2, padx=8, pady=3, sticky="ew")
